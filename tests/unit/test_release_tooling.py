# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import base64
import csv
import hashlib
import io
import subprocess
import sys
import tarfile
import zipfile
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))
import verify_release as release  # noqa: E402


def _source_tree(path: Path, version: str = "0.1.0", *, gui: bool = True) -> Path:
    (path / "src" / "ordifile").mkdir(parents=True)
    gui_contract = (
        '\nordifile-gui = "ordifile.desktop.app:main"\n'
        '\n[project.optional-dependencies]\ngui = ["PySide6-Essentials==6.11.2"]\n'
        if gui
        else "\n"
    )
    (path / "pyproject.toml").write_text(
        f'[project]\nname = "ordifile"\nversion = "{version}"\n'
        f'\n[project.scripts]\nordifile = "ordifile.cli.main:main"{gui_contract}',
        encoding="utf-8",
    )
    (path / "src" / "ordifile" / "_version.py").write_text(
        f'__version__ = "{version}"\n', encoding="utf-8"
    )
    for name in release.LICENSE_FILES:
        (path / name).write_text(f"{name} fixture\n", encoding="utf-8")
    for name in release.SDIST_MAINTAINER_FILES:
        target = path / name
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(f"{name} fixture\n", encoding="utf-8")
    return path


def _metadata(version: str, *, gui: bool = True) -> bytes:
    gui_metadata = (
        "Provides-Extra: gui\nRequires-Dist: pyside6-essentials==6.11.2; extra == 'gui'\n"
        if gui
        else ""
    )
    return (
        "Metadata-Version: 2.4\n"
        "Name: ordifile\n"
        f"Version: {version}\n"
        "License-Expression: Apache-2.0\n"
        "License-File: LICENSE\n"
        "License-File: NOTICE\n"
        "License-File: THIRD_PARTY_NOTICES.md\n"
        f"{gui_metadata}\n"
    ).encode()


def _record_digest(content: bytes) -> str:
    digest = base64.urlsafe_b64encode(hashlib.sha256(content).digest()).rstrip(b"=")
    return digest.decode("ascii")


def _write_wheel(
    path: Path,
    source: Path,
    *,
    extra: dict[str, bytes] | None = None,
    gui: bool = True,
) -> None:
    prefix = "ordifile-0.1.0.dist-info/"
    contents = {
        "ordifile/__init__.py": b'__version__ = "0.1.0"\n',
        f"{prefix}METADATA": _metadata("0.1.0", gui=gui),
        f"{prefix}WHEEL": b"Wheel-Version: 1.0\nTag: py3-none-any\n",
        f"{prefix}entry_points.txt": (
            b"[console_scripts]\n"
            b"ordifile = ordifile.cli.main:main\n"
            + (b"ordifile-gui = ordifile.desktop.app:main\n" if gui else b"")
        ),
    }
    for name in release.LICENSE_FILES:
        contents[f"{prefix}licenses/{name}"] = (source / name).read_bytes()
    contents.update(extra or {})
    record_name = f"{prefix}RECORD"
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    for name, content in contents.items():
        writer.writerow((name, f"sha256={_record_digest(content)}", str(len(content))))
    writer.writerow((record_name, "", ""))
    contents[record_name] = output.getvalue().encode()
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_STORED) as archive:
        for name, content in contents.items():
            info = zipfile.ZipInfo(name)
            info.orig_filename = name
            info.filename = name
            archive.writestr(info, content)


def _write_sdist(path: Path, source: Path, *, legacy: bool = False, gui: bool = True) -> None:
    root = "ordifile-0.1.0"
    contents = {
        f"{root}/PKG-INFO": _metadata("0.1.0", gui=gui),
        f"{root}/src/ordifile/__init__.py": b'__version__ = "0.1.0"\n',
    }
    for name in release.LICENSE_FILES:
        contents[f"{root}/{name}"] = (source / name).read_bytes()
    for name in release.SDIST_MAINTAINER_FILES:
        contents[f"{root}/{name}"] = (source / name).read_bytes()
    if legacy:
        contents[f"{root}/src/labconvert/__init__.py"] = b""
    with tarfile.open(path, "w:gz") as archive:
        for name, content in contents.items():
            info = tarfile.TarInfo(name)
            info.size = len(content)
            info.mode = 0o644
            archive.addfile(info, io.BytesIO(content))


def _write_sdist_with_aliases(
    path: Path,
    source: Path,
    first_name: str,
    second_name: str,
) -> None:
    complete = path.with_name("complete.tar.gz")
    _write_sdist(complete, source)
    with tarfile.open(complete, "r:gz") as original, tarfile.open(path, "w:gz") as output:
        for member in original.getmembers():
            extracted = original.extractfile(member) if member.isfile() else None
            output.addfile(member, extracted)
        for name, content in ((first_name, b"first"), (second_name, b"second")):
            info = tarfile.TarInfo(name)
            info.size = len(content)
            info.mode = 0o644
            output.addfile(info, io.BytesIO(content))


def _artifacts(source: Path, dist: Path) -> tuple[Path, Path]:
    dist.mkdir()
    wheel = dist / "ordifile-0.1.0-py3-none-any.whl"
    sdist = dist / "ordifile-0.1.0.tar.gz"
    _write_wheel(wheel, source)
    _write_sdist(sdist, source)
    return wheel, sdist


def _project_wheel(path: Path, project_root: Path) -> None:
    prefix = "ordifile-0.1.0.dist-info/"
    contents: dict[str, bytes] = {}
    for source_path in sorted((project_root / "src" / "ordifile").rglob("*.py")):
        relative = source_path.relative_to(project_root / "src").as_posix()
        contents[relative] = source_path.read_bytes()
    contents[f"{prefix}METADATA"] = _metadata("0.1.0")
    contents[f"{prefix}entry_points.txt"] = (
        b"[console_scripts]\n"
        b"ordifile = ordifile.cli.main:main\n"
        b"ordifile-gui = ordifile.desktop.app:main\n"
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_STORED) as archive:
        for name, content in contents.items():
            archive.writestr(name, content)


@pytest.mark.parametrize(
    "value",
    ("1.2", "01.2.3", "1.02.3", "1.2.03", "v1.2.3", "1.2.3-01", "1.2.3+"),
)
def test_strict_semver_rejects_noncanonical_values(value: str) -> None:
    with pytest.raises(release.ReleaseVerificationError, match="strict SemVer"):
        release.require_semver(value, field="test")


def test_source_version_requires_exact_tag_project_and_module_match(tmp_path: Path) -> None:
    source = _source_tree(tmp_path / "source")
    release.verify_source_version(source, "0.1.0", "v0.1.0")

    with pytest.raises(release.ReleaseVerificationError, match="tag must be exactly"):
        release.verify_source_version(source, "0.1.0", "0.1.0")
    with pytest.raises(release.ReleaseVerificationError, match="must be identical"):
        release.verify_source_version(source, "0.1.1")


def test_release_artifacts_and_checksums_are_verified_deterministically(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("GITHUB_REF_TYPE", "tag")
    monkeypatch.setenv("GITHUB_REF_NAME", "v9.9.9")
    source = _source_tree(tmp_path / "source")
    wheel, sdist = _artifacts(source, tmp_path / "dist")

    verified = release.verify_release(
        source_root=source,
        dist_dir=wheel.parent,
        expected_version="0.1.0",
        tag="v0.1.0",
        checksums=wheel.parent / "SHA256SUMS.txt",
        smoke=False,
    )
    first = verified.checksums.read_bytes()
    release.write_sha256sums((wheel, sdist), verified.checksums)

    assert verified.checksums.read_bytes() == first
    assert first.decode().splitlines() == [
        f"{hashlib.sha256(wheel.read_bytes()).hexdigest()}  {wheel.name}",
        f"{hashlib.sha256(sdist.read_bytes()).hexdigest()}  {sdist.name}",
    ]

    assert (
        release.main(
            [
                "--dist-dir",
                str(wheel.parent),
                "--source-root",
                str(source),
                "--expected-version",
                "0.1.0",
                "--tag",
                "v0.1.0",
                "--checksums",
                str(wheel.parent / "SHA256SUMS.txt"),
                "--skip-smoke",
            ]
        )
        == 0
    )


def test_wheel_rejects_legacy_package_and_record_tampering(tmp_path: Path) -> None:
    source = _source_tree(tmp_path / "source")
    legacy = tmp_path / "legacy.whl"
    _write_wheel(legacy, source, extra={"labconvert/__init__.py": b""})
    with pytest.raises(release.ReleaseVerificationError, match="legacy package"):
        release.verify_wheel(legacy, "0.1.0", source)

    valid = tmp_path / "valid.whl"
    tampered = tmp_path / "record.whl"
    _write_wheel(valid, source)
    with zipfile.ZipFile(valid) as archive:
        contents = {name: archive.read(name) for name in archive.namelist()}
    contents["ordifile/__init__.py"] = b'__version__ = "9.9.9"\n'
    with zipfile.ZipFile(tampered, "w", compression=zipfile.ZIP_STORED) as archive:
        for name, content in contents.items():
            archive.writestr(name, content)
    with pytest.raises(release.ReleaseVerificationError, match="RECORD mismatch"):
        release.verify_wheel(tampered, "0.1.0", source)


@pytest.mark.parametrize(
    ("first_name", "second_name"),
    (
        ("ordifile/alias.py", "ordifile//alias.py"),
        ("ordifile/alias.py", "ordifile/./alias.py"),
        ("ordifile/Alias.py", "ordifile/alias.py"),
        (
            "ordifile/e\N{COMBINING ACUTE ACCENT}.py",
            "ordifile/\N{LATIN SMALL LETTER E WITH ACUTE}.py",
        ),
        ("ordifile/", "ORDIFILE"),
    ),
    ids=("double-slash", "dot", "casefold", "unicode", "directory-slash"),
)
def test_wheel_rejects_lexical_and_normalized_aliases_with_exact_record(
    tmp_path: Path,
    first_name: str,
    second_name: str,
) -> None:
    source = _source_tree(tmp_path / "source")
    wheel = tmp_path / "aliases.whl"
    _write_wheel(
        wheel,
        source,
        extra={first_name: b"first bytes", second_name: b"different second bytes"},
    )

    with pytest.raises(release.ReleaseVerificationError, match="unsafe|duplicate"):
        release.verify_wheel(wheel, "0.1.0", source)


def test_wheel_rejects_raw_backslash_member_with_exact_record_on_every_platform(
    tmp_path: Path,
) -> None:
    source = _source_tree(tmp_path / "source")
    wheel = tmp_path / "raw-backslash.whl"
    raw_name = "ordifile\\raw_backslash.py"
    _write_wheel(wheel, source, extra={raw_name: b"raw backslash wheel bytes"})

    with zipfile.ZipFile(wheel) as archive:
        assert raw_name in {info.orig_filename for info in archive.infolist()}
        record = archive.read("ordifile-0.1.0.dist-info/RECORD").decode("utf-8")
        assert raw_name in record
    with pytest.raises(release.ReleaseVerificationError, match="unsafe archive member"):
        release.verify_wheel(wheel, "0.1.0", source)


@pytest.mark.parametrize(
    ("first_name", "second_name"),
    (
        (
            "ordifile-0.1.0/src/ordifile/alias.py",
            "ordifile-0.1.0/src/ordifile//alias.py",
        ),
        (
            "ordifile-0.1.0/src/ordifile/alias.py",
            "ordifile-0.1.0/src/ordifile/./alias.py",
        ),
        (
            "ordifile-0.1.0/src/ordifile/Alias.py",
            "ordifile-0.1.0/src/ordifile/alias.py",
        ),
    ),
    ids=("double-slash", "dot", "casefold"),
)
def test_sdist_rejects_lexical_and_casefold_aliases(
    tmp_path: Path,
    first_name: str,
    second_name: str,
) -> None:
    source = _source_tree(tmp_path / "source")
    sdist = tmp_path / "ordifile-0.1.0.tar.gz"
    _write_sdist_with_aliases(sdist, source, first_name, second_name)

    with pytest.raises(release.ReleaseVerificationError, match="unsafe|duplicate"):
        release.verify_sdist(sdist, "0.1.0", source)


@pytest.mark.parametrize(
    "unsafe_name",
    (
        "module.py:payload",
        "CON.py",
        "trailing.",
        "trailing ",
        "question?.py",
        "folder\\backslash.py",
        "control\x01.py",
        "delete\x7f.py",
        "e\N{COMBINING ACUTE ACCENT}.py",
    ),
    ids=(
        "ads",
        "reserved",
        "dot",
        "space",
        "question",
        "backslash",
        "control",
        "delete",
        "non-nfc",
    ),
)
def test_release_archives_reject_nonportable_single_member_names(
    tmp_path: Path,
    unsafe_name: str,
) -> None:
    source = _source_tree(tmp_path / "source")
    wheel = tmp_path / "unsafe.whl"
    _write_wheel(wheel, source, extra={f"ordifile/{unsafe_name}": b"unsafe bytes"})
    with pytest.raises(release.ReleaseVerificationError, match="unsafe"):
        release.verify_wheel(wheel, "0.1.0", source)

    sdist = tmp_path / "ordifile-0.1.0.tar.gz"
    _write_sdist_with_aliases(
        sdist,
        source,
        "ordifile-0.1.0/src/ordifile/safe.py",
        f"ordifile-0.1.0/src/ordifile/{unsafe_name}",
    )
    with pytest.raises(release.ReleaseVerificationError, match="unsafe"):
        release.verify_sdist(sdist, "0.1.0", source)


def test_wheel_rejects_wrong_entry_point_and_license(tmp_path: Path) -> None:
    source = _source_tree(tmp_path / "source")
    prefix = "ordifile-0.1.0.dist-info/"
    wrong_entry = tmp_path / "entry.whl"
    _write_wheel(
        wrong_entry,
        source,
        extra={f"{prefix}entry_points.txt": b"[console_scripts]\nold = old.cli:main\n"},
    )
    with pytest.raises(release.ReleaseVerificationError, match="console script"):
        release.verify_wheel(wrong_entry, "0.1.0", source)

    wrong_license = tmp_path / "license.whl"
    _write_wheel(
        wrong_license,
        source,
        extra={f"{prefix}licenses/NOTICE": b"different\n"},
    )
    with pytest.raises(release.ReleaseVerificationError, match="differs from source"):
        release.verify_wheel(wrong_license, "0.1.0", source)


def test_wheel_rejects_missing_or_unconditional_gui_dependency_metadata(tmp_path: Path) -> None:
    source = _source_tree(tmp_path / "source")
    prefix = "ordifile-0.1.0.dist-info/"
    missing_extra = tmp_path / "missing-extra.whl"
    metadata_without_extra = _metadata("0.1.0").replace(
        b"Provides-Extra: gui\nRequires-Dist: pyside6-essentials==6.11.2; extra == 'gui'\n",
        b"",
    )
    _write_wheel(
        missing_extra,
        source,
        extra={f"{prefix}METADATA": metadata_without_extra},
    )
    with pytest.raises(release.ReleaseVerificationError, match="gui extra"):
        release.verify_wheel(missing_extra, "0.1.0", source)

    unconditional = tmp_path / "unconditional-gui.whl"
    unconditional_metadata = _metadata("0.1.0").replace(
        b"Requires-Dist: pyside6-essentials==6.11.2; extra == 'gui'\n",
        b"Requires-Dist: pyside6-essentials==6.11.2; extra == 'gui'\n"
        b"Requires-Dist: pyside6-essentials==6.11.2\n",
    )
    _write_wheel(
        unconditional,
        source,
        extra={f"{prefix}METADATA": unconditional_metadata},
    )
    with pytest.raises(release.ReleaseVerificationError, match="unconditional"):
        release.verify_wheel(unconditional, "0.1.0", source)


def test_release_verifier_preserves_historical_cli_only_artifact_contract(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = _source_tree(tmp_path / "source", gui=False)
    dist = tmp_path / "dist"
    dist.mkdir()
    wheel = dist / "ordifile-0.1.0-py3-none-any.whl"
    sdist = dist / "ordifile-0.1.0.tar.gz"
    _write_wheel(wheel, source, gui=False)
    _write_sdist(sdist, source, gui=False)
    smoke_calls: list[bool] = []
    monkeypatch.setattr(
        release,
        "run_clean_wheel_smoke",
        lambda _wheel, *, expect_gui=True: smoke_calls.append(expect_gui),
    )

    release.verify_release(
        source_root=source,
        dist_dir=dist,
        expected_version="0.1.0",
    )

    assert smoke_calls == [False]


@pytest.mark.parametrize(
    "name",
    (
        "scripts/verify_release.py",
        "scripts/generate_demo_assets.sh",
        "tests/fixtures/synthetic/generate_xlsx.py",
    ),
)
def test_wheel_rejects_maintainer_scripts_and_fixture_generators(tmp_path: Path, name: str) -> None:
    source = _source_tree(tmp_path / "source")
    wheel = tmp_path / "maintainer-content.whl"
    _write_wheel(wheel, source, extra={name: b"must not ship in wheel\n"})

    with pytest.raises(release.ReleaseVerificationError, match="maintainer scripts"):
        release.verify_wheel(wheel, "0.1.0", source)


def test_sdist_rejects_legacy_package(tmp_path: Path) -> None:
    source = _source_tree(tmp_path / "source")
    sdist = tmp_path / "ordifile-0.1.0.tar.gz"
    _write_sdist(sdist, source, legacy=True)

    with pytest.raises(release.ReleaseVerificationError, match="legacy package"):
        release.verify_sdist(sdist, "0.1.0", source)


def test_sdist_requires_release_tools_and_fixture_generator(tmp_path: Path) -> None:
    source = _source_tree(tmp_path / "source")
    complete = tmp_path / "ordifile-0.1.0.tar.gz"
    _write_sdist(complete, source)
    release.verify_sdist(complete, "0.1.0", source)

    missing = tmp_path / "missing.tar.gz"
    with tarfile.open(complete, "r:gz") as original, tarfile.open(missing, "w:gz") as output:
        for member in original.getmembers():
            if member.name.endswith("/scripts/fetch_external_fixture.py"):
                continue
            extracted = original.extractfile(member) if member.isfile() else None
            output.addfile(member, extracted)
    with pytest.raises(release.ReleaseVerificationError, match="missing required maintainer"):
        release.verify_sdist(missing, "0.1.0", source)


def test_clean_wheel_smoke_runs_real_cli_and_reopens_workbook(tmp_path: Path) -> None:
    wheel = tmp_path / "ordifile-0.1.0-py3-none-any.whl"
    _project_wheel(wheel, PROJECT_ROOT)

    release.run_clean_wheel_smoke(wheel)


def test_release_cli_contract_returns_nonzero_for_invalid_version(tmp_path: Path) -> None:
    assert (
        release.main(
            [
                "--dist-dir",
                str(tmp_path),
                "--source-root",
                str(tmp_path),
                "--expected-version",
                "not-semver",
            ]
        )
        == 1
    )


def test_release_workflow_closes_openpyxl_workbook_explicitly() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    wheel_smoke = workflow.split(
        "      - name: Install and exercise the wheel outside the checkout", maxsplit=1
    )[1].split("      - name: Remove isolated job environment", maxsplit=1)[0]

    assert "workbook = load_workbook(" in workflow
    assert "with load_workbook(" not in workflow
    assert "finally:" in workflow
    assert "workbook.close()" in workflow
    assert wheel_smoke.index('"pip", "install"') < wheel_smoke.index(
        "from openpyxl import load_workbook"
    )
    assert 'sysconfig.get_path("scripts")' in wheel_smoke
    assert 'shutil.which("ordifile")' not in wheel_smoke


def test_release_workflow_resolves_every_installed_cli_without_path() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    smoke_blocks = {
        "wheel": workflow.split(
            "      - name: Install and exercise the wheel outside the checkout", maxsplit=1
        )[1].split("      - name: Remove isolated job environment", maxsplit=1)[0],
        "testpypi": workflow.split(
            "      - name: Download both files directly from TestPyPI and compare bytes",
            maxsplit=1,
        )[1].split("      - name: Remove isolated job environment", maxsplit=1)[0],
        "pypi": workflow.split(
            "      - name: Download both files directly from PyPI and compare bytes",
            maxsplit=1,
        )[1].split("      - name: Remove isolated job environment", maxsplit=1)[0],
    }

    for name, block in smoke_blocks.items():
        assert "import sysconfig" in block, name
        assert 'scripts = Path(sysconfig.get_path("scripts"))' in block, name
        assert '("ordifile.exe" if os.name == "nt" else "ordifile")' in block, name
        assert "if not executable.is_file():" in block, name
        assert block.index('"pip", "install"') < block.index(
            'scripts = Path(sysconfig.get_path("scripts"))'
        ), name
        assert 'shutil.which("ordifile")' not in block, name

    assert workflow.count('scripts = Path(sysconfig.get_path("scripts"))') == 3
    assert 'shutil.which("ordifile")' not in workflow


def test_release_workflow_revalidates_draft_before_final_publish() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    final_publish = workflow.split("  publish-github-release:", maxsplit=1)[1]

    assert "actions/download-artifact@" in final_publish
    assert "if len(expected) != 3:" in final_publish
    assert '"isDraft,isPrerelease,tagName,name,body,assets"' in final_publish
    assert 'release.get("isPrerelease") is not False' in final_publish
    assert 'release.get("name") != f"Ordifile {tag}"' in final_publish
    assert 'release.get("body")' in final_publish
    assert '(root / "release-notes.md").read_text(encoding="utf-8")' in final_publish
    assert 'f"repos/{repository}/git/ref/tags/{tag}"' in final_publish
    assert "Draft release bytes changed" in final_publish
    assert 'gh release edit "$GITHUB_REF_NAME" --draft=false' in final_publish


def test_release_notes_identify_their_filename_version() -> None:
    release_notes = PROJECT_ROOT / "docs" / "releases"
    for path in sorted(release_notes.glob("v*.md")):
        expected_heading = f"# Ordifile {path.stem}"
        first_line = path.read_text(encoding="utf-8").splitlines()[0]
        assert first_line == expected_heading or first_line.startswith(expected_heading + " ")


def test_release_workflow_validates_release_note_heading() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    build_gate = workflow.split(
        "      - name: Enforce release source, version, and tag policy", maxsplit=1
    )[1].split("      - name: Install release tooling and project", maxsplit=1)[0]

    assert 'notes.read_text(encoding="utf-8").splitlines()[0]' in build_gate
    assert 'expected_heading = f"# Ordifile v{version}"' in build_gate
    assert "not first_line.startswith(" in build_gate
    assert 'expected_heading + " "' in build_gate
    assert "Release notes are unreadable or empty" in build_gate


def test_release_workflow_rejects_stale_or_extra_publish_files() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    assert "Release artifact staging already exists; refusing stale payload files." in workflow
    test_publish = workflow.split("  publish-testpypi:", maxsplit=1)[1].split(
        "  verify-testpypi:", maxsplit=1
    )[0]
    production_publish = workflow.split("  publish-pypi:", maxsplit=1)[1].split(
        "  verify-pypi:", maxsplit=1
    )[0]
    for publish_job in (test_publish, production_publish):
        assert "actual_files != set(manifest)" in publish_job
        assert "Publish package set must be exactly" in publish_job


def test_release_workflow_rechecks_live_tag_and_uses_minimum_permissions() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    assert "persist-credentials: false" in workflow
    assert workflow.count("The live annotated tag no longer targets this workflow commit") == 2
    assert workflow.count("id-token: write") == 6
    attest_job = workflow.split("  attest:", maxsplit=1)[1].split(
        "\n  create-github-release:", maxsplit=1
    )[0]
    for permission in (
        "contents: read",
        "id-token: write",
        "attestations: write",
        "artifact-metadata: write",
    ):
        assert permission in attest_job
    for job_name in ("publish-testpypi", "publish-pypi"):
        job = workflow.split(f"  {job_name}:", maxsplit=1)[1].split(
            "\n  " + ("verify-testpypi" if job_name == "publish-testpypi" else "verify-pypi") + ":",
            maxsplit=1,
        )[0]
        assert "id-token: write" in job
        assert "contents: read" in job
        assert "attestations: write" not in job
        assert "contents: write" not in job


def test_release_promotion_is_allowlisted_hosted_and_never_rebuilds() -> None:
    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    verifier = (PROJECT_ROOT / "scripts" / "verify_release_promotion.py").read_text(
        encoding="utf-8"
    )
    audited_source = workflow + verifier
    promotion = workflow.split("  promote-validate-existing:", maxsplit=1)[1].split(
        "\n  smoke:", maxsplit=1
    )[0]

    for value in (
        "31980576873",
        "9272226213",
        "584321",
        "sha256:44a4040411ea5870d1dfb78e4a0d0969ccfbde666357441bd03c0ddf34de6216",
        "33e1b6ec4d6d822e1a0b532e0d075adc4d79c788",
        "ordifile-distributions-31980576873",
        "0d485620f46fb86cd37518ee9cd3cb38ecb4e421d2a96fc8666e4399616b4fa8",
        "14f71d8ebd4581c4c3001c724de5bb547b5274f165941af628d6f9b02e85ef39",
    ):
        assert value in audited_source
    assert "artifact-ids: 9272226213" in promotion
    assert "digest-mismatch: error" in promotion
    assert "merge-base --is-ancestor" in promotion
    assert "index-absent --index testpypi" in promotion
    assert "index-absent --index pypi" in promotion
    assert "--skip-smoke" in promotion
    assert "Build wheel and source distribution" not in promotion
    assert "actions/upload-artifact@" not in promotion
    assert "skip-existing" not in workflow
    assert "TWINE_PASSWORD" not in workflow
    assert "password:" not in workflow

    hosted_jobs = promotion.count("runs-on: ubuntu-latest")
    assert hosted_jobs == 8
    assert "runs-on: [self-hosted, dgx-spark]" not in promotion
    assert promotion.count("id-token: write") == 3
    assert promotion.count("environment:\n      name: testpypi") == 1
    assert promotion.count("environment:\n      name: pypi") == 1

    finalization = promotion.split("  promote-publish-github-release:", maxsplit=1)[1]
    assert "inputs.mode == 'finalize-existing'" in finalization
    assert (
        "for subject in release-artifact/packages/* release-artifact/SHA256SUMS.txt" in finalization
    )
    assert 'gh attestation verify "$subject"' in finalization
    assert "--deny-self-hosted-runners" in finalization
    assert "--source-ref refs/heads/main" in finalization
    assert 'release.get("body") != (root / "release-notes.md").read_text' in finalization
    assert "Require the v0.2.1 GitHub Release to be public" in finalization
    assert "pypa/gh-action-pypi-publish@" not in finalization
    assert "actions/attest@" not in finalization


def test_release_workflow_rejects_nested_annotated_tags(tmp_path: Path) -> None:
    repository = tmp_path / "repository"
    repository.mkdir()

    def git(*arguments: str) -> str:
        result = subprocess.run(
            ["git", *arguments],
            cwd=repository,
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip()

    git("init", "-b", "main")
    git("config", "user.name", "Ordifile Test")
    git("config", "user.email", "ordifile-test@users.noreply.github.com")
    (repository / "release.txt").write_text("release\n", encoding="utf-8")
    git("add", "release.txt")
    git("commit", "-m", "test: create release commit")
    git("tag", "-a", "v0.1.0", "-m", "Ordifile v0.1.0")
    git("tag", "-a", "nested", "-m", "nested tag", "v0.1.0")

    def immediate_target_type(reference: str) -> str:
        payload = git("cat-file", "-p", reference)
        return next(
            line.removeprefix("type ") for line in payload.splitlines() if line.startswith("type ")
        )

    direct = immediate_target_type("refs/tags/v0.1.0")
    nested = immediate_target_type("refs/tags/nested")
    assert direct == "commit"
    assert nested == "tag"

    workflow = (PROJECT_ROOT / ".github" / "workflows" / "release.yml").read_text(encoding="utf-8")
    assert 'payload = git("cat-file", "-p", reference)' in workflow
    assert 'target_types != ["commit"]' in workflow
