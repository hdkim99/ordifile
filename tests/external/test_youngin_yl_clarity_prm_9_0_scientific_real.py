from __future__ import annotations

import hashlib
import math
import os
import re
import stat
from decimal import ROUND_HALF_EVEN, Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from statistics import median
from zipfile import ZipFile, ZipInfo

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.models import DatasetBundle, SeriesKind, SignalSeries

EXPECTED_ARCHIVE_SIZE = 2_534_115
EXPECTED_ARCHIVE_SHA256 = "5cfcf67e5ac097a02a80b54b8ca9c509e309de0e3efa6a9894bd4977d3157847"
EXPECTED_MEMBER_COUNT = 20
EXPECTED_EXPANDED_BYTES = 8_848_218
EXPECTED_PRMS = 10
EXPECTED_STREAMS = 20
EXPECTED_POINTS_PER_DETECTOR = 131_760
EXPECTED_SCIENTIFIC_POINTS = 263_520
EXPECTED_DERIVED_PEAKS = 241
EXPECTED_RESULT_ROWS = 263
EXPECTED_FAIL_CLOSED_ROWS = 22
EXPECTED_AREA_EXACT_BY_DECIMAL_PLACES = {2: 241, 3: 230, 4: 241}
EXPECTED_AREA_WITHIN_1_PERCENT = 241
EXPECTED_AREA_WITHIN_5_PERCENT = 241
EXPECTED_AREA_MEDIAN_RELATIVE_ERROR = 4.034914044223078e-06
EXPECTED_AREA_P90_RELATIVE_ERROR = 2.563118239673121e-05
EXPECTED_AREA_MAX_RELATIVE_ERROR = 6.875806406999155e-05
PRINTED_SIGNAL_HALF_UNIT = Decimal("0.00005001")
AREA_QUANTA = {2: Decimal("0.01"), 3: Decimal("0.001"), 4: Decimal("0.0001")}
MAX_MEMBER_BYTES = 1_000_000
MAX_COMPRESSION_RATIO = 4.0


def _archive() -> Path:
    value = os.environ.get("ORDIFILE_YOUNGIN_9_0_CURVE_ARCHIVE")
    if not value:
        raise AssertionError("ORDIFILE_YOUNGIN_9_0_CURVE_ARCHIVE is required")
    return Path(value)


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_member(info: ZipInfo) -> bool:
    pure = PurePosixPath(info.filename)
    mode = (info.external_attr >> 16) & 0xFFFF
    return (
        not info.is_dir()
        and not (info.flag_bits & 0x1)
        and "\\" not in info.filename
        and not pure.is_absolute()
        and all(part not in {"", ".", ".."} for part in pure.parts)
        and (mode == 0 or stat.S_ISREG(mode))
    )


def _curve_blocks(data: bytes) -> tuple[tuple[tuple[str, str], ...], ...]:
    lines = data.decode("cp949", errors="strict").splitlines()
    header = "Time [min]\tVoltage [mV]"
    blocks: list[tuple[tuple[str, str], ...]] = []
    for index, line in enumerate(lines):
        if line != header:
            continue
        rows: list[tuple[str, str]] = []
        for candidate in lines[index + 1 :]:
            fields = candidate.split("\t")
            if len(fields) != 2:
                break
            try:
                Decimal(fields[0])
                Decimal(fields[1])
            except InvalidOperation:
                break
            rows.append((fields[0], fields[1]))
        if not rows:
            raise AssertionError("full-range curve has no numeric rows")
        blocks.append(tuple(rows))
    if len(blocks) != 2:
        raise AssertionError("curve export lacks the exact two full-range mV blocks")
    return tuple(blocks)


def _result_rows(data: bytes) -> tuple[tuple[str, Decimal, str, Decimal], ...]:
    rows: list[tuple[str, Decimal, str, Decimal]] = []
    for line in data.decode("cp949", errors="strict").splitlines():
        fields = line.split("\t")
        if (
            len(fields) < 9
            or not fields[3].strip().isdigit()
            or fields[4].strip() not in {"FID", "TCD"}
            or not fields[5].strip().isdigit()
        ):
            continue
        area_text = fields[7].strip()
        if re.fullmatch(r"[0-9]+\.[0-9]{4}", area_text) is None:
            raise AssertionError("official Result Area is not an exact nonnegative 4dp lexeme")
        try:
            official_area = Decimal(area_text)
            if not official_area.is_finite():
                raise AssertionError("official Result Area is not finite")
            rows.append((fields[4].strip(), Decimal(fields[6]), area_text, official_area))
        except InvalidOperation as error:
            raise AssertionError("composite Result row contains a non-numeric value") from error
    return tuple(rows)


def _derived_result_errors(
    bundle: DatasetBundle,
    rows: tuple[tuple[str, Decimal, str, Decimal], ...],
) -> tuple[int, list[float], dict[int, int], int]:
    by_detector = {
        detector: tuple(peak for peak in bundle.peaks if peak.detector == detector)
        for detector in {row[0] for row in rows}
    }
    positions = {detector: 0 for detector in by_detector}
    matched_rt = 0
    skipped = 0
    errors: list[float] = []
    exact_by_decimal_places = {places: 0 for places in AREA_QUANTA}
    for detector, retention_time, _official_area_text, official_area in rows:
        peaks = by_detector[detector]
        if not peaks:
            # The channel failed closed; its official rows are recorded as not compared.
            skipped += 1
            continue
        position = positions[detector]
        while (
            position < len(peaks)
            and Decimal(format(peaks[position].retention_time or 0.0, ".3f")) != retention_time
        ):
            position += 1
        assert position < len(peaks), "an official Result RT lacks its PRM marker candidate"
        peak = peaks[position]
        positions[detector] = position + 1
        matched_rt += 1
        assert peak.calculated_area is not None
        assert official_area != 0
        for places, quantum in AREA_QUANTA.items():
            rendered = Decimal.from_float(float(peak.calculated_area)).quantize(
                quantum, rounding=ROUND_HALF_EVEN
            )
            official = official_area.quantize(quantum, rounding=ROUND_HALF_EVEN)
            exact_by_decimal_places[places] += rendered == official
        errors.append(
            float(abs(Decimal(str(peak.calculated_area)) - official_area) / abs(official_area))
        )
    assert all(
        positions[detector] == len(peaks) for detector, peaks in by_detector.items() if peaks
    )
    return matched_rt, errors, exact_by_decimal_places, skipped


def _p90(values: list[float]) -> float:
    ordered = sorted(values)
    return ordered[math.floor(0.9 * (len(ordered) - 1))]


def _curve_matches(signal: SignalSeries, rows: tuple[tuple[str, str], ...]) -> bool:
    if len(signal.x_values) != len(rows) or len(signal.y_values) != len(rows):
        return False
    for index, ((time_text, response_text), x_value, y_value) in enumerate(
        zip(rows, signal.x_values, signal.y_values, strict=True)
    ):
        if time_text != format(index / 600, ".5f") or time_text != format(x_value, ".5f"):
            return False
        if abs(Decimal(response_text) - Decimal(str(y_value))) > PRINTED_SIGNAL_HALF_UNIT:
            return False
    return True


def test_owner_archive_validates_exact_9_0_scientific_family_and_millivolt_units(
    tmp_path: Path,
) -> None:
    archive = _archive()
    assert archive.stat().st_size == EXPECTED_ARCHIVE_SIZE
    assert _sha256(archive) == EXPECTED_ARCHIVE_SHA256

    prm_paths: dict[str, Path] = {}
    curves: dict[str, tuple[tuple[tuple[str, str], ...], ...]] = {}
    result_rows: dict[str, tuple[tuple[str, Decimal, str, Decimal], ...]] = {}
    private_names: set[str] = set()
    with ZipFile(archive) as source:
        infos = source.infolist()
        assert len(infos) == EXPECTED_MEMBER_COUNT
        assert sum(info.file_size for info in infos) == EXPECTED_EXPANDED_BYTES
        assert source.testzip() is None
        assert all(_safe_member(info) for info in infos)
        assert all(
            info.file_size <= MAX_MEMBER_BYTES
            and info.compress_size > 0
            and info.file_size / info.compress_size <= MAX_COMPRESSION_RATIO
            for info in infos
        )
        assert len({info.filename for info in infos}) == EXPECTED_MEMBER_COUNT
        assert len({info.filename.casefold() for info in infos}) == EXPECTED_MEMBER_COUNT
        for info in infos:
            data = source.read(info)
            digest = _sha256_bytes(data)
            private_names.add(PurePosixPath(info.filename).name)
            suffix = PurePosixPath(info.filename).suffix.casefold()
            if suffix == ".prm":
                path = tmp_path / f"source-{digest}.prm"
                path.write_bytes(data)
                prm_paths[digest] = path
            elif suffix == ".csv":
                curves[digest] = _curve_blocks(data)
                result_rows[digest] = _result_rows(data)
            else:
                raise AssertionError("archive contains an unexpected member type")

    assert len(prm_paths) == len(curves) == EXPECTED_PRMS
    assert len({path.read_bytes() for path in prm_paths.values()}) == EXPECTED_PRMS
    assert len(curves) == len(set(curves))

    matched_exports: set[str] = set()
    all_signals: list[SignalSeries] = []
    matched_result_rows = 0
    area_errors: list[float] = []
    fail_closed_rows = 0
    exact_area_by_decimal_places = {places: 0 for places in AREA_QUANTA}
    for path in prm_paths.values():
        inspected = inspect_file(path, experimental_derived_area=True)
        bundle = inspected.file.bundle
        assert bundle is not None
        assert bundle.peaks
        assert all(
            peak.status == "ordifile_derived_experimental"
            and peak.data_origin == "ordifile_marker_derived"
            and peak.derivation_method_id == "youngin-prm-marker-group-baseline-v4"
            and peak.area is None
            and peak.calculated_area is not None
            for peak in bundle.peaks
        )
        assert bundle.samples[0].detectors == ("FID", "TCD")
        signals = {signal.detector: signal for signal in bundle.signals}
        assert set(signals) == {"FID", "TCD"}
        assert all(
            signal.series_kind is SeriesKind.SCIENTIFIC_SIGNAL for signal in signals.values()
        )
        assert signals["FID"].x_unit == signals["TCD"].x_unit == "min"
        assert signals["FID"].y_unit == signals["TCD"].y_unit == "mV"
        matches = [
            export_digest
            for export_digest, blocks in curves.items()
            if _curve_matches(signals["FID"], blocks[0])
            and _curve_matches(signals["TCD"], blocks[1])
        ]
        assert len(matches) == 1
        matched_exports.add(matches[0])
        matched_rt, errors, exact, skipped = _derived_result_errors(bundle, result_rows[matches[0]])
        fail_closed_rows += skipped
        matched_result_rows += matched_rt
        area_errors.extend(errors)
        for places, count in exact.items():
            exact_area_by_decimal_places[places] += count
        all_signals.extend(signals.values())

    assert len(matched_exports) == EXPECTED_PRMS
    assert fail_closed_rows == EXPECTED_FAIL_CLOSED_ROWS
    assert matched_result_rows == len(area_errors)
    assert matched_result_rows + fail_closed_rows == EXPECTED_RESULT_ROWS
    assert exact_area_by_decimal_places == EXPECTED_AREA_EXACT_BY_DECIMAL_PLACES
    assert sum(error <= 0.01 for error in area_errors) == EXPECTED_AREA_WITHIN_1_PERCENT
    assert sum(error <= 0.05 for error in area_errors) == EXPECTED_AREA_WITHIN_5_PERCENT
    assert median(area_errors) == pytest.approx(EXPECTED_AREA_MEDIAN_RELATIVE_ERROR)
    assert _p90(area_errors) == pytest.approx(EXPECTED_AREA_P90_RELATIVE_ERROR)
    assert max(area_errors) == pytest.approx(EXPECTED_AREA_MAX_RELATIVE_ERROR)
    assert len(all_signals) == EXPECTED_STREAMS
    assert sum(len(signal.x_values) for signal in all_signals) == EXPECTED_SCIENTIFIC_POINTS
    assert (
        sum(len(signal.x_values) for signal in all_signals if signal.detector == "FID")
        == EXPECTED_POINTS_PER_DETECTOR
    )
    assert (
        sum(len(signal.x_values) for signal in all_signals if signal.detector == "TCD")
        == EXPECTED_POINTS_PER_DETECTOR
    )

    output = tmp_path / "scientific-9-0.xlsx"
    result = convert(
        tuple(prm_paths.values()),
        output,
        include_signals=True,
        experimental_derived_area=True,
    )
    assert result.success_count == EXPECTED_PRMS
    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        for sheet_name, detector in (("Signals_FID", "FID"), ("Signals_TCD", "TCD")):
            sheet = workbook[sheet_name]
            assert sheet.max_row == EXPECTED_POINTS_PER_DETECTOR + 1
            assert {
                (row[3], row[6], row[9]) for row in sheet.iter_rows(min_row=2, values_only=True)
            } == {(detector, "min", "mV")}
        peaks = tuple(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        headers = tuple(next(workbook["Peaks"].iter_rows(max_row=1, values_only=True)))
        assert len(peaks) == EXPECTED_DERIVED_PEAKS
        assert {row[headers.index("data_origin")] for row in peaks} == {"ordifile_marker_derived"}
        assert {row[headers.index("area")] for row in peaks} == {None}
        assert all(row[headers.index("calculated_area")] is not None for row in peaks)
        rendered = "\n".join(
            str(value)
            for sheet_name in ("Manifest", "Samples", "Metadata", "Import_Log")
            for row in workbook[sheet_name].iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        assert not any(name in rendered for name in private_names)
    finally:
        workbook.close()

    assert _sha256(archive) == EXPECTED_ARCHIVE_SHA256
