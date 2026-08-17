from __future__ import annotations

import hashlib
import math
import os
import struct
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters._shimadzu_labsolutions_result_ascii import read_result_ascii
from ordifile.api import convert, inspect_file

EXPECTED_ASCII_SIZE = 971_258
EXPECTED_ASCII_SHA256 = "46d1dcde188d7844c32abb89cda1f0d773cac480f6d6c93f2b6ca7149fdb9297"
EXPECTED_GCD_SIZE = 1_433_600
EXPECTED_GCD_SHA256 = "d670806265f994507ac99fc676f17098bf9b9d1c362c98df1cb31154ac7a5180"
EXPECTED_PEAK_COUNT = 83
EXPECTED_POINT_COUNT = 66_255
EXPECTED_RT_DIGEST = "c19d2d264c606a3bf5407e0c511bc2d49e4d6cb302ec48f4be8c0f405be39b34"
EXPECTED_START_DIGEST = "a0041097783b80adb6d24eb55892bb05c8a894fa1d8ac3b8d58b1170bccac2f3"
EXPECTED_END_DIGEST = "1ce215fd315b602fec4e4555a568a7b43e1a35ab2255ca6855ad72a5c3556873"
EXPECTED_AREA_DIGEST = "2a12071d874f79b02308ec8c86cf76be1c01efe26b62098926b4b31d56b38e3c"
EXPECTED_HEIGHT_DIGEST = "87a976f34205ee2a6b3a203fbedee09c07b0f86203116606345f696e00348c4c"
EXPECTED_ASCII_TIME_DIGEST = "5134dc0fa78155212116aa6f79f790223ce5058f678a7927dfe5a5aa932a52ab"
EXPECTED_ASCII_SIGNAL_DIGEST = "7fe6f13daa282a19fe26b5f92669fb7d6730dabd0359e54826cc4fb00227d75d"


def _fixture(environment: str) -> Path:
    value = os.environ.get(environment)
    if not value:
        raise AssertionError(f"{environment} is required")
    return Path(value)


def _text_digest(values: Iterable[str]) -> str:
    return hashlib.sha256("\n".join(values).encode("utf-8")).hexdigest()


def _ascii_chromatogram(path: Path) -> tuple[tuple[float, ...], tuple[int, ...]]:
    lines = path.read_text(encoding="ascii").splitlines()
    marker = lines.index("[Chromatogram (Ch1)]")
    if lines[marker + 5] != "R.Time (min)\tIntensity":
        raise AssertionError("external ASCII chromatogram header changed")
    declared = int(lines[marker + 2].split("\t", 1)[1])
    rows = tuple(line.split("\t") for line in lines[marker + 6 :] if line)
    if declared != EXPECTED_POINT_COUNT or len(rows) != declared:
        raise AssertionError("external ASCII chromatogram count changed")
    times = tuple(float(row[0]) for row in rows)
    intensities = tuple(int(row[1]) for row in rows)
    return times, intensities


def _private_values(path: Path) -> frozenset[str]:
    lines = path.read_text(encoding="ascii").splitlines()
    private_sections = {
        "[Sample Information]",
        "[Original Files]",
        "[File Description]",
    }
    values: set[str] = set()
    current = ""
    for line in lines:
        if line.startswith("[") and line.endswith("]"):
            current = line
            continue
        if current in private_sections and line:
            value = line.split("\t", 1)[-1].strip()
            if value:
                values.add(value)
        elif current == "[Header]" and line.startswith("Data File Name\t"):
            values.add(line.split("\t", 1)[1])
        elif current == "[File Information]" and line and not line.startswith("Type\t"):
            values.add(line.split("\t", 1)[-1])
    public_profile_values = frozenset(
        {"LabSolutions", "Data File", "GC-2014", "SFID1", "Unknown", "Standard"}
    )
    return frozenset(
        value
        for value in values
        if len(value) >= 4
        and any(character.isalpha() for character in value)
        and value not in public_profile_values
    )


def test_exact_result_ascii_peaks_same_run_gcd_and_workbook(tmp_path: Path) -> None:
    ascii_source = _fixture("ORDIFILE_SHIMADZU_RESULT_ASCII_FIXTURE")
    gcd_source = _fixture("ORDIFILE_SHIMADZU_GCD_FIXTURE")
    ascii_bytes = ascii_source.read_bytes()
    gcd_bytes = gcd_source.read_bytes()
    if (
        len(ascii_bytes) != EXPECTED_ASCII_SIZE
        or hashlib.sha256(ascii_bytes).hexdigest() != EXPECTED_ASCII_SHA256
        or len(gcd_bytes) != EXPECTED_GCD_SIZE
        or hashlib.sha256(gcd_bytes).hexdigest() != EXPECTED_GCD_SHA256
    ):
        raise AssertionError("external fixture identity changed")

    decoded = read_result_ascii(ascii_source)
    if len(decoded.peaks) != EXPECTED_PEAK_COUNT:
        raise AssertionError("external source peak count changed")
    sequences = (
        (tuple(peak.retention_time_text for peak in decoded.peaks), EXPECTED_RT_DIGEST),
        (tuple(peak.start_time_text for peak in decoded.peaks), EXPECTED_START_DIGEST),
        (tuple(peak.end_time_text for peak in decoded.peaks), EXPECTED_END_DIGEST),
        (tuple(peak.area_text for peak in decoded.peaks), EXPECTED_AREA_DIGEST),
        (tuple(peak.height_text for peak in decoded.peaks), EXPECTED_HEIGHT_DIGEST),
    )
    if any(_text_digest(values) != expected for values, expected in sequences):
        raise AssertionError("external source canonical peak digest changed")

    ascii_times, ascii_intensities = _ascii_chromatogram(ascii_source)
    time_digest = hashlib.sha256(b"".join(struct.pack("<d", value) for value in ascii_times))
    signal_digest = hashlib.sha256(
        b"".join(struct.pack(">q", value) for value in ascii_intensities)
    )
    if (
        time_digest.hexdigest() != EXPECTED_ASCII_TIME_DIGEST
        or signal_digest.hexdigest() != EXPECTED_ASCII_SIGNAL_DIGEST
    ):
        raise AssertionError("external ASCII same-run chromatogram digest changed")

    gcd_inspected = inspect_file(gcd_source)
    gcd_bundle = gcd_inspected.file.bundle
    if gcd_bundle is None or len(gcd_bundle.signals) != 1:
        raise AssertionError("paired native GCD did not produce one scientific signal")
    signal = gcd_bundle.signals[0]
    if (
        len(signal.x_values) != EXPECTED_POINT_COUNT
        or len(signal.y_values) != EXPECTED_POINT_COUNT
        or any(
            round(value, 5) != expected
            for value, expected in zip(signal.x_values, ascii_times, strict=True)
        )
        or any(
            round(value) != expected
            for value, expected in zip(signal.y_values, ascii_intensities, strict=True)
        )
    ):
        raise AssertionError("paired native GCD and same-run ASCII chromatogram disagree")
    gcd_metadata = {entry.key: entry.value for entry in gcd_bundle.metadata}
    if (
        gcd_metadata.get("software_version") != decoded.software_version
        or gcd_metadata.get("instrument_model") != decoded.instrument_model
        or gcd_metadata.get("sampling_interval") != decoded.chromatogram_interval_ms
        or signal.detector != "FID"
        or signal.channel != decoded.source_channel
    ):
        raise AssertionError("paired native GCD and result ASCII profile facts disagree")

    inspected = inspect_file(ascii_source)
    bundle = inspected.file.bundle
    if bundle is None or len(bundle.peaks) != EXPECTED_PEAK_COUNT:
        raise AssertionError("Ordifile did not return the exact external result count")
    if inspected.file.source.public_reference != f"source-{EXPECTED_ASCII_SHA256}":
        raise AssertionError("external result public identity is not hash-derived")
    for source_peak, canonical in zip(decoded.peaks, bundle.peaks, strict=True):
        if (
            canonical.peak_number != source_peak.peak_number
            or canonical.observation_order != source_peak.observation_order
            or canonical.retention_time != source_peak.retention_time
            or canonical.start_time != source_peak.start_time
            or canonical.end_time != source_peak.end_time
            or canonical.area != source_peak.area
            or canonical.height != source_peak.height
            or canonical.retention_time_unit != "min"
            or canonical.area_unit is not None
            or canonical.height_unit is not None
            or canonical.detector != "FID"
            or canonical.channel != "Ch1"
            or canonical.compound is not None
        ):
            raise AssertionError("Ordifile canonical result mapping changed")

    output = tmp_path / "shimadzu-result.xlsx"
    convert(ascii_source, output)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks_sheet = workbook["Peaks"]
        headers = tuple(next(peaks_sheet.values))
        rows = tuple(peaks_sheet.iter_rows(min_row=2, values_only=True))
        if len(rows) != EXPECTED_PEAK_COUNT:
            raise AssertionError("workbook peak rows were lost or duplicated")
        for row, canonical in zip(rows, bundle.peaks, strict=True):
            for column, expected in (
                ("retention_time", canonical.retention_time),
                ("area", canonical.area),
                ("height", canonical.height),
                ("start_time", canonical.start_time),
                ("end_time", canonical.end_time),
            ):
                actual = row[headers.index(column)]
                if (
                    type(actual) not in {int, float}
                    or expected is None
                    or not math.isclose(float(actual), expected, rel_tol=1e-14, abs_tol=1e-14)
                ):
                    raise AssertionError("workbook scientific numeric mapping changed")
        order_sheet = workbook["Peak_Order_Matrix"]
        if order_sheet.max_row != 2 or order_sheet.max_column != 7 + 2 * EXPECTED_PEAK_COUNT:
            raise AssertionError("workbook ordered peak matrix shape changed")
        if tuple(next(order_sheet.iter_rows(min_row=2, values_only=True)))[2:7] != (
            "Shimadzu",
            "FID",
            "Ch1",
            "min",
            None,
        ):
            raise AssertionError("workbook ordered peak matrix identity changed")
        workbook_strings = {
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        }
        if (
            _private_values(ascii_source) & workbook_strings
            or ascii_source.name in workbook_strings
        ):
            raise AssertionError("privacy-bearing source metadata reached the workbook")
    finally:
        workbook.close()
