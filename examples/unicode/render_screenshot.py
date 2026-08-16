# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

"""Render a worksheet from a generated Ordifile workbook into a PNG.

This is a documentation tool, not part of the verified package. It exists so the
example screenshot in ``docs/assets`` can be reproduced from the *actual* workbook
that the current CLI produces, rather than from a hand-drawn mock-up.

It reads one sheet with openpyxl (the same reader the integration tests use) and
draws the real cell values with matplotlib. Requires an extra dependency:

    python -m pip install matplotlib

Usage:

    ordifile convert examples/unicode --sort filename --output /tmp/unicode_result.xlsx
    python examples/unicode/render_screenshot.py /tmp/unicode_result.xlsx \
        docs/assets/unicode-example-samples.png --sheet Samples
"""

from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.font_manager as fm  # type: ignore[import-untyped]
import matplotlib.pyplot as plt  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]

# The example exercises Unicode filenames, so render with a CJK-capable font when
# one is available. The fallback is the default sans-serif (tofu boxes) which would
# defeat the point of the screenshot.
_CANDIDATE_FONTS = (
    "/System/Library/Fonts/AppleSDGothicNeo.ttc",
    "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
    "/Library/Fonts/AppleSDGothicNeo.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "C:\\Windows\\Fonts\\malgun.ttf",
)
for _candidate in _CANDIDATE_FONTS:
    if Path(_candidate).exists():
        fm.fontManager.addfont(_candidate)
        plt.rcParams["font.family"] = fm.FontProperties(fname=_candidate).get_name()
        break


def render_sheet(workbook_path: Path, sheet_name: str, output_path: Path) -> None:
    """Draw ``sheet_name`` of ``workbook_path`` to ``output_path`` as a PNG."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(
                f"{sheet_name!r} not in workbook {workbook_path}; have {workbook.sheetnames}"
            )
        rows = [list(row) for row in workbook[sheet_name].values]
    finally:
        workbook.close()

    if not rows:
        raise SystemExit(f"{sheet_name!r} in {workbook_path} is empty")

    column_count = max(len(row) for row in rows)
    table = [[cell if cell is not None else "" for cell in row] for row in rows]
    table = [row + [""] * (column_count - len(row)) for row in table]

    figure, axis = plt.subplots(figsize=(column_count * 1.6, len(table) * 0.9 + 1))
    axis.axis("off")
    rendered = axis.table(
        cellText=table,
        loc="center",
        cellLoc="left",
    )
    rendered.auto_set_font_size(False)
    rendered.set_fontsize(9)
    rendered.scale(1, 1.6)
    figure.tight_layout()
    figure.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(figure)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Generated .xlsx workbook")
    parser.add_argument("output", type=Path, help="Destination .png path")
    parser.add_argument(
        "--sheet",
        default="Samples",
        help="Worksheet to render (default: Samples)",
    )
    arguments = parser.parse_args()
    render_sheet(arguments.workbook, arguments.sheet, arguments.output)


if __name__ == "__main__":
    main()
