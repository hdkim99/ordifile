from __future__ import annotations

import runpy
from dataclasses import replace
from pathlib import Path
from typing import cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableMapping,
)


def _synthetic_agilent_result_xml() -> bytes:
    generator = runpy.run_path(
        str(
            Path(__file__).parents[1]
            / "fixtures"
            / "synthetic"
            / "generate_agilent_chemstation_result_xml.py"
        )
    )
    return cast(bytes, generator["synthetic_result_xml_bytes"]())


def test_required_workbook_structure_hashes_and_natural_order(tmp_path: Path) -> None:
    for name in ("sample_10.csv", "sample_1.csv", "sample_2.csv"):
        sample = name.removesuffix(".csv")
        (tmp_path / name).write_text(
            f"sample_id,retention_time,area,compound\n{sample},1,2,A\n{sample},2,3,A\n",
            encoding="utf-8",
        )
    output = tmp_path / "result.xlsx"
    result = convert(tmp_path, output)
    assert result.success_count == 3
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        required = {"Manifest", "Samples", "Peak_Matrix", "Peaks", "Metadata", "Import_Log"}
        assert required.issubset(workbook.sheetnames)
        sample_rows = list(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        assert [row[1] for row in sample_rows] == ["sample_1", "sample_2", "sample_10"]
        assert all(len(row[13]) == 64 for row in sample_rows)
        matrix = list(workbook["Peak_Matrix"].values)
        assert matrix[0][:3] == ("sample_id", "A_area", "A_area_2")
        assert matrix[1][1:] == (2, 3)
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["ordifile_version"] == "0.5.1"
        assert "labconvert_version" not in manifest
        assert manifest["original_modified"] == "No"
        assert manifest["sort_effective"] == "filename"
        assert manifest["result_summary_schema_version"] == 1
        assert manifest["sample_record_count"] == 3
        assert manifest["peak_record_count"] == 6
        assert manifest["execution_mode"] == "DIRECT"
    finally:
        workbook.close()


def test_peak_matrix_qualifiers_are_unambiguous_and_compound_only_stays_simple(
    tmp_path: Path,
) -> None:
    source = tmp_path / "qualified.csv"
    source.write_text(
        "sample_id,detector,channel,area,compound\n"
        "a,C,,1,A_B\n"
        "a,B_C,,2,A\n"
        "a,,,3,A\n"
        "a,B,C_D,4,A\n"
        "a,B_C,D,5,A\n",
        encoding="utf-8",
    )
    result = convert(source, tmp_path / "qualified.xlsx")
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        headers = tuple(next(workbook["Peak_Matrix"].values))
        assert "A_area" in headers
        assert len(headers) == len(set(value.casefold() for value in headers))
        qualified = [value for value in headers if value.startswith("q[")]
        assert len(qualified) == 4
        assert any("A%5FB" in value and "detector=C" in value for value in qualified)
        assert any("detector=B%5FC" in value for value in qualified)
        assert any("channel=C%5FD" in value for value in qualified)
    finally:
        workbook.close()


def test_peak_matrix_separates_conflicting_area_units(tmp_path: Path) -> None:
    mapped = tmp_path / "mapped.csv"
    mapped.write_text(
        "RT,Area,Compound,Detector,Channel\n1.0,999,compound-alpha,FID,FID1A\n",
        encoding="utf-8",
    )
    agilent = tmp_path / "agilent.xml"
    agilent.write_bytes(_synthetic_agilent_result_xml())
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
        area_unit="mV.s",
        compound_name_column=ColumnSelector("Compound", 3),
        detector_column=ColumnSelector("Detector", 4),
        channel_column=ColumnSelector("Channel", 5),
    )

    result = convert(
        (mapped, agilent),
        tmp_path / "mixed-units.xlsx",
        peak_table_mapping=mapping,
    )

    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        rows = tuple(workbook["Peak_Matrix"].iter_rows(values_only=True))
        headers = tuple(rows[0])
        mapped_header = (
            "q[compound=compound%2Dalpha][detector=FID][channel=FID1A][area_unit=mV%2Es]_area"
        )
        agilent_header = (
            "q[compound=compound%2Dalpha][detector=FID][channel=FID1A][area_unit=pA%2As]_area"
        )
        assert mapped_header in headers
        assert agilent_header in headers
        assert "q[compound=compound%2Dalpha][detector=FID][channel=FID1A]_area" not in headers
        mapped_index = headers.index(mapped_header)
        agilent_index = headers.index(agilent_header)
        assert {(row[mapped_index], row[agilent_index]) for row in rows[1:]} == {
            (999, None),
            (None, 100.5),
        }
    finally:
        workbook.close()

    unresolved = convert(
        (mapped, agilent),
        tmp_path / "mixed-unresolved-unit.xlsx",
        peak_table_mapping=replace(mapping, area_unit=None),
    )
    workbook = load_workbook(unresolved.output_path, read_only=True, data_only=False)
    try:
        headers = tuple(next(workbook["Peak_Matrix"].values))
        assert (
            "q[compound=compound%2Dalpha][detector=FID][channel=FID1A]"
            "[area_unit_status=unresolved]_area"
        ) in headers
        assert agilent_header in headers
    finally:
        workbook.close()


def test_samples_distinguish_same_basename_without_absolute_paths(tmp_path: Path) -> None:
    first = tmp_path / "one" / "sample.csv"
    second = tmp_path / "two" / "sample.csv"
    first.parent.mkdir()
    second.parent.mkdir()
    first.write_text("sample_id,area\na,1\n", encoding="utf-8")
    second.write_text("sample_id,area\nb,2\n", encoding="utf-8")
    result = convert((first, second), tmp_path / "result.xlsx")
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        paths = [row[3] for row in workbook["Samples"].iter_rows(min_row=2, values_only=True)]
        assert sorted(paths) == ["input_001/sample.csv", "input_002/sample.csv"]
        assert all(str(tmp_path) not in value for value in paths)
    finally:
        workbook.close()


def test_large_integers_are_written_as_exact_literals_with_manifest_policy(
    tmp_path: Path,
) -> None:
    exact = 9_007_199_254_740_993
    source = tmp_path / "exact.csv"
    source.write_text(
        f"sample_id,sequence,peak_number,area,compound\na,{exact},{exact},1,A\n",
        encoding="utf-8",
    )
    result = convert(source, tmp_path / "exact.xlsx")
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        samples = next(workbook["Samples"].iter_rows(min_row=2))
        peaks = next(workbook["Peaks"].iter_rows(min_row=2))
        assert samples[9].value == str(exact)
        assert samples[9].data_type == "s"
        assert peaks[4].value == str(exact)
        assert peaks[4].data_type == "s"
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["exact_integer_literal_count"] == 2
        assert "15 decimal digits" in manifest["exact_integer_policy"]
    finally:
        workbook.close()


def test_manifest_records_full_immutable_conversion_options(tmp_path: Path) -> None:
    source = tmp_path / "options.csv"
    source.write_text(
        "sample_id,area,time,signal\na,1,0,2\n",
        encoding="utf-8",
    )
    output = tmp_path / "options-result.xlsx"

    result = convert(
        source,
        output,
        recursive=True,
        extensions=(extension for extension in ("csv", "txt")),
        sort="filename",
        include_signals=True,
        adapter="generic_csv",
        sheet="Data",
        include_hidden_sheets=True,
        on_error="continue",
        overwrite=True,
        sidecar_mode="csv",
    )

    assert result.options.extensions == (".csv", ".txt")
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["conversion_options_policy"].startswith("Immutable snapshot")
        assert manifest["option_recursive"] == "True"
        assert manifest["option_extensions"] == ".csv; .txt"
        assert manifest["option_sort"] == "filename"
        assert manifest["option_include_signals"] == "True"
        assert manifest["option_adapter"] == "generic_csv"
        assert manifest["option_sheet"] == "Data"
        assert manifest["option_include_hidden_sheets"] == "True"
        assert manifest["option_on_error"] == "continue"
        assert manifest["option_overwrite"] == "True"
        assert manifest["option_sidecar_mode"] == "csv"
        assert manifest["option_output_name"] == output.name
    finally:
        workbook.close()
