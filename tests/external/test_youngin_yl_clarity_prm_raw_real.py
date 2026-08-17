from __future__ import annotations

import hashlib
import os
import re
import stat
import struct
from collections import Counter
from pathlib import Path, PurePosixPath
from zipfile import ZipFile, ZipInfo

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.models import SeriesKind

EXPECTED_ARCHIVE_SIZE = 3_083_937
EXPECTED_ARCHIVE_SHA256 = "4af61e1aa8abef3694a4c24a28203b0a1d382a11b6442c3e9b43653487f97fe5"
EXPECTED_MEMBER_COUNT = 24
EXPECTED_FILE_COUNT = 23
EXPECTED_UNCOMPRESSED_BYTES = 12_023_434
EXPECTED_BLOCK_COUNT = 43
EXPECTED_RECORD_COUNT = 563_240
EXPECTED_RECORD_HISTOGRAM = {11_970: 2, 11_980: 1, 13_180: 32, 13_190: 6, 13_210: 2}
EXPECTED_BATCH_DIGEST = "96d84945222d583c621459d25b57a2f5bb5cca9cf0cfa9f830811a5d483a3c5a"

_EMAIL = re.compile(r"(?i)\b[^\s@]+@[^\s@]+\.[^\s@]+\b")
_LOCAL_PATH = re.compile(r"(?i)(?:[a-z]:[\\/]|/users/|/home/|\\\\)")


def _archive() -> Path:
    value = os.environ.get("ORDIFILE_YOUNGIN_PRM_ARCHIVE")
    if not value:
        raise AssertionError("ORDIFILE_YOUNGIN_PRM_ARCHIVE is required")
    return Path(value)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_member(info: ZipInfo) -> bool:
    name = info.filename
    pure = PurePosixPath(name)
    mode = (info.external_attr >> 16) & 0xFFFF
    return (
        not info.is_dir()
        and not (info.flag_bits & 0x1)
        and "\\" not in name
        and not pure.is_absolute()
        and all(part not in {"", ".", ".."} for part in pure.parts)
        and (mode == 0 or stat.S_ISREG(mode))
    )


def _group(info: ZipInfo) -> str:
    stem = PurePosixPath(info.filename).stem.casefold()
    has_fid = "fid" in stem
    has_tcd = "tcd" in stem
    if has_fid and not has_tcd:
        return "FID_STD"
    if has_tcd and not has_fid:
        return "TCD_STD"
    return "MIXED_SAMPLE"


def _stage_private_archive(archive: Path, target: Path) -> tuple[tuple[Path, ...], frozenset[str]]:
    target.mkdir()
    with ZipFile(archive) as source:
        infos = source.infolist()
        assert len(infos) == EXPECTED_MEMBER_COUNT
        assert sum(info.file_size for info in infos) == EXPECTED_UNCOMPRESSED_BYTES
        assert source.testzip() is None
        files = [info for info in infos if not info.is_dir()]
        assert len(files) == EXPECTED_FILE_COUNT
        assert all(_safe_member(info) for info in files)
        assert len({info.filename for info in files}) == EXPECTED_FILE_COUNT
        assert len({info.filename.casefold() for info in files}) == EXPECTED_FILE_COUNT

        grouped: dict[str, list[ZipInfo]] = {
            "FID_STD": [],
            "TCD_STD": [],
            "MIXED_SAMPLE": [],
        }
        for info in files:
            grouped[_group(info)].append(info)
        assert {key: len(value) for key, value in grouped.items()} == {
            "FID_STD": 10,
            "TCD_STD": 10,
            "MIXED_SAMPLE": 3,
        }

        staged: list[Path] = []
        private_names: set[str] = set()
        for group in ("FID_STD", "TCD_STD", "MIXED_SAMPLE"):
            for ordinal, info in enumerate(
                sorted(grouped[group], key=lambda item: item.filename), 1
            ):
                private_names.add(PurePosixPath(info.filename).name)
                data = source.read(info)
                assert len(data) == info.file_size
                alias = target / f"{group}_{ordinal:03d}.PRM"
                alias.write_bytes(data)
                assert _sha256(alias) == hashlib.sha256(data).hexdigest()
                staged.append(alias)
    return tuple(staged), frozenset(private_names)


def _batch_digest(paths: tuple[Path, ...]) -> tuple[str, Counter[int], int, int, int, int]:
    digest = hashlib.sha256(b"ordifile-youngin-local-reference-v1\0")
    histogram: Counter[int] = Counter()
    block_count = 0
    record_count = 0
    dual_count = 0
    single_count = 0
    for path in paths:
        inspected = inspect_file(path)
        bundle = inspected.file.bundle
        assert bundle is not None
        signals = bundle.signals
        source_sha256 = inspected.file.source.sha256
        assert source_sha256 is not None
        assert bundle.samples[0].sample_id == f"PRM_{source_sha256[:16]}"
        assert not any(entry.key == "user_supplied_group" for entry in bundle.metadata)
        alias = path.stem.encode("ascii")
        digest.update(struct.pack(">H", len(alias)))
        digest.update(alias)
        digest.update(bytes.fromhex(source_sha256))
        digest.update(struct.pack(">H", len(signals)))
        labels = tuple(signal.channel for signal in signals)
        if labels == ("native_label_FID", "native_label_TCD"):
            dual_count += 1
        elif labels == ("native_label_TCD",):
            single_count += 1
        else:
            raise AssertionError("unsupported native-label profile reached the canonical bundle")
        for ordinal, signal in enumerate(signals, 1):
            assert signal.series_kind is SeriesKind.DECODED_RECORDS
            assert signal.detector is None
            assert signal.x_values == tuple(range(len(signal.y_values)))
            assert signal.x_label == "decoded_record_index"
            assert signal.x_unit is None
            assert signal.y_label == "decoded_raw_binary32"
            assert signal.y_unit is None
            digest.update(struct.pack(">H", ordinal))
            digest.update(struct.pack(">Q", len(signal.y_values)))
            for value in signal.y_values:
                assert type(value) is float
                digest.update(struct.pack(">f", value))
            histogram[len(signal.y_values)] += 1
            block_count += 1
            record_count += len(signal.y_values)
    return digest.hexdigest(), histogram, block_count, record_count, dual_count, single_count


def _signal_sheet_digest(sheet: object) -> tuple[str, int, set[str]]:
    digest = hashlib.sha256()
    rows = 0
    public_text: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):  # type: ignore[attr-defined]
        x_value = row[4]
        y_value = row[7]
        assert type(x_value) is int
        assert type(y_value) in {int, float}
        digest.update(struct.pack(">Qf", x_value, float(y_value)))
        rows += 1
        public_text.update(str(value) for value in row if type(value) is str)
    return digest.hexdigest(), rows, public_text


def test_all_local_prm_files_parse_deterministically_and_reopen_one_workbook(
    tmp_path: Path,
) -> None:
    archive = _archive()
    assert archive.stat().st_size == EXPECTED_ARCHIVE_SIZE
    assert _sha256(archive) == EXPECTED_ARCHIVE_SHA256
    staged, private_names = _stage_private_archive(archive, tmp_path / "staged")

    first = _batch_digest(staged)
    second = _batch_digest(staged)
    assert first == second
    batch_digest, histogram, block_count, record_count, dual_count, single_count = first
    assert batch_digest == EXPECTED_BATCH_DIGEST
    assert dict(histogram) == EXPECTED_RECORD_HISTOGRAM
    assert block_count == EXPECTED_BLOCK_COUNT
    assert record_count == EXPECTED_RECORD_COUNT
    assert dual_count == 20
    assert single_count == 3

    output = tmp_path / "youngin-prm-raw.xlsx"
    result = convert(staged, output, include_signals=True)
    assert result.success_count == EXPECTED_FILE_COUNT
    assert result.failure_count == 0
    assert (
        sum(
            len(signal.y_values)
            for item in result.files
            if item.bundle is not None
            for signal in item.bundle.signals
        )
        == EXPECTED_RECORD_COUNT
    )

    expected_by_channel: dict[str, bytearray] = {
        "native_label_FID": bytearray(),
        "native_label_TCD": bytearray(),
    }
    expected_rows: Counter[str] = Counter()
    for item in result.files:
        assert item.bundle is not None
        for signal in item.bundle.signals:
            assert signal.channel is not None
            for x_value, y_value in zip(signal.x_values, signal.y_values, strict=True):
                expected_by_channel[signal.channel].extend(struct.pack(">Qf", x_value, y_value))
                expected_rows[signal.channel] += 1

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        record_sheets = [
            sheet for sheet in workbook.worksheets if sheet.title.startswith("Signals_Records_")
        ]
        assert len(record_sheets) == 2
        public_text: set[str] = set()
        for sheet in record_sheets:
            workbook_digest, row_count, sheet_text = _signal_sheet_digest(sheet)
            public_text.update(sheet_text)
            if sheet.title.endswith("_FI"):
                channel = "native_label_FID"
            elif sheet.title.endswith("_TC"):
                channel = "native_label_TCD"
            else:
                raise AssertionError("unexpected decoded-record sheet label")
            assert workbook_digest == hashlib.sha256(expected_by_channel[channel]).hexdigest()
            assert row_count == expected_rows[channel]
        for sheet_name in ("Manifest", "Samples", "Metadata", "Import_Log"):
            for row in workbook[sheet_name].iter_rows(values_only=True):
                public_text.update(str(value) for value in row if type(value) is str)
        rendered = "\n".join(sorted(public_text))
        assert not any(name in rendered for name in private_names)
        assert not any(path.stem in rendered for path in staged)
        assert "user_supplied_group" not in rendered
        assert _EMAIL.search(rendered) is None
        assert _LOCAL_PATH.search(rendered) is None
    finally:
        workbook.close()

    assert _sha256(archive) == EXPECTED_ARCHIVE_SHA256
    output.unlink()
    for path in staged:
        path.unlink()
    staged[0].parent.rmdir()
    assert not output.exists()
