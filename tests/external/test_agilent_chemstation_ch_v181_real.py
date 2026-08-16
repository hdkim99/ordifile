from __future__ import annotations

import hashlib
import os
import shutil
import struct
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.models import SeriesKind

EXPECTED_SIZE = 298_146
EXPECTED_SHA256 = "9abeb86b09d54c10e81f46648804acc0319b6e1d014cee54034eae91331f97ef"
EXPECTED_SEQUENCE_SHA256 = "9d0adc5724779c8c3061da6e9523952401953d859922c4079b023072c6161667"
EXPECTED_FIRST_20 = (
    13086,
    13088,
    13101,
    13108,
    13106,
    13125,
    13149,
    13161,
    13139,
    13114,
    13117,
    13126,
    13125,
    13142,
    13178,
    13189,
    13161,
    13139,
    13133,
    13118,
)
EXPECTED_LAST_20 = (
    39493,
    39485,
    39488,
    39490,
    39521,
    39528,
    39532,
    39510,
    39495,
    39506,
    39500,
    39497,
    39501,
    39523,
    39517,
    39503,
    39496,
    39485,
    39467,
    39467,
)


def _fixture() -> Path:
    value = os.environ.get("ORDIFILE_AGILENT_CH_V181_FIXTURE")
    if not value:
        raise AssertionError("ORDIFILE_AGILENT_CH_V181_FIXTURE is required")
    return Path(value)


def _sequence_digest(values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for value in values:
        if type(value) is not int:
            raise AssertionError(f"decoded record is not an exact integer: {type(value).__name__}")
        digest.update(struct.pack(">q", value))
    return digest.hexdigest()


def test_exact_external_v181_fixture_and_workbook(tmp_path: Path) -> None:
    supplied_source = _fixture()
    assert supplied_source.stat().st_size == EXPECTED_SIZE
    assert hashlib.sha256(supplied_source.read_bytes()).hexdigest() == EXPECTED_SHA256
    source = tmp_path / "FID1A.CH"
    shutil.copyfile(supplied_source, source)
    assert hashlib.sha256(source.read_bytes()).hexdigest() == EXPECTED_SHA256

    inspected = inspect_file(source)
    bundle = inspected.file.bundle
    assert bundle is not None
    signal = bundle.signals[0]
    assert signal.series_kind is SeriesKind.DECODED_RECORDS
    assert len(signal.y_values) == 36_501
    assert signal.y_values[:20] == EXPECTED_FIRST_20
    assert signal.y_values[-20:] == EXPECTED_LAST_20
    assert min(signal.y_values) == 12_878
    assert max(signal.y_values) == 910_054
    assert signal.y_values[-1] == signal.y_values[-2]
    assert _sequence_digest(signal.y_values) == EXPECTED_SEQUENCE_SHA256
    assert signal.x_values == tuple(range(36_501))
    assert signal.x_unit is None
    assert signal.y_unit is None
    assert bundle.peaks == ()
    metadata = {entry.key: entry.value for entry in bundle.metadata}
    assert metadata["internal_version"] == 181
    assert metadata["payload_offset"] == 6_144
    assert metadata["absolute_record_count"] == 36_500
    assert metadata["ordinary_record_count"] == 1
    assert metadata["ambiguous_final_zero_ordinary_record_included"] is True
    assert metadata["header_f64_4732_candidate_hex"] == "0x1.1555555555555p-5"
    assert metadata["header_f64_4732_bytes_hex"] == "3fa1555555555555"

    output = tmp_path / "agilent-v181.xlsx"
    result = convert(source, output, include_signals=True)
    assert result.files[0].source.sha256 == EXPECTED_SHA256
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        sheet = workbook["Signals_Records_FID"]
        assert sheet.max_row == 36_502
        assert sheet.cell(2, 5).value == 0
        assert sheet.cell(36_502, 5).value == 36_500
        assert sheet.cell(36_502, 8).value == 39_467
        workbook_values = (
            row[0]
            for row in sheet.iter_rows(
                min_row=2,
                min_col=8,
                max_col=8,
                values_only=True,
            )
        )
        assert _sequence_digest(workbook_values) == EXPECTED_SEQUENCE_SHA256
        metadata_sheet = workbook["Metadata"]
        metadata_rows = {
            row[3]: row[4] for row in metadata_sheet.iter_rows(min_row=2, values_only=True)
        }
        assert metadata_rows["header_f64_4732_candidate_hex"] == "0x1.1555555555555p-5"
    finally:
        workbook.close()
