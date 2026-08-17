from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.cli.main import main
from ordifile.core.detection import SOURCE_IDENTITY_PROBE_REASON
from ordifile.core.models import FileStatus

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_chemstation_result_xml import (  # noqa: E402
    synthetic_result_xml_bytes,
)
from generate_shimadzu_gcsolution_gcd import synthetic_gcd_bytes  # noqa: E402
from generate_shimadzu_labsolutions_result_ascii import (  # noqa: E402
    synthetic_result_ascii_bytes,
)


def _workbook_values(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return tuple(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


def test_inspect_cli_and_result_bundle_hide_private_ascii_source(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    private_name = "private-person-labsolutions-result.txt"
    data = synthetic_result_ascii_bytes()
    source = tmp_path / private_name
    source.write_bytes(data)
    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.WARNING
    assert inspected.file.adapter_id == "shimadzu_labsolutions_result_ascii"
    assert inspected.file.source.public_reference == expected_reference
    assert inspected.file.source.path == Path(expected_reference)
    assert inspected.file.bundle is not None
    assert len(inspected.file.bundle.peaks) == 3
    assert inspected.file.bundle.signals == ()
    assert all(reason == SOURCE_IDENTITY_PROBE_REASON for _, _, reason in inspected.probes)

    assert main(["inspect", str(source), "--verbose"]) == 0
    output = capsys.readouterr().out
    assert "Peaks: 3" in output
    assert "Scientific signals: 0" in output
    assert private_name not in output


def test_standalone_result_workbook_preserves_order_with_unresolved_response_units(
    tmp_path: Path,
) -> None:
    private_name = "private-result.txt"
    source = tmp_path / private_name
    source.write_bytes(synthetic_result_ascii_bytes())
    output = tmp_path / "result.xlsx"

    result = convert(source, output)

    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].values)
        headers = peaks[0]
        rows = peaks[1:]
        assert len(rows) == 3
        assert [row[headers.index("peak_number")] for row in rows] == [1, 2, 3]
        assert [row[headers.index("observation_order")] for row in rows] == [1, 2, 3]
        assert {row[headers.index("manufacturer")] for row in rows} == {"Shimadzu"}
        assert {row[headers.index("detector")] for row in rows} == {"FID"}
        assert {row[headers.index("channel")] for row in rows} == {"Ch1"}
        assert {row[headers.index("retention_time_unit")] for row in rows} == {"min"}
        assert {row[headers.index("area_unit")] for row in rows} == {None}
        assert {row[headers.index("height_unit")] for row in rows} == {None}
        assert {row[headers.index("compound")] for row in rows} == {None}

        matrix = list(workbook["Peak_Order_Matrix"].values)
        assert matrix[0][:7] == (
            "sample_id",
            "source_file",
            "manufacturer",
            "detector",
            "channel",
            "retention_time_unit",
            "area_unit",
        )
        assert matrix[1][2:7] == ("Shimadzu", "FID", "Ch1", "min", None)
        assert matrix[1][7:] == (1.25, 100.5, 2.5, 200.75, 3.75, 300)
        all_values = tuple(
            value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        assert not any(private_name in str(value) for value in all_values)
        assert "Signals_FID" not in workbook.sheetnames
    finally:
        workbook.close()


def test_valid_and_corrupt_result_ascii_files_are_isolated(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    valid = synthetic_result_ascii_bytes()
    (inputs / "private-valid.txt").write_bytes(valid)
    (inputs / "private-corrupt.txt").write_bytes(valid[:-20])
    output = tmp_path / "isolated.xlsx"

    result = convert(inputs, output, extensions=(".txt",))

    assert result.failure_count == 1
    assert sum(item.bundle is not None for item in result.files) == 1
    assert all("private-" not in repr(item) for item in result.files)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 4
        assert workbook["Peak_Order_Matrix"].max_row == 2
        assert {
            row[4] for row in workbook["Import_Log"].iter_rows(min_row=2, values_only=True)
        } == {
            "warning",
            "failed",
        }
        assert not any("private-" in value for value in _workbook_values(output))
    finally:
        workbook.close()


def test_generic_txt_success_restores_relative_provenance_and_probe_evidence(
    tmp_path: Path,
) -> None:
    source = tmp_path / "ordinary-table.txt"
    source.write_text("sample_id;area\ngeneric;1\n", encoding="utf-8")

    inspected = inspect_file(source)

    assert inspected.file.adapter_id == "generic_semicolon"
    assert inspected.file.source.public_id is None
    assert inspected.file.source.public_reference == source.name
    probes = {adapter_id: reason for adapter_id, _confidence, reason in inspected.probes}
    assert probes["generic_semicolon"] != SOURCE_IDENTITY_PROBE_REASON
    assert probes["shimadzu_labsolutions_result_ascii"] == SOURCE_IDENTITY_PROBE_REASON


@pytest.mark.parametrize(
    "data",
    (
        synthetic_result_ascii_bytes(software_version="5.81"),
        synthetic_result_ascii_bytes()[:-20],
    ),
)
def test_private_malformed_shimadzu_txt_keeps_provisional_sha_identity(
    tmp_path: Path, data: bytes
) -> None:
    private_name = "private-malformed-result.txt"
    source = tmp_path / private_name
    source.write_bytes(data)

    inspected = inspect_file(source)

    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"
    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.source.public_reference == expected_reference
    assert inspected.file.source.path == Path(expected_reference)
    assert private_name not in repr(inspected)


@pytest.mark.parametrize(
    "data",
    (
        synthetic_result_ascii_bytes().replace(b"[Peak Table(Ch1)]", b"[Missing Peak Table]", 1),
        synthetic_result_ascii_bytes().replace(b"\r\n", b"\n"),
    ),
)
def test_identified_malformed_labsolutions_family_never_falls_through_to_generic_txt(
    tmp_path: Path,
    data: bytes,
) -> None:
    private_name = "private-family-corrupt.txt"
    source = tmp_path / private_name
    source.write_bytes(data)

    inspected = inspect_file(source)

    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"
    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.adapter_id == "shimadzu_labsolutions_result_ascii"
    assert inspected.file.source.public_reference == expected_reference
    assert private_name not in repr(inspected)


def test_mixed_agilent_shimadzu_results_generic_txt_and_native_gcd_remain_separate(
    tmp_path: Path,
) -> None:
    agilent = tmp_path / "private-agilent.xml"
    agilent.write_bytes(synthetic_result_xml_bytes())
    shimadzu = tmp_path / "private-shimadzu.txt"
    shimadzu.write_bytes(synthetic_result_ascii_bytes())
    generic = tmp_path / "ordinary.txt"
    generic.write_text("sample_id;area\ngeneric;1\n", encoding="utf-8")
    gcd = tmp_path / "native.gcd"
    gcd.write_bytes(synthetic_gcd_bytes())
    output = tmp_path / "mixed.xlsx"

    result = convert((agilent, shimadzu, generic, gcd), output, include_signals=True)

    assert result.failure_count == 0
    assert {item.adapter_id for item in result.files} == {
        "agilent_chemstation_result_xml",
        "shimadzu_labsolutions_result_ascii",
        "generic_semicolon",
        "shimadzu_gcsolution_gcd",
    }
    assert (
        next(
            item for item in result.files if item.adapter_id == "generic_semicolon"
        ).source.public_reference
        == "ordinary.txt"
    )
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        order_rows = list(workbook["Peak_Order_Matrix"].values)
        assert len(order_rows) == 3
        assert {row[2] for row in order_rows[1:]} == {"Agilent", "Shimadzu"}
        assert {row[3:5] for row in order_rows[1:]} == {("FID", "FID1A"), ("FID", "Ch1")}
        peaks = list(workbook["Peaks"].values)
        manufacturer = peaks[0].index("manufacturer")
        assert {row[manufacturer] for row in peaks[1:] if row[manufacturer] is not None} == {
            "Agilent",
            "Shimadzu",
        }
        assert "Signals_FID" in workbook.sheetnames
    finally:
        workbook.close()
