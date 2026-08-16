from __future__ import annotations

import hashlib
import math
import os
import struct
from collections.abc import Iterable
from datetime import UTC, datetime
from pathlib import Path
from typing import cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.models import SeriesKind

EXPECTED_SIZE = 1_433_600
EXPECTED_SHA256 = "d670806265f994507ac99fc676f17098bf9b9d1c362c98df1cb31154ac7a5180"
EXPECTED_POINT_COUNT = 66_255
EXPECTED_SIGNAL_SHA256 = "b836371e5f8171788b2f3ebd0a3a75d07bfeb7ee8eed081992a9016192987b9a"
EXPECTED_TIME_SHA256 = "18c335833a87d10e59e997623f82ddc0e8b73f00031522d5b2339ab3f3b119e2"
EXPECTED_PAIRS_SHA256 = "a1395b48d5f802b6772bf0351ee694bf63a89af10a822feea30baf4f28023f45"
EXPECTED_ASCII_ROUNDED_SIGNAL_SHA256 = (
    "7fe6f13daa282a19fe26b5f92669fb7d6730dabd0359e54826cc4fb00227d75d"
)
EXPECTED_ASCII_ROUNDED_TIME_SHA256 = (
    "5134dc0fa78155212116aa6f79f790223ce5058f678a7927dfe5a5aa932a52ab"
)
EXPECTED_FIRST_20 = (
    -361.8000053912401,
    -361.8000053912401,
    -361.90000539273024,
    -362.00000539422035,
    -362.10000539571047,
    -362.2000053972006,
    -362.3000053986907,
    -362.4000054001808,
    -362.50000540167093,
    -362.60000540316105,
    -362.70000540465117,
    -362.70000540465117,
    -362.8000054061413,
    -362.8000054061413,
    -362.9000054076314,
    -362.9000054076314,
    -362.9000054076314,
    -362.9000054076314,
    -362.9000054076314,
    -362.9000054076314,
)
EXPECTED_LAST_20 = (
    4328.7000645026565,
    4328.800064504147,
    4328.900064505637,
    4328.900064505637,
    4328.900064505637,
    4328.900064505637,
    4328.900064505637,
    4329.000064507127,
    4329.000064507127,
    4329.000064507127,
    4329.000064507127,
    4328.900064505637,
    4328.900064505637,
    4328.800064504147,
    4328.7000645026565,
    4328.600064501166,
    4328.500064499676,
    4328.00001525879,
    4328.300064496696,
    4328.200064495206,
)


def _fixture() -> Path:
    value = os.environ.get("ORDIFILE_SHIMADZU_GCD_FIXTURE")
    if not value:
        raise AssertionError("ORDIFILE_SHIMADZU_GCD_FIXTURE is required")
    return Path(value)


def _f64_digest(values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for value in values:
        if type(value) is not float:
            raise AssertionError(f"signal coordinate is not an exact float: {type(value).__name__}")
        digest.update(struct.pack(">d", value))
    return digest.hexdigest()


def _pair_digest(x_values: Iterable[object], y_values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for x_value, y_value in zip(x_values, y_values, strict=True):
        if type(x_value) is not float or type(y_value) is not float:
            raise AssertionError("signal pair is not two exact floats")
        digest.update(struct.pack(">dd", x_value, y_value))
    return digest.hexdigest()


def _rounded_i64_digest(values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for value in values:
        if type(value) is not float:
            raise AssertionError("signal value is not an exact float")
        digest.update(struct.pack(">q", round(value)))
    return digest.hexdigest()


def _ascii_rounded_time_digest(values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for value in values:
        if type(value) is not float:
            raise AssertionError("time value is not an exact float")
        digest.update(struct.pack("<d", round(value, 5)))
    return digest.hexdigest()


def _assert_excel_numeric_series(
    actual_values: Iterable[object], expected_values: Iterable[float]
) -> None:
    # XlsxWriter writes Excel numeric cells with 15 significant decimal digits.
    # This bound is tight enough to detect changed/reordered points while accepting
    # only that documented numeric serialization loss.
    for actual, expected in zip(actual_values, expected_values, strict=True):
        if type(actual) not in {int, float}:
            raise AssertionError(f"workbook signal value is not numeric: {type(actual).__name__}")
        numeric = cast(int | float, actual)
        assert math.isclose(float(numeric), expected, rel_tol=1e-14, abs_tol=1e-14)


def test_exact_external_gcd_fixture_reference_and_workbook(tmp_path: Path) -> None:
    supplied_source = _fixture()
    assert supplied_source.stat().st_size == EXPECTED_SIZE
    assert hashlib.sha256(supplied_source.read_bytes()).hexdigest() == EXPECTED_SHA256
    source = supplied_source

    inspected = inspect_file(source)
    bundle = inspected.file.bundle
    assert bundle is not None
    assert len(bundle.signals) == 1
    signal = bundle.signals[0]
    assert signal.series_kind is SeriesKind.SCIENTIFIC_SIGNAL
    assert signal.detector == "FID"
    assert signal.channel == "Ch1"
    assert signal.x_label == "retention_time"
    assert signal.x_unit == "min"
    assert signal.y_unit == "uV"
    assert len(signal.x_values) == EXPECTED_POINT_COUNT
    assert len(signal.y_values) == EXPECTED_POINT_COUNT
    assert signal.x_values[0] == 0.0003333333333333333
    assert signal.x_values[-1] == 44.169666666666664
    assert signal.y_values[:20] == EXPECTED_FIRST_20
    assert signal.y_values[-20:] == EXPECTED_LAST_20
    assert min(signal.y_values) == -395.7000058963895
    assert max(signal.y_values) == 347432.1051771417
    assert _f64_digest(signal.x_values) == EXPECTED_TIME_SHA256
    assert _f64_digest(signal.y_values) == EXPECTED_SIGNAL_SHA256
    assert _pair_digest(signal.x_values, signal.y_values) == EXPECTED_PAIRS_SHA256
    assert _rounded_i64_digest(signal.y_values) == EXPECTED_ASCII_ROUNDED_SIGNAL_SHA256
    assert _ascii_rounded_time_digest(signal.x_values) == EXPECTED_ASCII_ROUNDED_TIME_SHA256
    assert bundle.peaks == ()
    assert bundle.samples[0].acquired_at == datetime(2019, 7, 18, 23, 45, 56, 388_464, tzinfo=UTC)
    assert bundle.samples[0].acquired_at_reliable is True
    metadata = {entry.key: entry.value for entry in bundle.metadata}
    assert metadata["software_version"] == "5.82"
    assert metadata["point_count"] == EXPECTED_POINT_COUNT
    assert metadata["sampling_interval"] == 40
    assert metadata["initial_delay"] == 20.0
    assert metadata["axis_value_factor"] == 1.0
    assert metadata["correction_factor"] == 1.0
    assert metadata["gain_factor"] == 1.0
    assert metadata["instrument_model"] == "GC-2014"
    assert metadata["timestamp_status"] == "verified_utc_filetime"
    assert metadata["time_canonical_be_f64_sha256"] == EXPECTED_TIME_SHA256
    assert metadata["signal_canonical_be_f64_sha256"] == EXPECTED_SIGNAL_SHA256
    assert metadata["time_signal_pairs_be_f64_sha256"] == EXPECTED_PAIRS_SHA256

    output = tmp_path / "shimadzu-gcd.xlsx"
    result = convert(source, output, include_signals=True)
    assert result.files[0].source.sha256 == EXPECTED_SHA256
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        sheet = workbook["Signals_FID"]
        assert sheet.max_row == EXPECTED_POINT_COUNT + 1
        workbook_x = (
            row[0]
            for row in sheet.iter_rows(
                min_row=2,
                min_col=5,
                max_col=5,
                values_only=True,
            )
        )
        _assert_excel_numeric_series(workbook_x, signal.x_values)
        workbook_y = (
            row[0]
            for row in sheet.iter_rows(
                min_row=2,
                min_col=8,
                max_col=8,
                values_only=True,
            )
        )
        _assert_excel_numeric_series(workbook_y, signal.y_values)
    finally:
        workbook.close()
