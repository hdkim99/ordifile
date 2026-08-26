from __future__ import annotations

import hashlib
import sys
from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.acquisition.base import (
    ACQUISITION_PROVIDER_API_VERSION,
    AcquiredResultArtifact,
    AcquisitionAvailability,
    AcquisitionEnvironment,
    AcquisitionRequest,
    AcquisitionSource,
    ResultAcquisitionProviderDescriptor,
)
from ordifile.acquisition.coordinator import acquire_official_result
from ordifile.acquisition.registry import ResultAcquisitionRegistry
from ordifile.adapters.registry import create_registry
from ordifile.core.models import (
    BatchResult,
    DatasetBundle,
    FileResult,
    FileStatus,
    InstrumentMetadata,
    ResultAcquisitionMode,
    SampleRecord,
    SeriesKind,
    SignalSeries,
    SortDecision,
    SortMode,
    SourceFile,
)
from ordifile.exporters.excel import ExcelExporter

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_youngin_yl_clarity_result_csv import (  # noqa: E402
    synthetic_result_csv_bytes,
)


class _OfficialResultProvider:
    api_version: ClassVar[str] = ACQUISITION_PROVIDER_API_VERSION
    provider_id: ClassVar[str] = "synthetic_official_result"
    provider_version: ClassVar[str] = "1.0"
    descriptor: ClassVar[ResultAcquisitionProviderDescriptor] = ResultAcquisitionProviderDescriptor(
        provider_id,
        provider_version,
        "Synthetic official Result provider",
        ("youngin_yl_clarity_prm_raw",),
        "youngin_yl_clarity_result_csv",
    )

    def inspect_environment(self) -> AcquisitionEnvironment:
        return AcquisitionEnvironment(
            AcquisitionAvailability.AVAILABLE,
            "Synthetic YL-Clarity",
            "9.0.1.19",
        )

    def acquire(self, request: AcquisitionRequest, workspace: Path) -> AcquiredResultArtifact:
        assert request.staged_source.parent == workspace
        data = synthetic_result_csv_bytes()
        output = workspace / "private-official-result.csv"
        output.write_bytes(data)
        return AcquiredResultArtifact(
            output,
            self.descriptor.result_adapter_id,
            hashlib.sha256(data).hexdigest(),
            len(data),
        )


def _native_source(path: Path) -> tuple[SourceFile, DatasetBundle]:
    data = b"independently invented native chromatogram"
    path.write_bytes(data)
    digest = hashlib.sha256(data).hexdigest()
    public_reference = f"source-{digest}"
    source = SourceFile(
        path=path,
        relative_path=path.name,
        name=path.name,
        size=len(data),
        sha256=digest,
        modified_at=None,
        input_order=0,
        detected_format="youngin_yl_clarity_prm_raw",
        public_id=public_reference,
    )
    sample = SampleRecord(
        "SYNTHETIC_NATIVE",
        source,
        instrument=InstrumentMetadata("GC", "YoungIn"),
        channels=("Signal 1: TCD",),
        detectors=("TCD",),
    )
    signal = SignalSeries(
        sample.sample_id,
        public_reference,
        "Signal 1: TCD",
        "TCD",
        (0.0, 1.0 / 600.0),
        (1.0, 2.0),
        "retention_time",
        "min",
        "detector_response",
        "mV",
        SeriesKind.SCIENTIFIC_SIGNAL,
    )
    return source, DatasetBundle((source,), (sample,), (signal,))


def test_official_result_is_one_logical_workbook_source(tmp_path: Path) -> None:
    source, native = _native_source(tmp_path / "private-native.prm")
    providers = ResultAcquisitionRegistry()
    providers.register(_OfficialResultProvider())

    outcome = acquire_official_result(
        AcquisitionSource(
            source.path,
            source.public_reference,
            source.sha256 or "",
            source.size,
            "youngin_yl_clarity_prm_raw",
            "0.3.0",
        ),
        native,
        mode=ResultAcquisitionMode.AUTO,
        providers=providers,
        adapters=create_registry(include_external=False),
    )
    file_result = FileResult(
        source,
        FileStatus.SUCCESS,
        "youngin_yl_clarity_prm_raw",
        "0.3.0",
        outcome.bundle,
        outcome.issues,
        acquisitions=(outcome.record,),
    )
    batch = BatchResult(
        (file_result,),
        SortDecision(SortMode.INPUT_ORDER, SortMode.INPUT_ORDER, "test"),
    )
    output = tmp_path / "logical-source.xlsx"

    ExcelExporter().export(batch, output, include_signals=True)

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Samples"].max_row == 2
        assert workbook["Signals_TCD"].max_row == 3
        assert workbook["Peaks"].max_row == 3
        assert workbook["Import_Log"].max_row == 2
        assert {row[0] for row in workbook["Peaks"].iter_rows(min_row=2, values_only=True)} == {
            "SYNTHETIC_NATIVE"
        }
        assert {row[1] for row in workbook["Peaks"].iter_rows(min_row=2, values_only=True)} == {
            source.public_reference
        }
        headers = tuple(next(workbook["Import_Log"].values))
        row = tuple(next(workbook["Import_Log"].iter_rows(min_row=2, values_only=True)))
        values = dict(zip(headers, row, strict=True))
        assert values["result_acquisition_mode"] == "auto"
        assert values["result_acquisition_status"] == "success"
        assert values["result_acquisition_provider"] == "synthetic_official_result"
        assert values["result_acquisition_provider_version"] == "1.0"
        assert values["result_acquisition_adapter"] == "youngin_yl_clarity_result_csv"
        assert values["result_acquisition_adapter_version"] is not None
        assert values["result_acquisition_sha256"] is not None
        assert values["result_acquisition_peak_count"] == 2
        rendered = "\n".join(
            str(value)
            for sheet in workbook.worksheets
            for candidate in sheet.iter_rows(values_only=True)
            for value in candidate
            if value is not None
        )
        assert "private-official-result" not in rendered
    finally:
        workbook.close()
