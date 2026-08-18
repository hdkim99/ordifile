from __future__ import annotations

import hashlib
import os
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert

EXPECTED_IDENTITIES = {
    "ORDIFILE_AGILENT_RESULT_XML_FIXTURE": (
        98_084,
        "4c876bb5712b2d943b5ad32ce5854698018e0b82f2dfc10cc0971ffab9a7056f",
    ),
    "ORDIFILE_SHIMADZU_RESULT_ASCII_FIXTURE": (
        971_258,
        "46d1dcde188d7844c32abb89cda1f0d773cac480f6d6c93f2b6ca7149fdb9297",
    ),
    "ORDIFILE_YOUNGIN_RESULT_CSV_A_FIXTURE": (
        1_009,
        "d4987151cba83068b5143cf90c6b0e78f1fee6b3c9f04c38d8cb97441ddfadd7",
    ),
    "ORDIFILE_YOUNGIN_RESULT_CSV_B_FIXTURE": (
        1_802,
        "0ceb70ba51e41607a5a6ca4476c9b77e6e2bce41d56e47b865085bb3ea71f67b",
    ),
}


def _sources() -> tuple[Path, ...]:
    sources: list[Path] = []
    for variable, (size, sha256) in EXPECTED_IDENTITIES.items():
        value = os.environ.get(variable)
        if not value:
            raise AssertionError(f"{variable} is required")
        source = Path(value)
        data = source.read_bytes()
        if len(data) != size or hashlib.sha256(data).hexdigest() != sha256:
            raise AssertionError("external three-vendor fixture identity changed")
        sources.append(source)
    return tuple(sources)


def test_three_vendor_actual_results_produce_125_peaks_and_four_streams(
    tmp_path: Path,
) -> None:
    sources = _sources()
    before = tuple(hashlib.sha256(source.read_bytes()).hexdigest() for source in sources)
    output = tmp_path / "three-vendor-actual.xlsx"

    result = convert(sources, output, sort="input_order")

    if result.failure_count != 0 or result.success_count != 4:
        raise AssertionError("an actual Result source failed three-vendor conversion")
    if sum(len(item.bundle.peaks) for item in result.files if item.bundle is not None) != 125:
        raise AssertionError("three-vendor canonical peak total changed")
    if tuple(hashlib.sha256(source.read_bytes()).hexdigest() for source in sources) != before:
        raise AssertionError("an external Result source changed during conversion")

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].values)
        headers = peaks[0]
        rows = peaks[1:]
        if len(rows) != 125:
            raise AssertionError("three-vendor workbook peak rows were lost or duplicated")
        manufacturer_index = headers.index("manufacturer")
        counts = {
            manufacturer: sum(row[manufacturer_index] == manufacturer for row in rows)
            for manufacturer in ("Agilent", "Shimadzu", "YoungIn")
        }
        if counts != {"Agilent": 36, "Shimadzu": 83, "YoungIn": 6}:
            raise AssertionError("three-vendor manufacturer peak counts changed")
        order_rows = tuple(workbook["Peak_Order_Matrix"].iter_rows(min_row=2, values_only=True))
        if len(order_rows) != 4:
            raise AssertionError("three-vendor ordered peak stream count changed")
        units = {(row[2], row[5], row[6]) for row in order_rows}
        if units != {
            ("Agilent", "min", "pA*s"),
            ("Shimadzu", "min", None),
            ("YoungIn", "min", "mV.s"),
        }:
            raise AssertionError("three-vendor area-unit provenance changed")
        matrix_rows = tuple(workbook["Peak_Matrix"].values)
        if not all(
            all(value is None for value in row[1:])
            for row in matrix_rows[1:]
            if str(row[0]).startswith("YOUNGIN_RESULT_")
        ):
            raise AssertionError("YoungIn compound identity was inferred without source evidence")
    finally:
        workbook.close()
