# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import csv
import hashlib
import runpy
from datetime import datetime
from pathlib import Path
from typing import cast

import openpyxl  # type: ignore[import-untyped]
import pytest

from ordifile.adapters.base import ParseOptions
from ordifile.adapters.generic_csv import GenericCsvAdapter
from ordifile.api import (
    convert,
    inspect_file,
    inspect_inputs,
    list_peak_table_worksheets,
    preview_peak_table,
)
from ordifile.core.discovery import sha256_file
from ordifile.core.errors import OrdifileError, ParseError
from ordifile.core.models import FileStatus
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableImportSettings,
    PeakTableMapping,
    PeakTableTextEncoding,
)

HEADERS = ("Peak No.", "RT result", "Peak area", "Height", "Compound", "Private note")
ROWS = (
    ("8", "1.25", "10", "4", "A", "not exported"),
    ("9", "2.5", "20", "8", "B", "not exported"),
)


def synthetic_bytes(generator_name: str, function_name: str) -> bytes:
    generator = runpy.run_path(
        str(Path(__file__).parents[1] / "fixtures" / "synthetic" / generator_name)
    )
    return cast(bytes, generator[function_name]())


def peak_mapping(source_format: PeakTableFormat) -> PeakTableMapping:
    return PeakTableMapping(
        retention_time_column=ColumnSelector("RT result", 2),
        area_column=ColumnSelector("Peak area", 3),
        retention_time_unit="min",
        source_format=source_format,
        area_unit="mV.s",
        height_column=ColumnSelector("Height", 4),
        height_unit="mV",
        compound_name_column=ColumnSelector("Compound", 5),
        peak_index_column=ColumnSelector("Peak No.", 1),
        ignored_columns=(ColumnSelector("Private note", 6),),
    )


def write_text_table(path: Path, delimiter: str) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, delimiter=delimiter, lineterminator="\n")
        writer.writerow(HEADERS)
        writer.writerows(ROWS)


def write_xlsx_table(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(HEADERS)
    for row in ROWS:
        sheet.append(row)
    workbook.save(path)


@pytest.mark.parametrize(
    ("suffix", "delimiter", "source_format", "adapter_id"),
    (
        (".csv", ",", PeakTableFormat.CSV, "generic_csv"),
        (".tsv", "\t", PeakTableFormat.TSV, "generic_tsv"),
        (".txt", ";", PeakTableFormat.SEMICOLON, "generic_semicolon"),
    ),
)
@pytest.mark.researcher_acceptance
def test_explicit_text_mapping_creates_ordered_peaks(
    tmp_path: Path,
    suffix: str,
    delimiter: str,
    source_format: PeakTableFormat,
    adapter_id: str,
) -> None:
    source = tmp_path / f"private-name{suffix}"
    write_text_table(source, delimiter)

    result = inspect_file(source, peak_table_mapping=peak_mapping(source_format)).file

    assert result.status is FileStatus.SUCCESS
    assert result.adapter_id == adapter_id
    assert result.source.public_reference == f"source-{result.source.sha256}"
    assert result.bundle is not None
    assert tuple((peak.retention_time, peak.area) for peak in result.bundle.peaks) == (
        (1.25, 10.0),
        (2.5, 20.0),
    )
    assert tuple(peak.observation_order for peak in result.bundle.peaks) == (1, 2)
    assert all(peak.status == "user_supplied_mapping" for peak in result.bundle.peaks)
    assert "not exported" not in repr(result.bundle)


@pytest.mark.researcher_acceptance
def test_explicit_xlsx_mapping_reuses_audited_reader(tmp_path: Path) -> None:
    source = tmp_path / "private-name.xlsx"
    write_xlsx_table(source)

    preview = preview_peak_table(source, PeakTableFormat.XLSX, sheet="Results")
    result = inspect_file(
        source,
        sheet="Results",
        peak_table_mapping=peak_mapping(PeakTableFormat.XLSX),
    ).file

    assert preview.headers == HEADERS
    assert len(preview.rows) == 2
    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    assert len(result.bundle.peaks) == 2


@pytest.mark.researcher_acceptance
def test_explicit_header_row_uses_logical_csv_records_and_preserves_values(
    tmp_path: Path,
) -> None:
    source = tmp_path / "preamble.csv"
    source.write_text(
        'Instrument,"Line one\nLine two"\nGenerated,2026-08-23\n'
        "RT result,Peak area,Notes\n1.25E+1,3.4E+06,kept local\n",
        encoding="utf-8",
    )
    settings = PeakTableImportSettings(header_row=3)
    mapping = PeakTableMapping(
        ColumnSelector("RT result", 1),
        ColumnSelector("Peak area", 2),
        "min",
        PeakTableFormat.CSV,
        area_unit="AU",
        ignored_columns=(ColumnSelector("Notes", 3),),
        import_settings=settings,
    )

    preview = preview_peak_table(
        source,
        PeakTableFormat.CSV,
        import_settings=settings,
    )
    result = inspect_file(source, peak_table_mapping=mapping).file

    assert preview.headers == ("RT result", "Peak area", "Notes")
    assert preview.rows[0][:2] == ("1.25E+1", "3.4E+06")
    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    assert tuple((peak.retention_time, peak.area) for peak in result.bundle.peaks) == (
        (12.5, 3_400_000.0),
    )


@pytest.mark.parametrize(
    ("encoding", "header", "note"),
    (
        (PeakTableTextEncoding.CP949, "머무름시간", "한글 메모"),
        (PeakTableTextEncoding.WINDOWS_1252, "Retention", "Café note"),
    ),
)
@pytest.mark.researcher_acceptance
def test_explicit_text_encoding_is_selected_without_fallback(
    tmp_path: Path,
    encoding: PeakTableTextEncoding,
    header: str,
    note: str,
) -> None:
    source = tmp_path / "encoded.csv"
    source.write_bytes(f"{header},Area,Note\n1.5,20,{note}\n".encode(encoding.codec_name))
    settings = PeakTableImportSettings(encoding, 1)
    mapping = PeakTableMapping(
        ColumnSelector(header, 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
        ignored_columns=(ColumnSelector("Note", 3),),
        import_settings=settings,
    )

    preview = preview_peak_table(source, PeakTableFormat.CSV, import_settings=settings)
    result = inspect_file(source, peak_table_mapping=mapping).file

    assert preview.headers[0] == header
    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    assert result.bundle.peaks[0].retention_time == 1.5


def test_wrong_explicit_text_encoding_fails_clearly(tmp_path: Path) -> None:
    source = tmp_path / "encoded.csv"
    source.write_bytes("머무름시간,Area\n1,2\n".encode("cp949"))

    with pytest.raises(OrdifileError) as captured:
        preview_peak_table(source, PeakTableFormat.CSV)

    assert captured.value.code == "TEXT_ENCODING_UNSUPPORTED"
    assert "selected text encoding" in captured.value.message


def test_mapped_numeric_failure_reports_configured_source_row(tmp_path: Path) -> None:
    source = tmp_path / "preamble.csv"
    source.write_text(
        "Instrument,Synthetic\nGenerated,2026-08-23\nRT,Area\ninvalid,20\n",
        encoding="utf-8",
    )
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
        import_settings=PeakTableImportSettings(header_row=3),
    )

    with pytest.raises(ParseError) as captured:
        GenericCsvAdapter().parse(source, ParseOptions(peak_table_mapping=mapping))

    assert captured.value.code == "PEAK_MAPPING_NUMBER_INVALID"
    assert "Row 4" in captured.value.message


@pytest.mark.researcher_acceptance
def test_xlsx_header_row_and_explicit_worksheet_selection(tmp_path: Path) -> None:
    source = tmp_path / "multi-sheet.xlsx"
    workbook = openpyxl.Workbook()
    notes = workbook.active
    notes.title = "Notes"
    notes.append(("Not a result table",))
    results = workbook.create_sheet("Results")
    results.append(("Instrument", "Synthetic"))
    results.append(("Generated", "2026-08-23"))
    results.append(("RT", "Area"))
    results.append((1.25, 42))
    workbook.save(source)
    settings = PeakTableImportSettings(header_row=3)
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.XLSX,
        import_settings=settings,
    )

    assert list_peak_table_worksheets(source) == ("Notes", "Results")
    preview = preview_peak_table(
        source,
        PeakTableFormat.XLSX,
        sheet="Results",
        import_settings=settings,
    )
    result = inspect_file(source, sheet="Results", peak_table_mapping=mapping).file

    assert preview.headers == ("RT", "Area")
    assert preview.rows == (("1.25", "42"),)
    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    assert result.bundle.peaks[0].area == 42.0


def test_mapped_xlsx_private_sheet_title_is_not_persisted(tmp_path: Path) -> None:
    source = tmp_path / "private-source-name.xlsx"
    output = tmp_path / "mapped.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Private Sample Title"
    sheet.append(HEADERS)
    for row in ROWS:
        sheet.append(row)
    workbook.save(source)

    result = convert(
        source,
        output,
        sheet="Private Sample Title",
        peak_table_mapping=peak_mapping(PeakTableFormat.XLSX),
    )
    reopened = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        all_strings = {
            value
            for worksheet in reopened.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for value in row
            if isinstance(value, str)
        }
        manifest = {row[0]: row[1] for row in reopened["Manifest"].iter_rows(values_only=True)}
    finally:
        reopened.close()

    assert result.options.sheet == "USER_SELECTED"
    assert manifest["option_sheet"] == "USER_SELECTED"
    assert "Private Sample Title" not in all_strings
    assert source.name not in all_strings
    assert str(source) not in all_strings
    assert "not exported" not in all_strings


def test_text_mapping_rejects_xlsx_only_sheet_option(tmp_path: Path) -> None:
    source = tmp_path / "result.csv"
    write_text_table(source, ",")

    with pytest.raises(OrdifileError) as captured:
        inspect_file(
            source,
            sheet="Results",
            peak_table_mapping=peak_mapping(PeakTableFormat.CSV),
        )

    assert captured.value.code == "PEAK_MAPPING_SHEET_INVALID"


def test_delimited_preview_rejects_excess_columns_and_unsafe_header(tmp_path: Path) -> None:
    too_wide = tmp_path / "wide.csv"
    too_wide.write_text(",".join(f"c{index}" for index in range(1_025)) + "\n", encoding="utf-8")
    unsafe = tmp_path / "unsafe.csv"
    unsafe.write_text("RT\u202e,Area\n1,2\n", encoding="utf-8")

    with pytest.raises(OrdifileError) as wide_error:
        preview_peak_table(too_wide, PeakTableFormat.CSV)
    with pytest.raises(OrdifileError) as unsafe_error:
        preview_peak_table(unsafe, PeakTableFormat.CSV)

    assert wide_error.value.code == "PEAK_MAPPING_PREVIEW_COLUMN_LIMIT"
    assert unsafe_error.value.code == "PEAK_MAPPING_PREVIEW_HEADER_INVALID"


def test_delimited_preview_escapes_directional_row_text(tmp_path: Path) -> None:
    source = tmp_path / "preview.csv"
    source.write_text("RT,Area,Note\n1,2,left\u202eright\n", encoding="utf-8")

    preview = preview_peak_table(source, PeakTableFormat.CSV)

    assert preview.rows[0][2] == "left\\u202Eright"


def test_peak_table_preview_records_the_exact_source_snapshot(tmp_path: Path) -> None:
    source = tmp_path / "preview.csv"
    source.write_bytes(b"RT,Area\n1,2\n")

    preview = preview_peak_table(source, PeakTableFormat.CSV)

    assert preview.source_sha256 == hashlib.sha256(source.read_bytes()).hexdigest()


def test_peak_table_preview_fails_if_source_changes_during_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "preview.csv"
    source.write_bytes(b"RT,Area\n1,2\n")
    from ordifile import api

    real_hash = sha256_file

    def hash_then_change(path: Path) -> str:
        digest = real_hash(path)
        path.write_bytes(b"RT,Area\n3,4\n")
        return digest

    monkeypatch.setattr(api, "sha256_file", hash_then_change)

    with pytest.raises(OrdifileError) as captured:
        preview_peak_table(source, PeakTableFormat.CSV)

    assert captured.value.code == "PEAK_MAPPING_PREVIEW_SOURCE_CHANGED"


def test_delimited_preview_rejects_oversized_line(tmp_path: Path) -> None:
    source = tmp_path / "long.csv"
    source.write_text("RT,Area,Note\n1,2," + "x" * (256 * 1024) + "\n", encoding="utf-8")

    with pytest.raises(OrdifileError) as captured:
        preview_peak_table(source, PeakTableFormat.CSV)

    assert captured.value.code == "PEAK_MAPPING_PREVIEW_SIZE_LIMIT"


def test_xlsx_preview_rejects_excess_columns_and_non_text_header(tmp_path: Path) -> None:
    too_wide = tmp_path / "wide.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(tuple(f"c{index}" for index in range(1_025)))
    workbook.save(too_wide)
    numeric = tmp_path / "numeric.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append((1, "Area"))
    workbook.active.append((1, 2))
    workbook.save(numeric)

    with pytest.raises(OrdifileError) as wide_error:
        preview_peak_table(too_wide, PeakTableFormat.XLSX)
    with pytest.raises(OrdifileError) as header_error:
        preview_peak_table(numeric, PeakTableFormat.XLSX)

    assert wide_error.value.code == "PEAK_MAPPING_PREVIEW_COLUMN_LIMIT"
    assert header_error.value.code == "PEAK_MAPPING_HEADER_TYPE_UNSUPPORTED"


def test_xlsx_preview_rejects_header_only_total_text_over_budget(tmp_path: Path) -> None:
    source = tmp_path / "large-header.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(tuple(f"column-{index:04d}-" + "x" * 990 for index in range(1_024)))
    workbook.save(source)

    with pytest.raises(OrdifileError) as captured:
        preview_peak_table(source, PeakTableFormat.XLSX)

    assert captured.value.code == "PEAK_MAPPING_PREVIEW_SIZE_LIMIT"


@pytest.mark.parametrize(
    ("field", "value", "expected_code"),
    (
        ("RT result", "", "PEAK_MAPPING_VALUE_MISSING"),
        ("RT result", "NaN", "PEAK_MAPPING_NUMBER_NONFINITE"),
        ("RT result", "Inf", "PEAK_MAPPING_NUMBER_NONFINITE"),
        ("RT result", "not-a-number", "PEAK_MAPPING_NUMBER_INVALID"),
        ("Peak area", "", "PEAK_MAPPING_VALUE_MISSING"),
        ("Peak area", "NaN", "PEAK_MAPPING_NUMBER_NONFINITE"),
        ("Peak area", "-Inf", "PEAK_MAPPING_NUMBER_NONFINITE"),
        ("Peak area", "not-a-number", "PEAK_MAPPING_NUMBER_INVALID"),
    ),
)
def test_mapping_invalid_required_number_fails_whole_file(
    tmp_path: Path,
    field: str,
    value: str,
    expected_code: str,
) -> None:
    source = tmp_path / "bad.csv"
    rows = [list(ROWS[0]), list(ROWS[1])]
    rows[0][HEADERS.index(field)] = value
    with source.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(HEADERS)
        writer.writerows(rows)

    result = inspect_file(source, peak_table_mapping=peak_mapping(PeakTableFormat.CSV)).file

    assert result.status is FileStatus.FAILED
    assert result.bundle is None
    assert {issue.code for issue in result.issues} == {expected_code}


def test_optional_fields_and_user_provenance_are_preserved_without_verification(
    tmp_path: Path,
) -> None:
    source = tmp_path / "declared.csv"
    source.write_text(
        "RT,Area,Peak Name,Compound,Detector,Channel,Sample,Run,Acquired\n"
        "1.25,42,Peak A,Compound A,FID,Signal A,Sample A,Run A,2026-08-20T00:00:00Z\n",
        encoding="utf-8",
    )
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
        peak_name_column=ColumnSelector("Peak Name", 3),
        compound_name_column=ColumnSelector("Compound", 4),
        detector_column=ColumnSelector("Detector", 5),
        channel_column=ColumnSelector("Channel", 6),
        sample_id_column=ColumnSelector("Sample", 7),
        run_id_column=ColumnSelector("Run", 8),
        acquisition_time_column=ColumnSelector("Acquired", 9),
        manufacturer="User Manufacturer",
        software="User Software",
    )

    result = inspect_file(source, peak_table_mapping=mapping).file

    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    sample = result.bundle.samples[0]
    peak = result.bundle.peaks[0]
    metadata = {(entry.key, entry.value) for entry in result.bundle.metadata}
    assert sample.sample_id == "Sample A"
    assert sample.instrument.vendor == "User Manufacturer"
    assert sample.acquired_at is not None
    assert peak.compound == "Compound A"
    assert (peak.detector, peak.channel) == ("FID", "Signal A")
    assert ("run_id", "Run A") in metadata
    assert ("software", "User Software") in metadata
    assert ("manufacturer_verification_status", "USER_SUPPLIED") in metadata
    assert ("software_verification_status", "USER_SUPPLIED") in metadata
    assert any(
        entry.key.endswith("peak_name") and entry.value == "Peak A"
        for entry in result.bundle.metadata
    )


def test_mapping_mismatch_isolated_in_batch(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "bad.csv"
    write_text_table(good, ",")
    bad.write_text("RT result,wrong\n1,2\n", encoding="utf-8")

    result = inspect_inputs((good, bad), peak_table_mapping=peak_mapping(PeakTableFormat.CSV))

    assert result.success_count == 1
    assert result.failure_count == 1


def test_mapped_workbook_contains_pairs_and_provenance(tmp_path: Path) -> None:
    source = tmp_path / "result.csv"
    output = tmp_path / "mapped.xlsx"
    write_text_table(source, ",")
    mapping = peak_mapping(PeakTableFormat.CSV)

    result = convert(source, output, peak_table_mapping=mapping)
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        manifest = {row[0]: row[1] for row in workbook["Manifest"].iter_rows(values_only=True)}
        metadata_values = {
            row[4] for row in workbook["Metadata"].iter_rows(min_row=2, values_only=True)
        }
        order_rows = tuple(workbook["Peak_Order_Matrix"].iter_rows(values_only=True))
    finally:
        workbook.close()

    assert result.success_count == 1
    assert manifest["option_peak_table_mapping_sha256"] == mapping.semantic_sha256
    assert "USER_SUPPLIED" in metadata_values
    assert order_rows[1][-4:] == (1.25, 10, 2.5, 20)


def test_secondary_rt_and_interleaved_stream_order_are_explicit(tmp_path: Path) -> None:
    source = tmp_path / "two-dimensional.tsv"
    source.write_text(
        "RT1\tRT2\tArea\tDetector\n10\t1\t100\tA\n20\t2\t200\tB\n30\t3\t300\tA\n",
        encoding="utf-8",
    )
    mapping = PeakTableMapping(
        ColumnSelector("RT1", 1),
        ColumnSelector("Area", 3),
        "s",
        PeakTableFormat.TSV,
        detector_column=ColumnSelector("Detector", 4),
        secondary_retention_time_column=ColumnSelector("RT2", 2),
        secondary_retention_time_unit="s",
    )

    output = tmp_path / "two-dimensional.xlsx"
    batch = convert(source, output, peak_table_mapping=mapping)
    result = batch.files[0]
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        peaks_rows = tuple(workbook["Peaks"].iter_rows(values_only=True))
        matrix_rows = tuple(workbook["Peak_Order_Matrix_2D"].iter_rows(values_only=True))
    finally:
        workbook.close()

    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    assert tuple(peak.secondary_retention_time for peak in result.bundle.peaks) == (1.0, 2.0, 3.0)
    assert tuple(peak.observation_order for peak in result.bundle.peaks) == (1, 1, 2)
    secondary_index = peaks_rows[0].index("secondary_retention_time")
    assert tuple(row[secondary_index] for row in peaks_rows[1:]) == (1, 2, 3)
    matrix_by_detector = {row[3]: row for row in matrix_rows[1:]}
    assert matrix_by_detector["A"][8:14] == (10, 1, 100, 30, 3, 300)
    assert matrix_by_detector["B"][8:11] == (20, 2, 200)


def test_same_mapping_batches_three_files_into_one_workbook(tmp_path: Path) -> None:
    sources = tuple(tmp_path / f"run-{index}.csv" for index in range(3))
    for source in sources:
        write_text_table(source, ",")
    output = tmp_path / "batch.xlsx"

    result = convert(sources, output, peak_table_mapping=peak_mapping(PeakTableFormat.CSV))
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        peak_rows = workbook["Peaks"].max_row - 1
        order_streams = workbook["Peak_Order_Matrix"].max_row - 1
    finally:
        workbook.close()

    assert result.success_count == 3
    assert peak_rows == 6
    assert order_streams == 3


def test_mapped_xlsx_formula_is_not_used_as_scientific_value(tmp_path: Path) -> None:
    source = tmp_path / "formula.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(("RT", "Area"))
    sheet.append((1.0, "=20*2"))
    workbook.save(source)
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.XLSX,
    )

    result = inspect_file(source, sheet="Results", peak_table_mapping=mapping).file

    assert result.status is FileStatus.FAILED
    assert {issue.code for issue in result.issues} == {"PEAK_MAPPING_FORMULA_UNSUPPORTED"}


@pytest.mark.parametrize(
    ("header", "value", "mapping_field"),
    (
        ("Height", True, "height_column"),
        ("Detector", 7, "detector_column"),
        ("Peak Name", 9, "peak_name_column"),
        ("Height", datetime(2026, 8, 20), "height_column"),
    ),
)
def test_mapped_xlsx_rejects_incompatible_optional_cell_types(
    tmp_path: Path,
    header: str,
    value: object,
    mapping_field: str,
) -> None:
    source = tmp_path / "typed.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(("RT", "Area", header))
    sheet.append((1.0, 2.0, value))
    workbook.save(source)
    optional_selector = ColumnSelector(header, 3)
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.XLSX,
        height_column=(optional_selector if mapping_field == "height_column" else None),
        detector_column=(optional_selector if mapping_field == "detector_column" else None),
        peak_name_column=(optional_selector if mapping_field == "peak_name_column" else None),
    )

    result = inspect_file(source, sheet="Results", peak_table_mapping=mapping).file

    assert result.status is FileStatus.FAILED
    assert {issue.code for issue in result.issues} == {"PEAK_MAPPING_XLSX_CELL_TYPE_UNSUPPORTED"}


def test_mapped_xlsx_rejects_formula_beyond_header(tmp_path: Path) -> None:
    source = tmp_path / "formula-extra.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(("RT", "Area"))
    sheet.append((1.0, 2.0, "=1+1"))
    workbook.save(source)
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.XLSX,
    )

    result = inspect_file(source, sheet="Results", peak_table_mapping=mapping).file

    assert result.status is FileStatus.FAILED
    assert {issue.code for issue in result.issues} == {"PEAK_MAPPING_FORMULA_UNSUPPORTED"}


def test_mapped_xlsx_rejects_non_text_header(tmp_path: Path) -> None:
    source = tmp_path / "numeric-header.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append((1, "Area"))
    sheet.append((1.0, 2.0))
    workbook.save(source)
    mapping = PeakTableMapping(
        ColumnSelector("1", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.XLSX,
    )

    result = inspect_file(source, sheet="Results", peak_table_mapping=mapping).file

    assert result.status is FileStatus.FAILED
    assert {issue.code for issue in result.issues} == {"PEAK_MAPPING_HEADER_TYPE_UNSUPPORTED"}


def test_exact_vendor_owner_is_not_bypassed_by_generic_mapping(tmp_path: Path) -> None:
    source = tmp_path / "private-result.txt"
    source.write_bytes(
        synthetic_bytes(
            "generate_leco_chromatof_472_gcgc_result_txt.py",
            "synthetic_gcgc_result_bytes",
        )
    )
    unrelated_mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "s",
        PeakTableFormat.TSV,
        import_settings=PeakTableImportSettings(
            PeakTableTextEncoding.CP949,
            header_row=6,
        ),
    )

    result = inspect_file(source, peak_table_mapping=unrelated_mapping).file

    assert result.adapter_id == "leco_chromatof_gcxgc_result_txt"
    assert result.status is FileStatus.WARNING
    assert result.bundle is not None
    assert len(result.bundle.peaks) == 3
    assert "PEAK_MAPPING_NOT_APPLIED_EXACT_PROFILE" in {issue.code for issue in result.issues}


def test_four_exact_profiles_and_one_mapped_table_share_one_workbook(tmp_path: Path) -> None:
    sources = (
        (
            "private-agilent.xml",
            "generate_agilent_chemstation_result_xml.py",
            "synthetic_result_xml_bytes",
        ),
        (
            "private-shimadzu.txt",
            "generate_shimadzu_labsolutions_result_ascii.py",
            "synthetic_result_ascii_bytes",
        ),
        (
            "private-youngin.csv",
            "generate_youngin_yl_clarity_result_csv.py",
            "synthetic_result_csv_bytes",
        ),
        (
            "private-leco.txt",
            "generate_leco_chromatof_472_gcgc_result_txt.py",
            "synthetic_gcgc_result_bytes",
        ),
    )
    paths: list[Path] = []
    for filename, generator_name, function_name in sources:
        path = tmp_path / filename
        path.write_bytes(synthetic_bytes(generator_name, function_name))
        paths.append(path)
    mapped = tmp_path / "private-mapped.csv"
    write_text_table(mapped, ",")
    paths.append(mapped)
    output = tmp_path / "mixed.xlsx"

    result = convert(
        tuple(paths),
        output,
        sort="input_order",
        peak_table_mapping=peak_mapping(PeakTableFormat.CSV),
    )
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 14
        assert workbook["Peak_Order_Matrix"].max_row == 5
        assert workbook["Peak_Order_Matrix_2D"].max_row == 2
        assert workbook["Import_Log"].max_row == 6
    finally:
        workbook.close()

    assert result.failure_count == 0
    assert {item.adapter_id for item in result.files} == {
        "agilent_chemstation_result_xml",
        "shimadzu_labsolutions_result_ascii",
        "youngin_yl_clarity_result_csv",
        "leco_chromatof_gcxgc_result_txt",
        "generic_csv",
    }


def test_utf16_headerless_source_is_previewed_and_converted(tmp_path: Path) -> None:
    """A UTF-16 report table without a header record keeps every data row."""
    source = tmp_path / "report.csv"
    source.write_text('1,2.5,"BB",6.15\r\n2,3.5,"BV",310.5\r\n', encoding="utf-16")
    settings = PeakTableImportSettings(PeakTableTextEncoding.UTF16, 0)

    preview = preview_peak_table(source, PeakTableFormat.CSV, import_settings=settings)

    # Positional labels stand in for the absent header and no data record is consumed.
    assert preview.headers == ("1", "2", "3", "4")
    assert preview.rows[0] == ("1", "2.5", "BB", "6.15")

    mapping = PeakTableMapping(
        retention_time_column=ColumnSelector("2", 2),
        area_column=ColumnSelector("4", 4),
        retention_time_unit="min",
        ignored_columns=(ColumnSelector("1", 1), ColumnSelector("3", 3)),
        source_format=PeakTableFormat.CSV,
        import_settings=settings,
    )
    bundle = GenericCsvAdapter().parse(source, ParseOptions(peak_table_mapping=mapping))

    assert len(bundle.peaks) == 2
    assert [peak.retention_time for peak in bundle.peaks] == [2.5, 3.5]
    assert [peak.area for peak in bundle.peaks] == [6.15, 310.5]


def test_utf16_source_read_as_windows_1252_fails_closed(tmp_path: Path) -> None:
    """The single-unit encodings map every byte, so the mismatch must be refused."""
    source = tmp_path / "report.csv"
    source.write_text("peak,rt\r\n1,2.5\r\n", encoding="utf-16")

    with pytest.raises(ParseError) as caught:
        preview_peak_table(
            source,
            PeakTableFormat.CSV,
            import_settings=PeakTableImportSettings(PeakTableTextEncoding.WINDOWS_1252, 1),
        )

    assert caught.value.code == "PEAK_MAPPING_TEXT_ENCODING_MISMATCH"


def test_headerless_declaration_is_refused_for_worksheets() -> None:
    with pytest.raises(OrdifileError, match="delimited text"):
        PeakTableMapping(
            retention_time_column=ColumnSelector("1", 1),
            area_column=ColumnSelector("2", 2),
            retention_time_unit="min",
            source_format=PeakTableFormat.XLSX,
            import_settings=PeakTableImportSettings(PeakTableTextEncoding.UTF8, 0),
        )
