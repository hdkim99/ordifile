from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.detection import SOURCE_IDENTITY_PROBE_REASON
from ordifile.core.models import FileStatus

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_chemstation_result_xml import (  # noqa: E402
    synthetic_result_xml_bytes,
)
from generate_leco_chromatof_472_gcgc_result_txt import (  # noqa: E402
    synthetic_gcgc_result_bytes,
)
from generate_shimadzu_labsolutions_result_ascii import (  # noqa: E402
    synthetic_result_ascii_bytes,
)
from generate_youngin_yl_clarity_result_csv import (  # noqa: E402
    synthetic_result_csv_bytes,
)


def _workbook_values(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return tuple(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


def test_inspect_uses_private_source_alias_and_exact_adapter(tmp_path: Path) -> None:
    private_name = "private-person-gcxgc-result.txt"
    data = synthetic_gcgc_result_bytes()
    source = tmp_path / private_name
    source.write_bytes(data)
    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.WARNING
    assert inspected.file.adapter_id == "leco_chromatof_gcxgc_result_txt"
    assert inspected.file.source.public_reference == expected_reference
    assert inspected.file.source.path == Path(expected_reference)
    assert inspected.file.bundle is not None
    assert len(inspected.file.bundle.peaks) == 3
    assert all(reason == SOURCE_IDENTITY_PROBE_REASON for _, _, reason in inspected.probes)
    assert private_name not in repr(inspected)


@pytest.mark.researcher_acceptance
def test_result_only_workbook_preserves_dual_rt_area_height_and_extra_metadata(
    tmp_path: Path,
) -> None:
    private_name = "private-result.txt"
    source = tmp_path / private_name
    source.write_bytes(synthetic_gcgc_result_bytes())
    output = tmp_path / "result.xlsx"

    result = convert(source, output)

    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=False, data_only=False)
    try:
        peaks = tuple(workbook["Peaks"].values)
        headers = peaks[0]
        rows = peaks[1:]
        assert len(rows) == 3
        assert [row[headers.index("observation_order")] for row in rows] == [1, 2, 3]
        assert {row[headers.index("peak_number")] for row in rows} == {None}
        assert {row[headers.index("manufacturer")] for row in rows} == {"LECO"}
        assert [row[headers.index("retention_time")] for row in rows] == [120, 180, 240]
        assert [row[headers.index("secondary_retention_time")] for row in rows] == [
            0.45,
            0.875,
            1.25,
        ]
        assert [row[headers.index("area")] for row in rows] == [100_000, 250_000, 400_000]
        assert [row[headers.index("height")] for row in rows] == [5_000, 12_000, 18_000]
        assert {row[headers.index("retention_time_unit")] for row in rows} == {"s"}
        assert {row[headers.index("secondary_retention_time_unit")] for row in rows} == {"s"}
        assert {row[headers.index("area_unit")] for row in rows} == {"AU"}
        assert {row[headers.index("height_unit")] for row in rows} == {"AU"}
        assert "Peak_Order_Matrix" not in workbook.sheetnames
        order = tuple(workbook["Peak_Order_Matrix_2D"].values)
        assert len(order) == 2
        assert order[1][:8] == (
            rows[0][headers.index("sample_id")],
            rows[0][headers.index("source_file")],
            "LECO",
            None,
            None,
            "s",
            "s",
            "AU",
        )
        assert workbook["Peak_Order_Matrix_2D"].freeze_panes == "I2"
        assert workbook["Peak_Order_Matrix_2D"].auto_filter.ref == "A1:H2"
        assert order[1][8:] == (
            120,
            0.45,
            100_000,
            180,
            0.875,
            250_000,
            240,
            1.25,
            400_000,
        )
        metadata = tuple(workbook["Metadata"].values)
        assert any(row[3].endswith("_name") and row[4] == "Synthetic Alpha" for row in metadata[1:])
        assert any(
            row[3].endswith("_spectra") and row[4] == "43:999 58:250" for row in metadata[1:]
        )
        assert private_name not in _workbook_values(output)
    finally:
        workbook.close()


def test_unknown_name_remains_metadata_and_is_not_a_peak_matrix_identity(
    tmp_path: Path,
) -> None:
    source = tmp_path / "unknown.txt"
    source.write_bytes(synthetic_gcgc_result_bytes().replace(b"Synthetic Alpha", b"Unknown", 1))
    output = tmp_path / "unknown.xlsx"

    result = convert(source, output)

    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = tuple(workbook["Peaks"].values)
        compound_index = peaks[0].index("compound")
        assert peaks[1][compound_index] is None
        assert "Unknown" not in tuple(workbook["Peak_Matrix"].values)[0]
        metadata = tuple(workbook["Metadata"].values)
        assert any(row[3].endswith("_name") and row[4] == "Unknown" for row in metadata[1:])
    finally:
        workbook.close()


def test_malformed_family_never_falls_through_to_generic_tsv(tmp_path: Path) -> None:
    data = synthetic_gcgc_result_bytes().replace(b"1st Dimension Time (s)", b"retention_time", 1)
    private_name = "private-malformed.txt"
    source = tmp_path / private_name
    source.write_bytes(data)

    inspected = inspect_file(source)

    probes = {adapter_id: confidence for adapter_id, confidence, _reason in inspected.probes}
    assert probes["generic_tsv"] == 0.99
    assert probes["leco_chromatof_gcxgc_result_txt"] == 0.70
    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.adapter_id == "leco_chromatof_gcxgc_result_txt"
    assert inspected.file.source.public_reference == f"source-{hashlib.sha256(data).hexdigest()}"
    assert private_name not in repr(inspected)


def test_multiple_damaged_family_headers_cannot_expose_name_or_drop_rt2_via_generic(
    tmp_path: Path,
) -> None:
    private_name = "private-multiple-damage.txt"
    data = synthetic_gcgc_result_bytes()
    for old, new in (
        (b"1st Dimension Time (s)", b"retention_time"),
        (b"2nd Dimension Time (s)", b"secondary_time"),
        (b"Spectra", b"signal_pairs"),
        (b"Retention Index", b"index"),
    ):
        data = data.replace(old, new, 1)
    source = tmp_path / private_name
    source.write_bytes(data)

    inspected = inspect_file(source)

    probes = {adapter_id: confidence for adapter_id, confidence, _reason in inspected.probes}
    assert probes["generic_tsv"] == 0.99
    assert probes["leco_chromatof_gcxgc_result_txt"] == 0.70
    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.adapter_id == "leco_chromatof_gcxgc_result_txt"
    assert inspected.file.source.public_reference == f"source-{hashlib.sha256(data).hexdigest()}"
    assert private_name not in repr(inspected)


def test_unrelated_generic_tsv_retains_generic_adapter_and_relative_provenance(
    tmp_path: Path,
) -> None:
    source = tmp_path / "ordinary.txt"
    source.write_text(
        "sample_id\tretention_time\tarea\nordinary\t1\t2\n",
        encoding="utf-8",
    )

    inspected = inspect_file(source)

    assert inspected.file.adapter_id == "generic_tsv"
    assert inspected.file.source.public_id is None
    assert inspected.file.source.public_reference == source.name


def test_five_common_result_headers_do_not_steal_a_generic_tsv(tmp_path: Path) -> None:
    source = tmp_path / "ordinary-five-common-markers.txt"
    source.write_text(
        "sample_id\tretention_time\tName\tArea\tHeight\tSpectra\tRetention Index\n"
        "ordinary\t1\tPeak A\t2\t3\t4:5\t100\n",
        encoding="utf-8",
    )

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.SUCCESS
    assert inspected.file.adapter_id == "generic_tsv"
    assert inspected.file.source.public_id is None
    assert inspected.file.source.public_reference == source.name


def test_overlapping_private_family_markers_fail_as_ambiguous_without_name_leak(
    tmp_path: Path,
) -> None:
    private_name = "private-overlap.txt"
    data = synthetic_gcgc_result_bytes() + b"Application Name\tLabSolutions\r\n"
    source = tmp_path / private_name
    source.write_bytes(data)

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.source.public_reference == f"source-{hashlib.sha256(data).hexdigest()}"
    assert {issue.code for issue in inspected.file.issues} == {"FORMAT_AMBIGUOUS"}
    assert private_name not in repr(inspected)


def test_valid_and_corrupt_leco_result_files_are_isolated(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    valid = synthetic_gcgc_result_bytes()
    (inputs / "private-valid.txt").write_bytes(valid)
    (inputs / "private-corrupt.txt").write_bytes(valid[:-20])
    output = tmp_path / "isolated.xlsx"

    result = convert(inputs, output, extensions=(".txt",))

    assert result.failure_count == 1
    assert sum(item.bundle is not None for item in result.files) == 1
    assert all("private-" not in repr(item) for item in result.files)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 4
        assert workbook["Peak_Order_Matrix_2D"].max_row == 2
        assert not any("private-" in value for value in _workbook_values(output))
    finally:
        workbook.close()


@pytest.mark.researcher_acceptance
def test_four_vendor_synthetic_results_share_one_1d_and_2d_workbook(tmp_path: Path) -> None:
    agilent = tmp_path / "private-agilent.xml"
    agilent.write_bytes(synthetic_result_xml_bytes())
    shimadzu = tmp_path / "private-shimadzu.txt"
    shimadzu.write_bytes(synthetic_result_ascii_bytes())
    youngin = tmp_path / "private-youngin.csv"
    youngin.write_bytes(synthetic_result_csv_bytes())
    leco = tmp_path / "private-leco.txt"
    leco.write_bytes(synthetic_gcgc_result_bytes())
    output = tmp_path / "four-vendor.xlsx"

    result = convert((agilent, shimadzu, youngin, leco), output, sort="input_order")

    assert result.failure_count == 0
    assert {item.adapter_id for item in result.files} == {
        "agilent_chemstation_result_xml",
        "shimadzu_labsolutions_result_ascii",
        "youngin_yl_clarity_result_csv",
        "leco_chromatof_gcxgc_result_txt",
    }
    assert sum(len(item.bundle.peaks) for item in result.files if item.bundle is not None) == 11
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = tuple(workbook["Peaks"].values)
        manufacturer = peaks[0].index("manufacturer")
        assert {row[manufacturer] for row in peaks[1:]} == {
            "Agilent",
            "Shimadzu",
            "YoungIn",
            "LECO",
        }
        assert workbook["Peak_Order_Matrix"].max_row == 4
        assert workbook["Peak_Order_Matrix_2D"].max_row == 2
        assert workbook["Metadata"].max_row > 1
        assert workbook["Import_Log"].max_row == 5
    finally:
        workbook.close()
