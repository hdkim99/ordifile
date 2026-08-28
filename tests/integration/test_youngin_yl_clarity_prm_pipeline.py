from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file, plan_conversion
from ordifile.cli.main import main

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_youngin_yl_clarity_prm import synthetic_prm_bytes  # noqa: E402

MARKER_START = 0x20
MARKER_APEX = 0x50
MARKER_END = 0x80


def test_preflight_explains_when_calculated_area_is_not_selected(tmp_path: Path) -> None:
    source = tmp_path / "source.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            marker_records=(((MARKER_START, 0), (MARKER_APEX, 1), (MARKER_END, 2)),)
        )
    )

    normal = plan_conversion(source, tmp_path / "normal.xlsx")
    requested = plan_conversion(
        source,
        tmp_path / "requested.xlsx",
        experimental_derived_area=True,
    )

    assert "YOUNGIN_PRM_CALCULATED_AREA_NOT_REQUESTED" in normal.entries[0].issue_codes
    assert "YOUNGIN_PRM_EXPERIMENTAL_DERIVED_AREA_REQUESTED" in (requested.entries[0].issue_codes)


def test_preflight_does_not_offer_calculated_area_for_compatible_unknown_profile(
    tmp_path: Path,
) -> None:
    source = tmp_path / "compatible.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.2.0.0 FULL, SN: SYNTHETIC",
            marker_records=(((MARKER_START, 0), (MARKER_APEX, 1), (MARKER_END, 2)),),
        )
    )

    plan = plan_conversion(
        source,
        tmp_path / "compatible.xlsx",
        experimental_derived_area=True,
    )

    assert "YOUNGIN_PRM_DERIVED_AREA_PROFILE_UNAVAILABLE" in plan.entries[0].issue_codes
    assert "YOUNGIN_PRM_EXPERIMENTAL_DERIVED_AREA_REQUESTED" not in (plan.entries[0].issue_codes)


def test_inspect_and_cli_report_validated_9_0_scientific_signal(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "FID_STD_001.prm"
    source.write_bytes(synthetic_prm_bytes())

    inspected = inspect_file(source)
    assert inspected.file.adapter_id == "youngin_yl_clarity_prm_raw"
    assert inspected.file.source.detected_format == "youngin_yl_clarity_prm_raw"
    assert {issue.code for issue in inspected.file.issues} >= {
        "YOUNGIN_PRM_EXPERIMENTAL_SCIENTIFIC_SIGNAL"
    }

    assert main(["inspect", str(source)]) == 0
    output = capsys.readouterr().out
    assert "Scientific signals: 1" in output
    assert "Decoded record series: 0" in output
    assert "PRM profile: YL-Clarity 9.0.1.19 observed PRM scientific-signal profile" in output
    assert "Producer version: YL-Clarity 9.0.1.19" in output
    assert "Family: YL-Clarity PRM scientific family" in output
    assert "Compatibility: validated profile" in output
    assert "Channels: 1" in output
    assert "Scientific signal available: Yes" in output
    assert "Retention-time unit: min" in output
    assert "Signal units: TCD=mV" in output
    assert "Peak Result availability: unsupported" in output


def test_inspect_and_cli_report_validated_9_1_scientific_signals(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "scientific.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.1.0.76 FULL, SN: SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )

    inspected = inspect_file(source)
    assert inspected.file.adapter_id == "youngin_yl_clarity_prm_raw"
    assert {issue.code for issue in inspected.file.issues} >= {
        "YOUNGIN_PRM_EXPERIMENTAL_SCIENTIFIC_SIGNAL"
    }

    assert main(["inspect", str(source)]) == 0
    output = capsys.readouterr().out
    assert "Scientific signals: 2" in output
    assert "Decoded record series: 0" in output
    assert "PRM profile: YL-Clarity 9.1.0.76 observed PRM scientific-signal profile" in output
    assert "Producer version: YL-Clarity 9.1.0.76" in output
    assert "Compatibility: validated profile" in output
    assert "Channels: 2" in output
    assert "Scientific signal available: Yes" in output
    assert "Retention-time unit: min" in output
    assert "Signal units: FID=pA, TCD=mV" in output
    assert "Scientific signal points: 4" in output
    assert "Peak Result availability: unsupported" in output


def test_inspect_reports_compatible_unknown_signal_with_unresolved_units(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "compatible.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.2.0.0 FULL, SN: PRIVATE-SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )

    assert main(["inspect", str(source)]) == 0
    output = capsys.readouterr().out
    assert "Producer version: YL-Clarity 9.2.0.0" in output
    assert "Family: YL-Clarity PRM scientific family" in output
    assert "Compatibility: compatible unvalidated producer" in output
    assert "Retention-time unit: min" in output
    assert "Signal units: FID=unresolved, TCD=unresolved" in output
    assert "PRIVATE-SYNTHETIC" not in output


def test_valid_and_corrupt_prm_batch_isolated_and_workbook_reopens(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    first_data = synthetic_prm_bytes(channels=((1.0, 2.0, 3.0),))
    second_data = synthetic_prm_bytes(channels=((10.0, 20.0),))
    third_data = synthetic_prm_bytes(channels=((5.0,), (6.0, 7.0)))
    (inputs / "FID_STD_001.prm").write_bytes(first_data)
    (inputs / "TCD_STD_001.prm").write_bytes(second_data)
    (inputs / "MIXED_SAMPLE_001.prm").write_bytes(third_data)
    (inputs / "FID_STD_002.prm").write_bytes(synthetic_prm_bytes()[:-21])
    output = tmp_path / "result.xlsx"

    result = convert(inputs, output, extensions=(".prm",), include_signals=True)

    assert result.success_count == 3
    assert result.failure_count == 1
    assert output.is_file()
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        fid_record_rows = list(workbook["Signals_Records_native_label_FI"].values)
        tcd_record_rows = list(workbook["Signals_Records_native_label_TC"].values)
        tcd_scientific_rows = list(workbook["Signals_TCD"].values)
        assert fid_record_rows[0][-1] == "series_kind"
        assert len(fid_record_rows) - 1 == 1
        assert len(tcd_record_rows) - 1 == 2
        assert len(tcd_scientific_rows) - 1 == 5
        assert [row[7] for row in fid_record_rows[1:]] == [5.0]
        assert {row[-1] for row in fid_record_rows[1:] + tcd_record_rows[1:]} == {"decoded_records"}
        assert {row[9] for row in tcd_scientific_rows[1:]} == {"mV"}
        samples = list(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        assert {row[1] for row in samples if row[12] != "failed"} == {
            f"PRM_{hashlib.sha256(first_data).hexdigest()[:16]}",
            f"PRM_{hashlib.sha256(second_data).hexdigest()[:16]}",
            f"PRM_{hashlib.sha256(third_data).hexdigest()[:16]}",
        }
        metadata = list(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        assert not any(row[3] == "user_supplied_group" for row in metadata)
        import_log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in import_log} == {"warning", "failed"}
        assert all(row[9] for row in import_log)
    finally:
        workbook.close()


def test_workbook_does_not_expose_ignored_embedded_metadata(tmp_path: Path) -> None:
    source = tmp_path / "FID_STD_001.prm"
    private_marker = "PRIVATE_OPERATOR_AND_LOCAL_PATH"
    source.write_bytes(synthetic_prm_bytes(embedded_private_text=private_marker.encode("utf-8")))
    output = tmp_path / "privacy.xlsx"

    convert(source, output, include_signals=True)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert not any(
            private_marker in str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


def test_validated_9_1_profile_writes_scientific_signal_sheets(tmp_path: Path) -> None:
    source = tmp_path / "scientific.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.1.0.76 FULL, SN: SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )
    output = tmp_path / "scientific.xlsx"

    result = convert(source, output, include_signals=True)

    assert result.success_count == 1
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Signals_FID" in workbook.sheetnames
        assert "Signals_TCD" in workbook.sheetnames
        assert not any(name.startswith("Signals_Records_") for name in workbook.sheetnames)
        fid_rows = list(workbook["Signals_FID"].values)
        tcd_rows = list(workbook["Signals_TCD"].values)
        assert fid_rows[0] == (
            "sample_id",
            "source_file",
            "channel",
            "detector",
            "x",
            "x_label",
            "x_unit",
            "y",
            "y_label",
            "y_unit",
        )
        assert {row[2] for row in fid_rows[1:]} == {"native_label_FID"}
        assert {row[3] for row in fid_rows[1:]} == {"FID"}
        assert {row[5] for row in fid_rows[1:]} == {"retention_time"}
        assert {row[6] for row in fid_rows[1:]} == {"min"}
        assert {row[8] for row in fid_rows[1:]} == {"detector_response"}
        assert [row[9] for row in fid_rows[1:]] == ["pA", "pA"]
        assert [row[9] for row in tcd_rows[1:]] == ["mV", "mV"]
        assert [row[4] for row in fid_rows[1:]] == pytest.approx([0.0, 1 / 600])
        metadata = {
            row[3]: row[4] for row in workbook["Metadata"].iter_rows(min_row=2, values_only=True)
        }
        assert metadata["representation"] == "scientific_signal"
        assert metadata["producer_version"] == "YL-Clarity 9.1.0.76"
        assert metadata["peak_table_status"] == "unsupported"
        import_log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert import_log[0][2] == "youngin_yl_clarity_prm_raw"
        assert "YOUNGIN_PRM_EXPERIMENTAL_SCIENTIFIC_SIGNAL" in import_log[0][5]
        assert "Peaks" in workbook.sheetnames
        assert list(workbook["Peaks"].iter_rows(min_row=2, values_only=True)) == []
    finally:
        workbook.close()


def test_marker_derived_area_reaches_existing_peak_workbook_contract(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "derived-area.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            channels=((0.0, 1.0, 3.0, 1.0, 0.0),),
            marker_records=(((MARKER_START, 0), (MARKER_APEX, 1), (MARKER_END, 4)),),
        )
    )
    output = tmp_path / "derived-area.xlsx"

    result = convert(
        source,
        output,
        include_signals=True,
        experimental_derived_area=True,
    )

    assert result.success_count == 1
    assert main(["inspect", "--experimental-derived-area", str(source)]) == 0
    rendered = capsys.readouterr().out
    assert "Peak Result availability: not decoded from PRM" in rendered
    assert "Ordifile-calculated chromatographic Area: experimental" in rendered
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peak_rows = list(workbook["Peaks"].values)
        headers = peak_rows[0]
        row = peak_rows[1]
        assert row[headers.index("retention_time")] == pytest.approx(2 / 600)
        assert row[headers.index("area")] is None
        assert row[headers.index("area_unit")] is None
        assert row[headers.index("calculated_area")] == pytest.approx(0.5)
        assert row[headers.index("calculated_area_unit")] == "mV.s"
        assert row[headers.index("status")] == "ordifile_derived_experimental"
        assert row[headers.index("data_origin")] == "ordifile_marker_derived"
        assert row[headers.index("derivation_method_id")] == (
            "youngin-prm-marker-group-baseline-v4"
        )
        matrix_rows = list(workbook["Peak_Order_Matrix"].values)
        matrix_headers = matrix_rows[0]
        matrix_row = matrix_rows[1]
        assert matrix_row[matrix_headers.index("area_unit")] is None
        assert matrix_row[matrix_headers.index("calculated_area_unit")] == "mV.s"
        assert matrix_row[matrix_headers.index("data_origin")] == "ordifile_marker_derived"
        assert matrix_row[matrix_headers.index("peak_1_area")] is None
        assert matrix_row[matrix_headers.index("peak_1_calculated_area")] == pytest.approx(0.5)
        assert workbook["Signals_TCD"].max_row == 6
    finally:
        workbook.close()
