from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file


def _write_whitespace_table(path: Path, delimiter: str | None, row: list[str]) -> None:
    header = [
        "sample_id",
        "acquired_at",
        "sequence",
        "instrument",
        "vendor",
        "channel",
        "detector",
        "runtime",
        "peak_number",
        "retention_time",
        "retention_time_unit",
        "area",
        "compound",
        "compound_source",
        "time",
        "signal",
        "x_unit",
        "y_unit",
    ]
    if delimiter is None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(header)
        worksheet.append(row)
        workbook.save(path)
        return
    path.write_text(
        delimiter.join(header) + "\n" + delimiter.join(row) + "\n",
        encoding="utf-8",
    )


@pytest.mark.parametrize(
    ("suffix", "delimiter"),
    [("csv", ","), ("tsv", "\t"), ("txt", ";"), ("xlsx", None)],
)
def test_mapped_text_keeps_exact_whitespace_and_trimmed_parse_lexemes_are_audited(
    tmp_path: Path, suffix: str, delimiter: str | None
) -> None:
    source = tmp_path / f"whitespace.{suffix}"
    _write_whitespace_table(
        source,
        delimiter,
        [
            " sample ",
            " 2026-01-01T00:00:00Z ",
            " 7 ",
            " GC ",
            " Vendor ",
            " channel ",
            " detector ",
            " 10 ",
            " 1 ",
            " 1.5 ",
            " min ",
            " 20 ",
            " compound ",
            " library ",
            " 0 ",
            " 3 ",
            " seconds ",
            " mV ",
        ],
    )

    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    bundle = inspected.file.bundle
    sample = bundle.samples[0]
    peak = bundle.peaks[0]
    signal = bundle.signals[0]
    assert sample.sample_id == " sample "
    assert sample.instrument.instrument_type == " GC "
    assert sample.instrument.vendor == " Vendor "
    assert peak.channel == signal.channel == " channel "
    assert peak.detector == signal.detector == " detector "
    assert peak.retention_time_unit == " min "
    assert peak.compound == " compound "
    assert peak.compound_source == " library "
    assert signal.x_unit == " seconds "
    assert signal.y_unit == " mV "
    assert {
        (entry.key, entry.value)
        for entry in bundle.metadata
        if entry.key in {"acquired_at", "sequence", "runtime", "peak_number", "time"}
    } >= {
        ("acquired_at", " 2026-01-01T00:00:00Z "),
        ("sequence", " 7 "),
        ("runtime", " 10 "),
        ("peak_number", " 1 "),
        ("time", " 0 "),
    }
    assert any(issue.code == "PARSE_LEXEME_WHITESPACE_TRIMMED" for issue in inspected.file.issues)

    output = tmp_path / f"whitespace-{suffix}.xlsx"
    convert(source, output, include_signals=True)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        sample_row = next(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        peak_row = next(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        signal_sheet = next(name for name in workbook.sheetnames if name.startswith("Signals_"))
        signal_row = next(workbook[signal_sheet].iter_rows(min_row=2, values_only=True))
        assert sample_row[1] == " sample "
        assert sample_row[5:7] == (" GC ", " Vendor ")
        assert peak_row[2:4] == (" channel ", " detector ")
        assert peak_row[6] == " min "
        assert peak_row[9:11] == (" compound ", " library ")
        assert signal_row[2:4] == (" channel ", " detector ")
        assert signal_row[6] == " seconds "
        assert signal_row[9] == " mV "
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("suffix", "delimiter"),
    [("csv", ","), ("tsv", "\t"), ("txt", ";"), ("xlsx", None)],
)
def test_whitespace_only_mapped_text_is_metadata_not_canonical_semantics(
    tmp_path: Path, suffix: str, delimiter: str | None
) -> None:
    source = tmp_path / f"blank.{suffix}"
    whitespace = "   "
    _write_whitespace_table(
        source,
        delimiter,
        [
            whitespace,
            "2026-01-01T00:00:00Z",
            "1",
            whitespace,
            whitespace,
            whitespace,
            whitespace,
            "1",
            "1",
            "1",
            whitespace,
            "1",
            whitespace,
            whitespace,
            "0",
            "1",
            whitespace,
            whitespace,
        ],
    )

    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    bundle = inspected.file.bundle
    sample = bundle.samples[0]
    peak = bundle.peaks[0]
    signal = bundle.signals[0]
    assert sample.sample_id == "blank"
    assert sample.instrument.instrument_type is None
    assert sample.instrument.vendor is None
    assert peak.channel is peak.detector is None
    assert peak.retention_time_unit is peak.compound is peak.compound_source is None
    assert signal.x_unit is signal.y_unit is None
    whitespace_keys = {entry.key for entry in bundle.metadata if entry.value == whitespace}
    assert whitespace_keys >= {
        "sample_id",
        "instrument_type",
        "vendor",
        "channel",
        "detector",
        "retention_time_unit",
        "compound",
        "compound_source",
        "x_unit",
        "y_unit",
    }
    assert sum(issue.code == "MAPPED_TEXT_WHITESPACE_ONLY" for issue in inspected.file.issues) == 10

    output = tmp_path / f"blank-{suffix}.xlsx"
    convert(source, output, include_signals=True)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        metadata_rows = list(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        assert sum(row[4] == whitespace for row in metadata_rows) >= 10
    finally:
        workbook.close()


def test_exact_empty_mapped_values_are_absent_and_unknown_whitespace_is_preserved(
    tmp_path: Path,
) -> None:
    source = tmp_path / "empty-vs-whitespace.csv"
    source.write_text(
        "sample_id,channel,detector,compound,compound_source,x_unit,y_unit,area,"
        "  unknown header  \n"
        "sample,,,,,,,1,  raw value  \n",
        encoding="utf-8",
    )

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    bundle = inspected.file.bundle
    peak = bundle.peaks[0]
    assert peak.channel is peak.detector is peak.compound is peak.compound_source is None
    assert not any(issue.code == "MAPPED_TEXT_WHITESPACE_ONLY" for issue in inspected.file.issues)
    assert not any(
        entry.key in {"channel", "detector", "compound", "compound_source", "x_unit", "y_unit"}
        for entry in bundle.metadata
    )
    unknown = next(entry for entry in bundle.metadata if entry.key == "  unknown header  ")
    assert unknown.value == "  raw value  "

    result = convert(source, tmp_path / "empty-vs-whitespace.xlsx")
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        metadata = next(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        assert metadata[3:5] == ("  unknown header  ", "  raw value  ")
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("name", "delimiter", "adapter"),
    [
        ("sample.csv", ",", "generic_csv"),
        ("sample.tsv", "\t", "generic_tsv"),
        ("sample.txt", ";", "generic_semicolon"),
    ],
)
def test_verified_delimited_schemas(
    tmp_path: Path, name: str, delimiter: str, adapter: str
) -> None:
    source = tmp_path / name
    source.write_text(
        delimiter.join(
            [
                "sample_id",
                "acquired_at",
                "sequence",
                "instrument",
                "vendor",
                "channel",
                "detector",
                "runtime",
                "peak_number",
                "rt",
                "retention_time_unit",
                "area",
                "height",
                "compound",
                "compound_source",
                "x",
                "y",
                "x_unit",
                "y_unit",
                "unmapped_field",
            ]
        )
        + "\n"
        + delimiter.join(
            [
                "sample",
                "2026-01-01T00:00:00Z",
                "7",
                "GC",
                "Example",
                "A",
                "FID",
                "10",
                "1",
                "1.5",
                "min",
                "20",
                "4",
                "methanol",
                "export",
                "0",
                "3",
                "min",
                "mV",
                "preserved",
            ]
        )
        + "\n",
        encoding="utf-8-sig",
    )
    inspected = inspect_file(source)
    assert inspected.file.adapter_id == adapter
    assert inspected.file.bundle is not None
    assert inspected.file.bundle.metadata[0].key == "unmapped_field"
    assert inspected.file.bundle.signals[0].x_values == (0.0,)
    assert inspected.file.bundle.peaks[0].compound == "methanol"

    output = tmp_path / f"{adapter}.xlsx"
    result = convert(source, output, include_signals=True)
    assert result.success_count == 1
    assert any(name.startswith("Signals_FID") for name in result.sheets)


@pytest.mark.parametrize(
    ("content", "expected"),
    [
        (b"\xff\xfe\x00", "FORMAT_NOT_DETECTED"),
        (b"unknown,other\n1,2\n", "FORMAT_NOT_DETECTED"),
        (b"sample_id,sample_id,area\na,a,1\n", "DUPLICATE_HEADER"),
        (b"sample_id|area\na|1\n", "FORMAT_NOT_DETECTED"),
    ],
)
def test_bad_encoding_header_duplicates_and_delimiter_are_file_failures(
    tmp_path: Path, content: bytes, expected: str
) -> None:
    source = tmp_path / "bad.csv"
    source.write_bytes(content)
    result = convert(source, tmp_path / "result.xlsx")
    assert result.failure_count == 1
    assert result.files[0].issues[0].code == expected


def test_mixed_numeric_data_is_not_silently_discarded(tmp_path: Path) -> None:
    source = tmp_path / "mixed.csv"
    source.write_text(
        "sample_id,retention_time,area,compound\na,1.0,not-a-number,A\n",
        encoding="utf-8",
    )
    result = convert(source, tmp_path / "result.xlsx")
    assert result.warning_count == 1
    assert {issue.code for issue in result.files[0].issues} == {"INVALID_NUMBER"}
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        metadata = list(workbook["Metadata"].values)
        assert any("not-a-number" in row for row in metadata)
    finally:
        workbook.close()


def test_invalid_and_conflicting_sample_fields_and_signals_preserve_raw_provenance(
    tmp_path: Path,
) -> None:
    source = tmp_path / "provenance.csv"
    source.write_text(
        "sample_id,acquired_at,sequence,runtime,time,signal\n"
        "a,not-a-date,not-sequence,not-runtime,not-time,not-signal\n"
        "a,2026-01-01T00:00:00Z,2,3,1,2\n",
        encoding="utf-8",
    )
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    sample = inspected.file.bundle.samples[0]
    assert sample.acquired_at is None
    assert sample.sequence is None
    assert sample.runtime is None
    assert inspected.file.bundle.signals[0].x_values == (1.0,)
    preserved = {
        (entry.key, str(entry.value), entry.source) for entry in inspected.file.bundle.metadata
    }
    for field, first, later in (
        ("acquired_at", "not-a-date", "2026-01-01T00:00:00Z"),
        ("sequence", "not-sequence", "2"),
        ("runtime", "not-runtime", "3"),
        ("time", "not-time", None),
        ("signal", "not-signal", None),
    ):
        assert any(
            key == field and value == first and source_name == "table:row:2"
            for key, value, source_name in preserved
        )
        if later is not None:
            assert any(
                key == field and value == later and source_name == "table:row:3"
                for key, value, source_name in preserved
            )
    codes = {issue.code for issue in inspected.file.issues}
    assert {"INVALID_TIMESTAMP", "INVALID_NUMBER", "INCONSISTENT_SAMPLE_FIELD"} <= codes


def test_nonfinite_source_spelling_is_preserved_in_metadata(tmp_path: Path) -> None:
    source = tmp_path / "nonfinite.csv"
    source.write_text(
        "sample_id,runtime,retention_time,area,time,signal,compound\n"
        "a,Infinity,1,NaN,-Infinity,nan,A\n",
        encoding="utf-8",
    )
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    values = {(entry.key, entry.value) for entry in inspected.file.bundle.metadata}
    assert {
        ("runtime", "Infinity"),
        ("area", "NaN"),
        ("time", "-Infinity"),
        ("signal", "nan"),
    } <= values
    assert any(issue.code == "NONFINITE_NUMBER" for issue in inspected.file.issues)


def test_nonempty_extra_cells_are_preserved_positionally_and_empty_extras_are_ignored(
    tmp_path: Path,
) -> None:
    source = tmp_path / "extra.csv"
    source.write_text(
        "sample_id,area\na,1,raw-extra,=literal\na,2,,\n",
        encoding="utf-8",
    )
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    extras = [
        entry
        for entry in inspected.file.bundle.metadata
        if entry.key.startswith("unmapped_column_")
    ]
    assert [(entry.key, entry.value, entry.source) for entry in extras] == [
        ("unmapped_column_3", "raw-extra", "table:row:2:column:3"),
        ("unmapped_column_4", "=literal", "table:row:2:column:4"),
    ]
    assert [issue.code for issue in inspected.file.issues].count("EXTRA_CELLS_PRESERVED") == 1


@pytest.mark.parametrize(
    ("suffix", "delimiter"),
    (("csv", ","), ("tsv", "\t"), ("txt", ";"), ("xlsx", None)),
)
def test_whitespace_only_unknown_header_value_and_extra_cell_are_preserved_exactly(
    tmp_path: Path, suffix: str, delimiter: str | None
) -> None:
    source = tmp_path / f"unknown-whitespace.{suffix}"
    if delimiter is None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["sample_id", "area", "   "])
        worksheet.append(["a", 1, "   ", "  "])
        workbook.save(source)
    else:
        source.write_text(
            delimiter.join(["sample_id", "area", "   "])
            + "\n"
            + delimiter.join(["a", "1", "   ", "  "])
            + "\n",
            encoding="utf-8",
        )

    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    metadata = inspected.file.bundle.metadata
    unknown = next(entry for entry in metadata if entry.key == "   ")
    extra = next(entry for entry in metadata if entry.key == "unmapped_column_4")
    assert unknown.value == "   "
    assert extra.value == "  "
    assert {"UNKNOWN_VALUE_WHITESPACE_ONLY", "EXTRA_CELLS_PRESERVED"} <= {
        issue.code for issue in inspected.file.issues
    }


@pytest.mark.parametrize("bad_text", ("bad\x01", "bad_x000D_", "bad\ufdd0"))
def test_unrepresentable_csv_text_is_file_isolated_from_good_workbook(
    tmp_path: Path, bad_text: str
) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "bad.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad.write_text(
        f"sample_id,area,compound\nbad,3,{bad_text}\n",
        encoding="utf-8",
    )

    result = convert((good, bad), tmp_path / "isolated.xlsx")

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status.value == "failed")
    assert any(issue.code == "WORKBOOK_TEXT_UNREPRESENTABLE" for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("header", "bad_value"),
    (
        ("sample_id", "bad\x01"),
        ("channel", "channel_x000D_"),
        ("detector", "detector\ufdd0"),
    ),
)
def test_unrepresentable_sample_identity_text_is_file_isolated(
    tmp_path: Path, header: str, bad_value: str
) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "bad.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    if header == "sample_id":
        bad.write_text(
            f"sample_id,area,compound\n{bad_value},3,A\n",
            encoding="utf-8",
        )
    else:
        bad.write_text(
            f"sample_id,{header},area,compound\nbad,{bad_value},3,A\n",
            encoding="utf-8",
        )

    result = convert((good, bad), tmp_path / f"isolated-{header}.xlsx")

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status.value == "failed")
    assert any(issue.code == "WORKBOOK_TEXT_UNREPRESENTABLE" for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook.close()


def test_sample_id_at_excel_cell_limit_is_preserved_exactly(tmp_path: Path) -> None:
    sample_id = "s" * 32_767
    source = tmp_path / "boundary.csv"
    source.write_text(
        f"sample_id,area,compound\n{sample_id},2,A\n",
        encoding="utf-8",
    )

    result = convert(source, tmp_path / "boundary.xlsx")

    assert result.failure_count == 0
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        sample = next(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        assert sample[1] == sample_id
    finally:
        workbook.close()


@pytest.mark.parametrize("sidecar_mode", ("error", "csv"))
def test_overlong_mandatory_sample_id_is_file_isolated_from_good_input(
    tmp_path: Path, sidecar_mode: str
) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "bad.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad.write_text(
        f"sample_id,area,compound\n{'s' * 32_768},3,B\n",
        encoding="utf-8",
    )

    result = convert(
        (good, bad),
        tmp_path / f"overlong-{sidecar_mode}.xlsx",
        sidecar_mode=sidecar_mode,
    )

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status.value == "failed")
    assert any(issue.code == "WORKBOOK_CELL_TEXT_LIMIT" for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
    finally:
        workbook.close()


def test_joined_detector_channels_cell_limit_is_file_isolated(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "joined.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad.write_text(
        f"sample_id,channel,detector,area\nbad,{'c' * 20_000},{'d' * 20_000},3\n",
        encoding="utf-8",
    )

    result = convert((good, bad), tmp_path / "joined.xlsx")

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status.value == "failed")
    assert any(issue.code == "WORKBOOK_CELL_TEXT_LIMIT" for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 2
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "timestamp",
    (
        "0001-01-01T00:00:00+14:00",
        "9999-12-31T23:59:59-14:00",
    ),
)
def test_out_of_utc_range_timestamp_is_preserved_but_excluded_from_sorting(
    tmp_path: Path, timestamp: str
) -> None:
    boundary = tmp_path / "a-boundary.csv"
    good = tmp_path / "z-good.csv"
    boundary.write_text(
        f"sample_id,acquired_at,area\nboundary,{timestamp},1\n",
        encoding="utf-8",
    )
    good.write_text(
        "sample_id,acquired_at,area\ngood,2026-01-01T00:00:00Z,2\n",
        encoding="utf-8",
    )

    inspected = inspect_file(boundary)
    assert inspected.file.bundle is not None
    sample = inspected.file.bundle.samples[0]
    assert sample.acquired_at is not None
    assert sample.acquired_at_reliable is False
    assert any(
        entry.key in {"acquired_at", "acquired_at_unorderable_raw"} and entry.value == timestamp
        for entry in inspected.file.bundle.metadata
    )
    assert "ACQUIRED_AT_UTC_RANGE_UNORDERABLE" in {issue.code for issue in inspected.file.issues}

    explicit = convert(
        (boundary, good),
        tmp_path / "explicit-acquired.xlsx",
        sort="acquired_at",
    )
    assert explicit.failure_count == 0
    assert [item.source.name for item in explicit.files] == ["z-good.csv", "a-boundary.csv"]

    automatic = convert((boundary, good), tmp_path / "auto-acquired.xlsx")
    assert automatic.failure_count == 0
    assert automatic.sort.effective.value == "filename"
    assert [item.source.name for item in automatic.files] == ["a-boundary.csv", "z-good.csv"]


def test_large_integers_are_exact_and_lossy_float_lexemes_are_rejected(tmp_path: Path) -> None:
    source = tmp_path / "exact.csv"
    source.write_text(
        "sample_id,sequence,peak_number,retention_time,area,compound\n"
        "a,9007199254740993,9007199254740993,1,0.10000000000000001,A\n",
        encoding="utf-8",
    )
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sequence == 9_007_199_254_740_993
    peak = inspected.file.bundle.peaks[0]
    assert peak.peak_number == 9_007_199_254_740_993
    assert peak.area is None
    assert any(
        entry.key == "area" and entry.value == "0.10000000000000001"
        for entry in inspected.file.bundle.metadata
    )
    assert any(issue.code == "LOSSY_FLOAT_REJECTED" for issue in inspected.file.issues)


def test_finite_decimal_float_overflow_is_preserved_as_raw_metadata(tmp_path: Path) -> None:
    source = tmp_path / "overflow.csv"
    source.write_text(
        "sample_id,peak_number,area,compound\na,1,1e999,A\n",
        encoding="utf-8",
    )

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.peaks[0].area is None
    assert any(
        entry.key == "area" and entry.value == "1e999" for entry in inspected.file.bundle.metadata
    )
    assert any(issue.code == "LOSSY_FLOAT_REJECTED" for issue in inspected.file.issues)
    assert not any(issue.code == "NONFINITE_NUMBER" for issue in inspected.file.issues)


@pytest.mark.parametrize(
    "lexeme",
    [
        pytest.param("1e5000", id="exponent-5000"),
        pytest.param("1e1000000", id="exponent-million"),
        pytest.param("9" * 4_097, id="lexeme-over-4096"),
        pytest.param(" " * 4_096 + "1", id="whitespace-lexeme-over-4096"),
    ],
)
def test_oversized_sequence_and_peak_integer_are_bounded_before_int_construction(
    tmp_path: Path, lexeme: str
) -> None:
    source = tmp_path / "bounded.csv"
    source.write_text(
        f"sample_id,sequence,peak_number,area,compound\nbad,{lexeme},{lexeme},1,A\n",
        encoding="utf-8",
    )

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sequence is None
    assert inspected.file.bundle.peaks[0].peak_number is None
    assert {(entry.key, entry.value) for entry in inspected.file.bundle.metadata} >= {
        ("sequence", lexeme),
        ("peak_number", lexeme),
    }
    assert [issue.code for issue in inspected.file.issues].count("INTEGER_LIMIT_EXCEEDED") == 2


def test_mixed_batch_preserves_valid_data_and_oversized_integer_raw_value(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    bad_field = tmp_path / "oversized.csv"
    good.write_text(
        "sample_id,peak_number,area,compound\ngood,1,2,A\n",
        encoding="utf-8",
    )
    bad_field.write_text(
        "sample_id,peak_number,area,compound\nbounded,1e5000,3,B\n",
        encoding="utf-8",
    )

    result = convert((bad_field, good), tmp_path / "mixed.xlsx")

    assert result.success_count == 2
    assert result.failure_count == 0
    assert any(
        issue.code == "INTEGER_LIMIT_EXCEEDED" for item in result.files for issue in item.issues
    )
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert {(row[0], row[7]) for row in peaks} == {("good", 2), ("bounded", 3)}
        metadata = list(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        assert any(row[3:5] == ("peak_number", "1e5000") for row in metadata)
    finally:
        workbook.close()
