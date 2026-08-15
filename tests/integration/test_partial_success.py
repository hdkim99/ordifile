from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters.registry import create_registry
from ordifile.api import convert, inspect_file
from ordifile.core import discovery, pipeline
from ordifile.core.errors import OrdifileError
from ordifile.core.models import FileStatus, ProgressEvent
from ordifile.core.pipeline import run_pipeline
from ordifile.core.workbook_text import workbook_audit_display


def test_one_corrupt_file_does_not_discard_one_hundred_valid_files(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    for index in range(1, 101):
        (inputs / f"sample_{index}.csv").write_text(
            "sample_id,sequence,retention_time,area,compound\n"
            f"sample_{index},{index},1.0,{index},A\n",
            encoding="utf-8",
        )
    (inputs / "corrupt.csv").write_bytes(b"\xff\xfe")
    result = convert(inputs, tmp_path / "result.xlsx")
    assert result.success_count == 100
    assert result.failure_count == 1
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Samples"].max_row == 102
        assert workbook["Peaks"].max_row == 101
        log_status = [
            row[4] for row in workbook["Import_Log"].iter_rows(min_row=2, values_only=True)
        ]
        assert log_status.count("failed") == 1
        assert log_status.count("success") == 100
    finally:
        workbook.close()


@pytest.mark.skipif(os.name == "nt", reason="Windows forbids control characters in filenames.")
def test_unsafe_and_literal_escape_source_names_use_distinct_reversible_audit_display(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    control = tmp_path / "sample\x01.csv"
    literal = tmp_path / "sample_x000D_.csv"
    for source in (good, control, literal):
        source.write_text("area,compound\n2,A\n", encoding="utf-8")

    result = convert((good, control, literal), tmp_path / "escaped-paths.xlsx")

    assert result.success_count == 3
    assert result.failure_count == 0
    assert (
        sum(
            issue.code == "SOURCE_DISPLAY_ESCAPED" for item in result.files for issue in item.issues
        )
        == 2
    )
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        samples = list(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        displayed_paths = {row[2] for row in samples}
        expected_control = workbook_audit_display(control.name)
        expected_literal = workbook_audit_display(literal.name)
        assert expected_control in displayed_paths
        assert expected_literal in displayed_paths
        assert expected_control != expected_literal
        assert {row[1] for row in samples} >= {
            workbook_audit_display(control.stem),
            workbook_audit_display(literal.stem),
        }
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert all(row[9] for row in log)
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["source_display_escape_count"] == 2
        assert "~uXXXXXX;" in manifest["source_display_policy"]
        assert not any(
            "\x01" in str(value) or "_x000D_" in str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


@pytest.mark.skipif(os.name == "nt", reason="Windows forbids control characters in filenames.")
def test_unsafe_source_name_on_corrupt_file_keeps_good_workbook_and_audit_row(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "corrupt\x01.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad.write_bytes(b"\xff\xfe")

    result = convert((good, bad), tmp_path / "corrupt-path.xlsx")

    assert result.success_count == 1
    assert result.failure_count == 1
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        bad_log = next(row for row in log if row[4] == "failed")
        assert bad_log[0] == workbook_audit_display(bad.name)
        assert bad_log[9]
        assert "SOURCE_DISPLAY_ESCAPED" in bad_log[5]
        samples = list(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        assert any(row[2] == workbook_audit_display(bad.name) for row in samples)
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
    finally:
        workbook.close()


def test_duplicate_is_warning_not_failure_and_manifest_counts_it(tmp_path: Path) -> None:
    source = tmp_path / "sample.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    result = convert((source, source), tmp_path / "result.xlsx")
    assert result.success_count == 1
    assert result.failure_count == 0
    assert result.duplicate_count == 1
    assert [item.status for item in result.files].count(FileStatus.DUPLICATE) == 1
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["failure_count"] == 0
        assert manifest["duplicate_count"] == 1
    finally:
        workbook.close()


def test_on_error_stop_skips_parsing_after_first_failure(tmp_path: Path) -> None:
    bad = tmp_path / "bad.csv"
    bad.write_bytes(b"\xff\xfe")
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area\na,1\n", encoding="utf-8")
    registry = create_registry(include_external=False)
    stopped = run_pipeline((bad, good), registry, on_error="stop")
    assert [item.status for item in stopped.files] == [FileStatus.FAILED, FileStatus.SKIPPED]
    assert stopped.files[1].issues[0].code == "SKIPPED_AFTER_FAILURE"
    continued = run_pipeline((bad, good), registry, on_error="continue")
    assert continued.success_count == 1
    output = tmp_path / "stopped.xlsx"
    with pytest.raises(OrdifileError) as caught:
        convert((bad, good), output, on_error="stop", registry=registry)
    assert caught.value.code == "BATCH_FILE_FAILURE"
    assert not output.exists()


def test_pipeline_size_warning_and_hard_limit_are_preparse_and_hashed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "sample.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    registry = create_registry(include_external=False)
    monkeypatch.setattr(pipeline, "WARN_INPUT_FILE_BYTES", 1)
    monkeypatch.setattr(pipeline, "MAX_INPUT_FILE_BYTES", 10_000)
    warned = run_pipeline((source,), registry)
    assert warned.files[0].status is FileStatus.WARNING
    assert warned.files[0].issues[0].code == "INPUT_SIZE_WARNING"
    assert warned.files[0].source.sha256 is not None

    monkeypatch.setattr(pipeline, "MAX_INPUT_FILE_BYTES", 2)
    limited = run_pipeline((source,), registry)
    assert limited.files[0].status is FileStatus.FAILED
    assert limited.files[0].issues[0].code == "INPUT_SIZE_LIMIT"
    assert limited.files[0].source.sha256 == warned.files[0].source.sha256


def test_inspect_hashes_before_and_after_parse_without_adapter_duplicate_hash(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "sample.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    real_hash = discovery.sha256_file
    calls = 0

    def counting_hash(path: Path) -> str:
        nonlocal calls
        calls += 1
        return real_hash(path)

    monkeypatch.setattr(discovery, "sha256_file", counting_hash)
    monkeypatch.setattr(pipeline, "sha256_file", counting_hash)
    inspected = inspect_file(source, registry=create_registry(include_external=False))
    assert inspected.file.status is FileStatus.SUCCESS
    assert calls == 2


def test_inspect_rejects_directory_even_when_it_contains_one_file(tmp_path: Path) -> None:
    (tmp_path / "only.csv").write_text("sample_id,area\na,1\n", encoding="utf-8")
    with pytest.raises(OrdifileError) as caught:
        inspect_file(tmp_path)
    assert caught.value.code == "INSPECT_REQUIRES_FILE"


def test_invalid_sort_and_forced_adapter_typo_are_configuration_errors(tmp_path: Path) -> None:
    source = tmp_path / "sample.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    bad_sort_output = tmp_path / "sort.xlsx"
    with pytest.raises(OrdifileError) as bad_sort:
        convert(source, bad_sort_output, sort="not-a-sort")
    assert bad_sort.value.code == "SORT_MODE_INVALID"
    assert not bad_sort_output.exists()

    bad_adapter_output = tmp_path / "adapter.xlsx"
    with pytest.raises(OrdifileError) as bad_adapter:
        convert(source, bad_adapter_output, adapter="typo_adapter")
    assert bad_adapter.value.code == "ADAPTER_NOT_FOUND"
    assert not bad_adapter_output.exists()


def test_discovery_progress_event_reports_completed_total(tmp_path: Path) -> None:
    source = tmp_path / "sample.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    events: list[ProgressEvent] = []

    convert(source, tmp_path / "result.xlsx", progress=events.append)

    assert [event.stage for event in events] == [
        "discovery",
        "processing",
        "export_start",
        "export_complete",
    ]
    discovery_event, processing_event, export_start, export_complete = events
    assert (discovery_event.completed, discovery_event.total) == (1, 1)
    assert processing_event.completed == processing_event.total == 1
    assert processing_event.source_file == "sample.csv"
    assert processing_event.status is FileStatus.SUCCESS
    assert export_start.source_file == export_complete.source_file == "result.xlsx"
    assert (export_start.completed, export_complete.completed) == (0, 1)


def test_progress_callback_exception_is_not_reclassified_as_adapter_failure(
    tmp_path: Path,
) -> None:
    source = tmp_path / "sample.csv"
    output = tmp_path / "result.xlsx"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")

    class ProgressStopped(RuntimeError):
        pass

    def stop_after_processing(event: ProgressEvent) -> None:
        if event.stage == "processing":
            raise ProgressStopped

    with pytest.raises(ProgressStopped):
        convert(source, output, progress=stop_after_processing)
    assert not output.exists()


def test_all_failed_auto_sort_is_reflected_in_import_log(tmp_path: Path) -> None:
    inputs = []
    for name in ("sample_10.csv", "sample_1.csv", "sample_2.csv"):
        path = tmp_path / name
        path.write_bytes(b"\xff\xfe")
        inputs.append(path)

    result = convert(tuple(inputs), tmp_path / "failed.xlsx")

    assert result.sort.effective.value == "filename"
    assert [item.source.name for item in result.files] == [
        "sample_1.csv",
        "sample_2.csv",
        "sample_10.csv",
    ]
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        rows = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert [row[0] for row in rows] == [
            "sample_1.csv",
            "sample_2.csv",
            "sample_10.csv",
        ]
        assert [row[8] for row in rows] == [
            "sample_1.csv",
            "sample_2.csv",
            "sample_10.csv",
        ]
    finally:
        workbook.close()


def test_folder_discovery_excludes_own_workbook_sidecar_and_temp_with_audit_log(
    tmp_path: Path,
) -> None:
    source = tmp_path / "sample.csv"
    output = tmp_path / "result.xlsx"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    convert(source, output)
    sidecar = tmp_path / "result_Peaks_001.csv"
    temporary = tmp_path / ".ordifile_workbook_orphan.xlsx.tmp"
    sidecar.write_text("sample_id,area\nartifact,999\n", encoding="utf-8")
    temporary.write_bytes(b"temporary")

    result = convert(tmp_path, output, overwrite=True)

    assert result.success_count == 1
    excluded = [
        item
        for item in result.files
        if any(issue.code == "ORDIFILE_ARTIFACT_EXCLUDED" for issue in item.issues)
    ]
    assert {item.source.name for item in excluded} == {
        "result.xlsx",
        "result_Peaks_001.csv",
        ".ordifile_workbook_orphan.xlsx.tmp",
    }
    assert all(item.status is FileStatus.SKIPPED for item in excluded)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        log_rows = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        audited = [row for row in log_rows if row[4] == "skipped"]
        assert len(audited) == 3
        assert all("ORDIFILE_ARTIFACT_EXCLUDED" in row[5] for row in audited)
        peak_rows = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peak_rows) == 1
        assert peak_rows[0][0] == "a"
    finally:
        workbook.close()
