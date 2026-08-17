from __future__ import annotations

import hashlib
import math
import os
import xml.etree.ElementTree as stdlib_etree
from collections.abc import Iterable
from pathlib import Path

from defusedxml import ElementTree  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file

EXPECTED_SIZE = 98_084
EXPECTED_SHA256 = "4c876bb5712b2d943b5ad32ce5854698018e0b82f2dfc10cc0971ffab9a7056f"
EXPECTED_PEAK_COUNT = 36
EXPECTED_RT_DIGEST = "25104dd542e674f3e0d07d9c3dbfe8b019bc9b9b4b59bcc406b87300a00e9b9d"
EXPECTED_AREA_DIGEST = "db71fe58cf8646509cbd8dd2e34c0f8e566a7e4cf2043b3e49799a68115e9932"
EXPECTED_HEIGHT_DIGEST = "db939beb34b30313defecc864c511c666baa4da837914625f3ca51209fcf9c49"
EXPECTED_START_DIGEST = "cbd1a2091518a1f1f92557c94f3a24764483a5c3e547c13963572c50e9fb62bb"
EXPECTED_END_DIGEST = "8b68a48e59bbd33800198a1eae31eb79d6eda0209f64326cc9114d437588f976"


def _fixture() -> Path:
    value = os.environ.get("ORDIFILE_AGILENT_RESULT_XML_FIXTURE")
    if not value:
        raise AssertionError("ORDIFILE_AGILENT_RESULT_XML_FIXTURE is required")
    return Path(value)


def _text(parent: stdlib_etree.Element, tag: str) -> str:
    matches = parent.findall(tag)
    if len(matches) != 1 or matches[0].text is None:
        raise AssertionError("external source field cardinality changed")
    return matches[0].text


def _text_digest(values: Iterable[str]) -> str:
    return hashlib.sha256("\n".join(values).encode("utf-8")).hexdigest()


def _privacy_values(root: stdlib_etree.Element) -> frozenset[str]:
    selectors = (
        "./Acquisition/InstrumentName",
        "./Acquisition/MethodPath",
        "./Acquisition/MethodLastModifiedBy",
        "./Acquisition/BarCode",
        "./SampleInformation/SequencePath",
        "./SampleInformation/AcqInstName",
        "./SampleInformation/Method",
        "./SampleInformation/Operator",
        "./SampleInformation/SampleName",
        "./SampleInformation/SampleInfo",
        "./SampleInformation/LimsID",
        "./SampleInformation/LimsKField2",
        "./SampleInformation/LimsKField3",
        "./Chromatograms/Signal/Title",
        "./Chromatograms/Signal/Operator",
        "./Chromatograms/Signal/DateTime",
        "./Chromatograms/Signal/RawdataFile",
    )
    return frozenset(
        value
        for selector in selectors
        if (element := root.find(selector)) is not None
        if (value := (element.text or "").strip())
    )


def test_exact_external_result_xml_canonical_peaks_and_workbook(tmp_path: Path) -> None:
    source = _fixture()
    data = source.read_bytes()
    if len(data) != EXPECTED_SIZE or hashlib.sha256(data).hexdigest() != EXPECTED_SHA256:
        raise AssertionError("external fixture identity changed")
    root = ElementTree.fromstring(data)
    signal = root.find("./Chromatograms/Signal")
    group = root.find("./Results/ResultsGroup")
    if signal is None or group is None:
        raise AssertionError("external source result structure changed")
    integrations = signal.findall("IntegrationResults")
    source_peaks = group.findall("Peak")
    if len(integrations) != EXPECTED_PEAK_COUNT or len(source_peaks) != EXPECTED_PEAK_COUNT:
        raise AssertionError("external source result count changed")
    rt_text = tuple(_text(peak, "MeasRetTime") for peak in source_peaks)
    area_text = tuple(_text(peak, "Area") for peak in source_peaks)
    height_text = tuple(_text(peak, "Height") for peak in source_peaks)
    start_text = tuple(_text(row, "TimeStart") for row in integrations)
    end_text = tuple(_text(row, "TimeEnd") for row in integrations)
    if (
        _text_digest(rt_text) != EXPECTED_RT_DIGEST
        or _text_digest(area_text) != EXPECTED_AREA_DIGEST
        or _text_digest(height_text) != EXPECTED_HEIGHT_DIGEST
        or _text_digest(start_text) != EXPECTED_START_DIGEST
        or _text_digest(end_text) != EXPECTED_END_DIGEST
    ):
        raise AssertionError("external source canonical sequence digest changed")
    if any(
        (_text(row, "RetTime"), _text(row, "Area"), _text(row, "Height")) != (rt, area, height)
        for row, rt, area, height in zip(integrations, rt_text, area_text, height_text, strict=True)
    ):
        raise AssertionError("external source duplicate result rows disagree")

    inspected = inspect_file(source)
    bundle = inspected.file.bundle
    if bundle is None or len(bundle.peaks) != EXPECTED_PEAK_COUNT:
        raise AssertionError("Ordifile did not return the exact external peak count")
    if inspected.file.source.public_reference != f"source-{EXPECTED_SHA256}":
        raise AssertionError("external fixture public identity is not hash-derived")
    for index, (canonical, rt, area, height, start, end) in enumerate(
        zip(bundle.peaks, rt_text, area_text, height_text, start_text, end_text, strict=True),
        start=1,
    ):
        if (
            canonical.observation_order != index
            or canonical.peak_number is not None
            or canonical.retention_time != float(rt)
            or canonical.area != float(area)
            or canonical.height != float(height)
            or canonical.start_time != float(start)
            or canonical.end_time != float(end)
            or canonical.retention_time_unit != "min"
            or canonical.area_unit != "pA*s"
            or canonical.height_unit != "pA"
            or canonical.detector != "FID"
            or canonical.channel != "FID1A"
        ):
            raise AssertionError("Ordifile canonical result mapping changed")

    output = tmp_path / "external-result.xlsx"
    convert(source, output)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks_sheet = workbook["Peaks"]
        headers = tuple(next(peaks_sheet.values))
        rows = tuple(peaks_sheet.iter_rows(min_row=2, values_only=True))
        if len(rows) != EXPECTED_PEAK_COUNT:
            raise AssertionError("workbook peak rows were lost or duplicated")
        for row, canonical in zip(rows, bundle.peaks, strict=True):
            for column, expected in (
                ("retention_time", canonical.retention_time),
                ("area", canonical.area),
                ("height", canonical.height),
                ("start_time", canonical.start_time),
                ("end_time", canonical.end_time),
            ):
                actual = row[headers.index(column)]
                if (
                    type(actual) not in {int, float}
                    or expected is None
                    or not math.isclose(float(actual), expected, rel_tol=1e-14, abs_tol=1e-14)
                ):
                    raise AssertionError("workbook scientific numeric mapping changed")
        order_sheet = workbook["Peak_Order_Matrix"]
        if order_sheet.max_row != 2 or order_sheet.max_column != 7 + 2 * EXPECTED_PEAK_COUNT:
            raise AssertionError("workbook ordered peak matrix shape changed")
        order_values = tuple(next(order_sheet.iter_rows(min_row=2, values_only=True)))[7:]
        expected_values = tuple(
            value for peak in bundle.peaks for value in (peak.retention_time, peak.area)
        )
        if len(order_values) != len(expected_values) or any(
            type(actual) not in {int, float}
            or expected is None
            or not math.isclose(float(actual), expected, rel_tol=1e-14, abs_tol=1e-14)
            for actual, expected in zip(order_values, expected_values, strict=True)
        ):
            raise AssertionError("workbook ordered peak matrix values changed")
        workbook_strings = {
            value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell is not None
            if (value := str(cell))
        }
        if _privacy_values(root) & workbook_strings or source.name in workbook_strings:
            raise AssertionError("privacy-bearing source metadata reached the workbook")
    finally:
        workbook.close()
