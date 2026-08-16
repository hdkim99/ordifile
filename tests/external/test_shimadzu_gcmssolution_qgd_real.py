from __future__ import annotations

import hashlib
import math
import os
import struct
from collections.abc import Iterable
from pathlib import Path
from typing import cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.models import SeriesKind

EXPECTED_SIZE = 39_964_672
EXPECTED_SHA256 = "64b2faab81c0ad10bc36c57b23ed770751dbe5253f48d2a13b8b15df1de23f5d"
EXPECTED_POINT_COUNT = 16_800
EXPECTED_TIME_SHA256 = "5b05b696032d46547e1b914ae156057e8f5dd399f4e0d4c2a1af2da3013981f4"
EXPECTED_TIC_SHA256 = "ed6d7c404f2a898ca961358f47a5859e6a83c59847a0c8a7a7eade8591b84971"
EXPECTED_INDEX_SHA256 = "25959d70b88e30d1e368e367e4a79e604283cb0bf632971c08db08e06f93ec79"
EXPECTED_MS_STREAM_SHA256 = "45276fbcdb788e48b667f5331913433bbe25b1154083b97bbe5e627987ccc419"
EXPECTED_SCAN_SUMMARY_SHA256 = "0b7df4eab757360579796861a6b208f04732ce30917e67d689e3ffd4f8b870b8"
EXPECTED_FIRST_20 = (
    289_349,
    1_188_395,
    2_133_942,
    1_228_615,
    313_071,
    357_697,
    522_316,
    668_503,
    728_974,
    739_333,
    742_314,
    765_801,
    795_145,
    816_858,
    831_175,
    844_116,
    849_721,
    846_574,
    839_075,
    828_566,
)
EXPECTED_LAST_20 = (
    474_116,
    474_566,
    475_660,
    475_348,
    473_701,
    471_739,
    469_881,
    468_112,
    468_773,
    472_201,
    474_319,
    473_548,
    474_164,
    473_000,
    473_222,
    472_018,
    471_421,
    474_400,
    475_760,
    474_016,
)


def _fixture() -> Path:
    value = os.environ.get("ORDIFILE_SHIMADZU_QGD_FIXTURE")
    if not value:
        raise AssertionError("ORDIFILE_SHIMADZU_QGD_FIXTURE is required")
    return Path(value)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _f64_digest(values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for value in values:
        if type(value) is not float:
            raise AssertionError(f"time coordinate is not an exact float: {type(value).__name__}")
        digest.update(struct.pack(">d", value))
    return digest.hexdigest()


def _u64_digest(values: Iterable[object]) -> str:
    digest = hashlib.sha256()
    for value in values:
        if type(value) is not int:
            raise AssertionError(f"TIC value is not an exact integer: {type(value).__name__}")
        digest.update(struct.pack(">Q", value))
    return digest.hexdigest()


def _assert_excel_time(actual_values: Iterable[object], expected_values: Iterable[float]) -> None:
    for actual, expected in zip(actual_values, expected_values, strict=True):
        if type(actual) not in {int, float}:
            raise AssertionError(f"workbook time value is not numeric: {type(actual).__name__}")
        numeric = cast(int | float, actual)
        assert math.isclose(float(numeric), expected, rel_tol=1e-14, abs_tol=1e-14)


def test_exact_external_qgd_tic_reference_structure_and_workbook(tmp_path: Path) -> None:
    source = _fixture()
    assert source.stat().st_size == EXPECTED_SIZE
    assert _file_sha256(source) == EXPECTED_SHA256

    inspected = inspect_file(source)
    bundle = inspected.file.bundle
    assert bundle is not None
    assert len(bundle.signals) == 1
    signal = bundle.signals[0]
    assert signal.series_kind is SeriesKind.SCIENTIFIC_SIGNAL
    assert signal.detector == "MS"
    assert signal.channel == "TIC"
    assert signal.x_label == "retention_time"
    assert signal.x_unit == "min"
    assert signal.y_label == "raw_tic_intensity"
    assert signal.y_unit is None
    assert len(signal.x_values) == EXPECTED_POINT_COUNT
    assert len(signal.y_values) == EXPECTED_POINT_COUNT
    assert signal.x_values[0] == 4.0
    assert signal.x_values[-1] == 59.99666666666667
    assert signal.y_values[:20] == EXPECTED_FIRST_20
    assert signal.y_values[-20:] == EXPECTED_LAST_20
    assert min(signal.y_values) == 289_349
    assert max(signal.y_values) == 25_764_044
    assert sum(signal.y_values) == 9_258_016_526
    assert _f64_digest(signal.x_values) == EXPECTED_TIME_SHA256
    assert _u64_digest(signal.y_values) == EXPECTED_TIC_SHA256
    assert bundle.peaks == ()
    assert bundle.samples[0].acquired_at is None
    assert bundle.samples[0].acquired_at_reliable is False
    assert bundle.samples[0].runtime is None

    metadata = {entry.key: entry.value for entry in bundle.metadata}
    assert metadata["file_property_schema"] == "4.00"
    assert metadata["scan_count"] == EXPECTED_POINT_COUNT
    assert metadata["retention_time_start"] == 240_000
    assert metadata["retention_time_end"] == 3_599_800
    assert metadata["retention_time_interval"] == 200
    assert metadata["retention_time_canonical_be_f64_sha256"] == EXPECTED_TIME_SHA256
    assert metadata["tic_canonical_be_u64_sha256"] == EXPECTED_TIC_SHA256
    assert metadata["spectrum_index_canonical_be_u32_sha256"] == EXPECTED_INDEX_SHA256
    assert metadata["ms1_present"] is True
    assert metadata["ms1_long_row_count"] == 9_508_566
    assert metadata["ms1_points_per_scan_min"] == 356
    assert metadata["ms1_points_per_scan_max"] == 567
    assert metadata["ms1_intensity_widths_bytes"] == "2,3"
    assert metadata["ms1_stream_sha256"] == EXPECTED_MS_STREAM_SHA256
    assert metadata["ms1_scan_summary_sha256"] == EXPECTED_SCAN_SUMMARY_SHA256
    assert metadata["ms1_export_status"] == "unsupported"
    assert metadata["tic_signal_unit_status"] == "unknown"
    assert not any("path" in key or "operator" in key for key in metadata)

    warning_codes = {warning.code for warning in bundle.warnings}
    assert "SHIMADZU_QGD_EXPERIMENTAL_PROFILE" in warning_codes
    assert "QGD_MS1_NOT_EXPORTED" in warning_codes

    output = tmp_path / "shimadzu-qgd-tic.xlsx"
    result = convert(source, output, include_signals=True)
    assert result.files[0].source.sha256 == EXPECTED_SHA256
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Signals_MS" in workbook.sheetnames
        assert not any("MS1" in name for name in workbook.sheetnames)
        sheet = workbook["Signals_MS"]
        assert sheet.max_row == EXPECTED_POINT_COUNT + 1
        workbook_x = (
            row[0] for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True)
        )
        _assert_excel_time(workbook_x, signal.x_values)
        workbook_y = tuple(
            row[0] for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8, values_only=True)
        )
        assert workbook_y == signal.y_values
    finally:
        workbook.close()
    assert _file_sha256(source) == EXPECTED_SHA256
