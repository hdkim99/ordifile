from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert
from ordifile.core.errors import ExportLimitError
from ordifile.exporters import excel


def _peak_file(path: Path, rows: int) -> None:
    content = ["sample_id,retention_time,area,compound"]
    content.extend(f"sample,{index},{index * 10},C{index}" for index in range(rows))
    path.write_text("\n".join(content) + "\n", encoding="utf-8")


def test_row_heavy_sheets_split_without_truncation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(excel, "MAX_EXCEL_ROWS", 4)
    source = tmp_path / "peaks.csv"
    _peak_file(source, 8)
    output = tmp_path / "result.xlsx"
    result = convert(source, output)
    peak_sheets = [name for name in result.sheets if name.startswith("Peaks")]
    assert peak_sheets == ["Peaks", "Peaks_002", "Peaks_003"]
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        count = sum(max(workbook[name].max_row - 1, 0) for name in peak_sheets)
        assert count == 8
    finally:
        workbook.close()


def test_peak_matrix_columns_split_and_preserve_occurrences(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    dataset = excel._SheetData(
        "Peak_Matrix",
        ("sample_id", "A", "A_2", "B", "C"),
        (("sample", 1, 2, 3, 4),),
    )
    monkeypatch.setattr(excel, "MAX_EXCEL_COLUMNS", 3)
    segments = excel._column_segments(dataset)
    assert [item.headers for item in segments] == [
        ("sample_id", "A", "A_2"),
        ("sample_id", "B", "C"),
    ]
    assert segments[0].rows[0] == ("sample", 1, 2)


def test_peak_order_matrix_splits_only_between_atomic_rt_area_pairs(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fixed = (
        "sample_id",
        "source_file",
        "manufacturer",
        "detector",
        "channel",
        "retention_time_unit",
        "area_unit",
    )
    dataset = excel._SheetData(
        "Peak_Order_Matrix",
        (
            *fixed,
            "peak_1_rt",
            "peak_1_area",
            "peak_2_rt",
            "peak_2_area",
            "peak_3_rt",
            "peak_3_area",
        ),
        (("sample", "source", "Agilent", "FID", "FID1A", "min", "pA*s", 1, 10, 2, 20, 3, 30),),
    )
    monkeypatch.setattr(excel, "MAX_EXCEL_COLUMNS", 11)

    segments = excel._column_segments(dataset)

    assert [item.headers for item in segments] == [
        (*fixed, "peak_1_rt", "peak_1_area", "peak_2_rt", "peak_2_area"),
        (*fixed, "peak_3_rt", "peak_3_area"),
    ]
    assert segments[0].rows[0][-4:] == (1, 10, 2, 20)
    assert segments[1].rows[0][-2:] == (3, 30)


def test_peak_order_matrix_requires_room_for_identity_and_one_atomic_pair(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    dataset = excel._SheetData(
        "Peak_Order_Matrix",
        (
            "sample_id",
            "source_file",
            "manufacturer",
            "detector",
            "channel",
            "retention_time_unit",
            "area_unit",
            "peak_1_rt",
            "peak_1_area",
        ),
        (("sample", "source", "Agilent", "FID", "FID1A", "min", "pA*s", 1, 10),),
    )
    monkeypatch.setattr(excel, "MAX_EXCEL_COLUMNS", 8)
    with pytest.raises(ExportLimitError) as caught:
        excel._column_segments(dataset)
    assert caught.value.code == "EXCEL_COLUMN_LIMIT"


def test_peak_order_matrix_2d_splits_only_between_atomic_rt1_rt2_area_triples(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fixed = (
        "sample_id",
        "source_file",
        "manufacturer",
        "detector",
        "channel",
        "retention_time_unit",
        "secondary_retention_time_unit",
        "area_unit",
    )
    dataset = excel._SheetData(
        "Peak_Order_Matrix_2D",
        (
            *fixed,
            "peak_1_rt1",
            "peak_1_rt2",
            "peak_1_area",
            "peak_2_rt1",
            "peak_2_rt2",
            "peak_2_area",
        ),
        (("sample", "source", "LECO", None, None, "s", "s", "AU", 1, 0.1, 10, 2, 0.2, 20),),
    )
    monkeypatch.setattr(excel, "MAX_EXCEL_COLUMNS", 11)

    segments = excel._column_segments(dataset)

    assert [item.headers for item in segments] == [
        (*fixed, "peak_1_rt1", "peak_1_rt2", "peak_1_area"),
        (*fixed, "peak_2_rt1", "peak_2_rt2", "peak_2_area"),
    ]
    assert segments[0].rows[0][-3:] == (1, 0.1, 10)
    assert segments[1].rows[0][-3:] == (2, 0.2, 20)


def test_peak_order_matrix_2d_capacity_is_5458_atomic_triples() -> None:
    fixed_columns = 8
    atomic_columns = 3

    assert (excel.MAX_EXCEL_COLUMNS - fixed_columns) // atomic_columns == 5_458
    assert (excel.MAX_EXCEL_COLUMNS - fixed_columns) % atomic_columns == 2


def test_peak_order_matrix_2d_5459_peaks_create_two_column_segments() -> None:
    fixed = (
        "sample_id",
        "source_file",
        "manufacturer",
        "detector",
        "channel",
        "retention_time_unit",
        "secondary_retention_time_unit",
        "area_unit",
    )
    dynamic = tuple(
        header
        for index in range(1, 5_460)
        for header in (f"peak_{index}_rt1", f"peak_{index}_rt2", f"peak_{index}_area")
    )
    dataset = excel._SheetData("Peak_Order_Matrix_2D", (*fixed, *dynamic), ())

    segments = excel._column_segments(dataset)

    assert len(segments) == 2
    assert len(segments[0].headers) == 8 + 5_458 * 3
    assert segments[0].headers[-3:] == (
        "peak_5458_rt1",
        "peak_5458_rt2",
        "peak_5458_area",
    )
    assert segments[1].headers == (
        *fixed,
        "peak_5459_rt1",
        "peak_5459_rt2",
        "peak_5459_area",
    )


def test_peak_order_matrix_2d_sidecar_placeholder_keeps_identity_columns() -> None:
    headers = (
        "sample_id",
        "source_file",
        "manufacturer",
        "detector",
        "channel",
        "retention_time_unit",
        "secondary_retention_time_unit",
        "area_unit",
        "peak_1_rt1",
        "peak_1_rt2",
        "peak_1_area",
    )
    dataset = excel._SheetData("Peak_Order_Matrix_2D", headers, ())

    assert excel._sidecar_eligible(dataset)
    placeholder = excel._datasets_for_workbook((dataset,), (dataset,))[0]
    assert placeholder.headers == (*headers[:8], "sidecar_status")
    assert placeholder.rows == ()


def test_peak_order_matrix_sidecar_placeholder_keeps_identity_columns() -> None:
    headers = (
        "sample_id",
        "source_file",
        "manufacturer",
        "detector",
        "channel",
        "retention_time_unit",
        "area_unit",
        "peak_1_rt",
        "peak_1_area",
    )
    dataset = excel._SheetData("Peak_Order_Matrix", headers, ())
    assert excel._sidecar_eligible(dataset)
    placeholder = excel._datasets_for_workbook((dataset,), (dataset,))[0]
    assert placeholder.headers == (*headers[:7], "sidecar_status")
    assert placeholder.rows == ()


def test_impractical_sheet_plan_requires_explicit_sidecar(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(excel, "MAX_EXCEL_ROWS", 3)
    monkeypatch.setattr(excel, "MAX_WORKBOOK_SHEETS", 6)
    source = tmp_path / "peaks.csv"
    _peak_file(source, 8)
    with pytest.raises(ExportLimitError) as caught:
        convert(source, tmp_path / "error.xlsx")
    assert caught.value.code == "WORKBOOK_SHEET_LIMIT"

    result = convert(source, tmp_path / "sidecar.xlsx", sidecar_mode="csv")
    assert result.sidecars
    for record in result.sidecars:
        path = tmp_path / record.relative_path
        assert path.exists()
        assert len(record.sha256) == 64
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        manifest_rows = [
            row
            for name in workbook.sheetnames
            if name.startswith("Manifest")
            for row in workbook[name].values
        ]
        assert any(row[0] == "sidecar" and row[2] for row in manifest_rows)
    finally:
        workbook.close()
