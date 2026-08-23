# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

from pathlib import Path
from typing import get_type_hints

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from ordifile.adapters._mapped_table import (
    PeakMappingResolutionError,
    resolve_peak_table_mapping,
)
from ordifile.core.errors import OrdifileError
from ordifile.core.models import FileResult
from ordifile.core.peak_mapping import (
    MAX_PEAK_MAPPING_DRIFT_CANDIDATES,
    ColumnSelector,
    PeakMappingDriftCategory,
    PeakMappingDriftDiagnostic,
    PeakTableFormat,
    PeakTableImportSettings,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
    PeakTablePreview,
    clone_peak_table_mapping_profile,
)


def _profile(
    headers: tuple[str, ...] = ("Retention Time", "Area", "Height"),
    *,
    profile_id: str = "profile-11111111111111111111111111111111",
    rt_unit: str = "min",
) -> PeakTableMappingProfile:
    return PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector(headers[0], 1),
            ColumnSelector(headers[1], 2),
            rt_unit,
            PeakTableFormat.CSV,
            height_column=ColumnSelector(headers[2], 3) if len(headers) >= 3 else None,
            ignored_columns=tuple(
                ColumnSelector(header, index) for index, header in enumerate(headers[3:], start=4)
            ),
        ),
        profile_id=profile_id,
    )


def _diagnose(path: Path, profile: PeakTableMappingProfile) -> PeakMappingDriftDiagnostic:
    with pytest.raises(PeakMappingResolutionError) as captured:
        resolve_peak_table_mapping(path, PeakTableMappingSet((profile,)))
    assert captured.value.code == "PEAK_MAPPING_PROFILE_NOT_MATCHED"
    assert len(captured.value.diagnostics) == 1
    return captured.value.diagnostics[0]


def test_header_changes_are_unresolved_and_never_applied(tmp_path: Path) -> None:
    source = tmp_path / "changed.csv"
    source.write_text("RT,Peak Area,Height\n1,10,20\n", encoding="utf-8")

    diagnostic = _diagnose(source, _profile())

    assert PeakMappingDriftCategory.HEADER_CHANGED_UNRESOLVED in diagnostic.categories
    assert PeakMappingDriftCategory.REQUIRED_MAPPING_COLUMN_MISSING in diagnostic.categories
    assert diagnostic.changed_column_count == 2
    assert diagnostic.unresolved_required_roles == ("retention_time", "area")
    assert diagnostic.exact_position_matches == 1


@pytest.mark.parametrize(
    ("header", "category", "field", "expected"),
    (
        (
            "Retention Time,Area,Height,Note",
            PeakMappingDriftCategory.COLUMN_ADDED,
            "added_column_count",
            1,
        ),
        (
            "Retention Time,Area",
            PeakMappingDriftCategory.COLUMN_REMOVED,
            "removed_column_count",
            1,
        ),
        (
            "Area,Retention Time,Height",
            PeakMappingDriftCategory.COLUMN_REORDERED,
            "moved_column_count",
            2,
        ),
    ),
)
def test_structural_changes_have_fixed_categories(
    tmp_path: Path,
    header: str,
    category: PeakMappingDriftCategory,
    field: str,
    expected: int,
) -> None:
    source = tmp_path / "drift.csv"
    values = ",".join(str(index) for index in range(1, header.count(",") + 2))
    source.write_text(f"{header}\n{values}\n", encoding="utf-8")

    diagnostic = _diagnose(source, _profile())

    assert category in diagnostic.categories
    assert getattr(diagnostic, field) == expected
    if category is PeakMappingDriftCategory.COLUMN_REMOVED:
        assert diagnostic.unresolved_optional_roles == ("height",)
        assert PeakMappingDriftCategory.OPTIONAL_MAPPING_COLUMN_MISSING in diagnostic.categories
    if category is PeakMappingDriftCategory.COLUMN_REORDERED:
        assert diagnostic.unresolved_required_roles == ("retention_time", "area")
        assert PeakMappingDriftCategory.REQUIRED_MAPPING_COLUMN_MISSING not in diagnostic.categories


@pytest.mark.parametrize(
    "header",
    (
        "Note,Retention Time,Area,Height",
        "Area,Height",
    ),
)
def test_pure_leading_addition_or_removal_is_not_reported_as_reordering(
    tmp_path: Path,
    header: str,
) -> None:
    source = tmp_path / "shifted.csv"
    values = ",".join(str(index) for index in range(1, header.count(",") + 2))
    source.write_text(f"{header}\n{values}\n", encoding="utf-8")

    diagnostic = _diagnose(source, _profile())

    assert diagnostic.moved_column_count == 0
    assert PeakMappingDriftCategory.COLUMN_REORDERED not in diagnostic.categories


def test_duplicate_header_occurrences_are_distinct(tmp_path: Path) -> None:
    profile = _profile(("Retention Time", "Area", "Area"))
    source = tmp_path / "duplicate.csv"
    source.write_text("Retention Time,Area,Signal\n1,2,3\n", encoding="utf-8")

    diagnostic = _diagnose(source, profile)

    assert PeakMappingDriftCategory.DUPLICATE_HEADER_CHANGED in diagnostic.categories
    assert PeakMappingDriftCategory.OPTIONAL_MAPPING_COLUMN_MISSING in diagnostic.categories
    assert diagnostic.changed_column_count == 1
    assert diagnostic.unresolved_optional_roles == ("height",)


def test_measurement_values_and_filename_do_not_affect_exact_match(tmp_path: Path) -> None:
    first = tmp_path / "one.csv"
    second = tmp_path / "renamed.csv"
    first.write_text("Retention Time,Area,Height\n1,2,3\n", encoding="utf-8")
    second.write_text("Retention Time,Area,Height\n100,200,300\n", encoding="utf-8")
    mapping_set = PeakTableMappingSet((_profile(),))

    resolved_first = resolve_peak_table_mapping(first, mapping_set)
    resolved_second = resolve_peak_table_mapping(second, mapping_set)

    assert resolved_first.profile.profile_id == resolved_second.profile.profile_id
    assert resolved_first.adapter_id == resolved_second.adapter_id == "generic_csv"


def test_drift_diagnostics_do_not_read_an_oversized_measurement_row(tmp_path: Path) -> None:
    normal = tmp_path / "normal.csv"
    oversized = tmp_path / "oversized.csv"
    normal.write_text("RT,Area,Height\n1,2,3\n", encoding="utf-8")
    oversized.write_text(
        f"RT,Area,Height\n{'x' * 40_000},2,3\n",
        encoding="utf-8",
    )
    mapping_set = PeakTableMappingSet((_profile(),))

    diagnostics = []
    for source in (normal, oversized):
        with pytest.raises(PeakMappingResolutionError) as captured:
            resolve_peak_table_mapping(source, mapping_set)
        diagnostics.append(captured.value.diagnostics)

    assert diagnostics[0] == diagnostics[1]


def test_diagnostics_are_bounded_and_same_format_only(tmp_path: Path) -> None:
    source = tmp_path / "changed.csv"
    source.write_text("RT,Area,Height\n1,2,3\n", encoding="utf-8")
    profiles = tuple(
        _profile(profile_id=f"profile-{index:032x}", rt_unit=f"unit-{index}")
        for index in range(1, MAX_PEAK_MAPPING_DRIFT_CANDIDATES + 3)
    )
    xlsx = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("RT", 1),
            ColumnSelector("Area", 2),
            "min",
            PeakTableFormat.XLSX,
        ),
        profile_id="profile-ffffffffffffffffffffffffffffffff",
    )

    with pytest.raises(PeakMappingResolutionError) as captured:
        resolve_peak_table_mapping(source, PeakTableMappingSet((*profiles, xlsx)))

    assert len(captured.value.diagnostics) == MAX_PEAK_MAPPING_DRIFT_CANDIDATES
    assert {item.source_format for item in captured.value.diagnostics} == {PeakTableFormat.CSV}
    assert all(not hasattr(item, "headers") for item in captured.value.diagnostics)


def test_compatible_drift_is_not_hidden_by_bounded_incompatible_profiles(
    tmp_path: Path,
) -> None:
    source = tmp_path / "changed.csv"
    source.write_text("Retention Time,Peak Area,Height\n1,2,3\n", encoding="utf-8")
    unrelated = tuple(
        _profile(
            (f"Unrelated {index} A", f"Unrelated {index} B", "Other", "Note"),
            profile_id=f"profile-{index:032x}",
        )
        for index in range(1, MAX_PEAK_MAPPING_DRIFT_CANDIDATES + 2)
    )
    compatible = _profile(profile_id="profile-ffffffffffffffffffffffffffffffff")

    with pytest.raises(PeakMappingResolutionError) as captured:
        resolve_peak_table_mapping(
            source,
            PeakTableMappingSet((*unrelated, compatible)),
        )

    assert tuple(item.profile_id for item in captured.value.diagnostics) == (compatible.profile_id,)
    assert PeakMappingDriftCategory.INCOMPATIBLE_STRUCTURE not in (
        captured.value.diagnostics[0].categories
    )


def test_xlsx_exact_worksheet_rename_is_diagnostic_only(tmp_path: Path) -> None:
    source = tmp_path / "renamed.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "New Results"
    worksheet.append(("Retention Time", "Area"))
    worksheet.append((1, 2))
    workbook.save(source)
    profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Retention Time", 1),
            ColumnSelector("Area", 2),
            "min",
            PeakTableFormat.XLSX,
        ),
        worksheet_title="Old Results",
        profile_id="profile-11111111111111111111111111111111",
    )

    diagnostic = _diagnose(source, profile)

    assert diagnostic.categories == (
        PeakMappingDriftCategory.WORKSHEET_IDENTITY_CHANGED_UNRESOLVED,
    )
    assert diagnostic.total_difference_count == 1
    assert diagnostic.unresolved_required_roles == ()


def test_xlsx_diagnostic_uses_each_profiles_closest_observed_sheet(tmp_path: Path) -> None:
    source = tmp_path / "multiple.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Template A"
    first.append(("Retention Time", "Peak Area"))
    first.append((1, 2))
    second = workbook.create_sheet("Template B")
    second.append(("Elapsed", "Response"))
    second.append((3, 4))
    workbook.save(source)
    first_profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Retention Time", 1),
            ColumnSelector("Area", 2),
            "min",
            PeakTableFormat.XLSX,
        ),
        worksheet_title="Template A",
        profile_id="profile-11111111111111111111111111111111",
    )
    second_profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Time", 1),
            ColumnSelector("Response", 2),
            "min",
            PeakTableFormat.XLSX,
        ),
        worksheet_title="Template B",
        profile_id="profile-22222222222222222222222222222222",
    )

    with pytest.raises(PeakMappingResolutionError) as captured:
        resolve_peak_table_mapping(
            source,
            PeakTableMappingSet((first_profile, second_profile)),
        )

    diagnostic = next(
        item for item in captured.value.diagnostics if item.profile_id == first_profile.profile_id
    )
    assert (
        PeakMappingDriftCategory.WORKSHEET_IDENTITY_CHANGED_UNRESOLVED not in diagnostic.categories
    )
    assert diagnostic.changed_column_count == 1
    assert diagnostic.exact_position_matches == 1


def test_user_confirmed_repair_clones_profile_and_preserves_old_structure() -> None:
    parent = _profile()
    original = PeakTableMappingSet((parent,))
    preview = PeakTablePreview(
        PeakTableFormat.CSV,
        ("RT", "Peak Area", "Height"),
        (("1", "2", "3"),),
    )
    repaired = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Peak Area", 2),
        "min",
        PeakTableFormat.CSV,
        height_column=ColumnSelector("Height", 3),
    )

    updated = clone_peak_table_mapping_profile(
        original,
        parent_profile_id=parent.profile_id,
        observed_preview=preview,
        repaired_mapping=repaired,
        display_label="Repaired profile",
    )

    child = updated.profiles[-1]
    assert original.profiles == (parent,)
    assert updated.set_id == original.set_id
    assert updated.profiles[0] is parent
    assert child.profile_id != parent.profile_id
    assert child.semantic_sha256 != parent.semantic_sha256
    assert child.exact_structure_sha256 != parent.exact_structure_sha256
    assert original.match(PeakTableFormat.CSV, parent.mapping.declared_headers) == (parent,)
    assert updated.match(PeakTableFormat.CSV, preview.headers) == (child,)


def test_user_confirmed_repair_preserves_explicit_import_settings() -> None:
    parent = _profile()
    original = PeakTableMappingSet((parent,))
    settings = PeakTableImportSettings(header_row=4)
    preview = PeakTablePreview(
        PeakTableFormat.CSV,
        ("RT", "Peak Area", "Height"),
        (),
        import_settings=settings,
    )
    repaired = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Peak Area", 2),
        "min",
        PeakTableFormat.CSV,
        height_column=ColumnSelector("Height", 3),
        import_settings=settings,
    )

    updated = clone_peak_table_mapping_profile(
        original,
        parent_profile_id=parent.profile_id,
        observed_preview=preview,
        repaired_mapping=repaired,
        display_label="Preamble profile",
    )

    assert updated.profiles[-1].mapping.import_settings == settings
    assert updated.match(
        PeakTableFormat.CSV,
        preview.headers,
        import_settings=settings,
    ) == (updated.profiles[-1],)


def test_repair_rejects_import_settings_not_used_for_preview() -> None:
    parent = _profile()
    mapping_set = PeakTableMappingSet((parent,))
    preview = PeakTablePreview(
        PeakTableFormat.CSV,
        ("RT", "Peak Area", "Height"),
        (),
        import_settings=PeakTableImportSettings(header_row=3),
    )
    repaired = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Peak Area", 2),
        "min",
        PeakTableFormat.CSV,
        height_column=ColumnSelector("Height", 3),
    )

    with pytest.raises(OrdifileError) as captured:
        clone_peak_table_mapping_profile(
            mapping_set,
            parent_profile_id=parent.profile_id,
            observed_preview=preview,
            repaired_mapping=repaired,
            display_label="Rejected",
        )

    assert captured.value.code == "PEAK_MAPPING_REPAIR_IMPORT_SETTINGS_MISMATCH"


def test_repair_rejects_a_new_exact_match_ambiguity() -> None:
    parent = _profile()
    existing = _profile(
        ("RT", "Peak Area", "Height"),
        profile_id="profile-22222222222222222222222222222222",
    )
    mapping_set = PeakTableMappingSet((parent, existing))
    preview = PeakTablePreview(PeakTableFormat.CSV, existing.mapping.declared_headers, ())

    with pytest.raises(OrdifileError) as captured:
        clone_peak_table_mapping_profile(
            mapping_set,
            parent_profile_id=parent.profile_id,
            observed_preview=preview,
            repaired_mapping=existing.mapping,
            display_label="Ambiguous repair",
        )

    assert captured.value.code == "PEAK_MAPPING_REPAIR_AMBIGUOUS"
    assert mapping_set.profiles == (parent, existing)


def test_xlsx_repair_requires_one_explicitly_previewed_worksheet() -> None:
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.XLSX,
    )
    parent = PeakTableMappingProfile(mapping)
    mapping_set = PeakTableMappingSet((parent,))

    with pytest.raises(OrdifileError) as captured:
        clone_peak_table_mapping_profile(
            mapping_set,
            parent_profile_id=parent.profile_id,
            observed_preview=PeakTablePreview(PeakTableFormat.XLSX, ("RT", "Area"), ()),
            repaired_mapping=mapping,
            display_label="Missing worksheet",
        )

    assert captured.value.code == "PEAK_MAPPING_REPAIR_WORKSHEET_REQUIRED"
    assert mapping_set.profiles == (parent,)


def test_public_file_result_type_hints_resolve_drift_diagnostic() -> None:
    assert "mapping_diagnostics" in get_type_hints(FileResult)
