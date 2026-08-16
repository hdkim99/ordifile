from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.cli.main import main

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_shimadzu_gcsolution_gcd import synthetic_gcd_bytes  # noqa: E402


def test_inspect_reports_one_experimental_scientific_gcd_signal(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.gcd"
    source.write_bytes(synthetic_gcd_bytes())

    inspected = inspect_file(source)

    assert inspected.file.adapter_id == "shimadzu_gcsolution_gcd"
    assert inspected.file.source.detected_format == "shimadzu_gcsolution_gcd"
    assert inspected.file.bundle is not None
    assert len(inspected.file.bundle.signals) == 1
    assert {issue.code for issue in inspected.file.issues} == {"SHIMADZU_GCD_EXPERIMENTAL_PROFILE"}


def test_cli_distinguishes_gcd_scientific_signal_from_decoded_records(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "synthetic.gcd"
    source.write_bytes(synthetic_gcd_bytes())

    assert main(["inspect", str(source)]) == 0
    output = capsys.readouterr().out
    assert "Scientific signals: 1" in output
    assert "Decoded record series: 0" in output


def test_valid_and_corrupt_gcd_files_are_isolated_in_one_workbook(tmp_path: Path) -> None:
    valid = tmp_path / "valid.gcd"
    corrupt = tmp_path / "corrupt.gcd"
    values = tuple(50.0 + index / 10 for index in range(512))
    valid.write_bytes(synthetic_gcd_bytes(sample_id="valid-sample", values=values))
    corrupt.write_bytes(synthetic_gcd_bytes()[:-1])
    output = tmp_path / "result.xlsx"

    result = convert(tmp_path, output, extensions=(".gcd",), include_signals=True)

    assert result.success_count == 1
    assert result.failure_count == 1
    assert output.is_file()
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Signals_FID" in workbook.sheetnames
        rows = list(workbook["Signals_FID"].values)
        assert len(rows) == 513
        assert rows[0] == (
            "sample_id",
            "source_file",
            "channel",
            "detector",
            "x",
            "x_label",
            "x_unit",
            "y",
            "y_label",
            "y_unit",
        )
        assert rows[1][0] == "valid-sample"
        assert rows[1][4] == pytest.approx(20 / 60_000)
        assert rows[1][6] == "min"
        assert rows[1][7] == values[0]
        assert rows[1][9] == "uV"
        assert rows[-1][4] == pytest.approx((20 + 511 * 40) / 60_000)
        assert rows[-1][7] == values[-1]
        import_log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert len(import_log) == 2
        assert {row[4] for row in import_log} == {"warning", "failed"}
        failed = next(row for row in import_log if row[4] == "failed")
        # Removing the final CFB sector also removes the recoverable Shimadzu
        # stream identity, so discovery must not claim a vendor-specific match.
        assert failed[6] == "FORMAT_NOT_DETECTED"
    finally:
        workbook.close()
