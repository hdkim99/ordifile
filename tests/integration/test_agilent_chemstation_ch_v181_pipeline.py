from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.cli.main import main

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_ch_v181 import synthetic_v181_bytes  # noqa: E402


def test_inspect_reports_experimental_v181_with_structured_warning(tmp_path: Path) -> None:
    source = tmp_path / "FID1A.CH"
    source.write_bytes(synthetic_v181_bytes())

    inspected = inspect_file(source)

    assert inspected.file.adapter_id == "agilent_chemstation_ch_v181"
    assert inspected.file.source.detected_format == "agilent_chemstation_ch_v181"
    assert {issue.code for issue in inspected.file.issues} >= {
        "AGILENT_CH_V181_EXPERIMENTAL_RECORDS"
    }


def test_cli_inspect_distinguishes_decoded_records_from_scientific_signals(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "FID1A.CH"
    source.write_bytes(synthetic_v181_bytes())

    assert main(["inspect", str(source)]) == 0
    output = capsys.readouterr().out
    assert "Scientific signals: 0" in output
    assert "Decoded record series: 1" in output


def test_valid_and_corrupt_v181_files_are_isolated_in_one_workbook(tmp_path: Path) -> None:
    valid = tmp_path / "FID1A.CH"
    corrupt = tmp_path / "FID2A.CH"
    valid.write_bytes(
        synthetic_v181_bytes(records=(("absolute", 100), ("absolute", 105), ("relative", 0)))
    )
    corrupt.write_bytes(synthetic_v181_bytes()[:-1])
    output = tmp_path / "result.xlsx"

    result = convert(tmp_path, output, extensions=(".ch",), include_signals=True)

    assert result.success_count == 1
    assert result.failure_count == 1
    assert output.is_file()
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Signals_Records_FID" in workbook.sheetnames
        rows = list(workbook["Signals_Records_FID"].values)
        assert rows[0][-1] == "series_kind"
        assert [row[4] for row in rows[1:]] == [0, 1, 2]
        assert [row[7] for row in rows[1:]] == [100, 105, 105]
        assert {row[-1] for row in rows[1:]} == {"decoded_records"}
        assert list(workbook["Peaks"].values) == [
            (
                "sample_id",
                "source_file",
                "channel",
                "detector",
                "peak_number",
                "retention_time",
                "retention_time_unit",
                "area",
                "height",
                "compound",
                "compound_source",
                "status",
            )
        ]
        import_log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert len(import_log) == 2
        assert {row[4] for row in import_log} == {"warning", "failed"}
    finally:
        workbook.close()
