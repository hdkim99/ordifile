from __future__ import annotations

import hashlib
import os
import stat
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from zipfile import ZipFile, ZipInfo

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters._youngin_yl_clarity_result_csv import (
    YoungInResultCsvStructureError,
    read_result_csv,
)
from ordifile.api import convert, inspect_file
from ordifile.core.models import SeriesKind, SignalSeries

EXPECTED_ARCHIVE_SIZE = 1_604_581
EXPECTED_ARCHIVE_SHA256 = "fff59de802b01d1d78e393b66b026fc79b9b736c5cb79b2d72f2b3e841ae72db"
EXPECTED_MEMBER_COUNT = 10
EXPECTED_EXPANDED_BYTES = 4_405_612
EXPECTED_RECORDS_PER_CHANNEL = 13_800
EXPECTED_CURVE_POINTS = 138_000
PRINTED_SIGNAL_HALF_UNIT = Decimal("0.00005001")
MAX_MEMBER_BYTES = 1_000_000
MAX_COMPRESSION_RATIO = 4.0

EXPECTED_PAIRS = {
    "27b520f097d498c50607100f0d6fb94b8719768d2de7d782b40ffdb874555fc6": (
        "55a13d97463838cb0c4ba65cfea2a0e3217b95817255b158492fa55df2e37680",
        3,
    ),
    "5965b1ebf7f1315cc721448d394c44217e8a68518dd44621958fdb5f3205e85f": (
        "75533a14c09e09c1e06134e6fedbc3b10a035b1ae9d0246c9eacb330c80e8868",
        4,
    ),
    "9e35aceb898ad813f44ae104b8ae2f84c911d27b753922c9fce6f660c0e96f4a": (
        "83c8f59da3b736a2af02d6d3cab56b71eca8f7dd6d1319bc363d78fe0428d08d",
        5,
    ),
    "b627b3a39d4158a4fc0560ebd486a386a2eac4750a161324db6db8813cc984d8": (
        "845f9aaadb654010f0490eb40a1c89443b1a0eb059e66dcefb28ff8e5de25e3e",
        4,
    ),
    "c5196cbafc20fb666734038bac0ae039ed69e7a249eed88646421d341b633c2b": (
        "765a9266adbabcfb5bd15f41a18cc428ac86b4d03afe7ce65d6993c52f63d3c9",
        5,
    ),
}
EXPECTED_RESULT_DIGESTS = {
    "55a13d97463838cb0c4ba65cfea2a0e3217b95817255b158492fa55df2e37680": (
        "ad53e8806d17c82d38902738d1d47d96bddaade27513466322efa0f793149dd0",
        "7bbda3cac64c42363c50406362743c01df609c7abb13b63cc3a969392898d914",
        "8fa2f6bb4d637f66c281c7bb2ac503b101dbee90024422e8e7039641593ab8d5",
        "ad42c2b5c2a78fa8a7c1eba26828a1cc4d9fa9acce880229d0be7d1c386d55d7",
    ),
    "75533a14c09e09c1e06134e6fedbc3b10a035b1ae9d0246c9eacb330c80e8868": (
        "67497b776854008d38c2340e14925a64b36686230bccaa777db68f644196015f",
        "609ab1b5ef9c5331ea70c4ca9c67b67bb95135b52d05b759df36d398efd96336",
        "4049e57c3877c29eac5c4bef258b1b4027302cfb8a0cc6d391d11b9fe74c8886",
        "87514749477fcc0c8c549ebfcae5134093b87c5f51797cdbe7c99da685ca0146",
    ),
    "765a9266adbabcfb5bd15f41a18cc428ac86b4d03afe7ce65d6993c52f63d3c9": (
        "b5584a1464119498204d2bffb86944984de652acb277cb0d2a6810944ee13350",
        "86b87e2e05ac6059690928d33078d80694046ddfcb92d6c573aa285e9d55ef6a",
        "72d72499713046dc23d665add8e91636fc0d955190c4ae9a4eb48d20382b7bcf",
        "0cdc4de670f07bae1275825fa09b072511227e388a14599f4e7206fbd5bda078",
    ),
    "83c8f59da3b736a2af02d6d3cab56b71eca8f7dd6d1319bc363d78fe0428d08d": (
        "b5584a1464119498204d2bffb86944984de652acb277cb0d2a6810944ee13350",
        "7b93a8a7ac3ea34d4460694d7da6d5fafa56a1a960b984689bedf3562c825472",
        "46868419e31c47b4b7c8a0aacdd6b54b49f4da8594eb7946c6976296f2845200",
        "518cb39ffeaae3f5f13e91090850985aa0559d01589742238209cc6ed8b71497",
    ),
    "845f9aaadb654010f0490eb40a1c89443b1a0eb059e66dcefb28ff8e5de25e3e": (
        "67497b776854008d38c2340e14925a64b36686230bccaa777db68f644196015f",
        "11767309f4d4cb1c4ff2d825a34c5e48af402ec1dde45b5d9292fce503918580",
        "fcd2a24bdc4fc5a51f007798e0d7d0fd3564173860a30611bfa837db0dfcefd5",
        "4545c305f0133ceafc32ed506266a0d91b6c54aaa478cadd9e66882c52d19dca",
    ),
}


def _archive() -> Path:
    value = os.environ.get("ORDIFILE_YOUNGIN_EXPANDED_ARCHIVE")
    if not value:
        raise AssertionError("ORDIFILE_YOUNGIN_EXPANDED_ARCHIVE is required")
    return Path(value)


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_member(info: ZipInfo) -> bool:
    pure = PurePosixPath(info.filename)
    mode = (info.external_attr >> 16) & 0xFFFF
    return (
        not info.is_dir()
        and not (info.flag_bits & 0x1)
        and "\\" not in info.filename
        and not pure.is_absolute()
        and all(part not in {"", ".", ".."} for part in pure.parts)
        and (mode == 0 or stat.S_ISREG(mode))
    )


def _curve_blocks(data: bytes) -> dict[str, tuple[tuple[str, str], ...]]:
    lines = data.decode("cp949", errors="strict").splitlines()
    headers = {
        "Time [min]\tCurrent [pA]": "FID",
        "Time [min]\tVoltage [mV]": "TCD",
    }
    blocks: dict[str, tuple[tuple[str, str], ...]] = {}
    for index, line in enumerate(lines):
        detector = headers.get(line)
        if detector is None:
            continue
        rows: list[tuple[str, str]] = []
        for candidate in lines[index + 1 : index + 1 + EXPECTED_RECORDS_PER_CHANNEL]:
            fields = candidate.split("\t")
            if len(fields) != 2:
                raise AssertionError("full-range curve ended before its declared point count")
            try:
                Decimal(fields[0])
                Decimal(fields[1])
            except InvalidOperation as error:
                raise AssertionError("full-range curve contains a non-numeric point") from error
            rows.append((fields[0], fields[1]))
        if len(rows) != EXPECTED_RECORDS_PER_CHANNEL or detector in blocks:
            raise AssertionError("composite export has an unsupported curve layout")
        blocks[detector] = tuple(rows)
    if set(blocks) != {"FID", "TCD"}:
        raise AssertionError("composite export lacks the exact FID/TCD curve pair")
    return blocks


def _result_evidence(
    data: bytes,
) -> tuple[tuple[tuple[str, str, str, str], ...], int, bool]:
    lines = data.decode("cp949", errors="strict").splitlines()
    curve_headers = {"Time [min]\tCurrent [pA]", "Time [min]\tVoltage [mV]"}
    result_header_prefix = ("Signal No.", "Signal Name", "Peak No.", "Reten. time [min]")
    tcd_header_fields = (
        "Signal No.",
        "Signal Name",
        "Peak No.",
        "Reten. time [min]",
        "Area [mV.s]",
        "Height [mV]",
    )
    try:
        result_end = min(index for index, line in enumerate(lines) if line in curve_headers)
    except ValueError as error:
        raise AssertionError("composite export lacks a bounded Result section") from error
    first_fields = lines[0].split("\t")
    if len(first_fields) != 12 or tuple(first_fields[3:7]) != result_header_prefix:
        raise AssertionError("composite export lacks the exact Result start header")
    tcd_header_indexes = tuple(
        index
        for index, line in enumerate(lines[:result_end])
        if tuple(line.split("\t")[3:9]) == tcd_header_fields
    )
    if len(tcd_header_indexes) != 1 or tcd_header_indexes[0] == 0:
        raise AssertionError("composite export lacks one bounded TCD Result header")

    rows: list[tuple[str, str, str, str]] = []
    totals: list[tuple[str, str]] = []
    empty_fid_sections = 0
    for line in lines[:result_end]:
        fields = line.split("\t")
        if (
            len(fields) >= 6
            and fields[4].strip() == "FID"
            and fields[5].strip() == "No peak to report"
        ):
            empty_fid_sections += 1
        if len(fields) >= 12 and fields[4].strip() == "TCD" and fields[5].strip().isdigit():
            for value in fields[5:9]:
                try:
                    Decimal(value)
                except InvalidOperation as error:
                    raise AssertionError(
                        "Result row contains an invalid scientific value"
                    ) from error
            rows.append((fields[5], fields[6], fields[7], fields[8]))
        if len(fields) >= 9 and fields[4].strip() == "TCD" and fields[6].strip() == "Total":
            totals.append((fields[7], fields[8]))
    if len(totals) != 1:
        raise AssertionError("composite Result section lacks exactly one TCD Total row")
    try:
        displayed_area, displayed_height = (Decimal(value) for value in totals[0])
        exact_area = sum((Decimal(row[2]) for row in rows), Decimal(0))
        exact_height = sum((Decimal(row[3]) for row in rows), Decimal(0))
    except InvalidOperation as error:
        raise AssertionError("Result Total contains an invalid numeric value") from error
    return (
        tuple(rows),
        empty_fid_sections,
        (displayed_area != exact_area or displayed_height != exact_height),
    )


def _result_digests(rows: tuple[tuple[str, str, str, str], ...]) -> tuple[str, ...]:
    return tuple(
        hashlib.sha256("\n".join(row[index] for row in rows).encode("utf-8")).hexdigest()
        for index in range(4)
    )


def _curve_matches(signal: SignalSeries, rows: tuple[tuple[str, str], ...]) -> bool:
    if len(signal.x_values) != len(rows) or len(signal.y_values) != len(rows):
        return False
    for index, ((time_text, response_text), x_value, y_value) in enumerate(
        zip(rows, signal.x_values, signal.y_values, strict=True)
    ):
        if time_text != format(index / 600, ".5f") or time_text != format(x_value, ".5f"):
            return False
        if abs(Decimal(response_text) - Decimal(str(y_value))) > PRINTED_SIGNAL_HALF_UNIT:
            return False
    return True


def test_expanded_owner_archive_validates_exact_9_1_scientific_curves(
    tmp_path: Path,
) -> None:
    archive = _archive()
    assert archive.stat().st_size == EXPECTED_ARCHIVE_SIZE
    assert _sha256(archive) == EXPECTED_ARCHIVE_SHA256

    prm_paths: dict[str, Path] = {}
    csv_data: dict[str, bytes] = {}
    private_names: set[str] = set()
    with ZipFile(archive) as source:
        infos = source.infolist()
        assert len(infos) == EXPECTED_MEMBER_COUNT
        assert sum(info.file_size for info in infos) == EXPECTED_EXPANDED_BYTES
        assert source.testzip() is None
        assert all(_safe_member(info) for info in infos)
        assert all(
            info.file_size <= MAX_MEMBER_BYTES
            and info.compress_size > 0
            and info.file_size / info.compress_size <= MAX_COMPRESSION_RATIO
            for info in infos
        )
        assert len({info.filename for info in infos}) == EXPECTED_MEMBER_COUNT
        assert len({info.filename.casefold() for info in infos}) == EXPECTED_MEMBER_COUNT
        for info in infos:
            data = source.read(info)
            digest = _sha256_bytes(data)
            private_names.add(PurePosixPath(info.filename).name)
            suffix = PurePosixPath(info.filename).suffix.casefold()
            if suffix == ".prm":
                path = tmp_path / f"source-{digest}.prm"
                path.write_bytes(data)
                prm_paths[digest] = path
            elif suffix == ".csv":
                csv_data[digest] = data
            else:
                raise AssertionError("archive contains an unexpected member type")

    assert set(prm_paths) == set(EXPECTED_PAIRS)
    assert set(csv_data) == {expected[0] for expected in EXPECTED_PAIRS.values()}
    curves = {digest: _curve_blocks(data) for digest, data in csv_data.items()}
    result_evidence = {digest: _result_evidence(data) for digest, data in csv_data.items()}
    result_rows = {digest: evidence[0] for digest, evidence in result_evidence.items()}
    assert sum(len(rows) for rows in result_rows.values()) == 21
    assert sum(evidence[1] for evidence in result_evidence.values()) == 5
    assert sum(evidence[2] for evidence in result_evidence.values()) == 5
    assert {
        digest: _result_digests(rows) for digest, rows in result_rows.items()
    } == EXPECTED_RESULT_DIGESTS
    for digest, data in csv_data.items():
        composite = tmp_path / f"composite-{digest}.csv"
        composite.write_bytes(data)
        try:
            read_result_csv(composite)
        except YoungInResultCsvStructureError as error:
            assert error.code == "YOUNGIN_RESULT_CSV_HEADER_INVALID"
        else:
            raise AssertionError("composite grammar entered the standalone Result profile")

    matched_pairs: dict[str, str] = {}
    all_signals: list[SignalSeries] = []
    for prm_digest, path in prm_paths.items():
        inspected = inspect_file(path)
        bundle = inspected.file.bundle
        assert bundle is not None
        assert bundle.peaks == ()
        assert bundle.samples[0].detectors == ("FID", "TCD")
        signals = {signal.detector: signal for signal in bundle.signals}
        assert set(signals) == {"FID", "TCD"}
        assert all(
            signal.series_kind is SeriesKind.SCIENTIFIC_SIGNAL for signal in signals.values()
        )
        assert signals["FID"].x_unit == signals["TCD"].x_unit == "min"
        assert signals["FID"].y_unit == "pA"
        assert signals["TCD"].y_unit == "mV"
        matches = [
            csv_digest
            for csv_digest, blocks in curves.items()
            if _curve_matches(signals["FID"], blocks["FID"])
            and _curve_matches(signals["TCD"], blocks["TCD"])
        ]
        assert len(matches) == 1
        matched_pairs[prm_digest] = matches[0]
        all_signals.extend(signals.values())

    assert matched_pairs == {
        prm_digest: csv_digest for prm_digest, (csv_digest, _) in EXPECTED_PAIRS.items()
    }
    assert sum(len(signal.y_values) for signal in all_signals) == EXPECTED_CURVE_POINTS
    for prm_digest, (csv_digest, peak_rows) in EXPECTED_PAIRS.items():
        assert matched_pairs[prm_digest] == csv_digest
        assert len(result_rows[csv_digest]) == peak_rows

    output = tmp_path / "expanded-scientific.xlsx"
    result = convert(tuple(prm_paths.values()), output, include_signals=True)
    assert result.success_count == 5
    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Signals_FID"].max_row == 69_001
        assert workbook["Signals_TCD"].max_row == 69_001
        assert list(workbook["Peaks"].iter_rows(min_row=2, values_only=True)) == []
        rendered = "\n".join(
            str(value)
            for sheet_name in ("Manifest", "Samples", "Metadata", "Import_Log")
            for row in workbook[sheet_name].iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        assert not any(name in rendered for name in private_names)
    finally:
        workbook.close()

    assert _sha256(archive) == EXPECTED_ARCHIVE_SHA256
