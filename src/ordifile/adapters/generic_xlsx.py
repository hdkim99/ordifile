# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

"""Generic XLSX adapter backed by a bounded non-evaluating OOXML audit."""

from __future__ import annotations

import zipfile
from collections.abc import Iterable, Sequence
from dataclasses import replace
from pathlib import Path
from typing import Any, ClassVar

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from ordifile.adapters._tabular import (
    MAPPED_TEXT_FIELDS,
    is_compatible_header,
    parse_rows,
    semantic_headers,
)
from ordifile.adapters._xlsx_audit import (
    RawCell,
    SheetPart,
    XlsxAuditLimits,
    XlsxPackageAudit,
    audit_xlsx_package,
    capture_worksheet_cells,
)
from ordifile.adapters.base import AdapterDescriptor, DetectionResult, ParseOptions
from ordifile.core.errors import AdapterAmbiguityError, ParseError
from ordifile.core.models import DatasetBundle, Issue, MetadataEntry, Severity

MAX_XLSX_MEMBERS = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 1_000.0
_RATIO_MINIMUM_SIZE = 1024 * 1024

_NUMERIC_FIELDS = frozenset(
    {
        "sequence",
        "runtime",
        "peak_number",
        "retention_time",
        "area",
        "height",
        "time",
        "signal",
    }
)
_TEXT_FIELDS = frozenset(MAPPED_TEXT_FIELDS)
_STRING_CELL_TYPES = frozenset({"s", "str", "inlineStr"})


def preflight_xlsx(path: Path) -> XlsxPackageAudit:
    """Audit the package and worksheet streams before openpyxl interpretation."""
    return audit_xlsx_package(
        path,
        XlsxAuditLimits(
            MAX_XLSX_MEMBERS,
            MAX_XLSX_UNCOMPRESSED_BYTES,
            MAX_XLSX_COMPRESSION_RATIO,
            _RATIO_MINIMUM_SIZE,
        ),
    )


def _trimmed_header(values: Sequence[object]) -> list[object]:
    header = list(values)
    while header and (header[-1] is None or str(header[-1]) == ""):
        header.pop()
    return header


def _cell_metadata_key(column: int, header: Sequence[object], mapped: Sequence[str | None]) -> str:
    if column <= len(mapped) and mapped[column - 1] is not None:
        return str(mapped[column - 1])
    if column <= len(header) and str(header[column - 1]).strip():
        return str(header[column - 1])
    return f"unmapped_column_{column}"


def _formula_literal(cell: RawCell) -> str:
    return "=" + (cell.formula or "")


class GenericXlsxAdapter:
    """Read one unambiguous visible explicit-schema sheet from audited XLSX."""

    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "generic_xlsx"
    adapter_version: ClassVar[str] = "0.1.0"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        adapter_version,
        "Generic XLSX table",
        (".xlsx",),
        True,
        True,
        True,
        True,
    )

    def probe(self, path: Path) -> DetectionResult:
        """Check bounded package/worksheet structure; extensions remain supporting evidence."""
        try:
            preflight_xlsx(path)
        except ParseError as error:
            return DetectionResult(False, 0.0, error.message)
        reason = "audited OOXML package contains one safely mapped .xlsx workbook"
        confidence = 0.99
        if path.suffix.casefold() == ".xlsx":
            reason += "; extension is consistent"
            confidence = 1.0
        return DetectionResult(True, confidence, reason)

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        """Select and parse one audited sheet without evaluating formula caches."""
        package = preflight_xlsx(path)
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (KeyboardInterrupt, SystemExit, MemoryError):
            raise
        except (
            OSError,
            ValueError,
            KeyError,
            TypeError,
            InvalidFileException,
            zipfile.BadZipFile,
        ) as error:
            raise ParseError(
                "XLSX_OPEN_FAILED",
                f"Could not open audited XLSX workbook ({type(error).__name__}).",
            ) from error
        try:
            worksheets = list(workbook.worksheets)
            if len(worksheets) != len(package.sheets) or any(
                str(worksheet.title) != part.title
                for worksheet, part in zip(worksheets, package.sheets, strict=True)
            ):
                raise ParseError(
                    "XLSX_WORKBOOK_MAP_MISMATCH",
                    "openpyxl worksheet mapping does not match the audited package map.",
                )
            indexed = list(zip(worksheets, package.sheets, strict=True))
            if options.sheet is not None:
                indexed = [item for item in indexed if str(item[0].title) == options.sheet]
                if not indexed:
                    raise ParseError(
                        "XLSX_SHEET_NOT_FOUND", f"Sheet {options.sheet!r} does not exist."
                    )
            else:
                indexed = [
                    item
                    for item in indexed
                    if options.include_hidden_sheets or item[1].state == "visible"
                ]

            compatible: list[tuple[Any, SheetPart, list[object]]] = []
            invalid_schemas: list[ParseError] = []
            for worksheet, part in indexed:
                audit = part.worksheet
                worksheet.reset_dimensions()
                if audit.actual_max_row == 0 or audit.actual_max_column == 0:
                    header: list[object] = []
                else:
                    header_cells = next(
                        worksheet.iter_rows(
                            min_row=1,
                            max_row=1,
                            min_col=1,
                            max_col=audit.actual_max_column,
                        ),
                        (),
                    )
                    header = _trimmed_header([cell.value for cell in header_cells])
                try:
                    schema_is_compatible = is_compatible_header(header)
                    semantic_headers(header)
                except ParseError as error:
                    invalid_schemas.append(error)
                    continue
                if schema_is_compatible:
                    compatible.append((worksheet, part, header))

            if not compatible:
                if len(invalid_schemas) == 1:
                    raise invalid_schemas[0]
                if len(invalid_schemas) > 1:
                    raise AdapterAmbiguityError(
                        "XLSX_SHEET_AMBIGUOUS",
                        "Multiple candidate sheets contain invalid documented schemas; "
                        "choose one explicitly.",
                    )
                if options.sheet is not None:
                    raise ParseError(
                        "XLSX_SHEET_SCHEMA_UNRECOGNIZED",
                        f"Sheet {options.sheet!r} has no documented Ordifile schema.",
                    )
                raise ParseError(
                    "XLSX_NO_COMPATIBLE_SHEET",
                    "No compatible visible sheet with a documented Ordifile schema was found.",
                )
            if len(compatible) > 1:
                names = ", ".join(repr(item[1].title) for item in compatible)
                raise AdapterAmbiguityError(
                    "XLSX_SHEET_AMBIGUOUS",
                    f"Multiple compatible sheets were found ({names}); choose one explicitly.",
                )

            worksheet, part, header = compatible[0]
            captured = capture_worksheet_cells(path, package, part)
            raw_cells = {(cell.row, cell.column): cell for cell in captured.raw_cells}
            audited_header = list(header)
            for column_number in range(1, len(audited_header) + 1):
                raw_header = raw_cells.get((1, column_number))
                if raw_header is None:
                    continue
                if raw_header.cell_type == "inlineStr":
                    audited_header[column_number - 1] = raw_header.inline_text
                elif raw_header.cell_type == "s":
                    audited_header[column_number - 1] = raw_header.shared_text
            header = _trimmed_header(audited_header)
            if not is_compatible_header(header):
                raise ParseError(
                    "XLSX_AUDITED_HEADER_MISMATCH",
                    "The audited raw string header does not match the documented schema.",
                )
            mapped = semantic_headers(header)
            formula_cells = {
                (cell.row, cell.column) for cell in captured.raw_cells if cell.formula_present
            }
            cell_sources = {
                (cell.row, cell.column): f"sheet:{part.index}:cell:{cell.coordinate}"
                for cell in captured.raw_cells
            }
            date_cells: list[tuple[RawCell, str | None]] = []
            iso_date_cells: list[tuple[RawCell, str | None]] = []
            typed_mismatches: list[tuple[RawCell, str]] = []

            def selected_rows() -> Iterable[list[object]]:
                worksheet.reset_dimensions()
                rows = worksheet.iter_rows(
                    min_row=1,
                    max_row=captured.actual_max_row,
                    min_col=1,
                    max_col=captured.actual_max_column,
                )
                for row_number, row in enumerate(rows, start=1):
                    values: list[object] = []
                    for column_number, cell in enumerate(row, start=1):
                        raw = raw_cells.get((row_number, column_number))
                        semantic = (
                            mapped[column_number - 1] if column_number <= len(mapped) else None
                        )
                        if raw is None:
                            values.append(None)
                        elif raw.formula_present:
                            values.append(None)
                        elif semantic in _TEXT_FIELDS and raw.cell_type not in _STRING_CELL_TYPES:
                            typed_mismatches.append((raw, semantic))
                            if raw.cell_type == "d":
                                iso_date_cells.append((raw, semantic))
                            values.append(None)
                        elif semantic in _NUMERIC_FIELDS and raw.cell_type not in {
                            *_STRING_CELL_TYPES,
                            "n",
                        }:
                            typed_mismatches.append((raw, semantic))
                            if raw.cell_type == "d":
                                iso_date_cells.append((raw, semantic))
                            values.append(None)
                        elif semantic == "acquired_at" and not (
                            raw.cell_type in _STRING_CELL_TYPES
                            or raw.cell_type == "d"
                            or raw.cell_type == "n"
                            and bool(cell.is_date)
                        ):
                            typed_mismatches.append((raw, semantic))
                            values.append(None)
                        elif raw.cell_type == "inlineStr":
                            values.append(raw.inline_text)
                        elif raw.cell_type == "s":
                            values.append(raw.shared_text)
                        elif raw.cell_type == "d":
                            iso_date_cells.append((raw, semantic))
                            if semantic == "acquired_at" and raw.value is not None:
                                values.append(raw.value if "T" in raw.value else None)
                            elif semantic in _NUMERIC_FIELDS:
                                values.append(None)
                            else:
                                values.append(raw.value)
                        elif bool(cell.is_date):
                            date_cells.append((raw, semantic))
                            if semantic == "acquired_at":
                                values.append(cell.value)
                            elif semantic in _NUMERIC_FIELDS:
                                values.append(None)
                            elif raw.value_present:
                                values.append(raw.value)
                            else:
                                values.append(cell.value)
                        elif raw.cell_type == "n":
                            values.append(raw.value if raw.value_present else None)
                        else:
                            values.append(cell.value)
                    yield values

            bundle = parse_rows(
                path,
                selected_rows(),
                namespace=f"adapter:{self.adapter_id}:sheet:{part.title}",
                source_label=str(part.title),
                formula_cells=formula_cells,
                cell_sources=cell_sources,
            )
            sample_id = bundle.samples[0].sample_id
            namespace = f"adapter:{self.adapter_id}:sheet:{part.title}"
            extra_metadata: list[MetadataEntry] = []
            extra_issues: list[Issue] = []
            formula_extra_rows: set[int] = set()
            for raw in captured.raw_cells:
                if not raw.formula_present:
                    continue
                source = f"sheet:{part.index}:cell:{raw.coordinate}"
                key = _cell_metadata_key(raw.column, header, mapped)
                if raw.column > len(header):
                    formula_extra_rows.add(raw.row)
                extra_metadata.append(
                    MetadataEntry(
                        sample_id,
                        path.name,
                        namespace,
                        key,
                        _formula_literal(raw),
                        source=source,
                    )
                )
                extra_metadata.extend(
                    (
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            "xlsx_formula_lexeme",
                            raw.formula if raw.formula is not None else "",
                            source=source,
                        ),
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            "xlsx_cell_type",
                            raw.cell_type,
                            source=source,
                        ),
                    )
                )
                if raw.style_index is not None:
                    extra_metadata.append(
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            "xlsx_style_index",
                            raw.style_index,
                            source=source,
                        )
                    )
                if raw.value_present:
                    extra_metadata.append(
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            "xlsx_formula_cached_lexeme",
                            raw.value,
                            source=source,
                        )
                    )
                for attribute, value in raw.formula_attributes:
                    extra_metadata.append(
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            f"xlsx_formula_attribute:{attribute}",
                            value,
                            source=source,
                        )
                    )
            for row_number in sorted(formula_extra_rows):
                extra_issues.append(
                    Issue(
                        "EXTRA_CELLS_PRESERVED",
                        f"Row {row_number} contains formula cells beyond the header; they "
                        "were preserved as positional Metadata.",
                        Severity.WARNING,
                        f"sheet:{part.index}:row:{row_number}",
                    )
                )

            for raw, typed_semantic in typed_mismatches:
                source = f"sheet:{part.index}:cell:{raw.coordinate}"
                if raw.cell_type != "d":
                    extra_metadata.append(
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            typed_semantic,
                            raw.value if raw.value is not None else "",
                            source=source,
                        )
                    )
                extra_metadata.append(
                    MetadataEntry(
                        sample_id,
                        path.name,
                        namespace,
                        "xlsx_cell_type",
                        raw.cell_type,
                        source=source,
                    )
                )
                extra_issues.append(
                    Issue(
                        "XLSX_CELL_TYPE_FIELD_MISMATCH",
                        f"Cell {raw.coordinate} type {raw.cell_type!r} is incompatible with "
                        f"documented field {typed_semantic!r}; its raw lexeme and type were "
                        "preserved without canonical reinterpretation.",
                        Severity.WARNING,
                        source,
                    )
                )

            for raw, semantic in date_cells:
                source = f"sheet:{part.index}:cell:{raw.coordinate}"
                field = semantic or _cell_metadata_key(raw.column, header, mapped)
                if raw.value_present:
                    extra_metadata.append(
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            f"{field}_raw_serial",
                            raw.value,
                            source=source,
                        )
                    )
                extra_metadata.extend(
                    (
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            f"{field}_style_index",
                            raw.style_index if raw.style_index is not None else 0,
                            source=source,
                        ),
                        MetadataEntry(
                            sample_id,
                            path.name,
                            namespace,
                            f"{field}_workbook_epoch",
                            "1904" if package.date_1904 else "1900",
                            source=source,
                        ),
                    )
                )
                if semantic == "acquired_at":
                    extra_issues.append(
                        Issue(
                            "XLSX_DATE_SERIAL_UNRELIABLE",
                            f"Date-style cell {raw.coordinate} was decoded using the workbook "
                            "epoch, but remains unreliable for automatic sorting without a "
                            "timezone.",
                            Severity.WARNING,
                            source,
                        )
                    )
                elif semantic in _NUMERIC_FIELDS:
                    extra_issues.append(
                        Issue(
                            "XLSX_DATE_STYLE_NUMERIC_REJECTED",
                            f"Date-style cell {raw.coordinate} was not treated as numeric "
                            f"scientific field {semantic!r}; its raw serial was preserved.",
                            Severity.WARNING,
                            source,
                        )
                    )
            for raw, semantic in iso_date_cells:
                source = f"sheet:{part.index}:cell:{raw.coordinate}"
                field = semantic or _cell_metadata_key(raw.column, header, mapped)
                extra_metadata.append(
                    MetadataEntry(
                        sample_id,
                        path.name,
                        namespace,
                        f"{field}_raw_iso",
                        raw.value if raw.value is not None else "",
                        source=source,
                    )
                )
                if semantic == "acquired_at" and (raw.value is None or "T" not in raw.value):
                    extra_issues.append(
                        Issue(
                            "XLSX_ISO_DATE_TIMESTAMP_UNSUPPORTED",
                            f"ISO date cell {raw.coordinate} has no explicit time and was "
                            "preserved without assigning an acquisition timestamp.",
                            Severity.WARNING,
                            source,
                        )
                    )
                elif semantic in _NUMERIC_FIELDS and semantic != "acquired_at":
                    extra_issues.append(
                        Issue(
                            "XLSX_ISO_DATE_NUMERIC_REJECTED",
                            f"ISO date cell {raw.coordinate} was not treated as numeric "
                            f"scientific field {semantic!r}; its raw ISO lexeme was preserved.",
                            Severity.WARNING,
                            source,
                        )
                    )
            if captured.dimension_mismatch:
                actual = (
                    "empty"
                    if captured.physical_cells == 0
                    else f"{captured.actual_min_row},{captured.actual_min_column}:"
                    f"{captured.actual_max_row},{captured.actual_max_column}"
                )
                extra_issues.append(
                    Issue(
                        "XLSX_DIMENSION_MISMATCH",
                        "Declared worksheet dimension differs from audited physical cells; "
                        "audited actual bounds were used without trusting dimension metadata.",
                        Severity.WARNING,
                        f"sheet:{part.index}",
                        (("declared", captured.declared_dimension), ("actual", actual)),
                    )
                )
            return replace(
                bundle,
                metadata=bundle.metadata + tuple(extra_metadata),
                warnings=bundle.warnings + tuple(extra_issues),
            )
        except (KeyboardInterrupt, SystemExit, MemoryError, ParseError, AdapterAmbiguityError):
            raise
        except Exception as error:
            raise ParseError(
                "XLSX_PARSE_FAILED",
                f"Audited XLSX content could not be safely interpreted ({type(error).__name__}).",
            ) from error
        finally:
            workbook.close()
