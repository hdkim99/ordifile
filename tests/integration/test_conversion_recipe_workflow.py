# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import sys
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, convert_recipe, plan_conversion, plan_recipe
from ordifile.core.errors import OrdifileError
from ordifile.core.models import ConversionExecutionMode, SortMode
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
)
from ordifile.core.planning import ConversionPlanProblem, ConversionPlanRoute
from ordifile.core.recipe import ConversionRecipe

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_chemstation_result_xml import (  # noqa: E402
    synthetic_result_xml_bytes,
)
from generate_leco_chromatof_472_gcgc_result_txt import (  # noqa: E402
    synthetic_gcgc_result_bytes,
)
from generate_shimadzu_labsolutions_result_ascii import (  # noqa: E402
    synthetic_result_ascii_bytes,
)
from generate_youngin_yl_clarity_result_csv import (  # noqa: E402
    synthetic_result_csv_bytes,
)


def _recipe() -> ConversionRecipe:
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
        area_unit="mV.s",
    )
    profile = PeakTableMappingProfile(
        mapping,
        "Local template",
        profile_id="profile-11111111111111111111111111111111",
    )
    mapping_set = PeakTableMappingSet(
        (profile,),
        set_id="profile-set-22222222222222222222222222222222",
    )
    return ConversionRecipe(
        sort=SortMode.INPUT_ORDER,
        peak_table_mapping_set=mapping_set,
        display_label="Laboratory recipe",
    )


def _manifest(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return {
            str(row[0]): row[1]
            for row in workbook["Manifest"].iter_rows(min_row=2, values_only=True)
        }
    finally:
        workbook.close()


def _workbook_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return "\n".join(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


def test_recipe_routes_mapping_set_and_convert_runs_mandatory_preflight(tmp_path: Path) -> None:
    source = tmp_path / "template-a.csv"
    source.write_text("RT,Area\n1.5,10\n2.5,20\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    recipe = _recipe()

    plan = plan_recipe(source, output, recipe=recipe)
    result = convert_recipe(source, output, recipe=recipe)

    assert plan.entries[0].route is ConversionPlanRoute.USER_MAPPING_PROFILE
    assert plan.options.recipe_schema_version == 1
    assert plan.options.recipe_public_fingerprint_sha256 == recipe.public_fingerprint_sha256
    assert result.options.execution_mode is ConversionExecutionMode.REVALIDATED_PREFLIGHT
    assert result.options.conversion_recipe_schema_version == 1
    assert (
        result.options.conversion_recipe_public_fingerprint_sha256
        == recipe.public_fingerprint_sha256
    )
    assert result.files[0].bundle is not None
    assert [peak.area for peak in result.files[0].bundle.peaks] == [10.0, 20.0]
    manifest = _manifest(output)
    workbook_text = _workbook_text(output)
    assert manifest["conversion_recipe_schema_version"] == 1
    assert (
        manifest["conversion_recipe_public_fingerprint_sha256"] == recipe.public_fingerprint_sha256
    )
    assert recipe.semantic_sha256 not in workbook_text
    assert "Laboratory recipe" not in workbook_text
    assert "Local template" not in workbook_text


def test_recipe_identity_is_part_of_plan_without_changing_default_route(tmp_path: Path) -> None:
    source = tmp_path / "generic.csv"
    source.write_text("sample_id,retention_time,area\na,1,2\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"

    direct = plan_conversion(source, output)
    recipe_plan = plan_recipe(source, output, recipe=ConversionRecipe())

    assert direct.entries[0].route == recipe_plan.entries[0].route
    assert direct.public_summary_sha256 != recipe_plan.public_summary_sha256
    assert direct.options.recipe_schema_version is None
    assert recipe_plan.options.recipe_schema_version == 1


def test_recipe_plan_rejects_changed_effective_recipe_before_parse(tmp_path: Path) -> None:
    source = tmp_path / "template.csv"
    source.write_text("RT,Area\n1,10\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    recipe = _recipe()
    plan = plan_recipe(source, output, recipe=recipe)
    changed = replace(recipe, include_signals=True)

    with pytest.raises(OrdifileError) as captured:
        convert_recipe(source, output, recipe=changed, conversion_plan=plan)

    assert captured.value.code == "CONVERSION_PLAN_STALE"
    assert not output.exists()


def test_recipe_plan_rejects_mapping_set_repair_until_new_preflight(tmp_path: Path) -> None:
    source = tmp_path / "template.csv"
    source.write_text("RT,Area\n1,10\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    recipe = _recipe()
    plan = plan_recipe(source, output, recipe=recipe)
    assert recipe.peak_table_mapping_set is not None
    revised = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("RT", 1),
            ColumnSelector("Peak Area", 2),
            "min",
            PeakTableFormat.CSV,
        ),
        "Revised local template",
        profile_id="profile-33333333333333333333333333333333",
    )
    repaired_set = replace(
        recipe.peak_table_mapping_set,
        profiles=(*recipe.peak_table_mapping_set.profiles, revised),
    )
    changed = replace(recipe, peak_table_mapping_set=repaired_set)

    with pytest.raises(OrdifileError) as captured:
        convert_recipe(source, output, recipe=changed, conversion_plan=plan)

    assert captured.value.code == "CONVERSION_PLAN_STALE"
    assert not output.exists()


def test_recipe_mapping_profile_display_label_change_is_not_stale(
    tmp_path: Path,
) -> None:
    source = tmp_path / "template.csv"
    source.write_text("RT,Area\n1,10\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    recipe = _recipe()
    plan = plan_recipe(source, output, recipe=recipe)
    assert recipe.peak_table_mapping_set is not None
    profile = recipe.peak_table_mapping_set.profiles[0]
    renamed_profile = replace(profile, display_label="Renamed local label")
    renamed_set = replace(recipe.peak_table_mapping_set, profiles=(renamed_profile,))
    renamed_recipe = replace(recipe, peak_table_mapping_set=renamed_set)

    result = convert_recipe(
        source,
        output,
        recipe=renamed_recipe,
        conversion_plan=plan,
    )

    assert recipe.semantic_sha256 == renamed_recipe.semantic_sha256
    assert result.failure_count == 0
    assert output.is_file()


def test_recipe_reuses_same_profile_for_new_values_and_surfaces_drift(tmp_path: Path) -> None:
    first = tmp_path / "batch-one.csv"
    second = tmp_path / "batch-two.csv"
    drifted = tmp_path / "batch-three.csv"
    first.write_text("RT,Area\n1,10\n", encoding="utf-8")
    second.write_text("RT,Area\n9,999\n", encoding="utf-8")
    drifted.write_text("RT,Peak Area\n9,999\n", encoding="utf-8")
    recipe = _recipe()

    first_plan = plan_recipe(first, tmp_path / "first.xlsx", recipe=recipe)
    second_plan = plan_recipe(second, tmp_path / "second.xlsx", recipe=recipe)
    drift_plan = plan_recipe(drifted, tmp_path / "drift.xlsx", recipe=recipe)

    assert first_plan.entries[0].route is ConversionPlanRoute.USER_MAPPING_PROFILE
    assert second_plan.entries[0].route is ConversionPlanRoute.USER_MAPPING_PROFILE
    assert drift_plan.entries[0].route is ConversionPlanRoute.UNROUTED
    assert drift_plan.entries[0].problem is ConversionPlanProblem.MAPPING_SCHEMA_DRIFT
    assert drift_plan.entries[0].mapping_diagnostics


def test_recipe_embedded_mapping_set_never_takes_exact_adapter_ownership(
    tmp_path: Path,
) -> None:
    source = tmp_path / "youngin-result.csv"
    source.write_bytes(synthetic_result_csv_bytes())
    recipe = _recipe()

    plan = plan_recipe(source, tmp_path / "youngin.xlsx", recipe=recipe)

    assert plan.entries[0].route is ConversionPlanRoute.EXACT_ADAPTER
    assert plan.entries[0].adapter_id == "youngin_yl_clarity_result_csv"


def test_recipe_forced_generic_adapter_preserves_exact_owner_precedence(
    tmp_path: Path,
) -> None:
    exact_source = tmp_path / "youngin-result.csv"
    exact_source.write_bytes(synthetic_result_csv_bytes())
    generic_source = tmp_path / "generic.csv"
    generic_source.write_text(
        "sample_id,retention_time,area\nsample,1,10\n",
        encoding="utf-8",
    )
    recipe = ConversionRecipe(adapter="generic_csv")

    exact_output = tmp_path / "exact.xlsx"
    exact_plan = plan_recipe(exact_source, exact_output, recipe=recipe)
    generic_plan = plan_recipe(generic_source, tmp_path / "generic.xlsx", recipe=recipe)
    exact_result = convert_recipe(
        exact_source,
        exact_output,
        recipe=recipe,
        conversion_plan=exact_plan,
    )

    assert exact_plan.entries[0].route is ConversionPlanRoute.EXACT_ADAPTER
    assert exact_plan.entries[0].adapter_id == "youngin_yl_clarity_result_csv"
    assert exact_result.files[0].adapter_id == "youngin_yl_clarity_result_csv"
    assert generic_plan.entries[0].adapter_id == "generic_csv"


def test_recipe_single_mapping_private_digest_is_not_workbook_provenance(
    tmp_path: Path,
) -> None:
    private_header = "PRIVATE-RETENTION-HEADER"
    source = tmp_path / "mapped.csv"
    source.write_text(f"{private_header},Area\n1,10\n", encoding="utf-8")
    mapping = PeakTableMapping(
        ColumnSelector(private_header, 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
    )
    recipe = ConversionRecipe(peak_table_mapping=mapping)
    output = tmp_path / "mapped.xlsx"
    direct_output = tmp_path / "direct-mapped.xlsx"

    result = convert_recipe(source, output, recipe=recipe)
    convert(source, direct_output, peak_table_mapping=mapping)
    manifest = _manifest(output)
    recipe_workbook_text = _workbook_text(output)
    direct_workbook_text = _workbook_text(direct_output)

    assert result.options.peak_table_mapping_sha256 is None
    assert "option_peak_table_mapping_sha256" not in manifest
    assert private_header not in recipe_workbook_text
    assert mapping.semantic_sha256 not in recipe_workbook_text
    assert "mapping_sha256" not in recipe_workbook_text
    assert "mapping_selection_mode" not in recipe_workbook_text
    assert "mapping_profile_id" not in recipe_workbook_text
    assert "mapping_set_id" not in recipe_workbook_text
    assert mapping.semantic_sha256 in direct_workbook_text


def test_recipe_selected_xlsx_title_is_replaced_by_fixed_workbook_marker(
    tmp_path: Path,
) -> None:
    private_title = "PRIVATE-WORKSHEET-CANARY"
    source = tmp_path / "generic.xlsx"
    source_workbook = Workbook()
    source_sheet = source_workbook.active
    source_sheet.title = private_title
    source_sheet.append(("sample_id", "retention_time", "area"))
    source_sheet.append(("sample", 1, 10))
    source_workbook.save(source)
    source_workbook.close()
    recipe = ConversionRecipe(sheet=private_title)
    output = tmp_path / "result.xlsx"

    result = convert_recipe(source, output, recipe=recipe)
    workbook_text = _workbook_text(output)

    assert result.options.sheet == "USER_SELECTED"
    assert private_title not in workbook_text
    assert "USER_SELECTED" in workbook_text


@pytest.mark.parametrize("api_name", ["plan", "convert"])
def test_recipe_api_requires_a_recipe_object_without_output(
    tmp_path: Path,
    api_name: str,
) -> None:
    source = tmp_path / "generic.csv"
    source.write_text("sample_id,retention_time,area\na,1,2\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"

    with pytest.raises(OrdifileError) as captured:
        if api_name == "plan":
            plan_recipe(source, output, recipe=None)  # type: ignore[arg-type]
        else:
            convert_recipe(source, output, recipe=None)  # type: ignore[arg-type]

    assert captured.value.code == "OPTION_TYPE_INVALID"
    assert not output.exists()


def test_recipe_preserves_all_four_exact_result_adapter_owners(tmp_path: Path) -> None:
    sources = (
        ("agilent.xml", synthetic_result_xml_bytes(), "agilent_chemstation_result_xml"),
        (
            "shimadzu.txt",
            synthetic_result_ascii_bytes(),
            "shimadzu_labsolutions_result_ascii",
        ),
        ("youngin.csv", synthetic_result_csv_bytes(), "youngin_yl_clarity_result_csv"),
        (
            "leco.txt",
            synthetic_gcgc_result_bytes(),
            "leco_chromatof_gcxgc_result_txt",
        ),
    )
    paths: list[Path] = []
    for name, data, _adapter_id in sources:
        path = tmp_path / name
        path.write_bytes(data)
        paths.append(path)

    plan = plan_recipe(tuple(paths), tmp_path / "mixed.xlsx", recipe=_recipe())

    assert [entry.route for entry in plan.entries] == [ConversionPlanRoute.EXACT_ADAPTER] * 4
    assert [entry.adapter_id for entry in plan.entries] == [item[2] for item in sources]


def test_recipe_converts_one_mixed_exact_and_two_profile_laboratory_batch(
    tmp_path: Path,
) -> None:
    profile_a = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("RT", 1),
            ColumnSelector("Area", 2),
            "min",
            PeakTableFormat.CSV,
        ),
        "Template A",
        profile_id="profile-aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
    )
    profile_b = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Time", 1),
            ColumnSelector("Integrated", 2),
            "s",
            PeakTableFormat.CSV,
        ),
        "Template B",
        profile_id="profile-bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb",
    )
    recipe = ConversionRecipe(
        sort=SortMode.INPUT_ORDER,
        peak_table_mapping_set=PeakTableMappingSet((profile_a, profile_b)),
    )
    sources: list[Path] = []
    for index, text in enumerate(("RT,Area\n1,10\n", "RT,Area\n2,20\n"), start=1):
        source = tmp_path / f"template-a-{index}.csv"
        source.write_text(text, encoding="utf-8")
        sources.append(source)
    for index, text in enumerate(("Time,Integrated\n3,30\n", "Time,Integrated\n4,40\n"), start=1):
        source = tmp_path / f"template-b-{index}.csv"
        source.write_text(text, encoding="utf-8")
        sources.append(source)
    for name, data in (
        ("agilent.xml", synthetic_result_xml_bytes()),
        ("shimadzu.txt", synthetic_result_ascii_bytes()),
        ("youngin.csv", synthetic_result_csv_bytes()),
        ("leco.txt", synthetic_gcgc_result_bytes()),
    ):
        source = tmp_path / name
        source.write_bytes(data)
        sources.append(source)
    output = tmp_path / "mixed-laboratory-batch.xlsx"

    plan = plan_recipe(tuple(sources), output, recipe=recipe)
    result = convert_recipe(tuple(sources), output, recipe=recipe, conversion_plan=plan)

    assert plan.summary.total_inputs == 8
    assert plan.summary.exact_adapters == 4
    assert plan.summary.mapping_profiles == 4
    assert plan.summary.failed == 0
    assert result.failure_count == 0
    assert sum(len(item.bundle.peaks) for item in result.files if item.bundle is not None) == 15
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Peak_Order_Matrix" in workbook.sheetnames
        assert "Peak_Order_Matrix_2D" in workbook.sheetnames
        assert workbook["Peaks"].max_row == 16
    finally:
        workbook.close()


def test_recipe_api_rejects_non_recipe_objects_before_planning(tmp_path: Path) -> None:
    source = tmp_path / "template.csv"
    source.write_text("RT,Area\n1,10\n", encoding="utf-8")

    with pytest.raises(OrdifileError) as captured:
        plan_recipe(source, tmp_path / "result.xlsx", recipe=object())  # type: ignore[arg-type]

    assert captured.value.code == "OPTION_TYPE_INVALID"
