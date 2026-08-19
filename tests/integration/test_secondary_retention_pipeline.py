from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.core.models import (
    BatchResult,
    DatasetBundle,
    FileResult,
    FileStatus,
    InstrumentMetadata,
    PeakRecord,
    SampleRecord,
    SortDecision,
    SortMode,
    SourceFile,
)
from ordifile.core.validation import validate_bundle
from ordifile.exporters.excel import ExcelExporter

ONE_DIMENSIONAL_PEAK_HEADERS = (
    "sample_id",
    "source_file",
    "channel",
    "detector",
    "peak_number",
    "retention_time",
    "retention_time_unit",
    "area",
    "height",
    "compound",
    "compound_source",
    "status",
    "manufacturer",
    "observation_order",
    "start_time",
    "end_time",
    "area_unit",
    "height_unit",
)


def _file_result(
    path: Path,
    *,
    sample_id: str,
    vendor: str,
    peaks: tuple[PeakRecord, ...],
) -> FileResult:
    path.write_bytes(b"synthetic public fixture")
    source = SourceFile(path, path.name, path.name, path.stat().st_size, "a" * 64, None, 0)
    sample = SampleRecord(
        sample_id,
        source,
        instrument=InstrumentMetadata(instrument_type="GC", vendor=vendor),
    )
    bundle = DatasetBundle((source,), (sample,), peaks=peaks)
    assert validate_bundle(bundle) == ()
    return FileResult(source, FileStatus.SUCCESS, bundle=bundle)


def _peak(
    sample_id: str,
    source_file: str,
    order: int,
    *,
    retention_time: float,
    area: float,
    compound: str,
    secondary_retention_time: float | None = None,
) -> PeakRecord:
    return PeakRecord(
        sample_id,
        source_file,
        peak_number=order,
        retention_time=retention_time,
        retention_time_unit="s",
        area=area,
        height=area / 2,
        compound=compound,
        compound_source="reported",
        status="experimental",
        observation_order=order,
        area_unit="AU",
        height_unit="AU",
        secondary_retention_time=secondary_retention_time,
        secondary_retention_time_unit=("s" if secondary_retention_time is not None else None),
    )


def _batch(files: tuple[FileResult, ...]) -> BatchResult:
    return BatchResult(files, SortDecision(SortMode.INPUT_ORDER, SortMode.INPUT_ORDER, "test"))


def test_one_dimensional_workbook_contract_is_unchanged(tmp_path: Path) -> None:
    source_name = "one-dimensional.csv"
    file_result = _file_result(
        tmp_path / source_name,
        sample_id="one-dimensional",
        vendor="Agilent",
        peaks=(
            _peak(
                "one-dimensional",
                source_name,
                1,
                retention_time=1.0,
                area=10.0,
                compound="A",
            ),
        ),
    )
    output = tmp_path / "one-dimensional.xlsx"

    ExcelExporter().export(_batch((file_result,)), output)

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert tuple(next(workbook["Peaks"].values)) == ONE_DIMENSIONAL_PEAK_HEADERS
        assert "Peak_Order_Matrix_2D" not in workbook.sheetnames
        assert tuple(next(workbook["Peak_Order_Matrix"].values)) == (
            "sample_id",
            "source_file",
            "manufacturer",
            "detector",
            "channel",
            "retention_time_unit",
            "area_unit",
            "peak_1_rt",
            "peak_1_area",
        )
    finally:
        workbook.close()


def test_mixed_one_and_two_dimensional_workbook_preserves_atomic_coordinates(
    tmp_path: Path,
) -> None:
    one_name = "one-dimensional.csv"
    two_name = "two-dimensional.txt"
    one_dimensional = _file_result(
        tmp_path / one_name,
        sample_id="one-dimensional",
        vendor="Agilent",
        peaks=(
            _peak(
                "one-dimensional",
                one_name,
                1,
                retention_time=1.0,
                area=10.0,
                compound="A",
            ),
            _peak(
                "one-dimensional",
                one_name,
                2,
                retention_time=2.0,
                area=20.0,
                compound="B",
            ),
        ),
    )
    two_dimensional = _file_result(
        tmp_path / two_name,
        sample_id="two-dimensional",
        vendor="LECO",
        peaks=(
            _peak(
                "two-dimensional",
                two_name,
                1,
                retention_time=644.0,
                secondary_retention_time=1.170,
                area=4_422_843_053.0,
                compound="2-Octanone",
            ),
            _peak(
                "two-dimensional",
                two_name,
                2,
                retention_time=660.0,
                secondary_retention_time=0.775,
                area=4_061_466_196.0,
                compound="Decane",
            ),
        ),
    )
    output = tmp_path / "mixed.xlsx"

    ExcelExporter().export(_batch((one_dimensional, two_dimensional)), output)

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = tuple(workbook["Peaks"].values)
        assert peaks[0] == (
            *ONE_DIMENSIONAL_PEAK_HEADERS,
            "secondary_retention_time",
            "secondary_retention_time_unit",
        )
        assert peaks[1][-2:] == (None, None)
        assert peaks[3][-2:] == (1.170, "s")
        assert peaks[4][-2:] == (0.775, "s")

        one_matrix = tuple(workbook["Peak_Order_Matrix"].values)
        assert len(one_matrix) == 2
        assert one_matrix[1][2] == "Agilent"
        assert one_matrix[1][7:] == (1.0, 10.0, 2.0, 20.0)

        two_matrix = tuple(workbook["Peak_Order_Matrix_2D"].values)
        assert two_matrix[0][8:] == (
            "peak_1_rt1",
            "peak_1_rt2",
            "peak_1_area",
            "peak_2_rt1",
            "peak_2_rt2",
            "peak_2_area",
        )
        assert two_matrix[1][:8] == (
            "two-dimensional",
            two_name,
            "LECO",
            None,
            None,
            "s",
            "s",
            "AU",
        )
        assert two_matrix[1][8:] == (
            644.0,
            1.170,
            4_422_843_053.0,
            660.0,
            0.775,
            4_061_466_196.0,
        )
        compound_matrix = tuple(workbook["Peak_Matrix"].values)
        assert compound_matrix[0][:3] == ("sample_id", "A_area", "B_area")
        assert compound_matrix[1][:3] == ("one-dimensional", 10.0, 20.0)
    finally:
        workbook.close()


def test_two_dimensional_only_workbook_omits_one_dimensional_order_matrix(
    tmp_path: Path,
) -> None:
    source_name = "two-dimensional.txt"
    file_result = _file_result(
        tmp_path / source_name,
        sample_id="two-dimensional",
        vendor="LECO",
        peaks=(
            _peak(
                "two-dimensional",
                source_name,
                1,
                retention_time=644.0,
                secondary_retention_time=1.170,
                area=4_422_843_053.0,
                compound="2-Octanone",
            ),
        ),
    )
    output = tmp_path / "two-dimensional.xlsx"

    ExcelExporter().export(_batch((file_result,)), output)

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Peak_Order_Matrix" not in workbook.sheetnames
        assert "Peak_Order_Matrix_2D" in workbook.sheetnames
    finally:
        workbook.close()
