from __future__ import annotations

import hashlib
import sys
from pathlib import Path
from typing import ClassVar

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters.base import (
    AdapterDescriptor,
    DetectionResult,
    FormatAdapter,
    ParseOptions,
    SourceIdentityPolicy,
    SupportStatus,
)
from ordifile.adapters.registry import AdapterRegistry, create_registry
from ordifile.api import convert, inspect_file
from ordifile.cli.main import main
from ordifile.core.detection import SOURCE_IDENTITY_PROBE_REASON
from ordifile.core.errors import ParseError
from ordifile.core.models import (
    DatasetBundle,
    Issue,
    MetadataEntry,
    ProgressEvent,
    SampleRecord,
    SeriesKind,
    Severity,
    SignalSeries,
    SourceFile,
)
from ordifile.core.pipeline import (
    SOURCE_IDENTITY_ADAPTER_ERROR_MESSAGE,
    SOURCE_IDENTITY_UNEXPECTED_ERROR_MESSAGE,
    run_pipeline,
)

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_ch_v181 import synthetic_v181_bytes  # noqa: E402
from generate_shimadzu_gcmssolution_qgd import synthetic_qgd_bytes  # noqa: E402
from generate_shimadzu_gcsolution_gcd import synthetic_gcd_bytes  # noqa: E402
from generate_youngin_yl_clarity_prm import synthetic_prm_bytes  # noqa: E402


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _workbook_values(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return tuple(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


def _private_probe_reason(path: Path) -> str:
    return (
        f"private basename={path.name}\n"
        "control=\x1b[2JPRIVATE_PROBE_CONTROL; path=/synthetic-private/path"
    )


def _private_error_message(path: Path) -> str:
    return (
        f"private error basename={path.name}\n"
        "control=\x1b[2JPRIVATE_ERROR_CONTROL; path=/synthetic-private/error"
    )


def _private_error_details(path: Path) -> dict[str, str]:
    return {
        "private_basename": path.name,
        "private_control": "\x1b[2JPRIVATE_ERROR_DETAIL",
        "private_path": "/synthetic-private/error/detail",
    }


def test_source_file_public_reference_defaults_to_relative_path() -> None:
    source = SourceFile(Path("private.csv"), "relative.csv", "private.csv", 1, None, None, 0)
    assert source.public_reference == "relative.csv"
    assert source.public_id is None

    aliased = SourceFile(
        Path("private.csv"),
        "relative.csv",
        "private.csv",
        1,
        "a" * 64,
        None,
        0,
        public_id="source-" + "a" * 64,
    )
    assert aliased.public_reference == "source-" + "a" * 64


def test_private_prm_basename_is_absent_from_api_cli_progress_and_workbook(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    private_name = "private-person-instrument-path.PRM"
    data = synthetic_prm_bytes(channels=((1.0, 2.0, 3.0),))
    source = tmp_path / private_name
    source.write_bytes(data)
    digest = _sha256(data)
    public_reference = f"source-{digest}"

    inspected = inspect_file(source)
    assert inspected.file.source.public_reference == public_reference
    assert inspected.file.source.sha256 == digest
    assert inspected.file.source.path == Path(public_reference)
    assert inspected.file.source.relative_path == public_reference
    assert inspected.file.source.name == public_reference
    assert private_name not in repr(inspected)
    assert inspected.file.bundle is not None
    assert {item.source.public_reference for item in inspected.file.bundle.samples} == {
        public_reference
    }
    assert {item.source.path for item in inspected.file.bundle.samples} == {Path(public_reference)}
    assert {item.source_file for item in inspected.file.bundle.signals} == {public_reference}
    assert {item.source_file for item in inspected.file.bundle.metadata} == {public_reference}
    assert {item.source for item in inspected.file.issues} == {public_reference}

    assert main(["inspect", str(source), "--verbose"]) == 0
    terminal_output = capsys.readouterr().out
    assert public_reference in terminal_output
    assert private_name not in terminal_output

    events: list[ProgressEvent] = []
    output = tmp_path / "public-output.xlsx"
    result = convert(source, output, include_signals=True, progress=events.append)
    assert result.files[0].source.public_reference == public_reference
    assert result.files[0].source.path == Path(public_reference)
    assert result.files[0].source.relative_path == public_reference
    assert result.files[0].source.name == public_reference
    assert private_name not in repr(result)
    processing = [event for event in events if event.stage == "processing"]
    assert [event.source_file for event in processing] == [public_reference]
    assert private_name not in "\n".join(
        event.source_file or "" for event in events if event.source_file is not None
    )
    workbook_text = "\n".join(_workbook_values(output))
    assert public_reference in workbook_text
    assert digest in workbook_text
    assert private_name not in workbook_text


def test_corrupt_private_prm_uses_hash_alias_in_cli_and_workbook(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    private_name = "private-operator-corrupt.prm"
    data = synthetic_prm_bytes()[:-21]
    source = tmp_path / private_name
    source.write_bytes(data)
    public_reference = f"source-{_sha256(data)}"

    assert main(["inspect", str(source), "--verbose"]) == 1
    terminal_output = capsys.readouterr().out
    assert public_reference in terminal_output
    assert private_name not in terminal_output

    output = tmp_path / "corrupt.xlsx"
    result = convert(source, output, include_signals=True)
    assert result.failure_count == 1
    failed = result.files[0]
    assert failed.source.public_reference == public_reference
    assert {issue.source for issue in failed.issues} == {public_reference}
    workbook_text = "\n".join(_workbook_values(output))
    assert public_reference in workbook_text
    assert private_name not in workbook_text


def test_missing_prm_uses_input_order_alias_and_rebinds_discovery_issue(tmp_path: Path) -> None:
    missing = tmp_path / "private-missing.PRM"
    result = run_pipeline((missing,), create_registry(include_external=False))

    assert len(result.files) == 1
    failed = result.files[0]
    assert failed.source.public_reference == "source-input-000001"
    assert failed.source.sha256 is None
    assert {issue.source for issue in failed.issues} == {"source-input-000001"}


def test_forced_hash_policy_is_applied_before_wrong_extension_failure(tmp_path: Path) -> None:
    data = synthetic_prm_bytes()
    source = tmp_path / "private-wrong-extension.bin"
    source.write_bytes(data)
    expected = f"source-{_sha256(data)}"

    result = run_pipeline(
        (source,),
        create_registry(include_external=False),
        forced_adapter="youngin_yl_clarity_prm_raw",
    )

    failed = result.files[0]
    assert failed.source.public_reference == expected
    assert {issue.source for issue in failed.issues} == {expected}


def test_mixed_batch_keeps_generic_reference_and_hashes_only_prm(tmp_path: Path) -> None:
    csv_source = tmp_path / "ordinary.csv"
    csv_source.write_text("sample_id,area\ngeneric,1\n", encoding="utf-8")
    prm_data = synthetic_prm_bytes()
    prm_source = tmp_path / "private-mixed-source.prm"
    prm_source.write_bytes(prm_data)

    output = tmp_path / "mixed.xlsx"
    result = convert(
        (csv_source, prm_source),
        output,
        include_signals=True,
        sort="input_order",
    )

    assert [item.source.public_reference for item in result.files] == [
        "ordinary.csv",
        f"source-{_sha256(prm_data)}",
    ]
    assert result.files[0].source.path == csv_source.resolve()
    assert result.files[0].source.relative_path == "ordinary.csv"
    assert result.files[0].source.name == "ordinary.csv"
    assert result.files[1].source.path == Path(f"source-{_sha256(prm_data)}")
    assert "private-mixed-source.prm" not in repr(result.files[1])
    workbook_text = "\n".join(_workbook_values(output))
    assert "ordinary.csv" in workbook_text
    assert "private-mixed-source.prm" not in workbook_text


def test_two_prm_filename_sort_uses_public_hashes_not_private_basenames(tmp_path: Path) -> None:
    alphabetic_first = synthetic_prm_bytes(channels=((1.0,),))
    alphabetic_last = synthetic_prm_bytes(channels=((3.0,),))
    first = tmp_path / "a-private.prm"
    last = tmp_path / "z-private.prm"
    first.write_bytes(alphabetic_first)
    last.write_bytes(alphabetic_last)
    first_reference = f"source-{_sha256(alphabetic_first)}"
    last_reference = f"source-{_sha256(alphabetic_last)}"
    assert last_reference < first_reference

    output = tmp_path / "sorted.xlsx"
    result = convert((first, last), output, include_signals=True, sort="filename")

    assert [item.source.name for item in result.files] == [last_reference, first_reference]
    assert [item.source.public_reference for item in result.files] == [
        last_reference,
        first_reference,
    ]
    assert [item.sort_key for item in result.files] == [last_reference, first_reference]
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        sample_sources = [
            row[2] for row in workbook["Samples"].iter_rows(min_row=2, values_only=True)
        ]
        assert sample_sources == [last_reference, first_reference]
    finally:
        workbook.close()


class _AdapterAliasAttempt:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "source_identity_test"
    adapter_version: ClassVar[str] = "1"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        adapter_version,
        "Source identity test",
        (".identity-test",),
        True,
        False,
        True,
        True,
        SupportStatus.EXPERIMENTAL,
        (SeriesKind.DECODED_RECORDS,),
        SourceIdentityPolicy.SHA256_ALIAS,
    )

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(True, 1.0, _private_probe_reason(path))

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        del options
        data = path.read_bytes()
        digest = _sha256(data)
        adapter_source = SourceFile(
            path,
            "adapter-private.identity-test",
            "adapter-private.identity-test",
            len(data),
            digest,
            None,
            0,
            public_id="adapter-controlled-alias",
        )
        sample = SampleRecord("synthetic", adapter_source)
        signal = SignalSeries(
            "synthetic",
            "adapter-controlled-alias",
            "structural_channel_001",
            None,
            (0,),
            (1,),
            series_kind=SeriesKind.DECODED_RECORDS,
        )
        metadata = MetadataEntry(
            "synthetic",
            "adapter-controlled-alias",
            "adapter:source_identity_test",
            "structural_value",
            1,
        )
        warning = Issue(
            "SOURCE_IDENTITY_TEST_WARNING",
            "Synthetic warning.",
            Severity.WARNING,
            "adapter-controlled-alias",
        )
        return DatasetBundle(
            (adapter_source,),
            (sample,),
            signals=(signal,),
            metadata=(metadata,),
            warnings=(warning,),
        )


class _RelativeSuccess(_AdapterAliasAttempt):
    adapter_id: ClassVar[str] = "relative_success"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        _AdapterAliasAttempt.adapter_version,
        "Relative source identity test",
        (".identity-test",),
        True,
        False,
        True,
        True,
        SupportStatus.VERIFIED,
        (SeriesKind.DECODED_RECORDS,),
        SourceIdentityPolicy.RELATIVE_PATH,
    )


class _RelativeNonMatch:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "relative_nonmatch"
    adapter_version: ClassVar[str] = "1"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        adapter_version,
        "Relative non-match",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.VERIFIED,
    )

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(False, 0.0, _private_probe_reason(path))

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise AssertionError("the non-matching adapter must not parse")


class _RelativeMatch:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "relative_match"
    adapter_version: ClassVar[str] = "1"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        adapter_version,
        "Relative match",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.VERIFIED,
    )

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(True, 1.0, _private_probe_reason(path))

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise AssertionError("ambiguous adapters must not parse")


class _ShaNonMatch:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "sha_nonmatch"
    adapter_version: ClassVar[str] = "1"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        adapter_version,
        "SHA alias non-match",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.EXPERIMENTAL,
        (SeriesKind.DECODED_RECORDS,),
        SourceIdentityPolicy.SHA256_ALIAS,
    )

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(False, 0.0, _private_probe_reason(path))

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise AssertionError("the non-matching adapter must not parse")


class _RelativeCorrupt(_RelativeMatch):
    adapter_id: ClassVar[str] = "relative_corrupt"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        _RelativeMatch.adapter_version,
        "Relative corrupt source",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.VERIFIED,
    )

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise ParseError(
            "SOURCE_IDENTITY_TEST_RELATIVE_CORRUPT",
            _private_error_message(path),
            details=_private_error_details(path),
        )


class PrivateIdentifyingAdapterFailure(Exception):
    """Synthetic ordinary exception whose class name must be withheld for SHA aliases."""


class _RelativeOrdinaryCorrupt(_RelativeMatch):
    adapter_id: ClassVar[str] = "relative_ordinary_corrupt"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        _RelativeMatch.adapter_version,
        "Relative ordinary corrupt source",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.VERIFIED,
    )

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise PrivateIdentifyingAdapterFailure(_private_error_message(path))


class _RelativeStructuredFailure(_RelativeMatch):
    adapter_id: ClassVar[str] = "relative_structured_failure"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        _RelativeMatch.adapter_version,
        "Relative structured failure",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.VERIFIED,
    )

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise ParseError(
            "RELATIVE_STRUCTURED_FAILURE",
            "Relative structured message remains inspectable.",
            details={"evidence": "relative structured detail"},
        )


class _ShaCorrupt(_ShaNonMatch):
    adapter_id: ClassVar[str] = "sha_corrupt"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        _ShaNonMatch.adapter_version,
        "SHA alias corrupt source",
        (".identity-test",),
        False,
        False,
        False,
        True,
        SupportStatus.EXPERIMENTAL,
        (SeriesKind.DECODED_RECORDS,),
        SourceIdentityPolicy.SHA256_ALIAS,
    )

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(True, 1.0, _private_probe_reason(path))

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise ParseError(
            "SOURCE_IDENTITY_TEST_CORRUPT",
            "The synthetic source is structurally corrupt.",
        )


def test_selected_policy_overrides_mixed_extension_owners_and_adapter_alias(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    data = b"source identity fixture"
    source = tmp_path / "private.identity-test"
    source.write_bytes(data)
    expected = f"source-{_sha256(data)}"
    registry = AdapterRegistry()
    registry.register(_RelativeNonMatch())
    registry.register(_AdapterAliasAttempt())
    monkeypatch.setattr("ordifile.core.pipeline.WARN_INPUT_FILE_BYTES", 1)

    result = run_pipeline((source,), registry)

    item = result.files[0]
    assert item.source.public_reference == expected
    assert item.bundle is not None
    assert item.bundle.sources == (item.source,)
    assert {sample.source.public_reference for sample in item.bundle.samples} == {expected}
    assert {signal.source_file for signal in item.bundle.signals} == {expected}
    assert {entry.source_file for entry in item.bundle.metadata} == {expected}
    assert {issue.source for issue in item.issues} == {expected}
    assert "adapter-controlled-alias" not in repr(item.bundle)
    assert "private.identity-test" not in repr(item.issues)


def test_relative_success_restores_path_and_only_sha_owner_probe_is_redacted(
    tmp_path: Path,
) -> None:
    data = b"relative success with provisional SHA identity"
    source = tmp_path / "ordinary.identity-test"
    source.write_bytes(data)
    registry = AdapterRegistry()
    registry.register(_RelativeSuccess())
    registry.register(_ShaNonMatch())

    inspected = inspect_file(source, registry=registry)

    assert inspected.file.source.public_id is None
    assert inspected.file.source.public_reference == source.name
    assert inspected.file.source.path == source.resolve()
    assert inspected.probes == (
        ("relative_success", 1.0, _private_probe_reason(source)),
        ("sha_nonmatch", 0.0, SOURCE_IDENTITY_PROBE_REASON),
    )


@pytest.mark.parametrize(
    ("adapters", "expected_code"),
    (
        ((_RelativeMatch(), _AdapterAliasAttempt()), "SOURCE_IDENTITY_TEST_WARNING"),
        ((_RelativeNonMatch(), _ShaNonMatch()), "FORMAT_NOT_DETECTED"),
        (
            (_RelativeCorrupt(), _ShaNonMatch()),
            "SOURCE_IDENTITY_TEST_RELATIVE_CORRUPT",
        ),
        ((_RelativeNonMatch(), _ShaCorrupt()), "SOURCE_IDENTITY_TEST_CORRUPT"),
    ),
)
def test_any_hash_owner_hides_private_basename_on_detection_and_parse_failures(
    tmp_path: Path,
    adapters: tuple[FormatAdapter, FormatAdapter],
    expected_code: str,
) -> None:
    private_name = "private-person-and-machine.identity-test"
    data = b"mixed owner privacy fixture"
    source = tmp_path / private_name
    source.write_bytes(data)
    expected = f"source-{_sha256(data)}"
    registry = AdapterRegistry()
    for adapter in adapters:
        registry.register(adapter)

    inspected = inspect_file(source, registry=registry)

    assert inspected.file.source.public_reference == expected
    assert inspected.file.source.path == Path(expected)
    assert inspected.file.source.relative_path == expected
    assert inspected.file.source.name == expected
    assert {issue.code for issue in inspected.file.issues} == {expected_code}
    assert {issue.source for issue in inspected.file.issues} == {expected}
    assert private_name not in repr(inspected)


@pytest.mark.parametrize(
    ("adapters", "expected_cli_exit"),
    (
        ((_AdapterAliasAttempt(),), 0),
        ((_RelativeNonMatch(), _ShaNonMatch()), 1),
        ((_RelativeMatch(), _AdapterAliasAttempt()), 0),
        ((_RelativeCorrupt(), _ShaNonMatch()), 1),
    ),
)
def test_hash_policy_withholds_untrusted_probe_reasons_from_all_public_surfaces(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
    adapters: tuple[FormatAdapter, ...],
    expected_cli_exit: int,
) -> None:
    private_name = "private-probe-source.identity-test"
    data = b"probe reason privacy fixture"
    source = tmp_path / private_name
    source.write_bytes(data)
    expected_source = f"source-{_sha256(data)}"
    registry = AdapterRegistry()
    for adapter in adapters:
        registry.register(adapter)

    inspected = inspect_file(source, registry=registry)
    assert inspected.file.source.public_reference == expected_source
    assert all(reason == SOURCE_IDENTITY_PROBE_REASON for _, _, reason in inspected.probes)

    monkeypatch.setattr("ordifile.api.create_registry", lambda: registry)
    assert main(["inspect", str(source), "--verbose"]) == expected_cli_exit
    terminal_output = capsys.readouterr().out

    output = tmp_path / "probe-reason-privacy.xlsx"
    converted = convert(source, output, include_signals=True, registry=registry)
    workbook_text = "\n".join(_workbook_values(output))

    public_surfaces = (repr(inspected), terminal_output, repr(converted), workbook_text)
    for rendered in public_surfaces:
        assert private_name not in rendered
        assert "PRIVATE_PROBE_CONTROL" not in rendered
        assert "/synthetic-private/path" not in rendered
    if inspected.probes:
        assert SOURCE_IDENTITY_PROBE_REASON in repr(inspected)
        assert SOURCE_IDENTITY_PROBE_REASON in terminal_output
    else:
        assert SOURCE_IDENTITY_ADAPTER_ERROR_MESSAGE in repr(inspected.file.issues)
        assert SOURCE_IDENTITY_ADAPTER_ERROR_MESSAGE in terminal_output
        assert SOURCE_IDENTITY_ADAPTER_ERROR_MESSAGE in workbook_text


def test_relative_path_policy_preserves_existing_probe_reason_behavior(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "relative-probe-source.identity-test"
    source.write_bytes(b"relative probe reason fixture")
    registry = AdapterRegistry()
    registry.register(_RelativeSuccess())
    expected_reason = _private_probe_reason(source)

    inspected = inspect_file(source, registry=registry)
    assert inspected.file.source.public_reference == source.name
    assert inspected.probes == (("relative_success", 1.0, expected_reason),)

    monkeypatch.setattr("ordifile.api.create_registry", lambda: registry)
    assert main(["inspect", str(source), "--verbose"]) == 0
    terminal_output = capsys.readouterr().out
    assert source.name in terminal_output
    assert "PRIVATE_PROBE_CONTROL" in terminal_output
    assert "/synthetic-private/path" in terminal_output


@pytest.mark.parametrize(
    ("adapter", "expected_code", "expected_message"),
    (
        (
            _RelativeCorrupt(),
            "SOURCE_IDENTITY_TEST_RELATIVE_CORRUPT",
            SOURCE_IDENTITY_ADAPTER_ERROR_MESSAGE,
        ),
        (
            _RelativeOrdinaryCorrupt(),
            "ADAPTER_UNEXPECTED_ERROR",
            SOURCE_IDENTITY_UNEXPECTED_ERROR_MESSAGE,
        ),
    ),
)
def test_hash_policy_withholds_structured_and_ordinary_adapter_error_details(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
    adapter: FormatAdapter,
    expected_code: str,
    expected_message: str,
) -> None:
    private_name = "private-adapter-error.identity-test"
    data = b"adapter error privacy fixture"
    source = tmp_path / private_name
    source.write_bytes(data)
    expected_source = f"source-{_sha256(data)}"
    registry = AdapterRegistry()
    registry.register(adapter)
    registry.register(_ShaNonMatch())

    inspected = inspect_file(source, registry=registry)
    assert inspected.file.source.public_reference == expected_source
    assert inspected.file.status.value == "failed"
    issue = inspected.file.issues[0]
    assert issue.code == expected_code
    assert issue.message == expected_message
    assert issue.context == ()
    assert all(reason == SOURCE_IDENTITY_PROBE_REASON for _, _, reason in inspected.probes)

    monkeypatch.setattr("ordifile.api.create_registry", lambda: registry)
    assert main(["inspect", str(source), "--verbose"]) == 1
    terminal_output = capsys.readouterr().out

    output = tmp_path / "adapter-error-privacy.xlsx"
    converted = convert(source, output, registry=registry)
    workbook_text = "\n".join(_workbook_values(output))

    assert converted.files[0].issues[0].message == expected_message
    for rendered in (repr(inspected), terminal_output, repr(converted), workbook_text):
        assert expected_code in rendered
        assert expected_message in rendered
        assert private_name not in rendered
        assert "PRIVATE_ERROR_CONTROL" not in rendered
        assert "PRIVATE_ERROR_DETAIL" not in rendered
        assert "/synthetic-private/error" not in rendered
        assert "PrivateIdentifyingAdapterFailure" not in rendered


def test_relative_path_policy_preserves_structured_and_ordinary_error_behavior(
    tmp_path: Path,
) -> None:
    structured_source = tmp_path / "relative-structured.identity-test"
    structured_source.write_bytes(b"relative structured error fixture")
    structured_registry = AdapterRegistry()
    structured_registry.register(_RelativeStructuredFailure())

    structured = inspect_file(structured_source, registry=structured_registry)
    structured_issue = structured.file.issues[0]
    assert structured_issue.code == "RELATIVE_STRUCTURED_FAILURE"
    assert structured_issue.message == "Relative structured message remains inspectable."
    assert structured_issue.context == (("evidence", "relative structured detail"),)

    ordinary_source = tmp_path / "relative-ordinary.identity-test"
    ordinary_source.write_bytes(b"relative ordinary error fixture")
    ordinary_registry = AdapterRegistry()
    ordinary_registry.register(_RelativeOrdinaryCorrupt())

    ordinary = inspect_file(ordinary_source, registry=ordinary_registry)
    ordinary_issue = ordinary.file.issues[0]
    assert ordinary_issue.code == "ADAPTER_UNEXPECTED_ERROR"
    assert "PrivateIdentifyingAdapterFailure" in ordinary_issue.message


def test_generic_agilent_and_shimadzu_keep_default_relative_path_provenance(
    tmp_path: Path,
) -> None:
    fixtures = (
        ("ordinary.csv", b"sample_id,area\ngeneric,1\n", "generic_csv"),
        ("FID1A.CH", synthetic_v181_bytes(), "agilent_chemstation_ch_v181"),
        ("profile.gcd", synthetic_gcd_bytes(), "shimadzu_gcsolution_gcd"),
        ("profile.qgd", synthetic_qgd_bytes(), "shimadzu_gcmssolution_qgd"),
    )

    for filename, data, adapter_id in fixtures:
        source = tmp_path / filename
        source.write_bytes(data)
        inspected = inspect_file(source)

        assert inspected.file.adapter_id == adapter_id
        assert inspected.file.source.public_id is None
        assert inspected.file.source.public_reference == filename
        assert inspected.file.source.path == source.resolve()
        assert inspected.file.source.relative_path == filename
        assert inspected.file.source.name == filename
