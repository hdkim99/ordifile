from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.cli.main import main

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_shimadzu_gcmssolution_qgd import (  # noqa: E402
    RT_INTERVAL_MS,
    RT_START_MS,
    SCAN_COUNT,
    synthetic_qgd_bytes,
)


def test_inspect_reports_one_experimental_tic_and_ms1_limitation(tmp_path: Path) -> None:
    source = tmp_path / "synthetic.qgd"
    source.write_bytes(synthetic_qgd_bytes())

    inspected = inspect_file(source)

    assert inspected.file.adapter_id == "shimadzu_gcmssolution_qgd"
    assert inspected.file.source.detected_format == "shimadzu_gcmssolution_qgd"
    assert inspected.file.bundle is not None
    assert len(inspected.file.bundle.signals) == 1
    assert {issue.code for issue in inspected.file.issues} == {
        "SHIMADZU_QGD_EXPERIMENTAL_PROFILE",
        "QGD_MS1_NOT_EXPORTED",
    }


def test_cli_reports_qgd_tic_as_scientific_signal(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "synthetic.qgd"
    source.write_bytes(synthetic_qgd_bytes())

    assert main(["inspect", str(source)]) == 0

    output = capsys.readouterr().out
    assert "Scientific signals: 1" in output
    assert "Decoded record series: 0" in output
    assert "QGD_MS1_NOT_EXPORTED" in output


def test_valid_and_corrupt_qgd_files_are_isolated_without_tic_truncation(
    tmp_path: Path,
) -> None:
    valid = tmp_path / "valid.qgd"
    corrupt = tmp_path / "corrupt.qgd"
    valid.write_bytes(synthetic_qgd_bytes())
    corrupt.write_bytes(synthetic_qgd_bytes(header_scan_overrides={10: 11}))
    output = tmp_path / "result.xlsx"

    result = convert(tmp_path, output, extensions=(".qgd",), include_signals=True)

    assert result.success_count == 1
    assert result.failure_count == 1
    assert output.is_file()
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Signals_MS" in workbook.sheetnames
        assert all("MS1" not in sheet for sheet in workbook.sheetnames)
        sheet = workbook["Signals_MS"]
        assert sheet.max_row == SCAN_COUNT + 1
        rows = sheet.iter_rows(values_only=True)
        assert next(rows) == (
            "sample_id",
            "source_file",
            "channel",
            "detector",
            "x",
            "x_label",
            "x_unit",
            "y",
            "y_label",
            "y_unit",
        )
        first = next(rows)
        assert first[0] == "valid"
        assert first[2] == "TIC"
        assert first[3] == "MS"
        assert first[4] == pytest.approx(RT_START_MS / 60_000)
        assert first[6] == "min"
        assert first[7] == 1_000
        assert first[8] == "raw_tic_intensity"
        assert first[9] is None
        last = None
        row_count = 1
        for row in rows:
            last = row
            row_count += 1
        assert row_count == SCAN_COUNT
        assert last is not None
        assert last[4] == pytest.approx((RT_START_MS + (SCAN_COUNT - 1) * RT_INTERVAL_MS) / 60_000)
        assert last[7] == 1_000 + (SCAN_COUNT - 1) % 1_000

        import_log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert len(import_log) == 2
        assert {row[4] for row in import_log} == {"warning", "failed"}
        failed = next(row for row in import_log if row[4] == "failed")
        assert failed[6] == "SHIMADZU_QGD_MS1_INVALID"
    finally:
        workbook.close()
