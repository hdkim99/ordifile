# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import runpy
from pathlib import Path
from typing import ClassVar, cast

import openpyxl  # type: ignore[import-untyped]
import pytest

from ordifile.adapters.base import AdapterDescriptor, DetectionResult, ParseOptions
from ordifile.adapters.generic_csv import GenericCsvAdapter
from ordifile.adapters.registry import AdapterRegistry
from ordifile.api import convert, inspect_file, inspect_inputs, preview_peak_table
from ordifile.core.errors import OrdifileError
from ordifile.core.models import DatasetBundle, FileStatus, PeakRecord, SampleRecord, SourceFile
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
    clone_peak_table_mapping_profile,
)


def _csv_profile(
    *,
    profile_id: str = "profile-11111111111111111111111111111111",
    rt_unit: str = "min",
) -> PeakTableMappingProfile:
    return PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Template A RT", 1),
            ColumnSelector("Template A Area", 2),
            rt_unit,
            PeakTableFormat.CSV,
            area_unit="mV.s",
            ignored_columns=(ColumnSelector("Template A Note", 3),),
        ),
        "Template A",
        profile_id=profile_id,
    )


def _xlsx_profile() -> PeakTableMappingProfile:
    return PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Template B Time", 2),
            ColumnSelector("Template B Response Area", 3),
            "s",
            PeakTableFormat.XLSX,
            area_unit=None,
            peak_index_column=ColumnSelector("Template B Peak", 1),
        ),
        "Template B",
        profile_id="profile-22222222222222222222222222222222",
    )


def _write_csv(path: Path, values: tuple[tuple[str, str], ...]) -> None:
    rows = "".join(f"{rt},{area},local-only\n" for rt, area in values)
    path.write_text(
        "Template A RT,Template A Area,Template A Note\n" + rows,
        encoding="utf-8",
    )


def _write_xlsx(path: Path, values: tuple[tuple[int, float, float], ...]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Changing Run Sheet"
    worksheet.append(("Template B Peak", "Template B Time", "Template B Response Area"))
    for row in values:
        worksheet.append(row)
    workbook.save(path)


def _synthetic_leco_bytes() -> bytes:
    generator = runpy.run_path(
        str(
            Path(__file__).parents[1]
            / "fixtures/synthetic/generate_leco_chromatof_472_gcgc_result_txt.py"
        )
    )
    return cast(bytes, generator["synthetic_gcgc_result_bytes"]())


def _synthetic_bytes(generator_name: str, function_name: str) -> bytes:
    generator = runpy.run_path(
        str(Path(__file__).parents[1] / "fixtures" / "synthetic" / generator_name)
    )
    return cast(bytes, generator[function_name]())


def test_mapping_set_routes_multiple_generic_templates_into_one_workbook(tmp_path: Path) -> None:
    csv_one = tmp_path / "template-a-1.csv"
    csv_two = tmp_path / "template-a-2.csv"
    xlsx_one = tmp_path / "template-b-1.xlsx"
    xlsx_two = tmp_path / "template-b-2.xlsx"
    agilent = tmp_path / "exact-agilent.xml"
    shimadzu = tmp_path / "exact-shimadzu.txt"
    youngin = tmp_path / "exact-youngin.csv"
    leco = tmp_path / "exact-leco.txt"
    output = tmp_path / "combined.xlsx"
    _write_csv(csv_one, (("1.0", "10"), ("2.0", "20")))
    _write_csv(csv_two, (("3.0", "30"),))
    _write_xlsx(xlsx_one, ((1, 4.0, 40.0),))
    _write_xlsx(xlsx_two, ((1, 5.0, 50.0), (2, 6.0, 60.0)))
    agilent.write_bytes(
        _synthetic_bytes("generate_agilent_chemstation_result_xml.py", "synthetic_result_xml_bytes")
    )
    shimadzu.write_bytes(
        _synthetic_bytes(
            "generate_shimadzu_labsolutions_result_ascii.py", "synthetic_result_ascii_bytes"
        )
    )
    youngin.write_bytes(
        _synthetic_bytes("generate_youngin_yl_clarity_result_csv.py", "synthetic_result_csv_bytes")
    )
    leco.write_bytes(_synthetic_leco_bytes())
    mapping_set = PeakTableMappingSet((_csv_profile(), _xlsx_profile()))

    result = convert(
        (agilent, shimadzu, youngin, leco, csv_one, csv_two, xlsx_one, xlsx_two),
        output,
        peak_table_mapping_set=mapping_set,
    )
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = tuple(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        manifest = {row[0]: row[1] for row in workbook["Manifest"].iter_rows(values_only=True)}
        import_rows = tuple(workbook["Import_Log"].iter_rows(values_only=True))
        metadata_values = {
            row[4] for row in workbook["Metadata"].iter_rows(min_row=2, values_only=True)
        }
        workbook_strings = {
            value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for value in row
            if isinstance(value, str)
        }
    finally:
        workbook.close()

    assert result.failure_count == 0
    assert len(peaks) > 6
    assert {item.mapping_route for item in result.files} == {
        "EXACT_ADAPTER",
        "USER_MAPPING_PROFILE",
    }
    assert {
        item.mapping_profile_id
        for item in result.files
        if item.mapping_route == "USER_MAPPING_PROFILE"
    } == {
        _csv_profile().profile_id,
        _xlsx_profile().profile_id,
    }
    assert {item.adapter_id for item in result.files if item.mapping_route == "EXACT_ADAPTER"} == {
        "agilent_chemstation_result_xml",
        "shimadzu_labsolutions_result_ascii",
        "youngin_yl_clarity_result_csv",
        "leco_chromatof_gcxgc_result_txt",
    }
    assert manifest["option_peak_table_mapping_set_id"] == mapping_set.set_id
    assert manifest["option_peak_table_mapping_set_fingerprint"] == (
        mapping_set.structural_fingerprint_sha256
    )
    import_headers = tuple(import_rows[0])
    route_index = import_headers.index("conversion_route")
    assert "USER_MAPPING_PROFILE" in {row[route_index] for row in import_rows[1:]}
    assert "PROFILE_SET" in metadata_values
    assert "Changing Run Sheet" not in repr(import_rows)
    assert not any("Template A" in value for value in workbook_strings)
    assert not any("Template B" in value for value in workbook_strings)
    assert not any("Changing Run Sheet" in value for value in workbook_strings)
    assert _csv_profile().mapping.semantic_sha256 not in workbook_strings


def test_mapping_set_no_match_fails_without_generic_fallback(tmp_path: Path) -> None:
    source = tmp_path / "unmatched.csv"
    source.write_text("time,signal\n1,2\n", encoding="utf-8")

    result = inspect_file(
        source,
        peak_table_mapping_set=PeakTableMappingSet((_csv_profile(),)),
    ).file

    assert result.status is FileStatus.FAILED
    assert result.mapping_route == "NO_MAPPING_MATCH"
    assert {issue.code for issue in result.issues} == {"PEAK_MAPPING_PROFILE_NOT_MATCHED"}
    assert result.bundle is None
    assert len(result.mapping_diagnostics) == 1
    assert result.mapping_diagnostics[0].profile_id == _csv_profile().profile_id
    assert result.source.public_reference == f"source-{result.source.sha256}"
    assert source.name not in repr(result)


def test_mapping_set_reports_schema_drift_without_applying_a_candidate(tmp_path: Path) -> None:
    source = tmp_path / "changed.csv"
    source.write_text(
        "Template A Time,Template A Area,Template A Note\n1,2,local-only\n",
        encoding="utf-8",
    )

    result = inspect_file(
        source,
        peak_table_mapping_set=PeakTableMappingSet((_csv_profile(),)),
    ).file

    assert result.status is FileStatus.FAILED
    assert result.mapping_route == "SCHEMA_DRIFT_CANDIDATE"
    assert result.bundle is None
    assert result.mapping_profile_id is None
    assert len(result.mapping_diagnostics) == 1
    diagnostic = result.mapping_diagnostics[0]
    assert diagnostic.profile_id == _csv_profile().profile_id
    assert diagnostic.changed_column_count == 1
    assert diagnostic.unresolved_required_roles == ("retention_time",)
    assert source.name not in repr(result)


def test_schema_drift_import_log_contains_only_fixed_categories(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    drift = tmp_path / "private-drift.csv"
    output = tmp_path / "partial.xlsx"
    _write_csv(good, (("1", "2"),))
    drift.write_text(
        "Private Header Canary,Template A Area,Template A Note\n1,2,local-only\n",
        encoding="utf-8",
    )

    result = convert(
        (good, drift),
        output,
        peak_table_mapping_set=PeakTableMappingSet((_csv_profile(),)),
    )
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        rows = tuple(workbook["Import_Log"].iter_rows(values_only=True))
        strings = {
            value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for value in row
            if isinstance(value, str)
        }
    finally:
        workbook.close()

    headers = tuple(rows[0])
    route_index = headers.index("conversion_route")
    count_index = headers.index("mapping_diagnostic_candidates")
    category_index = headers.index("mapping_diagnostic_categories")
    drift_row = next(row for row in rows[1:] if row[route_index] == "SCHEMA_DRIFT_CANDIDATE")
    assert result.failure_count == 1
    assert drift_row[count_index] == 1
    assert "HEADER_CHANGED_UNRESOLVED" in drift_row[category_index]
    assert "Private Header Canary" not in strings
    assert "Template A RT" not in strings
    assert "Private Profile Label" not in strings
    assert drift.name not in strings


def test_repaired_profile_routes_old_and_new_templates_separately(tmp_path: Path) -> None:
    old = tmp_path / "old.csv"
    new = tmp_path / "new.csv"
    output = tmp_path / "old-new.xlsx"
    _write_csv(old, (("1", "10"),))
    new.write_text(
        "Template A Time,Template A Area,Template A Note\n2,20,local-only\n",
        encoding="utf-8",
    )
    original = PeakTableMappingSet((_csv_profile(),))
    preview = preview_peak_table(new, PeakTableFormat.CSV, row_limit=1)
    repaired_mapping = PeakTableMapping(
        ColumnSelector("Template A Time", 1),
        ColumnSelector("Template A Area", 2),
        "min",
        PeakTableFormat.CSV,
        area_unit="mV.s",
        ignored_columns=(ColumnSelector("Template A Note", 3),),
    )
    updated = clone_peak_table_mapping_profile(
        original,
        parent_profile_id=_csv_profile().profile_id,
        observed_preview=preview,
        repaired_mapping=repaired_mapping,
        display_label="Template A revised",
    )

    result = convert((old, new), output, peak_table_mapping_set=updated)

    assert result.failure_count == 0
    assert {item.mapping_profile_id for item in result.files} == {
        original.profiles[0].profile_id,
        updated.profiles[-1].profile_id,
    }
    assert original.profiles == (_csv_profile(),)
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 3
        assert workbook["Peak_Order_Matrix"].max_row == 3
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("suffix", "delimiter", "source_format", "adapter_id"),
    (
        (".tsv", "\t", PeakTableFormat.TSV, "generic_tsv"),
        (".txt", "\t", PeakTableFormat.TSV, "generic_tsv"),
        (".txt", ";", PeakTableFormat.SEMICOLON, "generic_semicolon"),
    ),
)
def test_mapping_set_reuses_existing_text_container_boundaries(
    tmp_path: Path,
    suffix: str,
    delimiter: str,
    source_format: PeakTableFormat,
    adapter_id: str,
) -> None:
    source = tmp_path / f"mapped{suffix}"
    source.write_text(
        delimiter.join(("RT", "Area")) + "\n" + delimiter.join(("1.5", "20")) + "\n",
        encoding="utf-8-sig",
    )
    profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("RT", 1),
            ColumnSelector("Area", 2),
            "min",
            source_format,
        )
    )

    result = inspect_file(
        source,
        peak_table_mapping_set=PeakTableMappingSet((profile,)),
    ).file

    assert result.status is FileStatus.SUCCESS
    assert result.adapter_id == adapter_id
    assert result.mapping_route == "USER_MAPPING_PROFILE"
    assert result.bundle is not None
    assert (result.bundle.peaks[0].retention_time, result.bundle.peaks[0].area) == (1.5, 20.0)


def test_mapping_set_ambiguous_profiles_fail_closed_and_isolate_sibling(tmp_path: Path) -> None:
    ambiguous = tmp_path / "ambiguous.csv"
    unmatched = tmp_path / "unmatched.csv"
    _write_csv(ambiguous, (("1", "2"),))
    unmatched.write_text("time,signal\n1,2\n", encoding="utf-8")
    mapping_set = PeakTableMappingSet(
        (
            _csv_profile(),
            _csv_profile(
                profile_id="profile-33333333333333333333333333333333",
                rt_unit="s",
            ),
        )
    )

    result = inspect_inputs((ambiguous, unmatched), peak_table_mapping_set=mapping_set)

    assert result.failure_count == 2
    routes = {item.mapping_route for item in result.files}
    assert routes == {"AMBIGUOUS_MAPPING_PROFILE", "NO_MAPPING_MATCH"}


def test_malformed_generic_table_keeps_mapping_validation_route(tmp_path: Path) -> None:
    source = tmp_path / "empty.csv"
    source.write_bytes(b"")

    result = inspect_file(
        source,
        peak_table_mapping_set=PeakTableMappingSet((_csv_profile(),)),
    ).file

    assert result.status is FileStatus.FAILED
    assert result.mapping_route == "MAPPING_VALIDATION_FAILED"
    assert {issue.code for issue in result.issues} == {"MISSING_HEADER"}
    assert result.mapping_diagnostics == ()


def test_mapping_set_single_visible_xlsx_profile_rejects_ambiguous_workbook(
    tmp_path: Path,
) -> None:
    source = tmp_path / "two-sheets.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Run One"
    first.append(("Template B Peak", "Template B Time", "Template B Response Area"))
    first.append((1, 1.0, 10.0))
    second = workbook.create_sheet("Run Two")
    second.append(("Template B Peak", "Template B Time", "Template B Response Area"))
    second.append((1, 2.0, 20.0))
    workbook.save(source)

    result = inspect_file(
        source,
        peak_table_mapping_set=PeakTableMappingSet((_xlsx_profile(),)),
    ).file

    assert result.status is FileStatus.FAILED
    assert result.mapping_route == "AMBIGUOUS_WORKSHEET"
    assert {issue.code for issue in result.issues} == {"PEAK_MAPPING_WORKSHEET_AMBIGUOUS"}


def test_mapping_set_exact_xlsx_sheet_profile_selects_named_structure(tmp_path: Path) -> None:
    source = tmp_path / "named-sheet.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Notes"
    workbook.active.append(("note",))
    results = workbook.create_sheet("Results")
    results.append(("Template B Peak", "Template B Time", "Template B Response Area"))
    results.append((1, 7.0, 70.0))
    workbook.save(source)
    profile = PeakTableMappingProfile(
        _xlsx_profile().mapping,
        "Template B fixed sheet",
        profile_id="profile-44444444444444444444444444444444",
        worksheet_title="Results",
    )

    result = inspect_file(
        source,
        peak_table_mapping_set=PeakTableMappingSet((profile,)),
    ).file

    assert result.status is FileStatus.SUCCESS
    assert result.bundle is not None
    assert result.bundle.peaks[0].retention_time == 7.0
    assert "Results" not in repr(result.source)


def test_exact_leco_adapter_owns_input_before_mapping_set(tmp_path: Path) -> None:
    source = tmp_path / "exact-profile.txt"
    source.write_bytes(_synthetic_leco_bytes())
    mapping_set = PeakTableMappingSet((_csv_profile(),))

    result = inspect_file(source, peak_table_mapping_set=mapping_set).file

    assert result.adapter_id == "leco_chromatof_gcxgc_result_txt"
    assert result.mapping_route == "EXACT_ADAPTER"
    assert result.mapping_profile_id is None
    assert result.mapping_diagnostics == ()
    assert result.bundle is not None
    assert any(issue.code == "PEAK_MAPPING_NOT_APPLIED_EXACT_PROFILE" for issue in result.issues)


def test_mapping_set_conflicts_are_rejected_before_input_discovery(tmp_path: Path) -> None:
    mapping = _csv_profile().mapping
    mapping_set = PeakTableMappingSet((_csv_profile(),))
    missing = tmp_path / "missing.csv"

    with pytest.raises(OrdifileError) as both:
        inspect_file(
            missing,
            peak_table_mapping=mapping,
            peak_table_mapping_set=mapping_set,
        )
    with pytest.raises(OrdifileError) as sheet:
        inspect_file(
            missing,
            sheet="Private title",
            peak_table_mapping_set=mapping_set,
        )
    with pytest.raises(OrdifileError) as adapter:
        inspect_file(
            missing,
            adapter="generic_csv",
            peak_table_mapping_set=mapping_set,
        )

    assert both.value.code == "PEAK_MAPPING_OPTION_CONFLICT"
    assert sheet.value.code == "PEAK_MAPPING_SET_SHEET_CONFLICT"
    assert adapter.value.code == "PEAK_MAPPING_ADAPTER_CONFLICT"


class _LowConfidenceExactCsvAdapter:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "external_exact_csv"
    adapter_version: ClassVar[str] = "1"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id,
        adapter_version,
        "Synthetic external exact CSV",
        (".csv",),
        True,
        True,
        False,
        True,
    )

    def probe(self, path: Path) -> DetectionResult:
        with path.open("rb") as stream:
            matched = stream.read(128).startswith(b"Template A RT,Template A Area,Template A Note")
        return DetectionResult(matched, 0.01 if matched else 0.0, "synthetic exact fixture")

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        del options
        source = SourceFile(path, path.name, path.name, path.stat().st_size, None, None, 0)
        sample = SampleRecord("external-exact", source)
        peak = PeakRecord(
            "external-exact",
            path.name,
            retention_time=9.0,
            retention_time_unit="min",
            area=99.0,
            observation_order=1,
        )
        return DatasetBundle((source,), (sample,), peaks=(peak,))


def test_external_exact_adapter_beats_higher_confidence_generic_profile(tmp_path: Path) -> None:
    source = tmp_path / "result.csv"
    _write_csv(source, (("1", "2"),))
    registry = AdapterRegistry()
    registry.register(GenericCsvAdapter())
    registry.register(_LowConfidenceExactCsvAdapter())

    result = inspect_file(
        source,
        registry=registry,
        peak_table_mapping_set=PeakTableMappingSet((_csv_profile(),)),
    ).file

    assert result.adapter_id == "external_exact_csv"
    assert result.mapping_route == "EXACT_ADAPTER"
    assert result.mapping_diagnostics == ()
    assert result.bundle is not None
    assert result.bundle.peaks[0].retention_time == 9.0
