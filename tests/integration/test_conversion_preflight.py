# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import runpy
from pathlib import Path
from typing import cast

import openpyxl  # type: ignore[import-untyped]
import pytest

from ordifile.api import convert_plan, plan_conversion
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
)
from ordifile.core.planning import (
    ConversionPlanEntryStatus,
    ConversionPlanProblem,
    ConversionPlanReadiness,
    ConversionPlanRoute,
)


def _synthetic_bytes(generator_name: str, function_name: str) -> bytes:
    generator = runpy.run_path(
        str(Path(__file__).parents[1] / "fixtures" / "synthetic" / generator_name)
    )
    return cast(bytes, generator[function_name]())


def _csv_profile() -> PeakTableMappingProfile:
    return PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Template A RT", 1),
            ColumnSelector("Template A Area", 2),
            "min",
            PeakTableFormat.CSV,
            area_unit="mV.s",
            ignored_columns=(ColumnSelector("Template A Note", 3),),
        ),
        "Template A",
        profile_id="profile-11111111111111111111111111111111",
    )


def _xlsx_profile() -> PeakTableMappingProfile:
    return PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Template B Time", 2),
            ColumnSelector("Template B Area", 3),
            "s",
            PeakTableFormat.XLSX,
            peak_index_column=ColumnSelector("Template B Peak", 1),
        ),
        "Template B",
        profile_id="profile-22222222222222222222222222222222",
    )


def _write_csv(path: Path, value: int) -> None:
    path.write_text(
        f"Template A RT,Template A Area,Template A Note\n1,{value},local-only\n",
        encoding="utf-8",
    )


def _write_xlsx(path: Path, value: int) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Local run"
    sheet.append(("Template B Peak", "Template B Time", "Template B Area"))
    sheet.append((1, 2.5, value))
    workbook.save(path)


@pytest.mark.researcher_acceptance
def test_mixed_preflight_routes_exact_profiles_and_failures_then_matches_conversion(
    tmp_path: Path,
) -> None:
    agilent = tmp_path / "agilent.xml"
    shimadzu = tmp_path / "shimadzu.txt"
    youngin = tmp_path / "youngin.csv"
    leco = tmp_path / "leco.txt"
    agilent.write_bytes(
        _synthetic_bytes("generate_agilent_chemstation_result_xml.py", "synthetic_result_xml_bytes")
    )
    shimadzu.write_bytes(
        _synthetic_bytes(
            "generate_shimadzu_labsolutions_result_ascii.py",
            "synthetic_result_ascii_bytes",
        )
    )
    youngin.write_bytes(
        _synthetic_bytes("generate_youngin_yl_clarity_result_csv.py", "synthetic_result_csv_bytes")
    )
    leco.write_bytes(
        _synthetic_bytes(
            "generate_leco_chromatof_472_gcgc_result_txt.py",
            "synthetic_gcgc_result_bytes",
        )
    )

    csv_one = tmp_path / "template-a-1.csv"
    csv_two = tmp_path / "template-a-2.csv"
    xlsx_one = tmp_path / "template-b-1.xlsx"
    xlsx_two = tmp_path / "template-b-2.xlsx"
    _write_csv(csv_one, 10)
    _write_csv(csv_two, 20)
    _write_xlsx(xlsx_one, 30)
    _write_xlsx(xlsx_two, 40)

    drifted = tmp_path / "drifted.csv"
    drifted.write_text(
        "Template A RT,Changed Area,Template A Note\n1,50,local-only\n",
        encoding="utf-8",
    )
    unmapped = tmp_path / "unmapped.tsv"
    unmapped.write_text("Unknown\tColumns\n1\t2\n", encoding="utf-8")
    unsupported = tmp_path / "unsupported.bin"
    unsupported.write_bytes(b"unsupported")
    output = tmp_path / "combined.xlsx"
    mapping_set = PeakTableMappingSet((_csv_profile(), _xlsx_profile()))

    plan = plan_conversion(
        (
            agilent,
            shimadzu,
            youngin,
            leco,
            csv_one,
            csv_two,
            xlsx_one,
            xlsx_two,
            drifted,
            unmapped,
            unsupported,
            csv_one,
        ),
        output,
        peak_table_mapping_set=mapping_set,
        on_error="continue",
    )

    assert plan.readiness is ConversionPlanReadiness.READY_WITH_KNOWN_FAILURES
    assert plan.summary.total_inputs == 12
    assert plan.summary.exact_adapters == 4
    assert plan.summary.mapping_profiles == 4
    assert plan.summary.drifted == 1
    assert plan.summary.unmapped == 1
    assert plan.summary.unsupported == 1
    assert plan.summary.duplicates == 1
    assert plan.summary.failed == 3
    assert [entry.route for entry in plan.entries[:4]] == [ConversionPlanRoute.EXACT_ADAPTER] * 4
    assert [entry.route for entry in plan.entries[4:8]] == [
        ConversionPlanRoute.USER_MAPPING_PROFILE
    ] * 4
    assert plan.entries[8].problem is ConversionPlanProblem.MAPPING_SCHEMA_DRIFT
    assert plan.entries[9].problem is ConversionPlanProblem.UNMAPPED_GENERIC_TABLE
    assert plan.entries[10].problem is ConversionPlanProblem.UNSUPPORTED_FORMAT
    assert plan.entries[11].problem is ConversionPlanProblem.DUPLICATE_INPUT
    assert plan.entries[11].status is ConversionPlanEntryStatus.DUPLICATE
    assert not output.exists()

    converted = convert_plan(plan)

    assert converted.success_count == 8
    assert converted.failure_count == 3
    assert converted.duplicate_count == 1
    assert output.is_file()
    assert {entry.adapter_id for entry in plan.entries[:4]} == {
        "agilent_chemstation_result_xml",
        "shimadzu_labsolutions_result_ascii",
        "youngin_yl_clarity_result_csv",
        "leco_chromatof_gcxgc_result_txt",
    }
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Peaks" in workbook.sheetnames
        assert "Peak_Order_Matrix" in workbook.sheetnames
        assert "Peak_Order_Matrix_2D" in workbook.sheetnames
        assert "Metadata" in workbook.sheetnames
        assert "Import_Log" in workbook.sheetnames
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        import_rows = tuple(workbook["Import_Log"].iter_rows(values_only=True))
        import_headers = tuple(import_rows[0])
        status_index = import_headers.index("status")
        route_index = import_headers.index("conversion_route")
        statuses = [row[status_index] for row in import_rows[1:]]
        routes = [row[route_index] for row in import_rows[1:]]
        assert statuses.count("success") + statuses.count("warning") == plan.summary.routable
        assert statuses.count("failed") == plan.summary.failed
        assert statuses.count("duplicate") == plan.summary.duplicates
        assert routes.count("EXACT_ADAPTER") == plan.summary.exact_adapters
        assert routes.count("USER_MAPPING_PROFILE") == plan.summary.mapping_profiles
        assert "SCHEMA_DRIFT_CANDIDATE" in routes
        assert "NO_MAPPING_MATCH" in routes
        assert manifest["execution_mode"] == "REVALIDATED_PREFLIGHT"
        assert manifest["failure_count"] == plan.summary.failed
        assert manifest["duplicate_count"] == plan.summary.duplicates
        assert manifest["conversion_plan_schema_version"] == plan.schema_version
        assert manifest["conversion_plan_public_summary_sha256"] == plan.public_summary_sha256
    finally:
        workbook.close()


def test_xlsx_preflight_surfaces_worksheet_ambiguity_without_switching_sheets(
    tmp_path: Path,
) -> None:
    source = tmp_path / "ambiguous.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "First local sheet"
    first.append(("Template B Peak", "Template B Time", "Template B Area"))
    second = workbook.create_sheet("Second local sheet")
    second.append(("Template B Peak", "Template B Time", "Template B Area"))
    workbook.save(source)

    plan = plan_conversion(
        source,
        tmp_path / "result.xlsx",
        peak_table_mapping_set=PeakTableMappingSet((_xlsx_profile(),)),
    )

    assert plan.entries[0].route is ConversionPlanRoute.UNROUTED
    assert plan.entries[0].problem is ConversionPlanProblem.WORKSHEET_AMBIGUOUS
    assert plan.summary.ambiguous == 1


def test_preflight_blocks_recognized_unsupported_and_malformed_exact_owners(
    tmp_path: Path,
) -> None:
    unsupported = tmp_path / "unsupported-agilent.xml"
    unsupported.write_bytes(
        _synthetic_bytes(
            "generate_agilent_chemstation_result_xml.py",
            "synthetic_result_xml_bytes",
        ).replace(
            "Rev. C.01.10 [201] Copyright © Agilent Technologies".encode("utf-16-le"),
            "Rev. C.01.09 [200] Copyright © Agilent Technologies".encode("utf-16-le"),
        )
    )
    malformed = tmp_path / "malformed-leco.txt"
    malformed.write_bytes(
        _synthetic_bytes(
            "generate_leco_chromatof_472_gcgc_result_txt.py",
            "synthetic_gcgc_result_bytes",
        ).replace(b"1st Dimension Time (s)", b"retention_time", 1)
    )

    plan = plan_conversion(
        (unsupported, malformed),
        tmp_path / "blocked.xlsx",
        on_error="continue",
    )

    assert plan.readiness is ConversionPlanReadiness.READY_WITH_KNOWN_FAILURES
    assert plan.summary.routable == 0
    assert plan.summary.failed == 2
    assert plan.entries[0].problem is ConversionPlanProblem.UNSUPPORTED_FORMAT
    assert plan.entries[0].issue_codes == ("AGILENT_RESULT_XML_VERSION_UNSUPPORTED",)
    assert plan.entries[1].problem is ConversionPlanProblem.MALFORMED_INPUT
    assert plan.entries[1].issue_codes == ("LECO_GCGC_RESULT_HEADER_INVALID",)
    assert not (tmp_path / "blocked.xlsx").exists()
