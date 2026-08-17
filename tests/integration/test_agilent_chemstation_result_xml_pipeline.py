from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters.agilent_chemstation_result_xml import (
    AgilentChemStationResultXmlAdapter,
)
from ordifile.adapters.base import ParseOptions
from ordifile.api import convert, inspect_file
from ordifile.cli.main import main
from ordifile.core.models import DatasetBundle, FileStatus

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_ch_v181 import synthetic_v181_bytes  # noqa: E402
from generate_agilent_chemstation_result_xml import (  # noqa: E402
    synthetic_result_xml_bytes,
)


def test_inspect_and_cli_expose_result_only_peaks_without_private_source(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    private_basename = "private-person-result.xml"
    source = tmp_path / private_basename
    data = synthetic_result_xml_bytes()
    source.write_bytes(data)

    inspected = inspect_file(source)
    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"
    assert inspected.file.adapter_id == "agilent_chemstation_result_xml"
    assert inspected.file.source.public_reference == expected_reference
    assert inspected.file.source.name == expected_reference
    assert inspected.file.source.path == Path(expected_reference)
    assert len(inspected.file.bundle.peaks) == 3  # type: ignore[union-attr]
    assert inspected.file.bundle.signals == ()  # type: ignore[union-attr]

    assert main(["inspect", str(source), "--verbose"]) == 0
    output = capsys.readouterr().out
    assert "Peaks: 3" in output
    assert "Scientific signals: 0" in output
    assert private_basename not in output


def test_standalone_result_workbook_preserves_order_and_units(tmp_path: Path) -> None:
    private_basename = "private-person-result.xml"
    source = tmp_path / private_basename
    data = synthetic_result_xml_bytes()
    source.write_bytes(data)
    output = tmp_path / "result.xlsx"

    result = convert(source, output)
    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert "Peak_Order_Matrix" in workbook.sheetnames
        peaks = list(workbook["Peaks"].values)
        headers = peaks[0]
        assert headers[12:] == (
            "manufacturer",
            "observation_order",
            "start_time",
            "end_time",
            "area_unit",
            "height_unit",
        )
        assert len(peaks) == 4
        assert {row[headers.index("manufacturer")] for row in peaks[1:]} == {"Agilent"}
        assert [row[headers.index("observation_order")] for row in peaks[1:]] == [1, 2, 3]
        assert {row[headers.index("area_unit")] for row in peaks[1:]} == {"pA*s"}
        assert {row[headers.index("height_unit")] for row in peaks[1:]} == {"pA"}

        matrix = list(workbook["Peak_Order_Matrix"].values)
        assert matrix[0][:7] == (
            "sample_id",
            "source_file",
            "manufacturer",
            "detector",
            "channel",
            "retention_time_unit",
            "area_unit",
        )
        assert matrix[0][7:] == (
            "peak_1_rt",
            "peak_1_area",
            "peak_2_rt",
            "peak_2_area",
            "peak_3_rt",
            "peak_3_area",
        )
        assert matrix[1][2:7] == ("Agilent", "FID", "FID1A", "min", "pA*s")
        assert matrix[1][7:] == (1.25, 100.5, 2.5, 200.75, 3.75, 300)
        all_values = (
            value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        assert not any(private_basename in str(value) for value in all_values)
    finally:
        workbook.close()


def test_valid_and_corrupt_result_files_are_isolated(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    (inputs / "private-valid.xml").write_bytes(synthetic_result_xml_bytes())
    (inputs / "private-corrupt.xml").write_bytes(synthetic_result_xml_bytes()[:-10])
    output = tmp_path / "isolated.xlsx"

    result = convert(inputs, output, extensions=(".xml",))
    assert result.failure_count == 1
    assert sum(item.bundle is not None for item in result.files) == 1
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 4
        assert workbook["Peak_Order_Matrix"].max_row == 2
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"warning", "failed"}
        assert not any(
            private in str(value)
            for row in log
            for value in row
            for private in ("private-valid", "private-corrupt")
        )
    finally:
        workbook.close()


def test_changed_then_restored_result_xml_is_excluded_by_adapter_read_hash(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "private-race.xml"
    original = synthetic_result_xml_bytes()
    changed = synthetic_result_xml_bytes(peaks=(("1", "2", "3", "0.5", "1.5", None),))
    source.write_bytes(original)
    original_parse = AgilentChemStationResultXmlAdapter.parse

    def parse_changed_then_restore(
        self: AgilentChemStationResultXmlAdapter, path: Path, options: ParseOptions
    ) -> DatasetBundle:
        path.write_bytes(changed)
        try:
            return original_parse(self, path, options)
        finally:
            path.write_bytes(original)

    monkeypatch.setattr(AgilentChemStationResultXmlAdapter, "parse", parse_changed_then_restore)
    output = tmp_path / "race.xlsx"

    result = convert(source, output)

    assert result.files[0].status is FileStatus.FAILED
    assert result.files[0].bundle is None
    assert {issue.code for issue in result.files[0].issues} == {"INPUT_CHANGED_DURING_PARSE"}
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 1
        assert workbook["Metadata"].max_row == 1
    finally:
        workbook.close()


def test_mixed_generic_raw_and_result_batch_keeps_existing_outputs(tmp_path: Path) -> None:
    result_xml = tmp_path / "private-result.xml"
    result_xml.write_bytes(synthetic_result_xml_bytes())
    raw_ch = tmp_path / "FID1A.CH"
    raw_ch.write_bytes(synthetic_v181_bytes())
    generic = tmp_path / "generic.csv"
    generic.write_text(
        "sample_id,detector,channel,retention_time,retention_time_unit,area,compound\n"
        "generic,FID,FID-A,1,min,2,generic-compound\n",
        encoding="utf-8",
    )
    output = tmp_path / "mixed.xlsx"

    result = convert((generic, raw_ch, result_xml), output, include_signals=True)
    assert result.failure_count == 0
    assert {item.adapter_id for item in result.files} == {
        "generic_csv",
        "agilent_chemstation_ch_v181",
        "agilent_chemstation_result_xml",
    }
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        order_rows = list(workbook["Peak_Order_Matrix"].values)
        assert len(order_rows) == 2
        assert order_rows[1][2:5] == ("Agilent", "FID", "FID1A")
        peak_matrix = list(workbook["Peak_Matrix"].values)
        assert len(peak_matrix[0]) >= 3
        peaks = list(workbook["Peaks"].values)
        compound_column = peaks[0].index("compound")
        assert {row[compound_column] for row in peaks[1:]} >= {
            "generic-compound",
            "compound-alpha",
        }
        assert "Signals_Records_FID" in workbook.sheetnames
    finally:
        workbook.close()
