# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

"""Create and execute a checkout-free, public-safe standalone smoke kit."""

from __future__ import annotations

import argparse
import hashlib
import json
import runpy
import sys
import tempfile
from collections.abc import Mapping, Sequence
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from ordifile import (
    ConversionRecipe,
    __version__,
    clone_peak_table_mapping_profile,
    load_conversion_recipe,
    save_conversion_recipe,
)
from ordifile.api import (
    convert,
    convert_plan,
    convert_recipe,
    get_format_report,
    inspect_inputs,
    list_formats,
    plan_conversion,
    plan_recipe,
    preview_peak_table,
)
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
    load_peak_table_mapping,
    load_peak_table_mapping_set,
    save_peak_table_mapping,
    save_peak_table_mapping_set,
)

SCHEMA_VERSION = 4
SCIENTIFIC_SHEETS = (
    "Samples",
    "Peak_Matrix",
    "Peaks",
    "Peak_Order_Matrix",
    "Metadata",
    "Import_Log",
)
CONDITIONAL_SCIENTIFIC_SHEETS = ("Peak_Order_Matrix_2D",)
GENERIC_NAME = "generic.csv"
GENERIC_BOM_NAME = "generic BOM.csv"
GENERIC_TSV_NAME = "generic.tsv"
GENERIC_SEMICOLON_NAME = "generic.txt"
GENERIC_XLSX_NAME = "generic.xlsx"
AGILENT_RAW_NAME = "FID1A.CH"
AGILENT_V179_RAW_NAME = "FID3A.CH"
AGILENT_RESULT_NAME = "agilent-result.xml"
SHIMADZU_GCD_NAME = "shimadzu.gcd"
SHIMADZU_QGD_NAME = "shimadzu.qgd"
SHIMADZU_RESULT_NAME = "shimadzu-result.txt"
LECO_GCXGC_RESULT_NAME = "leco-gcxgc-result.txt"
YOUNGIN_RAW_NAME = "youngin.prm"
YOUNGIN_NAME = "youngin-result.csv"
EXPECTED_NAME = "expected.json"
MAPPED_NAME = "explicit-mapped.csv"
MAPPING_NAME = "peak-mapping.json"
MAPPED_SET_CSV_NAME = "mapped-template-a.csv"
MAPPED_SET_XLSX_NAME = "mapped-template-b.xlsx"
MAPPING_SET_NAME = "peak-mapping-set.json"
RECIPE_NAME = "conversion-recipe.json"
CP949_PROBE_TEXT = "합성 보고서"
GENERATED_INPUTS = (
    (
        AGILENT_RAW_NAME,
        "generate_agilent_ch_v181.py",
        "synthetic_v181_bytes",
        "agilent_chemstation_ch_v181",
    ),
    (
        AGILENT_V179_RAW_NAME,
        "generate_agilent_ch_v179.py",
        "synthetic_v179_bytes",
        "agilent_chemstation_ch_v179",
    ),
    (
        AGILENT_RESULT_NAME,
        "generate_agilent_chemstation_result_xml.py",
        "synthetic_result_xml_bytes",
        "agilent_chemstation_result_xml",
    ),
    (
        SHIMADZU_GCD_NAME,
        "generate_shimadzu_gcsolution_gcd.py",
        "synthetic_gcd_bytes",
        "shimadzu_gcsolution_gcd",
    ),
    (
        SHIMADZU_QGD_NAME,
        "generate_shimadzu_gcmssolution_qgd.py",
        "synthetic_qgd_bytes",
        "shimadzu_gcmssolution_qgd",
    ),
    (
        SHIMADZU_RESULT_NAME,
        "generate_shimadzu_labsolutions_result_ascii.py",
        "synthetic_result_ascii_bytes",
        "shimadzu_labsolutions_result_ascii",
    ),
    (
        LECO_GCXGC_RESULT_NAME,
        "generate_leco_chromatof_472_gcgc_result_txt.py",
        "synthetic_gcgc_result_bytes",
        "leco_chromatof_gcxgc_result_txt",
    ),
    (
        YOUNGIN_RAW_NAME,
        "generate_youngin_yl_clarity_prm.py",
        "synthetic_prm_bytes",
        "youngin_yl_clarity_prm_raw",
    ),
    (
        YOUNGIN_NAME,
        "generate_youngin_yl_clarity_result_csv.py",
        "synthetic_result_csv_bytes",
        "youngin_yl_clarity_result_csv",
    ),
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_value(value: object) -> object:
    if value is None or type(value) in {bool, int, float, str}:
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def semantic_digest(path: Path) -> tuple[str, tuple[str, ...]]:
    """Hash ordered scientific cells while excluding container metadata."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        missing = tuple(sheet for sheet in SCIENTIFIC_SHEETS if sheet not in workbook.sheetnames)
        if missing:
            raise ValueError("The workbook is missing required scientific smoke sheets.")
        conditional_sheets = tuple(
            sheet
            for sheet in workbook.sheetnames
            if any(
                sheet == logical_name or sheet.startswith(f"{logical_name}_")
                for logical_name in CONDITIONAL_SCIENTIFIC_SHEETS
            )
        )
        digest_sheets = (*SCIENTIFIC_SHEETS, *conditional_sheets)
        payload = {
            sheet: [
                [_json_value(value) for value in row]
                for row in workbook[sheet].iter_rows(values_only=True)
            ]
            for sheet in digest_sheets
        }
        serialized = json.dumps(
            payload, ensure_ascii=True, allow_nan=False, sort_keys=True, separators=(",", ":")
        ).encode("ascii")
        return hashlib.sha256(serialized).hexdigest(), tuple(workbook.sheetnames)
    finally:
        workbook.close()


def validate_workbook_presentation(path: Path) -> None:
    """Verify the researcher-facing workbook presentation without reading values twice."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.active.title != "Samples":
            raise ValueError("The workbook does not open on the Samples sheet.")
        expected_freeze = {
            "Manifest": "A2",
            "Samples": "C2",
            "Peak_Matrix": "B2",
            "Peaks": "C2",
            "Metadata": "C2",
            "Import_Log": "B2",
        }
        for name, freeze_panes in expected_freeze.items():
            if workbook[name].freeze_panes != freeze_panes:
                raise ValueError("A workbook sheet has an unexpected frozen-header policy.")
        if workbook["Samples"].auto_filter.ref is None:
            raise ValueError("The Samples sheet has no researcher-facing filter.")
        manifest = {
            row[0]: row[1]
            for row in workbook["Manifest"].iter_rows(min_row=2, values_only=True)
            if row[0] is not None
        }
        if manifest.get("sample_record_count") != workbook["Samples"].max_row - 1:
            raise ValueError("The workbook sample summary differs from the Samples sheet.")
        if manifest.get("peak_record_count") != workbook["Peaks"].max_row - 1:
            raise ValueError("The workbook peak summary differs from the Peaks sheet.")
    finally:
        workbook.close()


def _write_json(path: Path, payload: Mapping[str, object]) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n",
        encoding="ascii",
    )
    temporary.replace(path)


def _generated_bytes(generator_root: Path, script: str, callable_name: str) -> bytes:
    generator = generator_root / script
    if not generator.is_file():
        raise ValueError("A synthetic standalone generator is unavailable.")
    original_path = tuple(sys.path)
    try:
        sys.path.insert(0, str(generator_root))
        namespace = runpy.run_path(str(generator))
    finally:
        sys.path[:] = original_path
    generate = namespace.get(callable_name)
    if not callable(generate):
        raise ValueError("A synthetic standalone generator has no supported callable.")
    value = generate()
    if not isinstance(value, bytes):
        raise TypeError("A synthetic standalone generator returned a non-byte value.")
    return value


def create_smoke_kit(output: Path, generator_root: Path) -> None:
    """Create public-safe inputs for every built-in adapter plus semantic evidence."""
    if output.exists():
        raise ValueError("The smoke-kit output already exists.")
    if not generator_root.is_dir():
        raise ValueError("The synthetic generator directory is unavailable.")
    output.mkdir(parents=True)
    generic = output / GENERIC_NAME
    generic.write_text(
        "sample_id,retention_time,area,compound\n"
        "generic-a,1.25,100.5,synthetic-a\n"
        "generic-b,2.50,150.75,synthetic-b\n",
        encoding="utf-8",
        newline="",
    )
    generic_bom = output / GENERIC_BOM_NAME
    generic_bom.write_text(
        "sample_id,retention_time,area,compound\n"
        "generic-bom-a,3.25,200.5,synthetic-bom-a\n"
        "generic-bom-b,4.50,250.75,synthetic-bom-b\n",
        encoding="utf-8-sig",
        newline="",
    )
    generic_tsv = output / GENERIC_TSV_NAME
    generic_tsv.write_text(
        "sample_id\tretention_time\tarea\tcompound\ngeneric-tsv\t3.75\t275.5\tsynthetic-tsv\n",
        encoding="utf-8",
        newline="",
    )
    generic_semicolon = output / GENERIC_SEMICOLON_NAME
    generic_semicolon.write_text(
        "sample_id;retention_time;area;compound\n"
        "generic-semicolon;4.75;375.5;synthetic-semicolon\n",
        encoding="utf-8",
        newline="",
    )
    xlsx_namespace = runpy.run_path(str(generator_root / "generate_xlsx.py"))
    generate_xlsx = xlsx_namespace.get("generate_fixture")
    if not callable(generate_xlsx):
        raise ValueError("The synthetic XLSX generator has no supported callable.")
    generic_xlsx = output / GENERIC_XLSX_NAME
    generate_xlsx(generic_xlsx)
    mapped_input = output / MAPPED_NAME
    mapped_input.write_text(
        "Peak,Declared RT,Declared Area,Note\n1,1.5,42,ignored\n2,2.75,84,ignored\n",
        encoding="utf-8",
        newline="",
    )
    mapping = PeakTableMapping(
        retention_time_column=ColumnSelector("Declared RT", 2),
        area_column=ColumnSelector("Declared Area", 3),
        retention_time_unit="min",
        source_format=PeakTableFormat.CSV,
        peak_index_column=ColumnSelector("Peak", 1),
        ignored_columns=(ColumnSelector("Note", 4),),
    )
    save_peak_table_mapping(mapping, output / MAPPING_NAME)
    mapped_set_csv = output / MAPPED_SET_CSV_NAME
    mapped_set_csv.write_text(
        "Template A RT,Template A Area\n1.0,10\n2.0,20\n",
        encoding="utf-8",
        newline="",
    )
    mapped_set_xlsx = output / MAPPED_SET_XLSX_NAME
    mapped_set_workbook = Workbook()
    mapped_set_sheet = mapped_set_workbook.active
    mapped_set_sheet.title = "Changing Run Sheet"
    mapped_set_sheet.append(("Template B Peak", "Template B Time", "Template B Area"))
    mapped_set_sheet.append((1, 3.0, 30.0))
    mapped_set_workbook.save(mapped_set_xlsx)
    mapping_set = PeakTableMappingSet(
        (
            PeakTableMappingProfile(
                PeakTableMapping(
                    ColumnSelector("Template A RT", 1),
                    ColumnSelector("Template A Area", 2),
                    "min",
                    PeakTableFormat.CSV,
                ),
                "Template A",
                profile_id="profile-11111111111111111111111111111111",
            ),
            PeakTableMappingProfile(
                PeakTableMapping(
                    ColumnSelector("Template B Time", 2),
                    ColumnSelector("Template B Area", 3),
                    "s",
                    PeakTableFormat.XLSX,
                    peak_index_column=ColumnSelector("Template B Peak", 1),
                ),
                "Template B",
                profile_id="profile-22222222222222222222222222222222",
            ),
        ),
        set_id="profile-set-33333333333333333333333333333333",
    )
    save_peak_table_mapping_set(mapping_set, output / MAPPING_SET_NAME)
    recipe = ConversionRecipe(
        peak_table_mapping_set=mapping_set,
        display_label="Synthetic mapping-set recipe",
    )
    save_conversion_recipe(recipe, output / RECIPE_NAME)

    generated_paths: list[Path] = []
    generated_expectations: dict[str, str] = {}
    for name, script, callable_name, adapter_id in GENERATED_INPUTS:
        data = _generated_bytes(generator_root, script, callable_name)
        if name == YOUNGIN_NAME:
            marker = CP949_PROBE_TEXT.encode("cp949")
            data = data.replace(b"Synthetic report", marker, 1)
            if marker not in data:
                raise ValueError("The synthetic YoungIn smoke input has no CP949-only marker.")
        path = output / name
        path.write_bytes(data)
        generated_paths.append(path)
        generated_expectations[name] = adapter_id

    inputs = (
        generic,
        generic_bom,
        generic_tsv,
        generic_semicolon,
        generic_xlsx,
        *generated_paths,
    )
    expected_adapters = {
        GENERIC_NAME: "generic_csv",
        GENERIC_BOM_NAME: "generic_csv",
        GENERIC_TSV_NAME: "generic_tsv",
        GENERIC_SEMICOLON_NAME: "generic_semicolon",
        GENERIC_XLSX_NAME: "generic_xlsx",
        **generated_expectations,
    }

    report = get_format_report()
    if report.load_errors:
        raise ValueError("External adapter loading failed in the smoke-kit environment.")
    adapter_ids = tuple(item.adapter_id for item in list_formats())
    with tempfile.TemporaryDirectory(prefix="ordifile-standalone-kit-") as temporary:
        workbook = Path(temporary) / "expected.xlsx"
        result = convert(inputs, workbook, sort="input_order")
        if result.failure_count:
            raise ValueError("Synthetic smoke-kit conversion failed.")
        digest, sheets = semantic_digest(workbook)
        mapped_workbook = Path(temporary) / "mapped-expected.xlsx"
        mapped_result = convert(mapped_input, mapped_workbook, peak_table_mapping=mapping)
        if mapped_result.failure_count:
            raise ValueError("Synthetic mapped standalone conversion failed.")
        mapped_digest, mapped_sheets = semantic_digest(mapped_workbook)
        mapped_set_workbook_path = Path(temporary) / "mapped-set-expected.xlsx"
        mapped_set_result = convert(
            (mapped_set_csv, mapped_set_xlsx),
            mapped_set_workbook_path,
            peak_table_mapping_set=mapping_set,
        )
        if mapped_set_result.failure_count:
            raise ValueError("Synthetic mapping-set standalone conversion failed.")
        mapped_set_digest, mapped_set_sheets = semantic_digest(mapped_set_workbook_path)
        recipe_workbook_path = Path(temporary) / "recipe-expected.xlsx"
        recipe_result = convert_recipe(
            (mapped_set_csv, mapped_set_xlsx),
            recipe_workbook_path,
            recipe=recipe,
        )
        if recipe_result.failure_count:
            raise ValueError("Synthetic recipe standalone conversion failed.")
        recipe_digest, recipe_sheets = semantic_digest(recipe_workbook_path)
    expected = {
        "schema_version": SCHEMA_VERSION,
        "ordifile_version": __version__,
        "adapter_ids": list(adapter_ids),
        "input_order": [path.name for path in inputs],
        "inputs": {
            path.name: {
                "sha256": _sha256(path),
                "adapter_id": expected_adapters[path.name],
            }
            for path in inputs
        },
        "scientific_sheets": list(SCIENTIFIC_SHEETS),
        "workbook_sheets": list(sheets),
        "semantic_sha256": digest,
        "mapped": {
            "input_sha256": _sha256(mapped_input),
            "mapping_file_sha256": _sha256(output / MAPPING_NAME),
            "mapping_semantic_sha256": mapping.semantic_sha256,
            "adapter_id": "generic_csv",
            "scientific_sheets": list(mapped_sheets),
            "semantic_sha256": mapped_digest,
        },
        "mapping_set": {
            "csv_input_sha256": _sha256(mapped_set_csv),
            "xlsx_input_sha256": _sha256(mapped_set_xlsx),
            "mapping_set_file_sha256": _sha256(output / MAPPING_SET_NAME),
            "mapping_set_fingerprint": mapping_set.structural_fingerprint_sha256,
            "scientific_sheets": list(mapped_set_sheets),
            "semantic_sha256": mapped_set_digest,
        },
        "recipe": {
            "file_sha256": _sha256(output / RECIPE_NAME),
            "public_fingerprint": recipe.public_fingerprint_sha256,
            "scientific_sheets": list(recipe_sheets),
            "semantic_sha256": recipe_digest,
        },
    }
    _write_json(output / EXPECTED_NAME, expected)


def _load_expected(kit: Path) -> dict[str, Any]:
    try:
        value = json.loads((kit / EXPECTED_NAME).read_text(encoding="ascii"))
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise ValueError("The standalone smoke expectation is unreadable.") from error
    if type(value) is not dict or value.get("schema_version") != SCHEMA_VERSION:
        raise ValueError("The standalone smoke expectation schema is invalid.")
    return cast(dict[str, Any], value)


def run_smoke(kit: Path, output: Path, report_path: Path) -> None:
    """Validate one packaged runtime using no checkout or external network service."""
    expected = _load_expected(kit)
    if expected.get("ordifile_version") != __version__:
        raise ValueError("The standalone runtime version does not match its smoke kit.")
    if output.exists() or report_path.exists():
        raise ValueError("Standalone smoke outputs must not already exist.")
    inputs_value = expected.get("inputs")
    if type(inputs_value) is not dict:
        raise ValueError("The standalone smoke input inventory is invalid.")
    paths: list[Path] = []
    expected_detection: list[str] = []
    input_order = expected.get("input_order")
    if (
        type(input_order) is not list
        or not input_order
        or not all(type(name) is str for name in input_order)
    ):
        raise ValueError("The standalone smoke input order is invalid.")
    for name in cast(list[str], input_order):
        item = inputs_value.get(name)
        if type(item) is not dict:
            raise ValueError("The standalone smoke input entry is invalid.")
        path = kit / name
        if not path.is_file() or _sha256(path) != item.get("sha256"):
            raise ValueError("A standalone smoke input failed its checksum.")
        adapter_id = item.get("adapter_id")
        if type(adapter_id) is not str:
            raise ValueError("A standalone smoke adapter expectation is invalid.")
        paths.append(path)
        expected_detection.append(adapter_id)

    codec_probe = CP949_PROBE_TEXT
    if codec_probe.encode("cp949").decode("cp949") != codec_probe:
        raise ValueError("The packaged runtime has no functioning CP949 codec.")
    format_report = get_format_report()
    if format_report.load_errors:
        raise ValueError("The packaged runtime reported external adapter load failures.")
    adapter_ids = tuple(item.adapter_id for item in list_formats())
    if list(adapter_ids) != expected.get("adapter_ids"):
        raise ValueError("The packaged adapter inventory differs from the source expectation.")
    mapped_expected = expected.get("mapped")
    if type(mapped_expected) is not dict:
        raise ValueError("The standalone mapped-input expectation is invalid.")
    mapped_input = kit / MAPPED_NAME
    mapping_path = kit / MAPPING_NAME
    if (
        not mapped_input.is_file()
        or _sha256(mapped_input) != mapped_expected.get("input_sha256")
        or not mapping_path.is_file()
        or _sha256(mapping_path) != mapped_expected.get("mapping_file_sha256")
    ):
        raise ValueError("The standalone mapped-input kit failed its checksum.")
    mapping = load_peak_table_mapping(mapping_path)
    if mapping.semantic_sha256 != mapped_expected.get("mapping_semantic_sha256"):
        raise ValueError("The standalone mapping semantic identity differs from the baseline.")
    mapped_set_expected = expected.get("mapping_set")
    mapped_set_path = kit / MAPPING_SET_NAME
    mapped_set_csv = kit / MAPPED_SET_CSV_NAME
    mapped_set_xlsx = kit / MAPPED_SET_XLSX_NAME
    if (
        type(mapped_set_expected) is not dict
        or not mapped_set_path.is_file()
        or _sha256(mapped_set_path) != mapped_set_expected.get("mapping_set_file_sha256")
        or not mapped_set_csv.is_file()
        or _sha256(mapped_set_csv) != mapped_set_expected.get("csv_input_sha256")
        or not mapped_set_xlsx.is_file()
        or _sha256(mapped_set_xlsx) != mapped_set_expected.get("xlsx_input_sha256")
    ):
        raise ValueError("The standalone mapping-set kit failed its checksum.")
    mapping_set = load_peak_table_mapping_set(mapped_set_path)
    if mapping_set.structural_fingerprint_sha256 != mapped_set_expected.get(
        "mapping_set_fingerprint"
    ):
        raise ValueError("The standalone mapping-set structure differs from the baseline.")
    recipe_expected = expected.get("recipe")
    recipe_path = kit / RECIPE_NAME
    if (
        type(recipe_expected) is not dict
        or not recipe_path.is_file()
        or _sha256(recipe_path) != recipe_expected.get("file_sha256")
    ):
        raise ValueError("The standalone recipe kit failed its checksum.")
    recipe = load_conversion_recipe(recipe_path)
    if recipe.public_fingerprint_sha256 != recipe_expected.get("public_fingerprint"):
        raise ValueError("The standalone recipe fingerprint differs from the baseline.")
    inspected = inspect_inputs(tuple(paths), sort="input_order")
    detected = [item.adapter_id for item in inspected.files]
    if detected != expected_detection or inspected.failure_count:
        raise ValueError("The packaged runtime did not detect the synthetic inputs exactly.")
    conversion_plan = plan_conversion(tuple(paths), output, sort="input_order")
    if (
        not conversion_plan.is_executable
        or conversion_plan.summary.routable != len(paths)
        or output.exists()
        or any(kit.glob(".ordifile_*"))
    ):
        raise ValueError("The packaged runtime preflight changed artifacts or routes.")
    result = convert_plan(conversion_plan)
    if result.failure_count:
        raise ValueError("The packaged standalone conversion failed.")
    output_sha256 = _sha256(output)
    try:
        convert(tuple(paths), output, sort="input_order")
    except (KeyboardInterrupt, SystemExit, MemoryError):
        raise
    except Exception as error:
        if getattr(error, "code", None) != "OUTPUT_EXISTS":
            raise ValueError("The existing-output refusal returned an unexpected error.") from None
    else:
        raise ValueError("The packaged runtime silently replaced an existing workbook.")
    if _sha256(output) != output_sha256:
        raise ValueError("The existing workbook changed during overwrite refusal.")
    digest, sheets = semantic_digest(output)
    if digest != expected.get("semantic_sha256"):
        raise ValueError("The packaged workbook differs from the source semantic baseline.")
    if list(sheets) != expected.get("workbook_sheets"):
        raise ValueError("The packaged workbook sheet inventory differs from the baseline.")
    validate_workbook_presentation(output)
    with tempfile.TemporaryDirectory(prefix="ordifile-standalone-mapping-") as temporary:
        mapped_output = Path(temporary) / "mapped.xlsx"
        mapped_result = convert(mapped_input, mapped_output, peak_table_mapping=mapping)
        if mapped_result.failure_count or tuple(
            item.adapter_id for item in mapped_result.files
        ) != (mapped_expected.get("adapter_id"),):
            raise ValueError("The packaged runtime did not apply explicit peak mapping exactly.")
        mapped_digest, mapped_sheets = semantic_digest(mapped_output)
    if mapped_digest != mapped_expected.get("semantic_sha256") or list(
        mapped_sheets
    ) != mapped_expected.get("scientific_sheets"):
        raise ValueError("The packaged mapped workbook differs from the source baseline.")
    with tempfile.TemporaryDirectory(prefix="ordifile-standalone-mapping-set-") as temporary:
        mapped_set_output = Path(temporary) / "mapped-set.xlsx"
        mapped_set_plan = plan_conversion(
            (mapped_set_csv, mapped_set_xlsx),
            mapped_set_output,
            peak_table_mapping_set=mapping_set,
        )
        if mapped_set_plan.summary.mapping_profiles != 2 or mapped_set_output.exists():
            raise ValueError("The packaged mapping-set preflight did not route exactly.")
        mapped_set_result = convert_plan(mapped_set_plan)
        if mapped_set_result.failure_count or {
            item.mapping_route for item in mapped_set_result.files
        } != {"USER_MAPPING_PROFILE"}:
            raise ValueError("The packaged runtime did not apply the mapping set exactly.")
        mapped_set_digest, mapped_set_sheets = semantic_digest(mapped_set_output)
        drift_input = Path(temporary) / "drift.csv"
        drift_input.write_text(
            "Template A Time,Template A Area\n4.0,40\n",
            encoding="utf-8",
            newline="",
        )
        drift_result = inspect_inputs((drift_input,), peak_table_mapping_set=mapping_set)
        drift_file = drift_result.files[0]
        if (
            drift_file.mapping_route != "SCHEMA_DRIFT_CANDIDATE"
            or drift_file.bundle is not None
            or len(drift_file.mapping_diagnostics) != 1
        ):
            raise ValueError("The packaged runtime did not fail closed on mapping drift.")
        drift_preview = preview_peak_table(drift_input, PeakTableFormat.CSV, row_limit=1)
        repaired_mapping = PeakTableMapping(
            ColumnSelector("Template A Time", 1),
            ColumnSelector("Template A Area", 2),
            "min",
            PeakTableFormat.CSV,
        )
        repaired_set = clone_peak_table_mapping_profile(
            mapping_set,
            parent_profile_id="profile-11111111111111111111111111111111",
            observed_preview=drift_preview,
            repaired_mapping=repaired_mapping,
            display_label="Template A revised",
        )
        if (
            mapping_set.profiles[0].mapping.retention_time_column.label != "Template A RT"
            or repaired_set.set_id != mapping_set.set_id
            or len(repaired_set.profiles) != len(mapping_set.profiles) + 1
        ):
            raise ValueError("The packaged repair changed its immutable parent profile.")
        repaired_output = Path(temporary) / "repaired.xlsx"
        repaired_result = convert(
            drift_input,
            repaired_output,
            peak_table_mapping_set=repaired_set,
        )
        if repaired_result.failure_count or repaired_result.files[0].mapping_route != (
            "USER_MAPPING_PROFILE"
        ):
            raise ValueError("The packaged runtime did not apply the confirmed repaired profile.")
    if mapped_set_digest != mapped_set_expected.get("semantic_sha256") or list(
        mapped_set_sheets
    ) != mapped_set_expected.get("scientific_sheets"):
        raise ValueError("The packaged mapping-set workbook differs from the source baseline.")
    with tempfile.TemporaryDirectory(prefix="ordifile-standalone-recipe-") as temporary:
        recipe_output = Path(temporary) / "recipe.xlsx"
        recipe_plan = plan_recipe(
            (mapped_set_csv, mapped_set_xlsx),
            recipe_output,
            recipe=recipe,
        )
        if recipe_plan.summary.mapping_profiles != 2 or recipe_output.exists():
            raise ValueError("The packaged recipe preflight did not route exactly.")
        recipe_result = convert_plan(recipe_plan)
        if recipe_result.failure_count:
            raise ValueError("The packaged recipe conversion failed.")
        recipe_digest, recipe_sheets = semantic_digest(recipe_output)
    if recipe_digest != recipe_expected.get("semantic_sha256") or list(
        recipe_sheets
    ) != recipe_expected.get("scientific_sheets"):
        raise ValueError("The packaged recipe workbook differs from the source baseline.")
    _write_json(
        report_path,
        {
            "schema_version": SCHEMA_VERSION,
            "status": "PASS",
            "ordifile_version": __version__,
            "adapter_ids": list(adapter_ids),
            "detected_adapter_ids": detected,
            "scientific_sheets": list(SCIENTIFIC_SHEETS),
            "semantic_sha256": digest,
            "mapped_semantic_sha256": mapped_digest,
            "mapping_set_semantic_sha256": mapped_set_digest,
            "mapping_drift_diagnostic": "PASS",
            "mapping_repair_clone": "PASS",
            "conversion_preflight": "PASS",
            "workbook_presentation": "PASS",
            "existing_output_preserved": True,
        },
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Ordifile standalone smoke tooling")
    subparsers = parser.add_subparsers(dest="command", required=True)
    make = subparsers.add_parser("make-kit")
    make.add_argument("--output", required=True, type=Path)
    make.add_argument("--generator-root", required=True, type=Path)
    run = subparsers.add_parser("run")
    run.add_argument("--kit", required=True, type=Path)
    run.add_argument("--output", required=True, type=Path)
    run.add_argument("--report", required=True, type=Path)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "make-kit":
            create_smoke_kit(args.output, args.generator_root)
        else:
            run_smoke(args.kit, args.output, args.report)
    except (KeyboardInterrupt, SystemExit, MemoryError):
        raise
    except Exception:
        print("Standalone smoke failed; no diagnostic path was published.")
        return 1
    print("Standalone smoke PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
