from __future__ import annotations

from datetime import datetime, timedelta, tzinfo
from pathlib import Path
from typing import Any, ClassVar, cast
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters.base import AdapterDescriptor, DetectionResult, ParseOptions
from ordifile.adapters.registry import AdapterRegistry, create_registry, load_external_adapters
from ordifile.api import convert
from ordifile.core.errors import OrdifileError
from ordifile.core.models import (
    DatasetBundle,
    FileStatus,
    InstrumentMetadata,
    Issue,
    MetadataEntry,
    PeakRecord,
    SampleRecord,
    Severity,
    SignalSeries,
    SourceFile,
)


class ExplodingText:
    def __str__(self) -> str:
        raise AssertionError("arbitrary plugin objects must not be stringified")


class EvilStr(str):
    def __str__(self) -> str:
        raise AssertionError("hostile string subclass must not reach export")


class EvilDateTime(datetime):
    def isoformat(self, *args: Any, **kwargs: Any) -> str:
        del args, kwargs
        raise AssertionError("hostile datetime subclass must not reach export")


class EvilInt(int):
    def __str__(self) -> str:
        raise AssertionError("hostile integer subclass must not reach export")


class BrokenTZ(tzinfo):
    def utcoffset(self, value: datetime | None) -> timedelta | None:
        del value
        raise RuntimeError("broken timezone hook")

    def dst(self, value: datetime | None) -> timedelta | None:
        del value
        return None

    def tzname(self, value: datetime | None) -> str | None:
        del value
        return "Broken"


class InstalledAdapter:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "installed_adapter"
    adapter_version: ClassVar[str] = "1.2.3"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id, adapter_version, "Installed", (".installed",), False, False, False, True
    )

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(False, 0.0, path.name)

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        raise NotImplementedError


class EntryPoint:
    name = "installed"

    def load(self) -> Any:
        return InstalledAdapter


def test_external_adapter_class_is_loaded_from_entry_point() -> None:
    registry = AdapterRegistry()
    load_external_adapters(registry, entry_points=[EntryPoint()])  # type: ignore[list-item]
    assert registry.get("installed_adapter").adapter_version == "1.2.3"


class BundleAdapter:
    api_version: ClassVar[str] = "1"
    adapter_id: ClassVar[str] = "bundle_test"
    adapter_version: ClassVar[str] = "1"
    descriptor: ClassVar[AdapterDescriptor] = AdapterDescriptor(
        adapter_id, adapter_version, "Bundle test", (".dat",), True, True, True, True
    )

    def __init__(
        self,
        *,
        failed: bool = False,
        mutate: bool = False,
        source_count: int = 1,
        raise_message: str | None = None,
        oversized_integer: bool = False,
        malformed: str | None = None,
        two_dimensional: bool = False,
    ) -> None:
        self.failed = failed
        self.mutate = mutate
        self.source_count = source_count
        self.raise_message = raise_message
        self.oversized_integer = oversized_integer
        self.malformed = malformed
        self.two_dimensional = two_dimensional

    def probe(self, path: Path) -> DetectionResult:
        return DetectionResult(path.suffix == ".dat", 1.0, "synthetic external fixture")

    def parse(self, path: Path, options: ParseOptions) -> DatasetBundle:
        del options
        if self.raise_message is not None:
            raise RuntimeError(self.raise_message)
        if self.malformed == "error_details":
            details: dict[object, object] = {
                ExplodingText(): ExplodingText(),
                "nested": {"private": "/private/machine/secret"},
                "unsafe": "bad\x01",
                "long": "/private/machine/secret" * 100,
                "oversized": 10**1_001,
                "nonfinite": float("inf"),
            }
            details.update({f"extra_{index}": index for index in range(40)})
            raise OrdifileError(
                "PLUGIN_STRUCTURED_FAILURE",
                "The plugin reported a structured failure.",
                details=cast(Any, details),
            )
        if self.malformed == "unsafe_structured_error":
            raise OrdifileError(
                "PLUGIN\x01ERROR",
                "unsafe_x000D_message",
                details={"unsafe": "secret\x01value"},
            )
        if self.malformed == "private_path_error":
            raise OrdifileError(
                "PLUGIN_PRIVATE_PATH",
                "Failed '/Users/example/My Project/private input.dat' and "
                "C:\\Users\\example\\Secret Project\\secret.dat; see "
                "https://example.test/help.",
                details={
                    "/Users/example/Private Key": "private key marker",
                    "posix": "/opt/private/My Instrument/instrument.dat",
                    "windows": "C:\\Users\\example\\Secret Project\\secret.dat",
                    "prefixed_windows": "note=C:\\Users\\example\\secret.dat",
                    "unc": "\\\\server\\Private Share\\secret.dat",
                    "prefixed_posix": "note=/opt/private/secret.dat",
                    "file_url_posix": "file:///Users/example/My%20Project/secret.dat",
                    "file_url_windows": "file://C:/Users/example/Secret%20Project/secret.dat",
                    "http_windows_query": (
                        "https://example.test/?path=C:\\Users\\example\\secret.dat"
                    ),
                    "http_windows_path": "http://example.test/C:/Users/example/secret.dat",
                    "http_nested_file": (
                        "https://example.test/?next=file:///Users/example/secret.dat"
                    ),
                    "http_remote": ("https://example.test/remote/path?next=/another/remote/path"),
                    "locator": "sheet:1:cell:A1",
                    "url": "https://example.test/reference/path",
                    "relative": "fixtures/example.dat",
                },
            )
        adapter_source = SourceFile(
            path,
            "/private/adapter/secret.dat",
            path.name,
            path.stat().st_size,
            None,
            None,
            0,
        )
        scientific_integer: Any = 10**1_000 if self.oversized_integer else None
        if self.malformed == "sequence_subclass":
            scientific_integer = EvilInt(7)
        runtime_value: Any = scientific_integer
        if self.malformed == "runtime_text":
            runtime_value = "ten"
        elif self.malformed == "runtime_nonfinite":
            runtime_value = float("nan")
        channels: Any = (
            ["A"]
            if self.malformed == "channels_list"
            else ("A", 1)
            if self.malformed == "channel_element"
            else ()
        )
        acquired_at: Any = "2026-01-01" if self.malformed == "acquired_text" else None
        if self.malformed == "datetime_subclass":
            acquired_at = EvilDateTime(2026, 1, 1)
        elif self.malformed in {"broken_tz_false", "broken_tz_true"}:
            acquired_at = datetime(2026, 1, 1, tzinfo=BrokenTZ())
        elif self.malformed == "zoneinfo_datetime":
            acquired_at = datetime(2026, 1, 1, tzinfo=ZoneInfo("Asia/Seoul"))
        acquired_reliable: Any = "yes" if self.malformed == "acquired_reliable_text" else False
        if self.malformed in {"broken_tz_true", "zoneinfo_datetime"}:
            acquired_reliable = True
        instrument: Any = "GC" if self.malformed == "instrument_text" else InstrumentMetadata()
        sample_id: Any = EvilStr("same") if self.malformed == "sample_id_subclass" else "same"
        if self.malformed == "channel_subclass":
            channels = (EvilStr("A"),)
        sample = SampleRecord(
            sample_id,
            adapter_source,
            acquired_at=acquired_at,
            acquired_at_reliable=acquired_reliable,
            sequence=scientific_integer,
            instrument=instrument,
            channels=channels,
            runtime=runtime_value,
        )
        if self.mutate:
            path.write_text("changed during parse", encoding="utf-8")
        signal = SignalSeries(
            "same",
            "/private/adapter/secret.dat",
            "A",
            "FID",
            cast(
                Any,
                [0.0]
                if self.malformed == "signal_list"
                else ("bad",)
                if self.malformed == "signal_element"
                else (scientific_integer if scientific_integer is not None else 0.0,),
            ),
            (scientific_integer if scientific_integer is not None else 1.0,),
        )
        signals: Any = [signal] if self.malformed == "bundle_list" else (signal,)
        metadata_value: Any = scientific_integer if scientific_integer is not None else "value"
        if self.malformed == "metadata_object":
            metadata_value = ExplodingText()
        elif self.malformed == "metadata_control":
            metadata_value = "bad\x01"
        elif self.malformed == "metadata_escape":
            metadata_value = "bad_x000D_"
        metadata_source = (
            None
            if self.malformed is None
            else {
                "metadata_source_file_posix": ("file:///Users/example/My%20Project/secret.dat"),
                "metadata_source_file_windows": (
                    "file://C:/Users/example/Secret%20Project/secret.dat"
                ),
                "metadata_source_http": "https://example.test/private/instrument.dat",
                "metadata_source_unc": "\\\\server\\Private Share\\secret.dat",
                "metadata_source_logical": "sheet:1:cell:D2",
            }.get(self.malformed)
        )
        warnings = (
            (Issue("PLUGIN_WARNING", "bad\x01", Severity.WARNING),)
            if self.malformed == "issue_control"
            else (Issue("PLUGIN_WARNING", "x" * 32_768, Severity.WARNING),)
            if self.malformed == "issue_overlong"
            else (
                Issue(
                    "PLUGIN_PRIVATE_PATH",
                    "Failed '/Users/example/My Project/private.dat'.",
                    Severity.WARNING,
                    "C:\\Users\\example\\Private Project\\input.dat",
                    (("\\\\server\\Private Share\\key", "/opt/My Project/value"),),
                ),
            )
            if self.malformed == "issue_private_path"
            else (
                Issue(
                    "PLUGIN_PRIVATE_PATH",
                    "https://example.test/?path=C:\\Users\\example\\private.dat",
                    Severity.WARNING,
                ),
            )
            if self.malformed == "issue_http_overlap"
            else (
                Issue(
                    "PLUGIN_REMOTE_REFERENCE",
                    "https://example.test/remote/path?next=/another/remote/path",
                    Severity.WARNING,
                ),
            )
            if self.malformed == "issue_http_remote"
            else ()
        )
        return DatasetBundle(
            tuple(adapter_source for _ in range(self.source_count)),
            (sample,),
            signals,
            (
                PeakRecord(
                    "same",
                    "/private/adapter/secret.dat",
                    peak_number=scientific_integer if scientific_integer is not None else 1,
                    retention_time=1.0 if self.two_dimensional else None,
                    retention_time_unit="s" if self.two_dimensional else None,
                    area=2.0,
                    compound="A",
                    observation_order=1 if self.two_dimensional else None,
                    area_unit="AU" if self.two_dimensional else None,
                    secondary_retention_time=0.5 if self.two_dimensional else None,
                    secondary_retention_time_unit="s" if self.two_dimensional else None,
                ),
            ),
            (
                MetadataEntry(
                    "same",
                    "/private/adapter/secret.dat",
                    "external",
                    "key",
                    metadata_value,
                    source=metadata_source,
                ),
            ),
            warnings=warnings,
            errors=(
                (Issue("ADAPTER_DATA_INVALID", "Synthetic invalid data.", Severity.ERROR),)
                if self.failed
                else (Issue("MISFILED_WARNING", "Synthetic mismatch.", Severity.WARNING),)
                if self.malformed == "issue_severity_mismatch"
                else ()
            ),
        )


def _registry(adapter: BundleAdapter) -> AdapterRegistry:
    registry = AdapterRegistry()
    registry.register(adapter)
    return registry


def test_external_adapter_api_v1_preserves_secondary_retention_coordinate(
    tmp_path: Path,
) -> None:
    source = tmp_path / "secondary.dat"
    source.write_bytes(b"fixture")

    result = convert(
        source,
        tmp_path / "secondary.xlsx",
        registry=_registry(BundleAdapter(two_dimensional=True)),
    )

    bundle = result.files[0].bundle
    assert bundle is not None
    assert bundle.peaks[0].secondary_retention_time == 0.5
    assert bundle.peaks[0].secondary_retention_time_unit == "s"
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        row = next(workbook["Peak_Order_Matrix_2D"].iter_rows(min_row=2, values_only=True))
        assert row[8:] == (1.0, 0.5, 2.0)
    finally:
        workbook.close()


def test_failed_bundle_scientific_data_is_excluded_from_workbook(tmp_path: Path) -> None:
    source = tmp_path / "failed.dat"
    source.write_bytes(b"fixture")
    result = convert(
        source,
        tmp_path / "failed.xlsx",
        include_signals=True,
        registry=_registry(BundleAdapter(failed=True)),
    )
    assert result.files[0].status is FileStatus.FAILED
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 1
        assert workbook["Peak_Matrix"].max_row == 1
        assert workbook["Metadata"].max_row == 1
        assert not any(name.startswith("Signals_") for name in workbook.sheetnames)
        sample = next(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        assert sample[1] == "failed"
        assert sample[5:12] == (None, None, "", None, None, None, 0)
        log = next(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert log[4] == "failed"
    finally:
        workbook.close()


@pytest.mark.parametrize("source_count", [0, 2])
def test_adapter_source_cardinality_survives_binding_and_fails_validation(
    tmp_path: Path, source_count: int
) -> None:
    source = tmp_path / f"sources-{source_count}.dat"
    source.write_bytes(b"fixture")
    result = convert(
        source,
        tmp_path / f"sources-{source_count}.xlsx",
        registry=_registry(BundleAdapter(source_count=source_count)),
    )
    assert result.files[0].status is FileStatus.FAILED
    assert any(issue.code == "SOURCE_COUNT_INVALID" for issue in result.files[0].issues)


def test_binding_replaces_external_absolute_source_refs_with_unique_relative_paths(
    tmp_path: Path,
) -> None:
    first = tmp_path / "one" / "sample.dat"
    second = tmp_path / "two" / "sample.dat"
    first.parent.mkdir()
    second.parent.mkdir()
    first.write_bytes(b"one")
    second.write_bytes(b"two")
    result = convert(
        (first, second),
        tmp_path / "safe.xlsx",
        include_signals=True,
        registry=_registry(BundleAdapter()),
    )
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        expected = {"input_001/sample.dat", "input_002/sample.dat"}
        for sheet_name, column in (("Peaks", 1), ("Metadata", 1), ("Import_Log", 0)):
            values = {
                row[column] for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True)
            }
            assert values == expected
        signal_name = next(name for name in workbook.sheetnames if name.startswith("Signals_"))
        signal_sources = {
            row[1] for row in workbook[signal_name].iter_rows(min_row=2, values_only=True)
        }
        assert signal_sources == expected
        assert not any(
            "/private/adapter" in str(value)
            for sheet_name in workbook.sheetnames
            for row in workbook[sheet_name].iter_rows(values_only=True)
            for value in row
        )
    finally:
        workbook.close()


def test_input_mutation_during_parse_is_detected_and_excluded(tmp_path: Path) -> None:
    source = tmp_path / "mutated.dat"
    source.write_bytes(b"before")
    result = convert(
        source, tmp_path / "mutated.xlsx", registry=_registry(BundleAdapter(mutate=True))
    )
    assert result.files[0].status is FileStatus.FAILED
    assert result.files[0].bundle is None
    assert any(issue.code == "INPUT_CHANGED_DURING_PARSE" for issue in result.files[0].issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 1
        assert workbook["Metadata"].max_row == 1
    finally:
        workbook.close()


def test_unexpected_adapter_error_does_not_expose_exception_text(tmp_path: Path) -> None:
    source = tmp_path / "error.dat"
    source.write_bytes(b"fixture")
    secret = "/private/machine/path/should-not-leak"

    result = convert(
        source,
        tmp_path / "error.xlsx",
        registry=_registry(BundleAdapter(raise_message=secret)),
    )

    issue = result.files[0].issues[0]
    assert issue.code == "ADAPTER_UNEXPECTED_ERROR"
    assert "RuntimeError" in issue.message
    assert secret not in issue.message


def test_external_oversized_integer_bundle_is_failed_and_scientific_data_is_excluded(
    tmp_path: Path,
) -> None:
    source = tmp_path / "oversized.dat"
    source.write_bytes(b"fixture")

    result = convert(
        source,
        tmp_path / "oversized.xlsx",
        include_signals=True,
        registry=_registry(BundleAdapter(oversized_integer=True)),
    )

    assert result.files[0].status is FileStatus.FAILED
    assert any(issue.code == "INTEGER_LIMIT_EXCEEDED" for issue in result.files[0].issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 1
        assert workbook["Metadata"].max_row == 1
        assert not any(name.startswith("Signals_") for name in workbook.sheetnames)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("malformed", "expected_code"),
    [
        ("runtime_text", "RUNTIME_TYPE_INVALID"),
        ("runtime_nonfinite", "RUNTIME_NONFINITE"),
        ("acquired_text", "ACQUIRED_AT_TYPE_INVALID"),
        ("acquired_reliable_text", "ACQUIRED_AT_RELIABILITY_TYPE_INVALID"),
        ("instrument_text", "INSTRUMENT_METADATA_TYPE_INVALID"),
        ("channels_list", "CANONICAL_TUPLE_TYPE_INVALID"),
        ("channel_element", "CANONICAL_TUPLE_ELEMENT_INVALID"),
        ("signal_list", "SIGNAL_TUPLE_TYPE_INVALID"),
        ("signal_element", "SIGNAL_VALUE_TYPE_INVALID"),
        ("bundle_list", "CANONICAL_CONTAINER_TYPE_INVALID"),
    ],
)
def test_malformed_external_canonical_bundle_is_file_failed_and_excluded(
    tmp_path: Path, malformed: str, expected_code: str
) -> None:
    source = tmp_path / f"{malformed}.dat"
    source.write_bytes(b"fixture")

    result = convert(
        source,
        tmp_path / f"{malformed}.xlsx",
        include_signals=True,
        registry=_registry(BundleAdapter(malformed=malformed)),
    )

    assert result.files[0].status is FileStatus.FAILED
    assert any(issue.code == expected_code for issue in result.files[0].issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 1
        assert workbook["Metadata"].max_row == 1
        assert not any(name.startswith("Signals_") for name in workbook.sheetnames)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "malformed",
    (
        "metadata_source_file_posix",
        "metadata_source_file_windows",
        "metadata_source_http",
        "metadata_source_unc",
    ),
)
def test_external_metadata_uri_or_absolute_source_is_failed_per_file_without_exposure(
    tmp_path: Path, malformed: str
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / f"{malformed}.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed=malformed))

    result = convert((good, bad), tmp_path / f"{malformed}.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    assert any(issue.code == "METADATA_SOURCE_UNSAFE" for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        audit_text = " ".join(
            "" if value is None else str(value)
            for row in workbook["Import_Log"].iter_rows(min_row=2, values_only=True)
            for value in row
        )
        assert "Users/example" not in audit_text
        assert "server\\Private" not in audit_text
        assert "example.test/private" not in audit_text
    finally:
        workbook.close()


def test_external_metadata_sheet_locator_remains_supported(tmp_path: Path) -> None:
    source = tmp_path / "logical.dat"
    source.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="metadata_source_logical"))

    result = convert(source, tmp_path / "logical.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 0
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        metadata = list(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        assert metadata[0][-1] == "sheet:1:cell:D2"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("malformed", "expected_code"),
    (
        ("metadata_object", "METADATA_VALUE_TYPE_INVALID"),
        ("metadata_control", "WORKBOOK_TEXT_UNREPRESENTABLE"),
        ("metadata_escape", "WORKBOOK_TEXT_UNREPRESENTABLE"),
        ("issue_control", "CANONICAL_ISSUE_TEXT_UNREPRESENTABLE"),
        ("issue_overlong", "CANONICAL_ISSUE_CELL_LIMIT"),
        ("issue_severity_mismatch", "CANONICAL_ISSUE_SEVERITY_INVALID"),
    ),
)
def test_bad_plugin_metadata_or_issue_is_isolated_from_good_builtin_workbook(
    tmp_path: Path, malformed: str, expected_code: str
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / f"{malformed}.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed=malformed))

    result = convert((good, bad), tmp_path / f"{malformed}.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    assert any(issue.code == expected_code for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook.close()


def test_plugin_error_details_are_bounded_without_invoking_object_stringification(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="error_details"))

    result = convert((good, bad), tmp_path / "safe-details.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    issue = next(item for item in failed.issues if item.code == "PLUGIN_STRUCTURED_FAILURE")
    assert len(issue.context) == 33
    context = dict(issue.context)
    assert context["detail_001"] == "[unsupported-detail-type]"
    assert context["nested"] == "[unsupported-detail-type]"
    assert context["unsafe"] == "[text-omitted-unrepresentable]"
    assert context["long"] == "[text-omitted-too-long]"
    assert context["oversized"] == "[integer-omitted-too-large]"
    assert context["nonfinite"] == "[nonfinite-float]"
    assert context["details_truncated"] == "true"
    assert not any("private" in value for _key, value in issue.context)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook.close()


def test_unsafe_plugin_error_code_and_message_are_replaced_per_file(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="unsafe_structured_error"))

    result = convert((good, bad), tmp_path / "safe-error.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    issue = next(item for item in failed.issues if item.code == "ADAPTER_ERROR_INVALID")
    assert issue.context == ()
    assert "unsafe_x000D_message" not in issue.message
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
    finally:
        workbook.close()


def test_plugin_structured_error_scrubs_machine_absolute_paths_per_file(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="private_path_error"))

    result = convert((good, bad), tmp_path / "private-path.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    issue = next(item for item in failed.issues if item.code == "PLUGIN_PRIVATE_PATH")
    assert issue.message == "[absolute-path-omitted]"
    context = dict(issue.context)
    assert context["[absolute-path-omitted]"] == "private key marker"
    assert context["posix"] == "[absolute-path-omitted]"
    assert context["windows"] == "[absolute-path-omitted]"
    assert context["prefixed_windows"] == "[absolute-path-omitted]"
    assert context["unc"] == "[absolute-path-omitted]"
    assert context["prefixed_posix"] == "[absolute-path-omitted]"
    assert context["file_url_posix"] == "[absolute-path-omitted]"
    assert context["file_url_windows"] == "[absolute-path-omitted]"
    assert context["http_windows_query"] == "[absolute-path-omitted]"
    assert context["http_windows_path"] == "[absolute-path-omitted]"
    assert context["http_nested_file"] == "[absolute-path-omitted]"
    assert context["http_remote"] == ("https://example.test/remote/path?next=/another/remote/path")
    assert context["locator"] == "sheet:1:cell:A1"
    assert context["url"] == "https://example.test/reference/path"
    assert context["relative"] == "fixtures/example.dat"
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        bad_log = next(row for row in log if row[4] == "failed")
        serialized = " ".join("" if value is None else str(value) for value in bad_log)
        assert "/Users/" not in serialized
        assert "C:\\Users\\" not in serialized
    finally:
        workbook.close()


def test_plugin_bundle_issue_with_private_paths_is_failed_without_exposure(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="issue_private_path"))

    result = convert((good, bad), tmp_path / "private-bundle-issue.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    issue = next(item for item in failed.issues if item.code == "CANONICAL_ISSUE_PRIVATE_PATH")
    assert "private text was omitted" in issue.message
    serialized_issues = repr(failed.issues)
    assert "/Users/" not in serialized_issues
    assert "C:\\Users\\" not in serialized_issues
    assert "\\\\server\\" not in serialized_issues
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook.close()


def test_plugin_bundle_issue_http_url_cannot_hide_windows_path(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="issue_http_overlap"))

    result = convert((good, bad), tmp_path / "http-overlap-issue.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    assert any(issue.code == "CANONICAL_ISSUE_PRIVATE_PATH" for issue in failed.issues)
    assert "Users" not in repr(failed.issues)


def test_plugin_bundle_issue_allows_normal_remote_http_paths(tmp_path: Path) -> None:
    source = tmp_path / "remote.dat"
    source.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed="issue_http_remote"))

    result = convert(source, tmp_path / "remote.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 0
    assert any(issue.code == "PLUGIN_REMOTE_REFERENCE" for issue in result.files[0].issues)


@pytest.mark.parametrize(
    ("malformed", "expected_code"),
    (
        ("sample_id_subclass", "SAMPLE_ID_TYPE_INVALID"),
        ("channel_subclass", "CANONICAL_TUPLE_ELEMENT_INVALID"),
        ("datetime_subclass", "ACQUIRED_AT_TYPE_INVALID"),
        ("sequence_subclass", "SEQUENCE_TYPE_INVALID"),
    ),
)
def test_hostile_builtin_subclasses_are_failed_per_plugin_file(
    tmp_path: Path, malformed: str, expected_code: str
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / f"{malformed}.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed=malformed))

    result = convert((good, bad), tmp_path / f"{malformed}.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    assert any(issue.code == expected_code for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
    finally:
        workbook.close()


@pytest.mark.parametrize("malformed", ("broken_tz_false", "broken_tz_true"))
def test_broken_timezone_hook_is_failed_per_plugin_file(tmp_path: Path, malformed: str) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / f"{malformed}.dat"
    bad.write_bytes(b"fixture")
    registry = create_registry(include_external=False)
    registry.register(BundleAdapter(malformed=malformed))

    result = convert((good, bad), tmp_path / f"{malformed}.xlsx", registry=registry)

    assert result.success_count == 1
    assert result.failure_count == 1
    failed = next(item for item in result.files if item.status is FileStatus.FAILED)
    assert any(issue.code == "ACQUIRED_AT_SERIALIZATION_INVALID" for issue in failed.issues)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 2
    finally:
        workbook.close()


def test_zoneinfo_datetime_is_normalized_to_hook_free_exact_offset(tmp_path: Path) -> None:
    try:
        ZoneInfo("Asia/Seoul")
    except ZoneInfoNotFoundError:
        pytest.skip("The host Python installation has no IANA timezone database.")
    source = tmp_path / "zoneinfo.dat"
    source.write_bytes(b"fixture")

    result = convert(
        source,
        tmp_path / "zoneinfo.xlsx",
        registry=_registry(BundleAdapter(malformed="zoneinfo_datetime")),
    )

    assert result.failure_count == 0
    assert result.files[0].bundle is not None
    sample = result.files[0].bundle.samples[0]
    assert sample.acquired_at is not None
    assert sample.acquired_at.isoformat() == "2026-01-01T00:00:00+09:00"
    assert sample.acquired_at_reliable is True
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        row = next(workbook["Samples"].iter_rows(min_row=2, values_only=True))
        assert row[8] == "2026-01-01T00:00:00+09:00"
    finally:
        workbook.close()
