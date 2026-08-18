from __future__ import annotations

import hashlib
import math
import os
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters._youngin_yl_clarity_result_csv import read_result_csv
from ordifile.api import convert, inspect_file

EXPECTED = (
    {
        "size": 1_009,
        "sha256": "d4987151cba83068b5143cf90c6b0e78f1fee6b3c9f04c38d8cb97441ddfadd7",
        "peaks": 2,
        "rt": "f839fe4904fafd0688cb0d1d5777c73b7b3ad5def2e55dc7d21dd540227a33c7",
        "area": "4cc8faaee92c77199d3105716fbdc2bc61f8e7152b6e7bb1d829dcadbaa4fb61",
        "height": "1262a6109a9b3afcba063e775d1d450dd4a39271e66d43dbba9a8e249370dab0",
    },
    {
        "size": 1_802,
        "sha256": "0ceb70ba51e41607a5a6ca4476c9b77e6e2bce41d56e47b865085bb3ea71f67b",
        "peaks": 4,
        "rt": "c6a2f265a171e92380998c0f6862cfbf20b4f20772ae71b2c1bd38156168874e",
        "area": "7f0cd6365208bbd2e595f80dc39d16024be5e6b7440bce89ba248a23e17a115f",
        "height": "7e29b557e851bed7a9b3ab7deccb617f3f17b8171124c4b1310d74ba956aea74",
    },
)


def _fixture(variable: str) -> Path:
    value = os.environ.get(variable)
    if not value:
        raise AssertionError(f"{variable} is required")
    return Path(value)


def _text_digest(values: Iterable[str]) -> str:
    return hashlib.sha256("\n".join(values).encode("utf-8")).hexdigest()


def _private_text_values(path: Path) -> frozenset[str]:
    safe = {
        "Signal No.",
        "Signal Name",
        "Peak No.",
        "Reten. time [min]",
        "Area [mV.s]",
        "Height [mV]",
        "Area [%]",
        "Height [%]",
        "W05 [min]",
        "FID",
        "TCD",
        "Total",
        "No peak to report",
        "Reten. Time [min]",
        "Response",
        "Amount [N/A]",
        "Amount% [%]",
        "Peak Type",
        "Compound Name",
    }
    fields = (
        field.strip()
        for line in path.read_bytes().decode("cp949").split("\r\n")
        for field in line.split("\t")
    )
    return frozenset(
        field
        for field in fields
        if len(field) >= 4 and field not in safe and any(character.isalpha() for character in field)
    )


def test_exact_owner_exports_full_canonical_comparison_and_workbook(tmp_path: Path) -> None:
    sources = (
        _fixture("ORDIFILE_YOUNGIN_RESULT_CSV_A_FIXTURE"),
        _fixture("ORDIFILE_YOUNGIN_RESULT_CSV_B_FIXTURE"),
    )
    decoded_documents = []
    for source, expected in zip(sources, EXPECTED, strict=True):
        data = source.read_bytes()
        if len(data) != expected["size"] or hashlib.sha256(data).hexdigest() != expected["sha256"]:
            raise AssertionError("external YoungIn Result export identity changed")
        decoded = read_result_csv(source)
        decoded_documents.append(decoded)
        if len(decoded.peaks) != expected["peaks"]:
            raise AssertionError("external YoungIn Result peak count changed")
        if (
            _text_digest(peak.retention_time_text for peak in decoded.peaks) != expected["rt"]
            or _text_digest(peak.area_text for peak in decoded.peaks) != expected["area"]
            or _text_digest(peak.height_text for peak in decoded.peaks) != expected["height"]
        ):
            raise AssertionError("external YoungIn Result scientific sequence changed")

        inspected = inspect_file(source)
        bundle = inspected.file.bundle
        if bundle is None or len(bundle.peaks) != expected["peaks"]:
            raise AssertionError("Ordifile did not preserve the external YoungIn peak rows")
        if inspected.file.source.public_reference != f"source-{expected['sha256']}":
            raise AssertionError("external YoungIn source identity is not SHA-derived")
        for source_peak, canonical in zip(decoded.peaks, bundle.peaks, strict=True):
            if (
                canonical.peak_number != source_peak.peak_number
                or canonical.observation_order != source_peak.observation_order
                or canonical.retention_time != source_peak.retention_time
                or canonical.area != source_peak.area
                or canonical.height != source_peak.height
                or canonical.retention_time_unit != "min"
                or canonical.area_unit != "mV.s"
                or canonical.height_unit != "mV"
                or canonical.detector is not None
                or canonical.compound is not None
                or canonical.start_time is not None
                or canonical.end_time is not None
            ):
                raise AssertionError("external YoungIn canonical mapping changed")

    if sum(len(document.peaks) for document in decoded_documents) != 6:
        raise AssertionError("external YoungIn aggregate peak count changed")
    output = tmp_path / "youngin-result.xlsx"
    convert(sources, output, sort="input_order")
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks_sheet = workbook["Peaks"]
        headers = tuple(next(peaks_sheet.values))
        rows = tuple(peaks_sheet.iter_rows(min_row=2, values_only=True))
        source_peaks = tuple(peak for document in decoded_documents for peak in document.peaks)
        if len(rows) != 6:
            raise AssertionError("external YoungIn workbook rows were lost or duplicated")
        for row, source_peak in zip(rows, source_peaks, strict=True):
            for column, expected_value in (
                ("retention_time", source_peak.retention_time),
                ("area", source_peak.area),
                ("height", source_peak.height),
            ):
                actual = row[headers.index(column)]
                if type(actual) not in {int, float} or not math.isclose(
                    float(actual), expected_value, rel_tol=1e-14, abs_tol=1e-14
                ):
                    raise AssertionError("external YoungIn workbook numeric mapping changed")
        order = tuple(workbook["Peak_Order_Matrix"].iter_rows(min_row=2, values_only=True))
        if len(order) != 2 or {row[2:7] for row in order} != {
            ("YoungIn", None, "Signal 1: TCD", "min", "mV.s"),
            ("YoungIn", None, "Signal 2: TCD", "min", "mV.s"),
        }:
            raise AssertionError("external YoungIn signal stream grouping changed")
        if workbook["Peak_Matrix"].max_column != 1:
            raise AssertionError("compound identity was inferred without source evidence")
        workbook_strings = {
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        }
        if any(_private_text_values(source) & workbook_strings for source in sources):
            raise AssertionError("private YoungIn metadata reached the workbook")
    finally:
        workbook.close()
