from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert
from ordifile.exporters import excel


def test_workbook_opens_on_samples_with_static_table_presentation(tmp_path: Path) -> None:
    one_dimensional = tmp_path / "one-dimensional.csv"
    one_dimensional.write_text(
        "sample_id,retention_time,area,detector,channel\nsample-1,1.25,250,FID,A\n",
        encoding="utf-8",
    )
    two_dimensional = tmp_path / "two-dimensional.csv"
    two_dimensional.write_text(
        "sample_id,retention_time,secondary_retention_time,area,detector,channel\n"
        "sample-2,2.5,0.75,500,MS,B\n",
        encoding="utf-8",
    )
    result = convert((one_dimensional, two_dimensional), tmp_path / "result.xlsx")

    workbook = load_workbook(result.output_path, read_only=False, data_only=False)
    try:
        assert workbook.sheetnames[:2] == ["Manifest", "Samples"]
        assert workbook.active.title == "Samples"
        assert workbook["Manifest"].freeze_panes == "A2"
        assert workbook["Samples"].freeze_panes == "C2"
        assert workbook["Peaks"].freeze_panes == "C2"
        assert workbook["Peak_Matrix"].freeze_panes == "B2"
        assert workbook["Metadata"].freeze_panes == "C2"
        assert workbook["Import_Log"].freeze_panes == "B2"
        assert workbook["Manifest"].auto_filter.ref is None
        assert workbook["Samples"].auto_filter.ref == "A1:N3"
        assert workbook["Peaks"].auto_filter.ref is not None
        assert workbook["Samples"]["A1"].font.bold is True
        assert workbook["Samples"]["A1"].fill.fill_type == "solid"
        assert 30 <= workbook["Samples"].column_dimensions["C"].width <= 31
        assert 28 <= workbook["Samples"].column_dimensions["I"].width <= 29
        peak_headers = tuple(cell.value for cell in workbook["Peaks"][1])
        retention_column = peak_headers.index("retention_time") + 1
        observation_column = peak_headers.index("observation_order") + 1
        observation_letter = workbook["Peaks"].cell(1, observation_column).column_letter
        assert 20 <= workbook["Peaks"].column_dimensions[observation_letter].width <= 21
        assert workbook["Peaks"].cell(2, retention_column).value == 1.25
        assert workbook["Peaks"].cell(2, retention_column).number_format == "General"
    finally:
        workbook.close()


def test_split_physical_sheets_retain_logical_presentation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(excel, "MAX_EXCEL_ROWS", 2)
    data = excel._SheetData("Peaks", ("sample_id", "area"), (("a", 1), ("b", 2)))

    physical = excel._physical_sheets((data,))

    assert [sheet.logical_name for sheet in physical] == ["Peaks", "Peaks"]
    assert [sheet.name for sheet in physical] == ["Peaks", "Peaks_002"]
    assert all(excel._presentation_for(sheet).freeze_columns == 2 for sheet in physical)
    one_dimensional = excel._PhysicalSheet(
        "Peak_Order_Matrix", "Peak_Order_Matrix", tuple(f"h{index}" for index in range(9)), ()
    )
    two_dimensional = excel._PhysicalSheet(
        "Peak_Order_Matrix_2D",
        "Peak_Order_Matrix_2D",
        tuple(f"h{index}" for index in range(10)),
        (),
    )
    signal = excel._PhysicalSheet(
        "Signals_FID",
        "Signals_FID",
        ("sample_id", "source_file", "detector", "channel", "x", "y"),
        (),
    )
    assert excel._presentation_for(one_dimensional).freeze_columns == 7
    assert excel._presentation_for(one_dimensional).filter_columns == 7
    assert excel._presentation_for(two_dimensional).freeze_columns == 8
    assert excel._presentation_for(two_dimensional).filter_columns == 8
    assert excel._presentation_for(signal).freeze_columns == 4
    assert excel._presentation_for(signal).filter_columns == 4
