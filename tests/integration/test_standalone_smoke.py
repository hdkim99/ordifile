# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts"))

import pytest  # noqa: E402
from openpyxl import load_workbook  # type: ignore[import-untyped]  # noqa: E402
from standalone.smoke import (  # noqa: E402
    CP949_PROBE_TEXT,
    EXPECTED_NAME,
    GENERIC_BOM_NAME,
    GENERIC_NAME,
    LECO_GCXGC_RESULT_NAME,
    SCIENTIFIC_SHEETS,
    YOUNGIN_NAME,
    create_smoke_kit,
    run_smoke,
    semantic_digest,
)

GENERATOR_ROOT = ROOT / "tests" / "fixtures" / "synthetic"


def test_public_smoke_kit_round_trips_generic_and_cp949_youngin(tmp_path: Path) -> None:
    kit = tmp_path / "kit"
    output = tmp_path / "합성 결과 workbook.xlsx"
    report = tmp_path / "합성 smoke report.json"
    create_smoke_kit(kit, GENERATOR_ROOT)
    assert CP949_PROBE_TEXT.encode("cp949") in (kit / YOUNGIN_NAME).read_bytes()
    assert not (kit / GENERIC_NAME).read_bytes().startswith(b"\xef\xbb\xbf")
    assert (kit / GENERIC_BOM_NAME).read_bytes().startswith(b"\xef\xbb\xbf")

    expected = json.loads((kit / EXPECTED_NAME).read_text(encoding="ascii"))
    assert expected["inputs"][YOUNGIN_NAME]["adapter_id"] == "youngin_yl_clarity_result_csv"
    assert expected["inputs"][LECO_GCXGC_RESULT_NAME]["adapter_id"] == (
        "leco_chromatof_gcxgc_result_txt"
    )
    assert {item["adapter_id"] for item in expected["inputs"].values()} == set(
        expected["adapter_ids"]
    )

    run_smoke(kit, output, report)

    evidence = json.loads(report.read_text(encoding="ascii"))
    assert evidence["status"] == "PASS"
    assert evidence["existing_output_preserved"] is True
    assert evidence["mapping_drift_diagnostic"] == "PASS"
    assert evidence["mapping_repair_clone"] == "PASS"
    assert evidence["conversion_preflight"] == "PASS"
    assert evidence["detected_adapter_ids"] == [
        expected["inputs"][name]["adapter_id"] for name in expected["input_order"]
    ]
    assert str(tmp_path) not in report.read_text(encoding="ascii")
    workbook = load_workbook(output, read_only=True)
    try:
        assert set(SCIENTIFIC_SHEETS) <= set(workbook.sheetnames)
        assert "Peak_Order_Matrix_2D" in workbook.sheetnames
    finally:
        workbook.close()


def test_semantic_baseline_is_independent_of_kit_directory_and_mtime(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"
    create_smoke_kit(first, GENERATOR_ROOT)
    create_smoke_kit(second, GENERATOR_ROOT)
    for path in second.iterdir():
        path.touch()
    expected_a = json.loads((first / EXPECTED_NAME).read_text(encoding="ascii"))
    expected_b = json.loads((second / EXPECTED_NAME).read_text(encoding="ascii"))
    assert expected_a["semantic_sha256"] == expected_b["semantic_sha256"]


def test_smoke_rejects_changed_input_before_conversion(tmp_path: Path) -> None:
    kit = tmp_path / "kit"
    create_smoke_kit(kit, GENERATOR_ROOT)
    source = kit / YOUNGIN_NAME
    source.write_bytes(source.read_bytes() + b"changed")
    with pytest.raises(ValueError, match="checksum"):
        run_smoke(kit, tmp_path / "result.xlsx", tmp_path / "report.json")


def test_semantic_digest_requires_all_scientific_sheets(tmp_path: Path) -> None:
    from openpyxl import Workbook

    output = tmp_path / "incomplete.xlsx"
    workbook = Workbook()
    workbook.save(output)
    workbook.close()
    with pytest.raises(ValueError, match="missing required scientific smoke sheets"):
        semantic_digest(output)


def test_semantic_digest_includes_conditional_two_dimensional_matrix(tmp_path: Path) -> None:
    from openpyxl import Workbook

    output = tmp_path / "two-dimensional.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = SCIENTIFIC_SHEETS[0]
    first.append(("value",))
    for sheet_name in SCIENTIFIC_SHEETS[1:]:
        workbook.create_sheet(sheet_name).append(("value",))
    two_dimensional = workbook.create_sheet("Peak_Order_Matrix_2D")
    two_dimensional.append(("peak_1_rt1", "peak_1_rt2", "peak_1_area"))
    two_dimensional.append((1.0, 0.1, 10.0))
    workbook.save(output)
    workbook.close()
    original, _sheets = semantic_digest(output)

    workbook = load_workbook(output)
    workbook["Peak_Order_Matrix_2D"]["B2"] = 0.2
    workbook.save(output)
    workbook.close()

    changed, _sheets = semantic_digest(output)
    assert changed != original


def test_semantic_digest_includes_split_two_dimensional_matrix(tmp_path: Path) -> None:
    from openpyxl import Workbook

    output = tmp_path / "split-two-dimensional.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = SCIENTIFIC_SHEETS[0]
    first.append(("value",))
    for sheet_name in SCIENTIFIC_SHEETS[1:]:
        workbook.create_sheet(sheet_name).append(("value",))
    workbook.create_sheet("Peak_Order_Matrix_2D").append(("first",))
    workbook.create_sheet("Peak_Order_Matrix_2D_002").append(("second",))
    workbook.save(output)
    workbook.close()
    original, _sheets = semantic_digest(output)

    workbook = load_workbook(output)
    workbook["Peak_Order_Matrix_2D_002"]["A1"] = "changed"
    workbook.save(output)
    workbook.close()

    changed, _sheets = semantic_digest(output)
    assert changed != original
