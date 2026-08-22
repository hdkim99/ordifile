# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import hashlib
import os
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from ordifile.adapters.base import AdapterDescriptor
from ordifile.api import FormatReport
from ordifile.cli.main import _terminal_safe, main
from ordifile.core.errors import OrdifileError
from ordifile.core.peak_mapping import (
    ColumnSelector,
    PeakTableFormat,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
    save_peak_table_mapping,
    save_peak_table_mapping_set,
)
from ordifile.core.recipe import ConversionRecipe, save_conversion_recipe


def _write_peak_table(path: Path, sample_id: str = "sample") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "sample_id,sequence,retention_time,area,compound,time,signal\n"
        f"{sample_id},1,1.5,20,methanol,0,3\n",
        encoding="utf-8",
    )


def test_help_exits_successfully(capsys: pytest.CaptureFixture[str]) -> None:
    with pytest.raises(SystemExit) as caught:
        main(["--help"])
    assert caught.value.code == 0
    output = capsys.readouterr().out
    assert "formats" in output
    assert "inspect" in output
    assert "convert" in output


def test_convert_uses_ordifile_default_output_name(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    source = tmp_path / "sample.csv"
    _write_peak_table(source)
    monkeypatch.chdir(tmp_path)

    assert main(["convert", str(source)]) == 0

    output = tmp_path / "Ordifile_Result.xlsx"
    assert output.is_file()
    assert "Ordifile_Result.xlsx" in capsys.readouterr().out


def test_inspect_and_convert_accept_same_explicit_peak_mapping(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "private-source.csv"
    source.write_text("Declared RT,Declared Area\n1.5,20\n", encoding="utf-8")
    mapping = PeakTableMapping(
        ColumnSelector("Declared RT", 1),
        ColumnSelector("Declared Area", 2),
        "min",
        PeakTableFormat.CSV,
    )
    mapping_path = tmp_path / "mapping.json"
    save_peak_table_mapping(mapping, mapping_path)
    output = tmp_path / "mapped.xlsx"

    assert main(["inspect", str(source), "--peak-mapping", str(mapping_path)]) == 0
    inspect_output = capsys.readouterr().out
    assert "Peaks: 1" in inspect_output
    assert source.name not in inspect_output
    assert (
        main(
            [
                "convert",
                str(source),
                "--peak-mapping",
                str(mapping_path),
                "--output",
                str(output),
            ]
        )
        == 0
    )
    assert output.is_file()


def test_cli_mapping_set_routes_batch_and_reports_privacy_safe_summary(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "private-source.csv"
    source.write_text(
        "Private RT Header,Private Area Header,Private Note Header\n"
        "1.5,20,Private Measurement Canary\n",
        encoding="utf-8",
    )
    profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Private RT Header", 1),
            ColumnSelector("Private Area Header", 2),
            "Private RT Unit",
            PeakTableFormat.CSV,
            area_unit="Private Area Unit",
            manufacturer="Private Manufacturer",
            software="Private Software",
            ignored_columns=(ColumnSelector("Private Note Header", 3),),
        ),
        "Private local profile label",
        profile_id="profile-11111111111111111111111111111111",
    )
    mapping_set_path = tmp_path / "private-set-name.json"
    save_peak_table_mapping_set(PeakTableMappingSet((profile,)), mapping_set_path)
    output = tmp_path / "mapped-set.xlsx"

    assert (
        main(
            [
                "convert",
                str(source),
                "--peak-mapping-set",
                str(mapping_set_path),
                "--output",
                str(output),
                "--dry-run",
                "--verbose",
            ]
        )
        == 0
    )
    dry_run = capsys.readouterr().out
    assert "Mapping profile: 1" in dry_run
    assert "Private local profile label" not in dry_run
    assert mapping_set_path.name not in dry_run
    for canary in (
        "Private RT Header",
        "Private Area Header",
        "Private Note Header",
        "Private Measurement Canary",
        "Private RT Unit",
        "Private Area Unit",
        "Private Manufacturer",
        "Private Software",
    ):
        assert canary not in dry_run
    assert not output.exists()

    assert (
        main(
            [
                "convert",
                str(source),
                "--peak-mapping-set",
                str(mapping_set_path),
                "--output",
                str(output),
            ]
        )
        == 0
    )
    rendered = capsys.readouterr().out

    assert output.is_file()
    assert "Mapping profiles used: 1" in rendered
    assert "Unmapped generic tables: 0" in rendered
    assert "Private local profile label" not in rendered
    assert mapping_set_path.name not in rendered


def test_cli_inspect_reports_privacy_safe_schema_drift_without_raw_headers(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "private-drift-source.csv"
    source.write_text("Private Header Canary,Declared Area\n1.5,20\n", encoding="utf-8")
    profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("Declared RT", 1),
            ColumnSelector("Declared Area", 2),
            "min",
            PeakTableFormat.CSV,
        ),
        "Private Profile Label",
        profile_id="profile-11111111111111111111111111111111",
    )
    mapping_set_path = tmp_path / "private-set.json"
    save_peak_table_mapping_set(PeakTableMappingSet((profile,)), mapping_set_path)

    assert (
        main(
            [
                "inspect",
                str(source),
                "--peak-mapping-set",
                str(mapping_set_path),
                "--verbose",
            ]
        )
        == 1
    )
    rendered = capsys.readouterr().out

    assert "Mapping route: SCHEMA_DRIFT_CANDIDATE" in rendered
    assert "Schema drift candidates: 1" in rendered
    assert "HEADER_CHANGED_UNRESOLVED" in rendered
    assert "Private Header Canary" not in rendered
    assert "Declared RT" not in rendered
    assert "Private Profile Label" not in rendered
    assert source.name not in rendered
    assert mapping_set_path.name not in rendered


def test_formats_distinguishes_verified_generic_and_experimental_capabilities(
    capsys: pytest.CaptureFixture[str],
) -> None:
    assert main(["formats"]) == 0
    output = capsys.readouterr().out
    for adapter_id in (
        "generic_csv",
        "generic_tsv",
        "generic_semicolon",
        "generic_xlsx",
    ):
        assert adapter_id in output
    assert "Verified adapters: 4" in output
    assert "agilent_chemstation_ch_v181" in output
    assert "agilent_chemstation_result_xml" in output
    assert "shimadzu_gcsolution_gcd" in output
    assert "shimadzu_labsolutions_result_ascii" in output
    assert "shimadzu_gcmssolution_qgd" in output
    assert "youngin_yl_clarity_prm_raw" in output
    assert "youngin_yl_clarity_result_csv" in output
    assert "leco_chromatof_gcxgc_result_txt" in output
    assert "Experimental" in output
    assert "Decoded records" in output
    assert "Scientific signals" in output
    assert "Experimental adapters: 8" in output
    assert "Fixture declarations: 0" in output
    assert "Experimental adapters expose only their explicitly documented capabilities" in output


def test_formats_includes_verified_external_descriptor_and_hides_unverified_one(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    descriptors = (
        AdapterDescriptor(
            "external_verified",
            "1.0.0",
            "External verified format",
            (".verified",),
            True,
            True,
            False,
            True,
        ),
        AdapterDescriptor(
            "external_unverified",
            "1.0.0",
            "External unverified format",
            (".unverified",),
            True,
            True,
            False,
            False,
        ),
    )
    monkeypatch.setattr(
        "ordifile.cli.main.get_format_report",
        lambda: FormatReport(descriptors, ("broken_plugin: RuntimeError",)),
    )

    assert main(["formats"]) == 0
    output = capsys.readouterr().out
    assert "external_verified" in output
    assert "Fixture declared" in output
    assert "external_unverified" not in output
    assert "External adapter load failures: 1" in output
    assert "broken_plugin: RuntimeError" in output


def test_formats_terminal_escapes_untrusted_descriptor_and_load_error_text(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    descriptor = AdapterDescriptor(
        "external\nverified",
        "1.0.0",
        "Name\x1b[2J\u202e",
        (".evil\r", ".literal\\n"),
        True,
        True,
        False,
        True,
    )
    monkeypatch.setattr(
        "ordifile.cli.main.get_format_report",
        lambda: FormatReport((descriptor,), ("broken\tplugin\x1b[31m",)),
    )

    assert main(["formats"]) == 0

    output = capsys.readouterr().out
    assert "external\\x0averified" in output
    assert "Name\\x1b[2J\\u202e" in output
    assert ".evil\\x0d" in output
    assert ".literal\\n" in output
    assert "broken\\x09plugin\\x1b[31m" in output
    assert "\x1b" not in output
    assert "\u202e" not in output


def test_inspect_reports_detection_hash_and_counts(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "sample.csv"
    _write_peak_table(source)

    assert main(["inspect", str(source), "--verbose"]) == 0

    output = capsys.readouterr().out
    assert "Status: success" in output
    assert "Detected format: generic_csv" in output
    assert "SHA-256:" in output
    assert "Samples: 1" in output
    assert "Peaks: 1" in output
    assert "Scientific signals: 1" in output
    assert "Decoded record series: 0" in output
    assert "Detection evidence:" in output


@pytest.mark.parametrize(
    ("unsafe_name", "escaped_name"),
    (
        pytest.param(
            "evil\nStatus: success.csv",
            "evil\\x0aStatus: success.csv",
            marks=pytest.mark.skipif(
                os.name == "nt", reason="Windows forbids control characters in filenames."
            ),
            id="newline-filename",
        ),
        pytest.param(
            "evil\x1b[2J.csv",
            "evil\\x1b[2J.csv",
            marks=pytest.mark.skipif(
                os.name == "nt", reason="Windows forbids control characters in filenames."
            ),
            id="escape-filename",
        ),
        ("evil\u202ereversed.csv", "evil\\u202ereversed.csv"),
    ),
)
def test_inspect_terminal_escapes_unsafe_filename_in_default_and_verbose_output(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    unsafe_name: str,
    escaped_name: str,
) -> None:
    source = tmp_path / unsafe_name
    _write_peak_table(source)

    for extra in ([], ["--verbose"]):
        assert main(["inspect", str(source), *extra]) in {0, 1}
        captured = capsys.readouterr()
        rendered = captured.out + captured.err
        assert f"File: {escaped_name}" in rendered
        assert "\x1b" not in rendered
        assert "\u202e" not in rendered
        assert "Traceback" not in rendered


@pytest.mark.parametrize(
    ("unsafe_name", "escaped_name"),
    (
        ("evil\nStatus: success.csv", "evil\\x0aStatus: success.csv"),
        ("evil\x1b[2J.csv", "evil\\x1b[2J.csv"),
        ("evil\u202ereversed.csv", "evil\\u202ereversed.csv"),
    ),
)
def test_terminal_renderer_escapes_unsafe_filename_text_without_filesystem(
    unsafe_name: str, escaped_name: str
) -> None:
    assert _terminal_safe(unsafe_name) == escaped_name


def test_inspect_terminal_escapes_issue_context_and_probe_evidence(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    issue = SimpleNamespace(
        severity=SimpleNamespace(value="warning"),
        code="WARN\nCODE",
        message="message\r\x1b[2J\u202e",
        context=(("key\t", "literal\\n\x85"),),
    )
    inspected = SimpleNamespace(
        file=SimpleNamespace(
            source=SimpleNamespace(
                relative_path="sample.csv",
                public_reference="sample.csv",
                detected_format="format\x1b[31m",
                sha256="a" * 64,
            ),
            status=SimpleNamespace(value="warning"),
            adapter_id="adapter\u202e",
            adapter_version="1\r2",
            bundle=SimpleNamespace(samples=(), peaks=(), signals=(), metadata=()),
            issues=(issue,),
        ),
        probes=(("probe\n", 0.5, "reason\x1b[2J"),),
    )
    monkeypatch.setattr("ordifile.cli.main.inspect_file", lambda *args, **kwargs: inspected)

    assert main(["inspect", "sample.csv", "--verbose"]) == 0

    output = capsys.readouterr().out
    assert "Detected format: format\\x1b[31m" in output
    assert "Adapter: adapter\\u202e" in output
    assert "Adapter version: 1\\x0d2" in output
    assert "[WARN\\x0aCODE]: message\\x0d\\x1b[2J\\u202e" in output
    assert "Context: key\\x09=literal\\n\\x85" in output
    assert "- probe\\x0a: confidence=0.500; reason\\x1b[2J" in output
    assert "\x1b" not in output
    assert "\u202e" not in output


def test_inspect_forwards_forced_adapter_and_sheet_selection(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    csv_source = tmp_path / "sample.csv"
    _write_peak_table(csv_source)
    assert main(["inspect", str(csv_source), "--adapter", "generic_csv"]) == 0
    assert "Adapter: generic_csv" in capsys.readouterr().out

    xlsx_source = tmp_path / "sheets.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "One"
    first.append(["sample_id", "area"])
    first.append(["one", 1])
    second = workbook.create_sheet("Two")
    second.append(["sample_id", "area"])
    second.append(["two", 2])
    workbook.save(xlsx_source)

    assert main(["inspect", str(xlsx_source), "--sheet", "Two"]) == 0
    assert "Status: success" in capsys.readouterr().out


def test_convert_success_writes_workbook_and_summary(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "sample.csv"
    output_path = tmp_path / "result.xlsx"
    _write_peak_table(source)

    exit_code = main(
        [
            "convert",
            str(source),
            "--output",
            str(output_path),
            "--include-signals",
            "--sort",
            "filename",
        ]
    )

    assert exit_code == 0
    assert output_path.is_file()
    output = capsys.readouterr().out
    assert "Status: success" in output
    assert "Output: result.xlsx" in output
    assert str(tmp_path) not in output
    assert "Successful files: 1" in output
    assert "Files with warnings: 0" in output
    assert "Failed files: 0" in output
    assert "Skipped files: 0" in output
    assert "Samples: 1" in output
    assert "Peaks: 1" in output
    assert "Scientific signal series: 1" in output
    assert "Structural record series: 0" in output
    assert "Sort used: filename" in output
    assert "Discovered files: 1" in output
    assert "Processed 1/1: success sample.csv" in output
    assert "Export started: result.xlsx" in output
    assert "Output ready: result.xlsx" in output
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        assert any(name.startswith("Signals_") for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_convert_terminal_escapes_progress_and_output_names(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "source\u202e.csv"
    output_path = tmp_path / "result\u202e.xlsx"
    _write_peak_table(source)

    assert main(["convert", str(source), "--output", str(output_path)]) == 0

    captured = capsys.readouterr()
    rendered = captured.out + captured.err
    assert "Processed 1/1: success source\\u202e.csv" in rendered
    assert "Export started: result\\u202e.xlsx" in rendered
    assert "Output ready: result\\u202e.xlsx" in rendered
    assert "\x1b" not in rendered
    assert "\u202e" not in rendered
    assert "Traceback" not in rendered


def test_convert_folder_recursive_and_extension_filter(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    inputs = tmp_path / "inputs"
    _write_peak_table(inputs / "sample_1.csv", "sample_1")
    _write_peak_table(inputs / "nested" / "sample_2.csv", "sample_2")
    (inputs / "nested" / "ignored.bin").write_bytes(b"ignored")
    output_path = tmp_path / "recursive.xlsx"

    exit_code = main(
        [
            "convert",
            str(inputs),
            "--recursive",
            "--extension",
            "csv",
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 0
    rendered = capsys.readouterr().out
    assert "Discovered files: 2" in rendered
    assert "Processed 1/2:" in rendered
    assert "Processed 2/2:" in rendered
    assert "Export started: recursive.xlsx" in rendered
    assert "Output ready: recursive.xlsx" in rendered
    assert "Successful files: 2" in rendered


def test_overwrite_requires_explicit_option_and_has_no_traceback(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "sample.csv"
    output_path = tmp_path / "result.xlsx"
    _write_peak_table(source)
    assert main(["convert", str(source), "--output", str(output_path)]) == 0
    capsys.readouterr()

    assert main(["convert", str(source), "--output", str(output_path)]) == 1

    captured = capsys.readouterr()
    assert "Error [OUTPUT_EXISTS]" in captured.err
    assert "--overwrite" in captured.err
    assert "Traceback" not in captured.err

    assert (
        main(
            [
                "convert",
                str(source),
                "--output",
                str(output_path),
                "--overwrite",
            ]
        )
        == 0
    )
    assert "Status: success" in capsys.readouterr().out


def test_convert_partial_success_returns_three(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    good = tmp_path / "good.csv"
    bad = tmp_path / "bad.csv"
    _write_peak_table(good)
    bad.write_bytes(b"\xff\xfe")
    output_path = tmp_path / "partial.xlsx"

    assert main(["convert", str(good), str(bad), "--output", str(output_path)]) == 3

    output = capsys.readouterr().out
    assert "Status: partial success" in output
    assert "Successful files: 1" in output
    assert "Failed files: 1" in output
    bad_sha256 = hashlib.sha256(bytes((0xFF, 0xFE))).hexdigest()
    safe_source = f"source-{bad_sha256}"
    assert f"{safe_source} [FORMAT_NOT_DETECTED]" in output
    assert "bad.csv" not in output
    assert output_path.is_file()


def test_convert_prints_warning_details_and_verbose_lists_all_warnings(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "warning.csv"
    source.write_text(
        "sample_id,acquired_at,area,compound\nsample,2026-01-01T00:00:00,not-a-number,methanol\n",
        encoding="utf-8",
    )

    assert main(["convert", str(source), "--output", str(tmp_path / "warning.xlsx")]) == 0
    basic = capsys.readouterr().out
    assert "Warning files:" in basic
    assert "warning.csv [TIMESTAMP_TIMEZONE_MISSING]" in basic
    assert "[INVALID_NUMBER]" not in basic

    assert (
        main(
            [
                "convert",
                str(source),
                "--output",
                str(tmp_path / "warning-verbose.xlsx"),
                "--verbose",
            ]
        )
        == 0
    )
    verbose = capsys.readouterr().out
    assert "warning.csv [TIMESTAMP_TIMEZONE_MISSING]" in verbose
    assert "[INVALID_NUMBER]" in verbose
    assert "Detection evidence:" in verbose
    assert "- warning.csv:" in verbose
    assert "generic_csv: confidence=" in verbose


def test_convert_all_failure_returns_one_and_preserves_failure_log(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    bad = tmp_path / "bad.csv"
    bad.write_bytes(b"\xff\xfe")
    output_path = tmp_path / "failure.xlsx"

    assert main(["convert", str(bad), "--output", str(output_path)]) == 1

    output = capsys.readouterr().out
    assert "Status: failed" in output
    assert "Successful files: 0" in output
    assert "Failed files: 1" in output
    assert output_path.is_file()
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        rows = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert rows[0][4] == "failed"
    finally:
        workbook.close()


def test_usage_and_configuration_errors_return_two(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    with pytest.raises(SystemExit) as caught:
        main([])
    assert caught.value.code == 2
    assert "the following arguments are required: command" in capsys.readouterr().err

    source = tmp_path / "sample.csv"
    _write_peak_table(source)
    assert main(["inspect", str(source), "--adapter", "not_registered"]) == 2
    assert "Error [ADAPTER_NOT_FOUND]" in capsys.readouterr().err


def test_argparse_terminal_escapes_invalid_user_value(
    capsys: pytest.CaptureFixture[str],
) -> None:
    with pytest.raises(SystemExit) as caught:
        main(["convert", "sample.csv", "--sort", "bad\n\x1b[2J\u202e\\n"])

    assert caught.value.code == 2
    error = capsys.readouterr().err
    assert "invalid choice: 'bad\\n\\\\x1b[2J\\\\u202e\\\\n'" in error
    assert "\x1b" not in error
    assert "\u202e" not in error


def test_on_error_stop_identifies_first_failed_file(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    bad = tmp_path / "bad.csv"
    good = tmp_path / "good.csv"
    bad.write_bytes(b"\xff\xfe")
    _write_peak_table(good)
    output = tmp_path / "stopped.xlsx"

    assert (
        main(
            [
                "convert",
                str(bad),
                str(good),
                "--output",
                str(output),
                "--on-error",
                "stop",
            ]
        )
        == 1
    )
    captured = capsys.readouterr()
    assert "Error [BATCH_FILE_FAILURE]" in captured.err
    bad_sha256 = hashlib.sha256(bytes((0xFF, 0xFE))).hexdigest()
    safe_source = f"source-{bad_sha256}"
    assert f"source_file={safe_source}" in captured.err
    assert "bad.csv" not in captured.err
    assert "error_code=FORMAT_NOT_DETECTED" in captured.err
    assert not output.exists()


def test_sidecar_csv_mode_is_forwarded(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    source = tmp_path / "long.csv"
    source.write_text(
        "sample_id,unmapped\nsample," + "x" * 32_768 + "\n",
        encoding="utf-8",
    )
    output = tmp_path / "sidecar.xlsx"

    assert (
        main(
            [
                "convert",
                str(source),
                "--output",
                str(output),
                "--sheet-mode",
                "sidecar-csv",
            ]
        )
        == 0
    )
    rendered = capsys.readouterr().out
    assert "Sidecars:" in rendered
    assert "sha256=" in rendered


def test_interrupt_and_unexpected_error_traceback_policy(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    def interrupt(_args: object) -> int:
        raise KeyboardInterrupt

    monkeypatch.setattr("ordifile.cli.main._run", interrupt)
    assert main(["formats"]) == 130
    assert "Interrupted." in capsys.readouterr().err

    def unexpected(_args: object) -> int:
        raise RuntimeError("synthetic failure")

    monkeypatch.setattr("ordifile.cli.main._run", unexpected)
    assert main(["formats"]) == 1
    ordinary = capsys.readouterr().err
    assert "Error [UNEXPECTED_ERROR]" in ordinary
    assert "Traceback" not in ordinary

    assert main(["inspect", "unused", "--verbose"]) == 1
    verbose = capsys.readouterr().err
    assert "Traceback" not in verbose
    assert "Diagnostic type: RuntimeError" in verbose
    assert "synthetic failure" not in verbose


def test_structured_error_terminal_escapes_code_message_and_details(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    def unsafe_error(_args: object) -> int:
        raise OrdifileError(
            "BAD\nCODE",
            "unsafe\rmessage\x1b[2J\u202e",
            details={"key\t": "literal\\n\x85"},
        )

    monkeypatch.setattr("ordifile.cli.main._run", unsafe_error)

    assert main(["formats"]) == 1

    error = capsys.readouterr().err
    assert "Error [BAD\\x0aCODE]: unsafe\\x0dmessage\\x1b[2J\\u202e" in error
    assert "Details: key\\x09=literal\\n\\x85" in error
    assert "\x1b" not in error
    assert "\u202e" not in error
    assert "Traceback" not in error


def test_terminal_renderer_preserves_readable_unicode_and_disambiguates_escapes() -> None:
    assert _terminal_safe("한글 sample.csv") == "한글 sample.csv"
    assert _terminal_safe(r"C:\Users\lab\new\result.xlsx") == r"C:\Users\lab\new\result.xlsx"
    assert _terminal_safe("literal\\n actual\n") == "literal\\n actual\\x0a"
    assert _terminal_safe("literal\\x1b actual\x1b") == "literal\\\\x1b actual\\x1b"
    assert _terminal_safe("literal\\u202e actual\u202e") == "literal\\\\u202e actual\\u202e"
    assert _terminal_safe("\x00\x7f\x85\u2066\u2028") == "\\x00\\x7f\\x85\\u2066\\u2028"


def test_convert_dry_run_prints_privacy_safe_plan_and_creates_no_artifact(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "private-canary-source.csv"
    output = tmp_path / "private-canary-output.xlsx"
    _write_peak_table(source)

    assert (
        main(
            [
                "convert",
                str(source),
                "--output",
                str(output),
                "--dry-run",
                "--verbose",
            ]
        )
        == 0
    )

    rendered = capsys.readouterr().out
    assert "Dry run: no workbook or sidecar was created." in rendered
    assert "Readiness: READY" in rendered
    assert "Generic input: 1" in rendered
    assert "Public plan-summary SHA-256:" in rendered
    assert "Sort result: deferred" in rendered
    assert "source-" in rendered
    assert "private-canary-source.csv" not in rendered
    assert "private-canary-output.xlsx" not in rendered
    assert not output.exists()
    assert not list(tmp_path.glob(".ordifile_*"))


def test_convert_dry_run_folder_partial_and_output_block_exit_semantics(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    folder = tmp_path / "inputs"
    good = folder / "nested" / "good.csv"
    unsupported = folder / "nested" / "unsupported.bin"
    _write_peak_table(good)
    unsupported.write_bytes(b"unsupported")
    output = tmp_path / "result.xlsx"

    assert (
        main(
            [
                "convert",
                str(folder),
                "--recursive",
                "--output",
                str(output),
                "--dry-run",
            ]
        )
        == 3
    )
    partial = capsys.readouterr().out
    assert "Readiness: READY_WITH_KNOWN_FAILURES" in partial
    assert "Inputs: 2" in partial
    assert "Unsupported: 1" in partial
    assert not output.exists()

    output.write_bytes(b"foreign")
    assert (
        main(
            [
                "convert",
                str(good),
                "--output",
                str(output),
                "--dry-run",
            ]
        )
        == 1
    )
    blocked = capsys.readouterr().out
    assert "Readiness: BLOCKED" in blocked
    assert "Output issue: OUTPUT_EXISTS" in blocked
    assert output.read_bytes() == b"foreign"

    assert (
        main(
            [
                "convert",
                str(good),
                "--output",
                str(output),
                "--dry-run",
                "--overwrite",
            ]
        )
        == 2
    )
    error = capsys.readouterr().err
    assert "CONVERSION_PLAN_OVERWRITE_UNSUPPORTED" in error
    assert output.read_bytes() == b"foreign"


def test_convert_dry_run_reflects_forced_adapter_without_parsing(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "source.csv"
    _write_peak_table(source)
    output = tmp_path / "result.xlsx"

    assert (
        main(
            [
                "convert",
                str(source),
                "--adapter",
                "generic_csv",
                "--output",
                str(output),
                "--dry-run",
                "--verbose",
            ]
        )
        == 0
    )
    rendered = capsys.readouterr().out
    assert "adapter=generic_csv" in rendered
    assert "route=GENERIC_INPUT" in rendered
    assert not output.exists()


def test_structured_input_error_has_actionable_message_without_traceback(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    missing = tmp_path / "missing.csv"

    assert main(["inspect", str(missing)]) == 1

    captured = capsys.readouterr()
    assert "Error [INPUT_NOT_FOUND]: The input path does not exist." in captured.err
    assert "Traceback" not in captured.err


def test_cli_recipe_dry_run_and_conversion_reuse_embedded_mapping_set(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = tmp_path / "private-source.csv"
    source.write_text("RT,Area\n1.5,20\n", encoding="utf-8")
    profile = PeakTableMappingProfile(
        PeakTableMapping(
            ColumnSelector("RT", 1),
            ColumnSelector("Area", 2),
            "min",
            PeakTableFormat.CSV,
        ),
        "PRIVATE-LOCAL-LABEL",
    )
    recipe = ConversionRecipe(
        peak_table_mapping_set=PeakTableMappingSet((profile,)),
        display_label="PRIVATE-RECIPE-LABEL",
    )
    recipe_path = tmp_path / "private-laboratory-recipe.json"
    save_conversion_recipe(recipe, recipe_path)
    output = tmp_path / "result.xlsx"

    assert (
        main(
            [
                "convert",
                str(source),
                "--recipe",
                str(recipe_path),
                "--output",
                str(output),
                "--dry-run",
            ]
        )
        == 0
    )
    dry_run = capsys.readouterr().out
    assert "Mapping profile: 1" in dry_run
    assert "PRIVATE-LOCAL-LABEL" not in dry_run
    assert "PRIVATE-RECIPE-LABEL" not in dry_run
    assert recipe_path.name not in dry_run
    assert not output.exists()

    assert (
        main(
            [
                "convert",
                str(source),
                "--recipe",
                str(recipe_path),
                "--output",
                str(output),
            ]
        )
        == 0
    )
    rendered = capsys.readouterr().out
    assert "Peaks: 1" in rendered
    assert output.is_file()
    assert "PRIVATE-LOCAL-LABEL" not in rendered
    assert "PRIVATE-RECIPE-LABEL" not in rendered
    assert recipe_path.name not in rendered


@pytest.mark.parametrize(
    "option",
    [
        ("--recursive",),
        ("--sort", "input_order"),
        ("--extension", "csv"),
        ("--include-signals",),
        ("--sheet-mode", "sidecar-csv"),
        ("--on-error", "stop"),
        ("--sheet", "Results"),
        ("--include-hidden-sheets",),
        ("--overwrite",),
        ("--adapter", "generic_csv"),
        ("--peak-mapping", "unused-mapping.json"),
        ("--peak-mapping-set", "unused-mapping-set.json"),
    ],
)
def test_cli_recipe_rejects_separate_behavior_options(
    option: tuple[str, ...],
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    source = tmp_path / "source.csv"
    _write_peak_table(source)
    recipe_path = tmp_path / "recipe.json"
    save_conversion_recipe(ConversionRecipe(), recipe_path)

    assert (
        main(
            [
                "convert",
                str(source),
                "--recipe",
                str(recipe_path),
                "--output",
                str(tmp_path / "result.xlsx"),
                *option,
            ]
        )
        == 2
    )
    captured = capsys.readouterr()
    assert "CONVERSION_RECIPE_OPTION_CONFLICT" in captured.err
