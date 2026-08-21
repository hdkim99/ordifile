from __future__ import annotations

import csv
import os
import shutil
import tempfile
import unicodedata
from decimal import Decimal
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

import ordifile.api as api_module
from ordifile.api import convert, convert_plan, inspect_file, plan_conversion
from ordifile.core.errors import ExportError, ExportLimitError, OrdifileError
from ordifile.core.models import (
    BatchResult,
    DatasetBundle,
    FileResult,
    FileStatus,
    Issue,
    MetadataEntry,
    SampleRecord,
    Severity,
    SortDecision,
    SortMode,
    SourceFile,
)
from ordifile.exporters import excel
from ordifile.exporters.excel import ExcelExporter


def test_formula_looking_strings_are_literal_and_unmodified(tmp_path: Path) -> None:
    source = tmp_path / "formula.csv"
    source.write_text(
        "sample_id,retention_time,area,compound,unknown\nsample,1.0,2.0,=DANGEROUS(),@source\n",
        encoding="utf-8",
    )
    output = tmp_path / "result.xlsx"
    convert(source, output)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2))
        assert peaks[0][9].value == "=DANGEROUS()"
        assert peaks[0][9].data_type == "s"
        metadata = list(workbook["Metadata"].values)
        assert any("@source" in row for row in metadata)
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["literal_string_policy"].startswith("All XLSX strings")
        assert manifest["formula_like_literal_count"] >= 2
    finally:
        workbook.close()


def test_overwrite_and_input_output_protection(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    convert(source, output)
    with pytest.raises(ExportError) as exists:
        convert(source, output)
    assert exists.value.code == "OUTPUT_EXISTS"
    convert(source, output, overwrite=True)
    with pytest.raises(ExportError) as same:
        convert(source, source, overwrite=True)
    assert same.value.code == "OUTPUT_IS_INPUT"


@pytest.mark.parametrize("invalid", ("false", 0, 1, None))
def test_invalid_overwrite_type_never_changes_existing_output(
    tmp_path: Path, invalid: object
) -> None:
    source = tmp_path / "source.csv"
    output = tmp_path / "result.xlsx"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output.write_bytes(b"existing-workbook-sentinel")

    with pytest.raises(OrdifileError) as caught:
        convert(source, output, overwrite=invalid)  # type: ignore[arg-type]

    assert caught.value.code == "OPTION_TYPE_INVALID"
    assert output.read_bytes() == b"existing-workbook-sentinel"


def test_exact_boolean_overwrite_false_and_true_keep_expected_semantics(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    output = tmp_path / "result.xlsx"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output.write_bytes(b"existing-workbook-sentinel")

    with pytest.raises(ExportError) as caught:
        convert(source, output, overwrite=False)
    assert caught.value.code == "OUTPUT_EXISTS"
    assert output.read_bytes() == b"existing-workbook-sentinel"

    convert(source, output, overwrite=True)
    assert output.read_bytes().startswith(b"PK")


@pytest.mark.skipif(os.name == "nt", reason="POSIX directory mode policy")
def test_export_rejects_non_sticky_shared_writable_output_directory(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    shared = tmp_path / "shared"
    shared.mkdir()
    shared.chmod(0o777)
    output = shared / "result.xlsx"

    with pytest.raises(ExportError) as caught:
        convert(source, output)

    assert caught.value.code == "OUTPUT_DIRECTORY_UNSAFE"
    assert not output.exists()
    assert not list(shared.glob(".ordifile_*"))


def test_xlsxwriter_internal_temporaries_stay_in_private_transaction_directory(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    observed_directories: list[str | os.PathLike[str] | None] = []
    real_mkstemp = tempfile.mkstemp

    def tracked_mkstemp(*args: Any, **kwargs: Any) -> tuple[int, str]:
        observed_directories.append(kwargs.get("dir"))
        return cast(tuple[int, str], real_mkstemp(*args, **kwargs))

    monkeypatch.setattr(tempfile, "mkstemp", tracked_mkstemp)
    convert(source, output)

    assert len(observed_directories) >= 2
    assert all(directory is not None for directory in observed_directories)
    resolved = {Path(directory).resolve() for directory in observed_directories if directory}
    assert len(resolved) == 1
    transaction_directory = resolved.pop()
    assert transaction_directory.parent == tmp_path.resolve()
    assert transaction_directory.name.startswith(".ordifile_transaction_")
    assert not transaction_directory.exists()


@pytest.mark.parametrize("option", ("recursive", "include_signals", "include_hidden_sheets"))
@pytest.mark.parametrize("invalid", (0, 1, "false", None))
def test_convert_rejects_non_boolean_public_options_before_output(
    tmp_path: Path, option: str, invalid: object
) -> None:
    source = tmp_path / "source.csv"
    output = tmp_path / "result.xlsx"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    kwargs: dict[str, Any] = {option: invalid}

    with pytest.raises(OrdifileError) as caught:
        convert(source, output, **kwargs)

    assert caught.value.code == "OPTION_TYPE_INVALID"
    assert not output.exists()


@pytest.mark.parametrize("invalid", (0, 1, "false", None))
def test_inspect_rejects_non_boolean_include_hidden_before_path_access(
    tmp_path: Path, invalid: object
) -> None:
    missing = tmp_path / "missing.csv"
    with pytest.raises(OrdifileError) as caught:
        inspect_file(missing, include_hidden_sheets=invalid)  # type: ignore[arg-type]
    assert caught.value.code == "OPTION_TYPE_INVALID"


@pytest.mark.parametrize(
    ("kwargs", "expected"),
    (
        ({"on_error": False}, "ON_ERROR_INVALID"),
        ({"sidecar_mode": 1}, "SIDECAR_MODE_INVALID"),
        ({"sort": 1}, "SORT_MODE_INVALID"),
        ({"extensions": ".csv"}, "OPTION_TYPE_INVALID"),
        ({"extensions": (".csv", 1)}, "OPTION_TYPE_INVALID"),
        ({"adapter": 1}, "OPTION_TYPE_INVALID"),
        ({"sheet": 1}, "OPTION_TYPE_INVALID"),
        ({"progress": 1}, "OPTION_TYPE_INVALID"),
        ({"registry": object()}, "OPTION_TYPE_INVALID"),
    ),
)
def test_invalid_public_configuration_is_rejected_before_output(
    tmp_path: Path, kwargs: dict[str, Any], expected: str
) -> None:
    source = tmp_path / "source.csv"
    output = tmp_path / "result.xlsx"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")

    with pytest.raises(OrdifileError) as caught:
        convert(source, output, **kwargs)

    assert caught.value.code == expected
    assert not output.exists()


@pytest.mark.parametrize(
    "extensions",
    (
        ("",),
        ("bad\x01",),
        ("../csv",),
        ("a" * 33,),
        (".CSV", "csv"),
        tuple(f".{index:02d}{'a' * 29}" for index in range(33)),
        tuple(f".{index:02d}{'a' * 29}" for index in range(32)),
    ),
)
def test_invalid_extension_filters_are_rejected_before_discovery_or_output_mutation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    extensions: tuple[str, ...],
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "sentinel.xlsx"
    output.write_bytes(b"sentinel")

    def unexpected_pipeline(*args: Any, **kwargs: Any) -> None:
        del args, kwargs
        raise AssertionError("invalid extensions must fail before discovery and hashing")

    monkeypatch.setattr(api_module, "run_pipeline", unexpected_pipeline)
    with pytest.raises(OrdifileError) as caught:
        convert(source, output, extensions=extensions, overwrite=True)

    assert caught.value.code == "EXTENSIONS_INVALID"
    assert output.read_bytes() == b"sentinel"


def test_cell_text_over_limit_is_rejected_without_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(excel, "MAX_EXCEL_CELL_CHARACTERS", 10)
    source = tmp_path / "long.csv"
    source.write_text("sample_id,area\na_very_long_id,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    with pytest.raises(ExportLimitError) as caught:
        convert(source, output)
    assert caught.value.code == "EXCEL_CELL_TEXT_LIMIT"
    assert not output.exists()


def test_nan_and_infinity_are_visible_literal_strings(tmp_path: Path) -> None:
    source = tmp_path / "numbers.csv"
    source.write_text(
        "sample_id,retention_time,area,compound\na,1,nan,A\na,2,inf,B\n",
        encoding="utf-8",
    )
    output = tmp_path / "result.xlsx"
    convert(source, output)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        areas = [row[7] for row in workbook["Peaks"].iter_rows(min_row=2, values_only=True)]
        assert areas == ["nan", "inf"]
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["nonfinite_literal_count"] == 4  # Peaks and Peak_Matrix each preserve both.
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("name", "code"),
    [
        ("CON.xlsx", "WINDOWS_OUTPUT_NAME_RESERVED"),
        ("prn.report.xlsx", "WINDOWS_OUTPUT_NAME_RESERVED"),
        ("COM9.xlsx", "WINDOWS_OUTPUT_NAME_RESERVED"),
        ("LPT1.xlsx", "WINDOWS_OUTPUT_NAME_RESERVED"),
        ("bad:name.xlsx", "WINDOWS_OUTPUT_NAME_INVALID"),
        ("bad|name.xlsx", "WINDOWS_OUTPUT_NAME_INVALID"),
        ("bad\x01name.xlsx", "WINDOWS_OUTPUT_NAME_INVALID"),
        ("trailing.xlsx. ", "WINDOWS_OUTPUT_NAME_INVALID"),
        ("result.xls", "OUTPUT_EXTENSION_INVALID"),
    ],
)
def test_portable_output_name_preflight(tmp_path: Path, name: str, code: str) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    with pytest.raises(ExportError) as caught:
        convert(source, tmp_path / name)
    assert caught.value.code == code


def test_output_directory_and_unicode_path_length_policy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    directory = tmp_path / "directory.xlsx"
    directory.mkdir()
    with pytest.raises(ExportError) as caught:
        convert(source, directory, overwrite=True)
    assert caught.value.code == "OUTPUT_IS_DIRECTORY"

    unicode_output = tmp_path / "한글 결과.xlsx"
    resolved_length = len(str(unicode_output.resolve(strict=False)))
    monkeypatch.setattr(excel, "MAX_PORTABLE_OUTPUT_PATH_CHARACTERS", resolved_length)
    excel._validate_output_path(unicode_output)
    monkeypatch.setattr(excel, "MAX_PORTABLE_OUTPUT_PATH_CHARACTERS", resolved_length - 1)
    with pytest.raises(ExportError) as too_long:
        excel._validate_output_path(unicode_output)
    assert too_long.value.code == "OUTPUT_PATH_TOO_LONG"


def _write_long_formula_source(path: Path, area: int) -> None:
    compound = "=" + "X" * 33_000
    path.write_text(
        f"sample_id,retention_time,area,compound\na,1,{area},{compound}\n",
        encoding="utf-8",
    )


def _inject_late_publish_collision(
    monkeypatch: pytest.MonkeyPatch,
    target: Path,
    canary: bytes,
) -> None:
    real_rename = excel._rename_no_replace

    def collide_rename(source_path: Path, destination_path: Path) -> None:
        if Path(destination_path) == target:
            target.write_bytes(canary)
        real_rename(Path(source_path), Path(destination_path))

    monkeypatch.setattr(excel, "_rename_no_replace", collide_rename)


@pytest.mark.parametrize("interrupt", [False, True])
def test_transaction_restores_workbook_and_sidecars_on_finalize_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, interrupt: bool
) -> None:
    source = tmp_path / "source.csv"
    output = tmp_path / "result.xlsx"
    _write_long_formula_source(source, -5)
    initial = convert(source, output, sidecar_mode="csv")
    artifact_paths = [output, *(tmp_path / item.relative_path for item in initial.sidecars)]
    original = {path: path.read_bytes() for path in artifact_paths}
    _write_long_formula_source(source, -6)

    real_replace = os.replace
    call_count = 0
    fail_at = len(artifact_paths) + 2

    def failing_replace(source_path: Path, destination_path: Path) -> None:
        nonlocal call_count
        call_count += 1
        if call_count == fail_at:
            if interrupt:
                raise KeyboardInterrupt
            raise OSError("injected finalize failure")
        real_replace(source_path, destination_path)

    monkeypatch.setattr(os, "replace", failing_replace)
    expected = KeyboardInterrupt if interrupt else ExportError
    with pytest.raises(expected):
        convert(source, output, sidecar_mode="csv", overwrite=True)
    assert {path: path.read_bytes() for path in artifact_paths} == original
    assert not list(tmp_path.glob(".ordifile_*"))


def test_sidecar_input_collision_is_rejected_even_with_overwrite(tmp_path: Path) -> None:
    colliding_input = tmp_path / "Result_Peak_Matrix_001.csv"
    colliding_input.write_text("sample_id,area\ncollision,1\n", encoding="utf-8")
    long_source = tmp_path / "long.csv"
    _write_long_formula_source(long_source, -5)
    with pytest.raises(ExportError) as caught:
        convert(
            (colliding_input, long_source),
            tmp_path / "Result.xlsx",
            sidecar_mode="csv",
            overwrite=True,
        )
    assert caught.value.code == "SIDECAR_IS_INPUT"
    assert colliding_input.read_text(encoding="utf-8").startswith("sample_id")


def test_long_scientific_values_use_explicit_sidecars_and_count_formula_escapes(
    tmp_path: Path,
) -> None:
    source = tmp_path / "long.csv"
    _write_long_formula_source(source, -5)
    with pytest.raises(ExportLimitError) as caught:
        convert(source, tmp_path / "error.xlsx")
    assert caught.value.code == "EXCEL_CELL_TEXT_LIMIT"

    result = convert(source, tmp_path / "result.xlsx", sidecar_mode="csv")
    assert sum(item.formula_escape_count for item in result.sidecars) == 2
    rows: list[list[str]] = []
    for record in result.sidecars:
        with (tmp_path / record.relative_path).open(encoding="utf-8", newline="") as stream:
            rows.extend(csv.reader(stream))
    flattened = [value for row in rows for value in row]
    assert "-5.0" in flattened
    assert "'-5.0" not in flattened
    assert any(value.startswith("'=") for value in flattened)
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        assert next(workbook["Peak_Matrix"].values) == ("sample_id", "sidecar_status")
        manifest_rows = [
            row
            for name in workbook.sheetnames
            if name.startswith("Manifest")
            for row in workbook[name].values
            if row[0] == "sidecar"
        ]
        assert sum(row[5] for row in manifest_rows) == 2
    finally:
        workbook.close()


def test_long_unknown_metadata_value_is_sidecar_data_not_truncated(tmp_path: Path) -> None:
    long_value = "Z" * 33_000
    source = tmp_path / "metadata.csv"
    source.write_text(f"sample_id,unknown\na,{long_value}\n", encoding="utf-8")
    result = convert(source, tmp_path / "result.xlsx", sidecar_mode="csv")
    metadata_sidecar = next(item for item in result.sidecars if "Metadata" in item.relative_path)
    assert long_value in (tmp_path / metadata_sidecar.relative_path).read_text(encoding="utf-8")


def test_existing_hardlink_output_alias_is_rejected_without_modifying_input(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    original = b"sample_id,area\na,1\n"
    source.write_bytes(original)
    output = tmp_path / "result.xlsx"
    try:
        os.link(source, output)
    except OSError:
        pytest.skip("hardlinks unavailable")
    with pytest.raises(ExportError) as caught:
        convert(source, output, overwrite=True)
    assert caught.value.code == "OUTPUT_IS_INPUT"
    assert source.read_bytes() == original


def test_sidecar_hardlink_alias_is_rejected_even_with_overwrite(tmp_path: Path) -> None:
    protected = tmp_path / "protected.csv"
    original = b"sample_id,area\nprotected,1\n"
    protected.write_bytes(original)
    long_source = tmp_path / "long.csv"
    _write_long_formula_source(long_source, -5)
    sidecar_alias = tmp_path / "Result_Peak_Matrix_001.csv"
    try:
        os.link(protected, sidecar_alias)
    except OSError:
        pytest.skip("hardlinks unavailable")
    with pytest.raises(ExportError) as caught:
        convert(
            (protected, long_source),
            tmp_path / "Result.xlsx",
            sidecar_mode="csv",
            overwrite=True,
        )
    assert caught.value.code == "SIDECAR_IS_INPUT"
    assert protected.read_bytes() == original


def test_casefold_and_unicode_normalization_output_aliases_are_rejected(tmp_path: Path) -> None:
    composed = tmp_path / "Résult.xlsx"
    source_workbook = Workbook()
    worksheet = source_workbook.active
    worksheet.append(["sample_id", "area"])
    worksheet.append(["a", 1])
    source_workbook.save(composed)
    original = composed.read_bytes()

    case_variant = tmp_path / "RÉSULT.XLSX"
    with pytest.raises(ExportError) as case_error:
        convert(composed, case_variant, overwrite=True)
    assert case_error.value.code == "OUTPUT_IS_INPUT"
    assert composed.read_bytes() == original

    decomposed_name = unicodedata.normalize("NFD", composed.name)
    normalized_variant = tmp_path / decomposed_name
    with pytest.raises(ExportError) as unicode_error:
        convert(composed, normalized_variant, overwrite=True)
    assert unicode_error.value.code == "OUTPUT_IS_INPUT"
    assert composed.read_bytes() == original


def test_finalization_rechecks_alias_created_after_preflight(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    original = b"sample_id,area\na,1\n"
    source.write_bytes(original)
    output = tmp_path / "result.xlsx"
    real_write = excel._write_physical

    def write_then_race(workbook: object, sheets: object) -> tuple[int, int, int]:
        result = real_write(workbook, sheets)  # type: ignore[arg-type]
        os.link(source, output)
        return result

    monkeypatch.setattr(excel, "_write_physical", write_then_race)
    with pytest.raises(ExportError) as caught:
        convert(source, output, overwrite=True)
    assert caught.value.code == "OUTPUT_IS_INPUT"
    assert source.read_bytes() == original


def test_no_overwrite_finalization_preserves_late_foreign_workbook(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    canary = b"foreign-workbook"
    _inject_late_publish_collision(monkeypatch, output, canary)
    with pytest.raises(ExportError) as caught:
        convert(source, output)
    assert caught.value.code == "OUTPUT_COLLISION"
    assert output.read_bytes() == canary
    assert not list(tmp_path.glob(".ordifile_*"))


def test_no_overwrite_finalization_preserves_late_foreign_sidecar(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    _write_long_formula_source(source, -5)
    output = tmp_path / "result.xlsx"
    sidecar = tmp_path / "result_Peak_Matrix_001.csv"
    canary = b"foreign-sidecar"
    _inject_late_publish_collision(monkeypatch, sidecar, canary)
    with pytest.raises(ExportError) as caught:
        convert(source, output, sidecar_mode="csv")
    assert caught.value.code == "OUTPUT_COLLISION"
    assert sidecar.read_bytes() == canary
    assert not output.exists()
    assert not list(tmp_path.glob(".ordifile_*"))


def test_no_overwrite_keeps_owned_sidecar_before_later_collision(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    _write_long_formula_source(source, -5)
    output = tmp_path / "result.xlsx"
    first_sidecar = tmp_path / "result_Peak_Matrix_001.csv"
    second_sidecar = tmp_path / "result_Peaks_002.csv"
    canary = b"foreign-second-sidecar"
    _inject_late_publish_collision(monkeypatch, second_sidecar, canary)

    with pytest.raises(ExportError) as caught:
        convert(source, output, sidecar_mode="csv")

    assert caught.value.code == "OUTPUT_TRANSACTION_INCOMPLETE"
    assert first_sidecar.exists()
    assert second_sidecar.read_bytes() == canary
    assert not output.exists()
    assert not list(tmp_path.glob(".ordifile_*"))


def test_planned_conversion_reports_incomplete_multi_artifact_publication(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    _write_long_formula_source(source, -5)
    output = tmp_path / "result.xlsx"
    first_sidecar = tmp_path / "result_Peak_Matrix_001.csv"
    second_sidecar = tmp_path / "result_Peaks_002.csv"
    canary = b"foreign-second-sidecar"
    plan = plan_conversion(source, output, sidecar_mode="csv")
    _inject_late_publish_collision(monkeypatch, second_sidecar, canary)

    with pytest.raises(ExportError) as caught:
        convert_plan(plan)

    assert caught.value.code == "OUTPUT_TRANSACTION_INCOMPLETE"
    assert first_sidecar.exists()
    assert second_sidecar.read_bytes() == canary
    assert not output.exists()


def test_no_overwrite_never_deletes_swapped_foreign_sidecar(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    _write_long_formula_source(source, -5)
    output = tmp_path / "result.xlsx"
    first_sidecar = tmp_path / "result_Peak_Matrix_001.csv"
    second_sidecar = tmp_path / "result_Peaks_002.csv"
    first_canary = b"foreign-first-sidecar"
    second_canary = b"foreign-second-sidecar"
    real_rename = excel._rename_no_replace

    def collide_after_swap(
        source_path: Path,
        destination_path: Path,
    ) -> None:
        if Path(destination_path) == second_sidecar:
            first_sidecar.unlink()
            first_sidecar.write_bytes(first_canary)
            second_sidecar.write_bytes(second_canary)
        real_rename(Path(source_path), Path(destination_path))

    monkeypatch.setattr(excel, "_rename_no_replace", collide_after_swap)
    with pytest.raises(ExportError) as caught:
        convert(source, output, sidecar_mode="csv")

    assert caught.value.code == "OUTPUT_TRANSACTION_INCOMPLETE"
    assert first_sidecar.read_bytes() == first_canary
    assert second_sidecar.read_bytes() == second_canary
    assert not output.exists()
    assert not list(tmp_path.glob(".ordifile_*"))


def test_no_overwrite_never_unlinks_a_recreated_temporary_name(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    canary = b"foreign-temporary"
    recreated: Path | None = None
    real_rename = excel._rename_no_replace

    def rename_then_recreate(source_path: Path, destination_path: Path) -> None:
        nonlocal recreated
        real_rename(Path(source_path), Path(destination_path))
        recreated = Path(source_path)
        recreated.write_bytes(canary)

    monkeypatch.setattr(excel, "_rename_no_replace", rename_then_recreate)
    convert(source, output)

    assert output.exists()
    assert recreated is not None
    assert recreated.read_bytes() == canary
    shutil.rmtree(recreated.parent)


@pytest.mark.skipif(os.name == "nt", reason="directory-descriptor cleanup is POSIX-only")
def test_transaction_cleanup_preserves_a_replaced_foreign_directory(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    canary = b"foreign-directory"
    moved_owned_directory: Path | None = None
    foreign_directory: Path | None = None

    def exchange_directory_then_fail(source_path: Path, destination_path: Path) -> None:
        nonlocal moved_owned_directory, foreign_directory
        del destination_path
        foreign_directory = Path(source_path).parent
        moved_owned_directory = foreign_directory.with_name(foreign_directory.name + "_moved")
        foreign_directory.rename(moved_owned_directory)
        foreign_directory.mkdir()
        (foreign_directory / "canary.bin").write_bytes(canary)
        raise OSError("injected finalization failure")

    monkeypatch.setattr(excel, "_rename_no_replace", exchange_directory_then_fail)
    with pytest.raises(ExportError) as caught:
        convert(source, output)

    assert caught.value.code == "OUTPUT_FINALIZATION_FAILED"
    assert foreign_directory is not None
    assert (foreign_directory / "canary.bin").read_bytes() == canary
    assert moved_owned_directory is not None
    assert not list(moved_owned_directory.iterdir())
    foreign_directory.joinpath("canary.bin").unlink()
    foreign_directory.rmdir()
    moved_owned_directory.rmdir()
    assert not output.exists()


@pytest.mark.skipif(os.name == "nt", reason="POSIX temporary symlink regression")
def test_workbook_writes_retained_descriptor_and_rejects_replaced_temp_name(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"
    foreign_target = tmp_path / "foreign.bin"
    canary = b"foreign-target"
    foreign_target.write_bytes(canary)
    exchanged: Path | None = None
    real_fdopen = os.fdopen

    def exchange_before_fd_write(
        descriptor: int,
        mode: str = "r",
        *args: Any,
        **kwargs: Any,
    ) -> Any:
        nonlocal exchanged
        if mode == "w+b":
            transaction_directory = next(tmp_path.glob(".ordifile_transaction_*"))
            exchanged = next(transaction_directory.iterdir())
            exchanged.unlink()
            exchanged.symlink_to(foreign_target)
        return real_fdopen(descriptor, mode, *args, **kwargs)

    monkeypatch.setattr(os, "fdopen", exchange_before_fd_write)
    with pytest.raises(ExportError) as caught:
        convert(source, output)

    assert caught.value.code == "OUTPUT_TEMP_CHANGED"
    assert foreign_target.read_bytes() == canary
    assert not output.exists()
    assert exchanged is not None
    assert exchanged.is_symlink()
    exchanged.unlink()
    exchanged.parent.rmdir()


def test_export_planning_rejects_oversized_int_before_decimal_string_conversion(
    tmp_path: Path,
) -> None:
    class ExplodingStringInt(int):
        def __str__(self) -> str:
            raise ValueError("must not stringify")

    input_path = tmp_path / "source.dat"
    input_path.write_bytes(b"fixture")
    source = SourceFile(input_path, input_path.name, input_path.name, 7, "a" * 64, None, 0)
    oversized = ExplodingStringInt(10**1_000)
    sample = SampleRecord("sample", source, sequence=oversized)
    bundle = DatasetBundle((source,), (sample,))
    file_result = FileResult(source, FileStatus.SUCCESS, bundle=bundle)
    batch = BatchResult(
        (file_result,),
        SortDecision(SortMode.AUTO, SortMode.FILENAME, "test"),
    )
    output = tmp_path / "result.xlsx"

    with pytest.raises(ExportLimitError) as caught:
        ExcelExporter().export(batch, output)

    assert caught.value.code == "INTEGER_LIMIT_EXCEEDED"
    assert not output.exists()


def test_sidecar_safety_escapes_stringified_non_numeric_values_only() -> None:
    class FormulaLikeValue:
        def __str__(self) -> str:
            return "=external-text"

    assert excel._sidecar_safe(FormulaLikeValue()) == ("'=external-text", 1)
    assert excel._sidecar_safe(-5) == ("-5", 0)
    assert excel._sidecar_safe(-5.0) == ("-5.0", 0)
    assert excel._sidecar_safe(Decimal("-5.25")) == ("-5.25", 0)
    assert excel._sidecar_safe(True) == ("True", 0)


def test_manifest_separates_warning_and_error_summaries(tmp_path: Path) -> None:
    warning_source = tmp_path / "warning.csv"
    failed_source = tmp_path / "failed.csv"
    warning_source.write_text(
        "sample_id,acquired_at,area\na,2026-01-01T00:00:00,1\n",
        encoding="utf-8",
    )
    failed_source.write_bytes(b"\xff\xfe")

    result = convert((warning_source, failed_source), tmp_path / "summary.xlsx")
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert "TIMESTAMP_TIMEZONE_MISSING" in manifest["warning_summary"]
        assert "FORMAT_NOT_DETECTED" not in manifest["warning_summary"]
        assert "FORMAT_NOT_DETECTED" in manifest["error_summary"]
        assert "TIMESTAMP_TIMEZONE_MISSING" not in manifest["error_summary"]
    finally:
        workbook.close()


def _metadata_batch(input_path: Path, value: object) -> BatchResult:
    input_path.write_bytes(b"fixture")
    source = SourceFile(input_path, input_path.name, input_path.name, 7, "a" * 64, None, 0)
    sample = SampleRecord("sample", source)
    metadata = MetadataEntry("sample", source.name, "test", "raw", value)
    bundle = DatasetBundle((source,), (sample,), metadata=(metadata,))
    file_result = FileResult(source, FileStatus.SUCCESS, bundle=bundle)
    return BatchResult(
        (file_result,),
        SortDecision(SortMode.AUTO, SortMode.FILENAME, "test"),
    )


@pytest.mark.parametrize(
    "value",
    [
        pytest.param("bad\x01text", id="xml-control"),
        pytest.param("bad\rtext", id="carriage-return-normalization"),
        pytest.param("bad\ud800text", id="unpaired-surrogate"),
        pytest.param("bad\ufdd0text", id="unicode-noncharacter"),
        pytest.param("bad\U0010fffftext", id="plane-noncharacter"),
        pytest.param("bad_x000D_text", id="reserved-xlsx-escape-token"),
    ],
)
def test_workbook_unrepresentable_text_fails_before_output(tmp_path: Path, value: str) -> None:
    output = tmp_path / "invalid.xlsx"
    batch = _metadata_batch(tmp_path / "source.dat", value)

    with pytest.raises(ExportLimitError) as caught:
        ExcelExporter().export(batch, output)

    assert caught.value.code == "WORKBOOK_TEXT_UNREPRESENTABLE"
    assert not output.exists()


def test_representable_workbook_text_reopens_exactly(tmp_path: Path) -> None:
    value = "  한글\tline one\nline two 😀 =literal  "
    batch = _metadata_batch(tmp_path / "source.dat", value)
    output = tmp_path / "exact.xlsx"

    ExcelExporter().export(batch, output)

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        metadata = next(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        assert metadata[4] == value
    finally:
        workbook.close()


def test_unexpected_planning_exception_is_structured_and_writes_nothing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "result.xlsx"

    def fail_planning(_result: object) -> object:
        raise ValueError("synthetic planner failure")

    monkeypatch.setattr(excel, "_samples_data", fail_planning)
    with pytest.raises(ExportError) as caught:
        convert(source, output)
    assert caught.value.code == "WORKBOOK_PLANNING_FAILED"
    assert not output.exists()


def test_manifest_issue_summaries_are_deterministically_bounded(tmp_path: Path) -> None:
    input_path = tmp_path / "source.dat"
    input_path.write_bytes(b"fixture")
    files: list[FileResult] = []
    for index in range(250):
        source = SourceFile(
            input_path,
            f"warning-{index}.dat",
            f"warning-{index}.dat",
            7,
            "a" * 64,
            None,
            index,
        )
        files.append(
            FileResult(
                source,
                FileStatus.WARNING,
                issues=(Issue(f"WARNING_{index:04d}", "Synthetic warning.", Severity.WARNING),),
            )
        )
    for index in range(250):
        source = SourceFile(
            input_path,
            f"error-{index}.dat",
            f"error-{index}.dat",
            7,
            "a" * 64,
            None,
            index + 250,
        )
        files.append(
            FileResult(
                source,
                FileStatus.FAILED,
                issues=(Issue(f"ERROR_{index:04d}", "Synthetic error.", Severity.ERROR),),
            )
        )
    batch = BatchResult(
        tuple(files),
        SortDecision(SortMode.AUTO, SortMode.FILENAME, "test"),
    )
    output = tmp_path / "summaries.xlsx"

    ExcelExporter().export(batch, output)

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        manifest = dict(
            (row[0], row[1]) for row in workbook["Manifest"].iter_rows(values_only=True)
        )
        assert manifest["warning_summary_omitted_count"] == 150
        assert manifest["error_summary_omitted_count"] == 150
        assert manifest["warning_summary"].startswith("WARNING_0000; WARNING_0001")
        assert manifest["error_summary"].startswith("ERROR_0000; ERROR_0001")
        assert len(manifest["warning_summary"]) < 32_767
        assert len(manifest["error_summary"]) < 32_767
    finally:
        workbook.close()


@pytest.mark.parametrize("invalid", ("false", 0, 1, None))
def test_direct_exporter_rejects_invalid_overwrite_without_changing_existing_output(
    tmp_path: Path, invalid: object
) -> None:
    batch = _metadata_batch(tmp_path / "source.dat", "value")
    output = tmp_path / "direct.xlsx"
    output.write_bytes(b"direct-export-sentinel")

    with pytest.raises(ExportError) as caught:
        ExcelExporter().export(batch, output, overwrite=invalid)  # type: ignore[arg-type]

    assert caught.value.code == "EXPORT_CONFIGURATION_INVALID"
    assert output.read_bytes() == b"direct-export-sentinel"


def test_direct_exporter_exact_false_and_true_overwrite_semantics(tmp_path: Path) -> None:
    batch = _metadata_batch(tmp_path / "source.dat", "value")
    output = tmp_path / "direct.xlsx"
    output.write_bytes(b"direct-export-sentinel")

    with pytest.raises(ExportError) as caught:
        ExcelExporter().export(batch, output, overwrite=False)
    assert caught.value.code == "OUTPUT_EXISTS"
    assert output.read_bytes() == b"direct-export-sentinel"

    ExcelExporter().export(batch, output, overwrite=True)
    assert output.read_bytes().startswith(b"PK")


@pytest.mark.parametrize(
    ("kwargs", "expected"),
    (
        ({"include_signals": 1}, "EXPORT_CONFIGURATION_INVALID"),
        ({"include_signals": "false"}, "EXPORT_CONFIGURATION_INVALID"),
        ({"sidecar_mode": 1}, "SIDECAR_MODE_INVALID"),
        ({"sidecar_mode": "invalid"}, "SIDECAR_MODE_INVALID"),
    ),
)
def test_direct_exporter_rejects_invalid_options_before_writing(
    tmp_path: Path, kwargs: dict[str, Any], expected: str
) -> None:
    batch = _metadata_batch(tmp_path / "source.dat", "value")
    output = tmp_path / "direct.xlsx"

    with pytest.raises(ExportError) as caught:
        ExcelExporter().export(batch, output, **kwargs)

    assert caught.value.code == expected
    assert not output.exists()


def test_direct_exporter_rejects_invalid_result_and_output_runtime_types(tmp_path: Path) -> None:
    batch = _metadata_batch(tmp_path / "source.dat", "value")
    output = tmp_path / "direct.xlsx"

    with pytest.raises(ExportError) as bad_result:
        ExcelExporter().export(object(), output)  # type: ignore[arg-type]
    assert bad_result.value.code == "EXPORT_INPUT_INVALID"
    with pytest.raises(ExportError) as bad_output:
        ExcelExporter().export(batch, str(output))  # type: ignore[arg-type]
    assert bad_output.value.code == "EXPORT_INPUT_INVALID"
    assert not output.exists()
