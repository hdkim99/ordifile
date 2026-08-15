from __future__ import annotations

import os
import re
import zipfile
from collections.abc import Callable
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.datetime import CALENDAR_MAC_1904  # type: ignore[import-untyped]

from labconvert.adapters import _xlsx_audit, generic_xlsx
from labconvert.adapters.generic_xlsx import preflight_xlsx
from labconvert.api import convert, inspect_file
from labconvert.core.errors import ParseError
from labconvert.core.models import MetadataEntry


def _sheet(workbook: Workbook, name: str, *, hidden: bool = False) -> None:
    worksheet = workbook.create_sheet(name)
    worksheet.append(["sample_id", "retention_time", "area", "compound", "unknown"])
    worksheet.append([name, 1.0, 2.0, "A", "raw"])
    if hidden:
        worksheet.sheet_state = "hidden"


def _rewrite_archive(
    path: Path,
    replacements: dict[str, bytes | Callable[[bytes], bytes]],
    additions: tuple[tuple[str, bytes], ...] = (),
) -> None:
    temporary = path.with_name(path.name + ".rewrite")
    with (
        zipfile.ZipFile(path) as source,
        zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED) as target,
    ):
        for info in source.infolist():
            data = source.read(info)
            replacement = replacements.get(info.filename)
            if replacement is not None:
                data = replacement(data) if callable(replacement) else replacement
            target.writestr(info, data)
        for name, data in additions:
            target.writestr(name, data)
    os.replace(temporary, path)


def _worksheet_xml(*rows: str, dimension: str = "A1:B2") -> bytes:
    body = "".join(rows)
    return (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="{dimension}"/><sheetData>{body}</sheetData></worksheet>'
    ).encode()


def _inline_cell(coordinate: str, value: str) -> str:
    return f'<c r="{coordinate}" t="inlineStr"><is><t>{value}</t></is></c>'


def _numeric_cell(coordinate: str, value: str, *, style: int | None = None) -> str:
    style_attribute = "" if style is None else f' s="{style}"'
    return f'<c r="{coordinate}" t="n"{style_attribute}><v>{value}</v></c>'


def _typed_cell(coordinate: str, cell_type: str, value: str) -> str:
    return f'<c r="{coordinate}" t="{cell_type}"><v>{value}</v></c>'


def _row(number: int, *cells: str) -> str:
    return f'<row r="{number}">{"".join(cells)}</row>'


def _workbook_with_sheet_xml(path: Path, xml: bytes) -> None:
    workbook = Workbook()
    workbook.save(path)
    _rewrite_archive(path, {"xl/worksheets/sheet1.xml": xml})


def _add_shared_strings(path: Path, *items: str) -> None:
    xml = (
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{len(items)}" uniqueCount="{len(items)}">' + "".join(items) + "</sst>"
    ).encode()
    content_type = (
        b'<Override PartName="/xl/sharedStrings.xml" '
        b'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.'
        b'sharedStrings+xml"/>'
    )
    relationship = (
        b'<Relationship Id="rIdSharedStringsTest" '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
        b'sharedStrings" Target="sharedStrings.xml"/>'
    )
    _rewrite_archive(
        path,
        {
            "[Content_Types].xml": lambda data: data.replace(
                b"</Types>", content_type + b"</Types>", 1
            ),
            "xl/_rels/workbook.xml.rels": lambda data: data.replace(
                b"</Relationships>", relationship + b"</Relationships>", 1
            ),
        },
        (("xl/sharedStrings.xml", xml),),
    )


def _replace_workbook_properties(path: Path, replacement: bytes) -> None:
    def replace(data: bytes) -> bytes:
        updated, count = re.subn(rb"<workbookPr\b[^>]*/>", replacement, data, count=1)
        assert count == 1
        return updated

    _rewrite_archive(path, {"xl/workbook.xml": replace})


def test_one_visible_compatible_sheet_is_selected_and_hidden_ignored(tmp_path: Path) -> None:
    source = tmp_path / "table.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "Visible")
    _sheet(workbook, "Hidden", hidden=True)
    workbook.save(source)
    inspected = inspect_file(source)
    assert inspected.file.adapter_id == "generic_xlsx"
    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sample_id == "Visible"


def test_explicit_sheet_name_can_select_hidden_sheet(tmp_path: Path) -> None:
    source = tmp_path / "hidden-explicit.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "Visible")
    _sheet(workbook, "Hidden", hidden=True)
    workbook.save(source)

    inspected = inspect_file(source, sheet="Hidden")

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sample_id == "Hidden"


def test_multiple_visible_compatible_sheets_require_selection(tmp_path: Path) -> None:
    source = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "One")
    _sheet(workbook, "Two")
    workbook.save(source)
    failed = convert(source, tmp_path / "failed.xlsx")
    assert failed.failure_count == 1
    assert failed.files[0].issues[0].code == "XLSX_SHEET_AMBIGUOUS"
    selected = inspect_file(source, sheet="Two")
    assert selected.file.bundle is not None
    assert selected.file.bundle.samples[0].sample_id == "Two"


def test_formula_cell_is_preserved_as_literal_and_warned(tmp_path: Path) -> None:
    source = tmp_path / "formula.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["sample_id", "retention_time", "area", "compound"])
    worksheet.append(["sample", 1, "=1+2", "A"])
    workbook.save(source)
    output = tmp_path / "result.xlsx"
    result = convert(source, output)
    assert {issue.code for issue in result.files[0].issues} == {"FORMULA_PRESERVED"}
    exported = load_workbook(output, read_only=True, data_only=False)
    try:
        metadata = list(exported["Metadata"].iter_rows(min_row=2))
        formula_cells = [cell for row in metadata for cell in row if cell.value == "=1+2"]
        assert formula_cells and formula_cells[0].data_type == "s"
    finally:
        exported.close()


def test_empty_and_corrupt_xlsx_are_isolated(tmp_path: Path) -> None:
    empty = tmp_path / "empty.xlsx"
    Workbook().save(empty)
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"PK not really xlsx")
    result = convert((empty, corrupt), tmp_path / "result.xlsx")
    assert result.failure_count == 2
    assert {item.issues[0].code for item in result.files} == {
        "XLSX_NO_COMPATIBLE_SHEET",
        "FORMAT_NOT_DETECTED",
    }


def _minimal_archive(path: Path, *extra_members: tuple[str, bytes]) -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", b"types")
        archive.writestr("xl/workbook.xml", b"workbook")
        for name, content in extra_members:
            archive.writestr(name, content)


def test_xlsx_preflight_rejects_path_traversal(tmp_path: Path) -> None:
    source = tmp_path / "traversal.xlsx"
    _minimal_archive(source, ("../escape.xml", b"no"))
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_PATH_TRAVERSAL"


def test_xlsx_preflight_rejects_encrypted_flag(tmp_path: Path) -> None:
    source = tmp_path / "encrypted.xlsx"
    _minimal_archive(source)
    content = bytearray(source.read_bytes())
    for signature, flag_offset in ((b"PK\x03\x04", 6), (b"PK\x01\x02", 8)):
        start = 0
        while (position := content.find(signature, start)) >= 0:
            flags = int.from_bytes(
                content[position + flag_offset : position + flag_offset + 2], "little"
            )
            content[position + flag_offset : position + flag_offset + 2] = (flags | 1).to_bytes(
                2, "little"
            )
            start = position + 4
    source.write_bytes(content)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_ENCRYPTED"


def test_xlsx_preflight_limits_members_size_and_ratio(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "bounded.xlsx"
    _minimal_archive(source, ("xl/payload.bin", b"A" * 4096))
    monkeypatch.setattr(generic_xlsx, "MAX_XLSX_MEMBERS", 2)
    with pytest.raises(ParseError) as members:
        preflight_xlsx(source)
    assert members.value.code == "XLSX_MEMBER_LIMIT"

    monkeypatch.setattr(generic_xlsx, "MAX_XLSX_MEMBERS", 10)
    monkeypatch.setattr(generic_xlsx, "MAX_XLSX_UNCOMPRESSED_BYTES", 5)
    with pytest.raises(ParseError) as size:
        preflight_xlsx(source)
    assert size.value.code == "XLSX_SIZE_LIMIT"

    monkeypatch.setattr(generic_xlsx, "MAX_XLSX_UNCOMPRESSED_BYTES", 100_000)
    monkeypatch.setattr(generic_xlsx, "_RATIO_MINIMUM_SIZE", 1)
    monkeypatch.setattr(generic_xlsx, "MAX_XLSX_COMPRESSION_RATIO", 2.0)
    with pytest.raises(ParseError) as ratio:
        preflight_xlsx(source)
    assert ratio.value.code == "XLSX_COMPRESSION_RATIO"


@pytest.mark.parametrize("digits", ("١", "１"))
def test_xlsx_rejects_non_ascii_shared_string_and_style_indices(
    tmp_path: Path, digits: str
) -> None:
    shared = tmp_path / "shared-index.xlsx"
    _workbook_with_sheet_xml(
        shared,
        _worksheet_xml(
            _row(1, _typed_cell("A1", "s", digits)),
            dimension="A1",
        ),
    )
    with pytest.raises(ParseError) as shared_error:
        preflight_xlsx(shared)
    assert shared_error.value.code == "XLSX_SHARED_STRING_INDEX_INVALID"

    style = tmp_path / "style-index.xlsx"
    xml = _worksheet_xml(
        f'<row r="1"><c r="A1" t="n" s="{digits}"><v>1</v></c></row>',
        dimension="A1",
    )
    _workbook_with_sheet_xml(style, xml)
    with pytest.raises(ParseError) as style_error:
        preflight_xlsx(style)
    assert style_error.value.code == "XLSX_STYLE_INDEX_INVALID"


@pytest.mark.parametrize("digits", ("١", "１"))
def test_xlsx_rejects_non_ascii_sheet_count_coordinate_and_formula_indices(
    tmp_path: Path, digits: str
) -> None:
    sheet_id = tmp_path / "sheet-id.xlsx"
    Workbook().save(sheet_id)
    _rewrite_archive(
        sheet_id,
        {
            "xl/workbook.xml": lambda data: re.sub(
                rb'sheetId="[0-9]+"',
                f'sheetId="{digits}"'.encode(),
                data,
                count=1,
            )
        },
    )
    with pytest.raises(ParseError) as sheet_error:
        preflight_xlsx(sheet_id)
    assert sheet_error.value.code == "XLSX_SHEET_ID_INVALID"

    shared_count = tmp_path / "shared-count.xlsx"
    Workbook().save(shared_count)
    shared_xml = (
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{digits}" uniqueCount="0"></sst>'
    ).encode()
    _rewrite_archive(shared_count, {}, (("xl/sharedStrings.xml", shared_xml),))
    with pytest.raises(ParseError) as count_error:
        preflight_xlsx(shared_count)
    assert count_error.value.code == "XLSX_SHARED_STRING_COUNT_INVALID"

    style_count = tmp_path / "style-count.xlsx"
    Workbook().save(style_count)
    _rewrite_archive(
        style_count,
        {
            "xl/styles.xml": lambda data: re.sub(
                rb'(<cellXfs count=")[0-9]+',
                rb"\g<1>" + digits.encode(),
                data,
                count=1,
            )
        },
    )
    with pytest.raises(ParseError) as style_count_error:
        preflight_xlsx(style_count)
    assert style_count_error.value.code == "XLSX_STYLE_COUNT_INVALID"

    coordinate = tmp_path / "coordinate.xlsx"
    coordinate_xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="A1"/><sheetData><row r="{digits}"><c r="A{digits}" t="n">'
        "<v>1</v></c></row></sheetData></worksheet>"
    ).encode()
    _workbook_with_sheet_xml(coordinate, coordinate_xml)
    with pytest.raises(ParseError) as coordinate_error:
        preflight_xlsx(coordinate)
    assert coordinate_error.value.code in {
        "XLSX_ROW_COORDINATE_REQUIRED",
        "XLSX_CELL_COORDINATE_REQUIRED",
    }

    formula = tmp_path / "formula-index.xlsx"
    formula_xml = _worksheet_xml(
        f'<row r="1"><c r="A1" t="n"><f si="{digits}">1+1</f><v>2</v></c></row>',
        dimension="A1",
    )
    _workbook_with_sheet_xml(formula, formula_xml)
    with pytest.raises(ParseError) as formula_error:
        preflight_xlsx(formula)
    assert formula_error.value.code == "XLSX_FORMULA_INDEX_INVALID"


@pytest.mark.parametrize("date_table_first", (False, True))
def test_xlsx_styles_rejects_duplicate_cell_xfs_regardless_of_table_order(
    tmp_path: Path, date_table_first: bool
) -> None:
    source = tmp_path / "duplicate-cell-xfs.xlsx"
    Workbook().save(source)

    def duplicate_table(data: bytes) -> bytes:
        match = re.search(rb"<cellXfs\b[^>]*>.*?</cellXfs>", data, flags=re.DOTALL)
        assert match is not None
        normal = match.group(0)
        date_style = re.sub(rb'numFmtId="[0-9]+"', b'numFmtId="14"', normal, count=1)
        replacement = date_style + normal if date_table_first else normal + date_style
        return data[: match.start()] + replacement + data[match.end() :]

    _rewrite_archive(source, {"xl/styles.xml": duplicate_table})

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_STYLE_TABLE_INVALID"


@pytest.mark.parametrize("declared", ("0", "2"))
def test_xlsx_styles_requires_declared_count_to_match_direct_xf_children(
    tmp_path: Path, declared: str
) -> None:
    source = tmp_path / "style-count-mismatch.xlsx"
    Workbook().save(source)
    _rewrite_archive(
        source,
        {
            "xl/styles.xml": lambda data: re.sub(
                rb'(<cellXfs count=")[0-9]+',
                rb"\g<1>" + declared.encode(),
                data,
                count=1,
            )
        },
    )

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_STYLE_COUNT_MISMATCH"


def test_xlsx_extra_formula_cell_is_preserved_with_position(tmp_path: Path) -> None:
    source = tmp_path / "extra-formula.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["sample_id", "area"])
    worksheet.append(["a", 1, "=1+1"])
    workbook.save(source)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    extra = next(
        entry for entry in inspected.file.bundle.metadata if entry.key == "unmapped_column_3"
    )
    assert extra.value == "=1+1"
    assert extra.source == "sheet:1:cell:C2"
    assert {issue.code for issue in inspected.file.issues} >= {
        "EXTRA_CELLS_PRESERVED",
        "FORMULA_PRESERVED",
    }


def test_xlsx_rejects_duplicate_normalized_zip_member(tmp_path: Path) -> None:
    source = tmp_path / "duplicate-member.xlsx"
    Workbook().save(source)
    with zipfile.ZipFile(source) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml")
    _rewrite_archive(source, {}, (("xl/worksheets/./sheet1.xml", sheet),))
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_DUPLICATE_MEMBER"


@pytest.mark.parametrize(
    ("content_type", "expected"),
    (
        (
            "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            "XLSX_MACRO_CONTENT_UNSUPPORTED",
        ),
        (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
            "XLSX_WORKBOOK_CONTENT_TYPE",
        ),
        ("application/octet-stream", "XLSX_WORKBOOK_CONTENT_TYPE"),
    ),
)
def test_xlsx_rejects_non_xlsx_workbook_content_types(
    tmp_path: Path, content_type: str, expected: str
) -> None:
    source = tmp_path / "wrong-main-type.xlsx"
    Workbook().save(source)

    def replace_type(data: bytes) -> bytes:
        return re.sub(
            rb"application/vnd\.openxmlformats-officedocument\.spreadsheetml\.sheet\.main\+xml",
            content_type.encode(),
            data,
        )

    _rewrite_archive(source, {"[Content_Types].xml": replace_type})
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == expected


def test_xlsx_rejects_vba_member_even_with_xlsx_main_type(tmp_path: Path) -> None:
    source = tmp_path / "vba.xlsx"
    Workbook().save(source)
    _rewrite_archive(source, {}, (("xl/vbaProject.bin", b"not executable"),))
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_MACRO_CONTENT_UNSUPPORTED"


def test_xlsx_rejects_unsafe_worksheet_relationship(tmp_path: Path) -> None:
    source = tmp_path / "unsafe-rel.xlsx"
    Workbook().save(source)

    def unsafe_target(data: bytes) -> bytes:
        return data.replace(b"/xl/worksheets/sheet1.xml", b"../outside.xml")

    _rewrite_archive(source, {"xl/_rels/workbook.xml.rels": unsafe_target})
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_WORKSHEET_RELATIONSHIP"


@pytest.mark.parametrize(
    ("member", "mutate"),
    (
        (
            "[Content_Types].xml",
            lambda data: data.replace(
                b"http://schemas.openxmlformats.org/package/2006/content-types",
                b"urn:evil-content-types",
                1,
            ),
        ),
        (
            "_rels/.rels",
            lambda data: data.replace(
                b"http://schemas.openxmlformats.org/package/2006/relationships",
                b"urn:evil-relationships",
                1,
            ),
        ),
        (
            "[Content_Types].xml",
            lambda data: data.replace(
                b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
                b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types" '
                b'xmlns:evil="urn:evil">',
            ).replace(b"<Override ", b"<evil:Override ", 1),
        ),
        (
            "_rels/.rels",
            lambda data: data.replace(
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
                b'relationships">',
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
                b'relationships" xmlns:evil="urn:evil">',
            ).replace(b"<Relationship ", b"<evil:Relationship ", 1),
        ),
        (
            "xl/workbook.xml",
            lambda data: data.replace(
                b'<workbook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships" ',
                b'<workbook xmlns:evil="urn:evil" xmlns:r="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships" ',
            ).replace(b"<sheet ", b"<evil:sheet ", 1),
        ),
        (
            "xl/workbook.xml",
            lambda data: data.replace(
                b'<workbook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships" ',
                b'<workbook xmlns:evil="urn:evil" xmlns:r="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships" ',
            ).replace(b"r:id=", b"evil:id=", 1),
        ),
    ),
)
def test_xlsx_rejects_spoofed_package_semantic_namespaces(
    tmp_path: Path, member: str, mutate: Callable[[bytes], bytes]
) -> None:
    source = tmp_path / "namespace-package.xlsx"
    Workbook().save(source)
    _rewrite_archive(source, {member: mutate})
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_NAMESPACE_INVALID"


@pytest.mark.parametrize(
    "xml",
    (
        (
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:evil="urn:evil"><dimension ref="A1"/><sheetData><row r="1">'
            b'<evil:c r="A1" t="n"><evil:v>7</evil:v></evil:c>'
            b"</row></sheetData></worksheet>"
        ),
        (
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:evil="urn:evil"><dimension ref="A1"/><sheetData><row r="1">'
            b'<c r="A1" t="n"><evil:v>7</evil:v></c>'
            b"</row></sheetData></worksheet>"
        ),
        (
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:evil="urn:evil"><evil:dimension ref="A1"/><sheetData/></worksheet>'
        ),
        (
            b'<evil:worksheet xmlns:evil="urn:evil"><evil:dimension ref="A1"/>'
            b"<evil:sheetData/></evil:worksheet>"
        ),
    ),
)
def test_xlsx_rejects_spoofed_worksheet_semantic_namespaces(tmp_path: Path, xml: bytes) -> None:
    source = tmp_path / "namespace-sheet.xlsx"
    _workbook_with_sheet_xml(source, xml)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_NAMESPACE_INVALID"


@pytest.mark.parametrize(
    "cell",
    (
        '<c r="A1" t="inlineStr"><is><t>text</t></is><v>7</v></c>',
        '<c r="A1" t="n"><is><t>text</t></is></c>',
        '<c r="A1" t="inlineStr"><is><t>one</t></is><is><t>two</t></is></c>',
    ),
)
def test_xlsx_rejects_mixed_or_duplicate_inline_value_forms(tmp_path: Path, cell: str) -> None:
    source = tmp_path / "mixed-cell.xlsx"
    _workbook_with_sheet_xml(source, _worksheet_xml(_row(1, cell), dimension="A1"))
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code in {"XLSX_CELL_VALUE_MIXED", "XLSX_INLINE_STRING_INVALID"}


@pytest.mark.parametrize(
    ("xml", "expected"),
    (
        (
            _worksheet_xml('<row><c r="A1" t="inlineStr"><is><t>sample_id</t></is></c></row>'),
            "XLSX_ROW_COORDINATE_REQUIRED",
        ),
        (
            _worksheet_xml('<row r="1"><c t="inlineStr"><is><t>sample_id</t></is></c></row>'),
            "XLSX_CELL_COORDINATE_REQUIRED",
        ),
        (
            _worksheet_xml(_row(1, _inline_cell("A2", "sample_id"))),
            "XLSX_CELL_ROW_MISMATCH",
        ),
        (
            _worksheet_xml(_row(2, _inline_cell("A2", "a")), _row(1, _inline_cell("A1", "b"))),
            "XLSX_ROW_ORDER_INVALID",
        ),
        (
            _worksheet_xml(
                _row(1, _inline_cell("A1", "sample_id"), _inline_cell("A1", "duplicate"))
            ),
            "XLSX_CELL_ORDER_INVALID",
        ),
        (
            _worksheet_xml(_row(1, _inline_cell("B1", "area"), _inline_cell("A1", "sample_id"))),
            "XLSX_CELL_ORDER_INVALID",
        ),
        (
            _worksheet_xml(_row(1, _inline_cell("XFE1", "sample_id"))),
            "XLSX_COORDINATE_OUT_OF_RANGE",
        ),
        (
            _worksheet_xml(_row(1, _inline_cell("a1", "sample_id"))),
            "XLSX_CELL_COORDINATE_REQUIRED",
        ),
        (
            _worksheet_xml(_row(250_001, _inline_cell("A250001", "sample_id"))),
            "XLSX_PROJECTED_ROW_LIMIT",
        ),
        (
            _worksheet_xml(_row(1_000, _inline_cell("XFD1000", "sample_id"))),
            "XLSX_PROJECTED_CELL_LIMIT",
        ),
        (
            _worksheet_xml(_row(1, _inline_cell("A1", "sample_id")), dimension="not-a-range"),
            "XLSX_DIMENSION_INVALID",
        ),
        (
            (
                b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                b'<dimension ref="A1"/><c r="A1" t="n"><v>1</v></c><sheetData/>'
                b"</worksheet>"
            ),
            "XLSX_CELL_STRUCTURE_INVALID",
        ),
        (
            _worksheet_xml(_row(1, '<c r="A1" t="n"><v>1</v><v>2</v></c>'), dimension="A1"),
            "XLSX_CELL_VALUE_DUPLICATE",
        ),
        (
            _worksheet_xml(_row(1, '<c r="A1"><f>1</f><f>2</f><v>3</v></c>'), dimension="A1"),
            "XLSX_CELL_FORMULA_DUPLICATE",
        ),
    ),
)
def test_xlsx_coordinate_and_bound_audit_rejects_malformed_sheets(
    tmp_path: Path, xml: bytes, expected: str
) -> None:
    source = tmp_path / f"{expected}.xlsx"
    _workbook_with_sheet_xml(source, xml)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == expected


def test_xlsx_requires_single_sheet_data(tmp_path: Path) -> None:
    source = tmp_path / "two-sheet-data.xlsx"
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<dimension ref="A1"/><sheetData/><sheetData/></worksheet>'
    )
    _workbook_with_sheet_xml(source, xml)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_SHEETDATA_INVALID"


def test_xlsx_practical_caps_are_enforced_without_large_fixtures(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "caps.xlsx"
    xml = _worksheet_xml(
        _row(1, _inline_cell("A1", "sample_id")),
        _row(2, _inline_cell("A2", "sample")),
        dimension="A1:A2",
    )
    _workbook_with_sheet_xml(source, xml)
    monkeypatch.setattr(_xlsx_audit, "MAX_XLSX_PHYSICAL_ROWS", 1)
    with pytest.raises(ParseError) as rows:
        preflight_xlsx(source)
    assert rows.value.code == "XLSX_PHYSICAL_ROW_LIMIT"

    monkeypatch.setattr(_xlsx_audit, "MAX_XLSX_PHYSICAL_ROWS", 10)
    monkeypatch.setattr(_xlsx_audit, "MAX_XLSX_PHYSICAL_CELLS", 1)
    with pytest.raises(ParseError) as cells:
        preflight_xlsx(source)
    assert cells.value.code == "XLSX_PHYSICAL_CELL_LIMIT"

    monkeypatch.setattr(_xlsx_audit, "MAX_XLSX_PHYSICAL_CELLS", 10)
    monkeypatch.setattr(_xlsx_audit, "MAX_XLSX_RAW_CELL_LEXEME", 5)
    with pytest.raises(ParseError) as lexeme:
        preflight_xlsx(source)
    assert lexeme.value.code == "XLSX_CELL_LEXEME_LIMIT"


def test_xlsx_xml_depth_is_bounded(tmp_path: Path) -> None:
    source = tmp_path / "deep.xlsx"
    nesting = "<x>" * 129 + "</x>" * 129
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="A1"/>{nesting}<sheetData/></worksheet>'
    ).encode()
    _workbook_with_sheet_xml(source, xml)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_XML_DEPTH_LIMIT"


@pytest.mark.parametrize("dimension", ("A1:B1", "A1:XFD1048576"))
def test_xlsx_dimension_mismatch_uses_actual_cells_without_truncation_or_fill(
    tmp_path: Path, dimension: str
) -> None:
    source = tmp_path / "dimension-mismatch.xlsx"
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", "area"),
            _inline_cell("C1", "compound"),
        ),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _numeric_cell("B2", "2.5"),
            _inline_cell("C2", "methanol"),
        ),
        dimension=dimension,
    )
    _workbook_with_sheet_xml(source, xml)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    assert inspected.file.bundle.peaks[0].area == 2.5
    mismatch = next(
        issue for issue in inspected.file.issues if issue.code == "XLSX_DIMENSION_MISMATCH"
    )
    assert dict(mismatch.context)["declared"] == dimension


def test_xlsx_valid_sparse_bounded_sheet_preserves_far_column(tmp_path: Path) -> None:
    source = tmp_path / "sparse.xlsx"
    xml = _worksheet_xml(
        _row(1, _inline_cell("A1", "sample_id"), _inline_cell("B1", "area")),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _numeric_cell("B2", "1"),
            _inline_cell("XFD2", "far raw"),
        ),
        dimension="A1:XFD2",
    )
    _workbook_with_sheet_xml(source, xml)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    far = next(
        entry for entry in inspected.file.bundle.metadata if entry.key == "unmapped_column_16384"
    )
    assert far.value == "far raw"


def test_xlsx_raw_numeric_lexemes_drive_loss_and_exact_integer_policy(tmp_path: Path) -> None:
    source = tmp_path / "raw-numbers.xlsx"
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", "area"),
            _inline_cell("C1", "sequence"),
            _inline_cell("D1", "peak_number"),
            _inline_cell("E1", "compound"),
        ),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _numeric_cell("B2", "0.10000000000000001"),
            _numeric_cell("C2", "9007199254740993"),
            _numeric_cell("D2", "9007199254740993"),
            _inline_cell("E2", "A"),
        ),
        dimension="A1:E2",
    )
    _workbook_with_sheet_xml(source, xml)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    bundle = inspected.file.bundle
    assert bundle.samples[0].sequence == 9_007_199_254_740_993
    assert bundle.peaks[0].peak_number == 9_007_199_254_740_993
    assert bundle.peaks[0].area is None
    raw_area = next(entry for entry in bundle.metadata if entry.key == "area")
    assert raw_area.value == "0.10000000000000001"
    assert raw_area.source == "sheet:1:cell:B2"
    assert "LOSSY_FLOAT_REJECTED" in {issue.code for issue in inspected.file.issues}


def test_xlsx_exponent_and_nonfinite_lexemes_remain_explicit(tmp_path: Path) -> None:
    source = tmp_path / "bounded-numbers.xlsx"
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", "sequence"),
            _inline_cell("C1", "peak_number"),
            _inline_cell("D1", "area"),
            _inline_cell("E1", "compound"),
        ),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _numeric_cell("B2", "1e5000"),
            _numeric_cell("C2", "1e1000000"),
            '<c r="D2" t="str"><v>NaN</v></c>',
            _inline_cell("E2", "A"),
        ),
        dimension="A1:E2",
    )
    _workbook_with_sheet_xml(source, xml)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    bundle = inspected.file.bundle
    assert bundle.samples[0].sequence is None
    assert bundle.peaks[0].peak_number is None
    assert {
        entry.value for entry in bundle.metadata if entry.key in {"sequence", "peak_number"}
    } == {
        "1e5000",
        "1e1000000",
    }
    assert {"INTEGER_LIMIT_EXCEEDED", "NONFINITE_NUMBER"} <= {
        issue.code for issue in inspected.file.issues
    }


def test_xlsx_nonfinite_numeric_storage_is_rejected_before_openpyxl(tmp_path: Path) -> None:
    source = tmp_path / "numeric-nan.xlsx"
    xml = _worksheet_xml(
        _row(1, _inline_cell("A1", "sample_id"), _inline_cell("B1", "area")),
        _row(2, _inline_cell("A2", "sample"), _numeric_cell("B2", "NaN")),
        dimension="A1:B2",
    )
    _workbook_with_sheet_xml(source, xml)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_NUMERIC_LEXEME_NONFINITE"


@pytest.mark.parametrize(
    "lexeme",
    ("0", "+10", "-10", ".5", "1.", "-0.25", "+1.25E-3", "1e+2"),
)
def test_xlsx_numeric_lexeme_accepts_ascii_sign_decimal_and_exponent_boundaries(
    tmp_path: Path, lexeme: str
) -> None:
    source = tmp_path / "valid-number.xlsx"
    xml = _worksheet_xml(_row(1, _numeric_cell("A1", lexeme)), dimension="A1")
    _workbook_with_sheet_xml(source, xml)

    preflight_xlsx(source)


@pytest.mark.parametrize(
    "lexeme",
    ("1_0", "١٢", "1 0", " 10", "10 ", "1\t0", "++1", ".", "1e"),
)
def test_xlsx_numeric_lexeme_rejects_non_ascii_and_whitespace_forms(
    tmp_path: Path, lexeme: str
) -> None:
    source = tmp_path / "invalid-number.xlsx"
    xml = _worksheet_xml(_row(1, _numeric_cell("A1", lexeme)), dimension="A1")
    _workbook_with_sheet_xml(source, xml)

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_NUMERIC_LEXEME_INVALID"


def test_invalid_xlsx_numeric_lexeme_is_isolated_from_good_csv(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad-number.xlsx"
    xml = _worksheet_xml(
        _row(1, _inline_cell("A1", "sample_id"), _inline_cell("B1", "area")),
        _row(2, _inline_cell("A2", "bad"), _numeric_cell("B2", "1_0")),
        dimension="A1:B2",
    )
    _workbook_with_sheet_xml(bad, xml)

    result = convert((good, bad), tmp_path / "mixed-number.xlsx")

    assert result.success_count == 1
    assert result.failure_count == 1
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("lexeme", "expected", "reliable"),
    (
        ("2026-01-02T03:04:05", datetime(2026, 1, 2, 3, 4, 5), False),
        (
            "2026-01-02T03:04:05+09:00",
            datetime.fromisoformat("2026-01-02T03:04:05+09:00"),
            True,
        ),
        ("2026-01-02T03:04:05Z", datetime.fromisoformat("2026-01-02T03:04:05+00:00"), True),
    ),
)
def test_xlsx_iso_date_cell_uses_raw_lexeme_without_openpyxl_offset_loss(
    tmp_path: Path, lexeme: str, expected: datetime, reliable: bool
) -> None:
    source = tmp_path / "iso-date.xlsx"
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", "acquired_at"),
        ),
        _row(2, _inline_cell("A2", "sample"), _typed_cell("B2", "d", lexeme)),
        dimension="A1:B2",
    )
    _workbook_with_sheet_xml(source, xml)

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    sample = inspected.file.bundle.samples[0]
    assert sample.acquired_at == expected
    assert sample.acquired_at_reliable is reliable
    metadata = inspected.file.bundle.metadata
    raw = next(entry for entry in metadata if entry.key == "acquired_at_raw_iso")
    assert raw.value == lexeme
    assert raw.source == "sheet:1:cell:B2"
    assert not any(
        entry.key
        in {
            "acquired_at_raw_serial",
            "acquired_at_style_index",
            "acquired_at_workbook_epoch",
        }
        for entry in metadata
    )


def test_xlsx_iso_date_only_and_numeric_field_are_preserved_without_serial_provenance(
    tmp_path: Path,
) -> None:
    source = tmp_path / "iso-date-preserved.xlsx"
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", "acquired_at"),
            _inline_cell("C1", "area"),
        ),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _typed_cell("B2", "d", "2026-01-02"),
            _typed_cell("C2", "d", "2026-01-03T01:02:03Z"),
        ),
        dimension="A1:C2",
    )
    _workbook_with_sheet_xml(source, xml)

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].acquired_at is None
    assert inspected.file.bundle.peaks == ()
    metadata = {entry.key: entry.value for entry in inspected.file.bundle.metadata}
    assert metadata["acquired_at_raw_iso"] == "2026-01-02"
    assert metadata["area_raw_iso"] == "2026-01-03T01:02:03Z"
    assert not any(
        key.endswith("_raw_serial") or key.endswith("_workbook_epoch") for key in metadata
    )
    assert {
        "XLSX_ISO_DATE_TIMESTAMP_UNSUPPORTED",
        "XLSX_ISO_DATE_NUMERIC_REJECTED",
        "XLSX_CELL_TYPE_FIELD_MISMATCH",
    } <= {issue.code for issue in inspected.file.issues}


@pytest.mark.parametrize(
    "lexeme", ("2026-13-01", "2026-01-02 03:04:05", "2026-01-02T03:04:05+25:00")
)
def test_xlsx_iso_date_rejects_invalid_calendar_and_unsupported_forms(
    tmp_path: Path, lexeme: str
) -> None:
    source = tmp_path / "invalid-iso.xlsx"
    xml = _worksheet_xml(
        _row(1, _typed_cell("A1", "d", lexeme)),
        dimension="A1",
    )
    _workbook_with_sheet_xml(source, xml)

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_ISO_DATE_LEXEME_INVALID"


@pytest.mark.parametrize(
    "semantic",
    (
        "sample_id",
        "instrument_type",
        "vendor",
        "channel",
        "detector",
        "retention_time_unit",
        "compound",
        "compound_source",
        "x_unit",
        "y_unit",
    ),
)
@pytest.mark.parametrize(
    ("cell_type", "raw_value"),
    (
        ("b", "1"),
        ("e", "#VALUE!"),
        ("d", "2026-01-02T03:04:05Z"),
        ("n", "42"),
    ),
)
def test_xlsx_mapped_text_fields_reject_typed_cell_reinterpretation(
    tmp_path: Path, semantic: str, cell_type: str, raw_value: str
) -> None:
    source = tmp_path / f"typed-{semantic}-{cell_type}.xlsx"
    if semantic == "sample_id":
        header = _row(1, _inline_cell("A1", semantic))
        data = _row(2, _typed_cell("A2", cell_type, raw_value))
        dimension = "A1:A2"
        coordinate = "A2"
    else:
        header = _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", semantic),
        )
        data = _row(
            2,
            _inline_cell("A2", "sample"),
            _typed_cell("B2", cell_type, raw_value),
        )
        dimension = "A1:B2"
        coordinate = "B2"
    _workbook_with_sheet_xml(source, _worksheet_xml(header, data, dimension=dimension))

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sample_id == (
        source.stem if semantic == "sample_id" else "sample"
    )
    assert "XLSX_CELL_TYPE_FIELD_MISMATCH" in {issue.code for issue in inspected.file.issues}
    source_locator = f"sheet:1:cell:{coordinate}"
    typed = [
        entry
        for entry in inspected.file.bundle.metadata
        if entry.key == "xlsx_cell_type" and entry.source == source_locator
    ]
    assert len(typed) == 1
    assert typed[0].value == cell_type
    raw_key = f"{semantic}_raw_iso" if cell_type == "d" else semantic
    preserved = next(
        entry
        for entry in inspected.file.bundle.metadata
        if entry.key == raw_key and entry.source == source_locator
    )
    assert preserved.value == raw_value


@pytest.mark.parametrize(
    ("semantic", "cell_type", "raw_value"),
    (
        ("sequence", "b", "1"),
        ("area", "e", "#VALUE!"),
        ("signal", "d", "2026-01-02T03:04:05Z"),
        ("acquired_at", "b", "0"),
    ),
)
def test_xlsx_numeric_and_timestamp_fields_reject_incompatible_cell_types(
    tmp_path: Path, semantic: str, cell_type: str, raw_value: str
) -> None:
    source = tmp_path / f"typed-scientific-{semantic}.xlsx"
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", semantic),
        ),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _typed_cell("B2", cell_type, raw_value),
        ),
        dimension="A1:B2",
    )
    _workbook_with_sheet_xml(source, xml)

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert "XLSX_CELL_TYPE_FIELD_MISMATCH" in {issue.code for issue in inspected.file.issues}
    assert any(
        entry.key == "xlsx_cell_type"
        and entry.value == cell_type
        and entry.source == "sheet:1:cell:B2"
        for entry in inspected.file.bundle.metadata
    )


@pytest.mark.parametrize(("epoch_1904", "epoch_label"), ((False, "1900"), (True, "1904")))
def test_xlsx_date_style_acquired_at_preserves_serial_style_and_epoch(
    tmp_path: Path, epoch_1904: bool, epoch_label: str
) -> None:
    source = tmp_path / f"date-{epoch_label}.xlsx"
    workbook = Workbook()
    if epoch_1904:
        workbook.epoch = CALENDAR_MAC_1904
    worksheet = workbook.active
    worksheet.append(["sample_id", "acquired_at"])
    worksheet.append(["sample", datetime(2026, 1, 2, 3, 4, 5)])
    workbook.save(source)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    sample = inspected.file.bundle.samples[0]
    assert sample.acquired_at == datetime(2026, 1, 2, 3, 4, 5)
    assert sample.acquired_at_reliable is False
    metadata = {entry.key: entry for entry in inspected.file.bundle.metadata}
    assert metadata["acquired_at_raw_serial"].source == "sheet:1:cell:B2"
    assert metadata["acquired_at_style_index"].value >= 1
    assert metadata["acquired_at_workbook_epoch"].value == epoch_label
    assert "XLSX_DATE_SERIAL_UNRELIABLE" in {issue.code for issue in inspected.file.issues}


@pytest.mark.parametrize(
    ("date1904_lexeme", "epoch_1904", "epoch_label"),
    (
        ("0", False, "1900"),
        ("false", False, "1900"),
        ("1", True, "1904"),
        ("true", True, "1904"),
    ),
)
def test_xlsx_date1904_accepts_only_exact_schema_boolean_lexemes(
    tmp_path: Path, date1904_lexeme: str, epoch_1904: bool, epoch_label: str
) -> None:
    source = tmp_path / f"date1904-{date1904_lexeme}.xlsx"
    workbook = Workbook()
    if epoch_1904:
        workbook.epoch = CALENDAR_MAC_1904
    worksheet = workbook.active
    worksheet.append(["sample_id", "acquired_at"])
    worksheet.append(["sample", datetime(2026, 1, 2, 3, 4, 5)])
    workbook.save(source)
    _replace_workbook_properties(
        source,
        f'<workbookPr date1904="{date1904_lexeme}"/>'.encode(),
    )

    audit = preflight_xlsx(source)
    assert audit.date_1904 is epoch_1904
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    sample = inspected.file.bundle.samples[0]
    assert sample.acquired_at == datetime(2026, 1, 2, 3, 4, 5)
    metadata = {entry.key: entry.value for entry in inspected.file.bundle.metadata}
    assert metadata["acquired_at_workbook_epoch"] == epoch_label


@pytest.mark.parametrize("date1904_lexeme", ("yes", " 1 ", "True", "FALSE", ""))
def test_xlsx_date1904_rejects_non_schema_boolean_lexemes(
    tmp_path: Path, date1904_lexeme: str
) -> None:
    source = tmp_path / "invalid-date1904.xlsx"
    workbook = Workbook()
    workbook.save(source)
    _replace_workbook_properties(
        source,
        f'<workbookPr date1904="{date1904_lexeme}"/>'.encode(),
    )

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == "XLSX_DATE1904_INVALID"


@pytest.mark.parametrize(
    ("replacement", "expected_code", "add_namespace"),
    (
        (
            b'<workbookPr date1904="0"/><workbookPr date1904="1"/>',
            "XLSX_WORKBOOK_PROPERTIES_INVALID",
            False,
        ),
        (
            b'<workbookPr date1904="0" date1904="1"/>',
            "XLSX_XML_INVALID",
            False,
        ),
        (
            b'<workbookPr date1904="0" evil:date1904="1"/>',
            "XLSX_NAMESPACE_INVALID",
            True,
        ),
    ),
)
def test_xlsx_date1904_rejects_duplicate_property_and_attribute_boundaries(
    tmp_path: Path, replacement: bytes, expected_code: str, add_namespace: bool
) -> None:
    source = tmp_path / "duplicate-date1904.xlsx"
    workbook = Workbook()
    workbook.save(source)
    _replace_workbook_properties(source, replacement)
    if add_namespace:
        _rewrite_archive(
            source,
            {
                "xl/workbook.xml": lambda data: data.replace(
                    b"<workbook ",
                    b'<workbook xmlns:evil="urn:evil" ',
                    1,
                )
            },
        )

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == expected_code


def test_invalid_date1904_workbook_is_isolated_from_good_builtin_input(
    tmp_path: Path,
) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad = tmp_path / "bad-date1904.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["sample_id", "area", "compound"])
    worksheet.append(["bad", 3, "B"])
    workbook.save(bad)
    _replace_workbook_properties(bad, b'<workbookPr date1904="yes"/>')

    result = convert((good, bad), tmp_path / "mixed-date1904.xlsx")

    assert result.success_count == 1
    assert result.failure_count == 1
    workbook_result = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook_result["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook_result["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook_result.close()


def test_xlsx_date_style_non_time_numeric_is_not_reinterpreted(tmp_path: Path) -> None:
    source = tmp_path / "date-area.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["sample_id", "area", "compound"])
    worksheet.append(["sample", datetime(2026, 1, 2), "A"])
    workbook.save(source)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    assert inspected.file.bundle.peaks[0].area is None
    raw = next(entry for entry in inspected.file.bundle.metadata if entry.key == "area_raw_serial")
    assert raw.source == "sheet:1:cell:B2"
    assert "XLSX_DATE_STYLE_NUMERIC_REJECTED" in {issue.code for issue in inspected.file.issues}


def test_xlsx_formula_and_cache_are_preserved_by_coordinate_but_not_parsed(tmp_path: Path) -> None:
    source = tmp_path / "formula-cache.xlsx"
    formula = '<c r="C2"><f ref="C2">1+2</f><v>3</v></c>'
    xml = _worksheet_xml(
        _row(
            1,
            _inline_cell("A1", "sample_id"),
            _inline_cell("B1", "retention_time"),
            _inline_cell("C1", "area"),
            _inline_cell("D1", "compound"),
        ),
        _row(
            2,
            _inline_cell("A2", "sample"),
            _numeric_cell("B2", "1"),
            formula,
            _inline_cell("D2", "A"),
        ),
        dimension="A1:D2",
    )
    _workbook_with_sheet_xml(source, xml)
    inspected = inspect_file(source)
    assert inspected.file.bundle is not None
    assert inspected.file.bundle.peaks[0].area is None
    by_key: dict[str, list[MetadataEntry]] = {}
    for entry in inspected.file.bundle.metadata:
        by_key.setdefault(entry.key, []).append(entry)
    assert by_key["area"][0].value == "=1+2"
    assert by_key["area"][0].source == "sheet:1:cell:C2"
    assert by_key["xlsx_formula_lexeme"][0].value == "1+2"
    assert by_key["xlsx_cell_type"][0].value == "n"
    assert by_key["xlsx_formula_cached_lexeme"][0].value == "3"
    assert by_key["xlsx_formula_attribute:ref"][0].value == "C2"


@pytest.mark.parametrize(
    ("attributes", "expected_code"),
    (
        (f'ref="{"A" * 32_768}"', "XLSX_FORMULA_ATTRIBUTE_LIMIT"),
        ('ref="A0:B2"', "XLSX_FORMULA_REFERENCE_INVALID"),
        ('xmlns:evil="urn:evil" evil:ref="A1"', "XLSX_NAMESPACE_INVALID"),
        ('unknown="value"', "XLSX_FORMULA_ATTRIBUTE_INVALID"),
    ),
    ids=("overlong-ref", "invalid-ref", "evil-namespace", "unknown-attribute"),
)
def test_xlsx_formula_attributes_are_strictly_bounded_before_openpyxl(
    tmp_path: Path, attributes: str, expected_code: str
) -> None:
    source = tmp_path / f"{expected_code}.xlsx"
    formula = f'<c r="A1"><f {attributes}>1+2</f></c>'
    _workbook_with_sheet_xml(
        source,
        _worksheet_xml(_row(1, formula), dimension="A1"),
    )

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)

    assert caught.value.code == expected_code


def test_xlsx_valid_shared_formula_attributes_are_preserved(tmp_path: Path) -> None:
    source = tmp_path / "valid-shared-formula.xlsx"
    xml = _worksheet_xml(
        _row(2, '<c r="C2"><f t="shared" ref="C2:C3" si="0" aca="false">1+2</f></c>'),
        _row(3, '<c r="C3"><f t="shared" si="0"/></c>'),
        dimension="C2:C3",
    )
    _workbook_with_sheet_xml(source, xml)

    package = preflight_xlsx(source)
    captured = _xlsx_audit.capture_worksheet_cells(source, package, package.sheets[0])

    assert captured.raw_cells[0].formula_attributes == (
        ("aca", "false"),
        ("ref", "C2:C3"),
        ("si", "0"),
        ("t", "shared"),
    )
    assert captured.raw_cells[1].formula_attributes == (("si", "0"), ("t", "shared"))


def test_xlsx_shared_string_direct_rich_and_index_boundaries(tmp_path: Path) -> None:
    direct = tmp_path / "shared-direct-boundary.xlsx"
    _workbook_with_sheet_xml(
        direct,
        _worksheet_xml(_row(1, '<c r="A1" t="s"><v>0</v></c>'), dimension="A1"),
    )
    _add_shared_strings(direct, f"<si><t>{'a' * 32_767}</t></si>")
    assert preflight_xlsx(direct).shared_string_count == 1

    rich = tmp_path / "shared-rich-boundary.xlsx"
    Workbook().save(rich)
    _add_shared_strings(
        rich,
        f"<si><r><t>{'a' * 16_000}</t></r><r><rPr><b/></rPr><t>{'b' * 16_767}</t></r></si>",
    )
    assert preflight_xlsx(rich).shared_string_count == 1


@pytest.mark.parametrize(
    "item",
    (
        f"<si><t>{'a' * 32_768}</t></si>",
        f"<si><r><t>{'a' * 20_000}</t></r><r><t>{'b' * 20_000}</t></r></si>",
    ),
    ids=("direct-overlong", "rich-aggregate-overlong"),
)
def test_xlsx_shared_string_logical_text_cannot_exceed_cell_boundary(
    tmp_path: Path, item: str
) -> None:
    source = tmp_path / "shared-overlong.xlsx"
    Workbook().save(source)
    _add_shared_strings(source, item)

    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)

    assert caught.value.code == "XLSX_SHARED_STRING_LIMIT"


def test_xlsx_sheet_name_31_character_boundary_is_audited(tmp_path: Path) -> None:
    valid = tmp_path / "sheet-name-31.xlsx"
    Workbook().save(valid)
    _rewrite_archive(
        valid,
        {
            "xl/workbook.xml": lambda data: data.replace(
                b'name="Sheet"', f'name="{"S" * 31}"'.encode(), 1
            )
        },
    )
    assert preflight_xlsx(valid).sheets[0].title == "S" * 31

    invalid = tmp_path / "sheet-name-32.xlsx"
    Workbook().save(invalid)
    _rewrite_archive(
        invalid,
        {
            "xl/workbook.xml": lambda data: data.replace(
                b'name="Sheet"', f'name="{"S" * 32}"'.encode(), 1
            )
        },
    )
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(invalid)
    assert caught.value.code == "XLSX_SHEET_NAME_INVALID"


def test_xlsx_audited_string_failures_are_isolated_from_good_csv(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")

    bad_formula = tmp_path / "bad-formula.xlsx"
    _workbook_with_sheet_xml(
        bad_formula,
        _worksheet_xml(
            _row(1, f'<c r="A1"><f ref="{"A" * 32_768}">1</f></c>'),
            dimension="A1",
        ),
    )
    bad_shared = tmp_path / "bad-shared.xlsx"
    Workbook().save(bad_shared)
    _add_shared_strings(bad_shared, f"<si><t>{'x' * 32_768}</t></si>")
    bad_title = tmp_path / "bad-title.xlsx"
    Workbook().save(bad_title)
    _rewrite_archive(
        bad_title,
        {
            "xl/workbook.xml": lambda data: data.replace(
                b'name="Sheet"', f'name="{"T" * 32}"'.encode(), 1
            )
        },
    )

    result = convert(
        (good, bad_formula, bad_shared, bad_title),
        tmp_path / "audited-string-isolation.xlsx",
    )

    assert result.success_count == 1
    assert result.failure_count == 3
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
        log = list(workbook["Import_Log"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in log} == {"success", "failed"}
    finally:
        workbook.close()


def test_xlsx_formula_exported_literal_boundary_includes_equals_prefix(tmp_path: Path) -> None:
    valid = tmp_path / "formula-literal-32767.xlsx"
    valid_formula = "1" * 32_766
    _workbook_with_sheet_xml(
        valid,
        _worksheet_xml(
            _row(
                1,
                _inline_cell("A1", "sample_id"),
                _inline_cell("B1", "area"),
                _inline_cell("C1", "compound"),
            ),
            _row(
                2,
                _inline_cell("A2", "sample"),
                f'<c r="B2"><f>{valid_formula}</f></c>',
                _inline_cell("C2", "A"),
            ),
            dimension="A1:C2",
        ),
    )
    package = preflight_xlsx(valid)
    captured = _xlsx_audit.capture_worksheet_cells(valid, package, package.sheets[0])
    formula_cell = next(cell for cell in captured.raw_cells if cell.formula_present)
    assert formula_cell.formula == valid_formula
    assert len("=" + (formula_cell.formula or "")) == 32_767
    result = convert(valid, tmp_path / "formula-literal-boundary.xlsx")
    assert result.success_count == 1
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        metadata_values = {
            row[4] for row in workbook["Metadata"].iter_rows(min_row=2, values_only=True)
        }
        assert "=" + valid_formula in metadata_values
    finally:
        workbook.close()

    invalid = tmp_path / "formula-literal-32768.xlsx"
    _workbook_with_sheet_xml(
        invalid,
        _worksheet_xml(
            _row(1, f'<c r="A1"><f>{"1" * 32_767}</f></c>'),
            dimension="A1",
        ),
    )
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(invalid)
    assert caught.value.code == "XLSX_FORMULA_LITERAL_LIMIT"


def test_xlsx_valid_inline_rich_text_is_reconstructed_for_mapping(tmp_path: Path) -> None:
    source = tmp_path / "inline-rich.xlsx"
    rich_header = (
        '<c r="A1" t="inlineStr"><is><r><rPr><b/></rPr><t>sample</t></r><r><t>_id</t></r></is></c>'
    )
    rich_value = (
        '<c r="A2" t="inlineStr"><is><r><t>sample</t></r>'
        "<r><rPr><i/></rPr><t> rich</t></r></is></c>"
    )
    _workbook_with_sheet_xml(
        source,
        _worksheet_xml(
            _row(1, rich_header, _inline_cell("B1", "area"), _inline_cell("C1", "compound")),
            _row(2, rich_value, _numeric_cell("B2", "2"), _inline_cell("C2", "A")),
            dimension="A1:C2",
        ),
    )

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sample_id == "sample rich"


@pytest.mark.parametrize(
    ("inline_payload", "expected_code"),
    (
        ("<t>sample</t><t>_id</t>", "XLSX_STRING_STRUCTURE_INVALID"),
        ("<wrapper><t>sample_id</t></wrapper>", "XLSX_STRING_STRUCTURE_INVALID"),
        (
            '<evil:t xmlns:evil="urn:evil">sample_id</evil:t>',
            "XLSX_NAMESPACE_INVALID",
        ),
    ),
)
def test_xlsx_inline_string_rejects_lossy_or_spoofed_structures(
    tmp_path: Path, inline_payload: str, expected_code: str
) -> None:
    source = tmp_path / f"inline-{expected_code}.xlsx"
    cell = f'<c r="A1" t="inlineStr"><is>{inline_payload}</is></c>'
    _workbook_with_sheet_xml(
        source,
        _worksheet_xml(_row(1, cell), dimension="A1"),
    )
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == expected_code


def test_xlsx_valid_shared_rich_text_is_reconstructed_for_mapping(tmp_path: Path) -> None:
    source = tmp_path / "shared-rich-mapping.xlsx"
    _workbook_with_sheet_xml(
        source,
        _worksheet_xml(
            _row(
                1,
                '<c r="A1" t="s"><v>0</v></c>',
                '<c r="B1" t="s"><v>1</v></c>',
                '<c r="C1" t="s"><v>2</v></c>',
            ),
            _row(
                2,
                '<c r="A2" t="s"><v>3</v></c>',
                _numeric_cell("B2", "2"),
                '<c r="C2" t="s"><v>4</v></c>',
            ),
            dimension="A1:C2",
        ),
    )
    _add_shared_strings(
        source,
        "<si><r><t>sample</t></r><r><rPr><b/></rPr><t>_id</t></r></si>",
        "<si><t>area</t></si>",
        "<si><t>compound</t></si>",
        "<si><r><t>shared</t></r><r><t> sample</t></r></si>",
        "<si><t>A</t></si>",
    )

    inspected = inspect_file(source)

    assert inspected.file.bundle is not None
    assert inspected.file.bundle.samples[0].sample_id == "shared sample"
    assert inspected.file.bundle.peaks[0].compound == "A"


@pytest.mark.parametrize(
    ("item", "expected_code"),
    (
        ("<si><t>sample</t><t>_id</t></si>", "XLSX_STRING_STRUCTURE_INVALID"),
        (
            "<si><wrapper><t>sample_id</t></wrapper></si>",
            "XLSX_STRING_STRUCTURE_INVALID",
        ),
        (
            '<si xmlns:evil="urn:evil"><evil:t>sample_id</evil:t></si>',
            "XLSX_NAMESPACE_INVALID",
        ),
    ),
)
def test_xlsx_shared_string_rejects_lossy_or_spoofed_structures(
    tmp_path: Path, item: str, expected_code: str
) -> None:
    source = tmp_path / f"shared-{expected_code}.xlsx"
    Workbook().save(source)
    _add_shared_strings(source, item)
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == expected_code


def test_xlsx_formula_and_string_failures_remain_file_isolated(tmp_path: Path) -> None:
    good = tmp_path / "good.csv"
    good.write_text("sample_id,area,compound\ngood,2,A\n", encoding="utf-8")
    bad_formula = tmp_path / "bad-formula-literal.xlsx"
    _workbook_with_sheet_xml(
        bad_formula,
        _worksheet_xml(
            _row(1, f'<c r="A1"><f>{"1" * 32_767}</f></c>'),
            dimension="A1",
        ),
    )
    bad_inline = tmp_path / "bad-inline.xlsx"
    _workbook_with_sheet_xml(
        bad_inline,
        _worksheet_xml(
            _row(1, '<c r="A1" t="inlineStr"><is><t>A</t><t>B</t></is></c>'),
            dimension="A1",
        ),
    )
    bad_shared = tmp_path / "bad-shared-structure.xlsx"
    Workbook().save(bad_shared)
    _add_shared_strings(bad_shared, "<si><t>A</t><t>B</t></si>")

    result = convert(
        (good, bad_formula, bad_inline, bad_shared),
        tmp_path / "string-isolation.xlsx",
    )

    assert result.success_count == 1
    assert result.failure_count == 3
    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(peaks) == 1
        assert peaks[0][0] == "good"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("cell", "expected"),
    (
        ('<c r="A1" t="s"><v>99</v></c>', "XLSX_SHARED_STRING_INDEX_INVALID"),
        (_numeric_cell("A1", "1", style=999), "XLSX_STYLE_INDEX_INVALID"),
        ('<c r="A1" t="unsupported"><v>1</v></c>', "XLSX_CELL_TYPE_UNSUPPORTED"),
    ),
)
def test_xlsx_validates_cell_type_shared_string_and_style_indices(
    tmp_path: Path, cell: str, expected: str
) -> None:
    source = tmp_path / f"{expected}.xlsx"
    _workbook_with_sheet_xml(source, _worksheet_xml(_row(1, cell), dimension="A1"))
    with pytest.raises(ParseError) as caught:
        preflight_xlsx(source)
    assert caught.value.code == expected


def test_xlsx_mixed_good_and_adversarial_file_isolates_bad_input(tmp_path: Path) -> None:
    good = tmp_path / "good.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["sample_id", "area", "compound"])
    worksheet.append(["good", 2, "A"])
    workbook.save(good)
    bad = tmp_path / "bad.xlsx"
    bad_xml = _worksheet_xml(
        _row(1, _inline_cell("A1", "sample_id"), _inline_cell("A1", "duplicate"))
    )
    _workbook_with_sheet_xml(bad, bad_xml)
    output = tmp_path / "result.xlsx"
    result = convert((good, bad), output, adapter="generic_xlsx")
    assert result.success_count == 1
    assert result.failure_count == 1
    assert output.exists()
    exported = load_workbook(output, read_only=True, data_only=False)
    try:
        samples = list(exported["Samples"].iter_rows(values_only=True))
        assert any(row[1] == "good" and row[12] in {"success", "warning"} for row in samples[1:])
        peaks = list(exported["Peaks"].iter_rows(values_only=True))
        assert any(row[7] == 2 for row in peaks[1:])
    finally:
        exported.close()
