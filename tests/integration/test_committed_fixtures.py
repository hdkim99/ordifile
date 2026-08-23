from __future__ import annotations

import hashlib
import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile import (
    load_conversion_recipe,
    load_peak_table_mapping,
    load_peak_table_mapping_set,
)
from ordifile.api import convert, convert_recipe, inspect_file, plan_recipe

FIXTURES = Path(__file__).parents[1] / "fixtures" / "synthetic"
EXAMPLES = Path(__file__).parents[2] / "examples" / "basic"
PILOT = Path(__file__).parents[2] / "examples" / "pilot"
EXPECTED_XLSX_SHA256 = "df36be466410ed5c02d75723ec59ed6c896fad2be51ec4bb8f6edcab937fce96"


@pytest.mark.researcher_acceptance
def test_public_safe_pilot_recipe_reopens_with_documented_values(tmp_path: Path) -> None:
    mapping = load_peak_table_mapping(PILOT / "template-a.mapping.json")
    mapping_set = load_peak_table_mapping_set(PILOT / "templates.mapping-set.json")
    recipe = load_conversion_recipe(PILOT / "laboratory.recipe.json")
    assert recipe.peak_table_mapping_set is not None
    assert recipe.peak_table_mapping_set.to_dict() == mapping_set.to_dict()
    assert mapping.to_dict() == mapping_set.profiles[0].mapping.to_dict()

    inputs = PILOT / "inputs"
    source_bytes = {path: path.read_bytes() for path in sorted(inputs.iterdir())}
    output = tmp_path / "Pilot_Result.xlsx"
    plan = plan_recipe(inputs, output, recipe=recipe)
    result = convert_recipe(inputs, output, recipe=recipe, conversion_plan=plan)

    assert plan.summary.mapping_profiles == 2
    assert plan.summary.failed == 0
    assert result.success_count == 2
    assert result.failure_count == 0
    assert {path: path.read_bytes() for path in source_bytes} == source_bytes
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook.active.title == "Samples"
        assert workbook.sheetnames == [
            "Manifest",
            "Samples",
            "Peak_Matrix",
            "Peak_Order_Matrix",
            "Peaks",
            "Metadata",
            "Import_Log",
        ]
        peak_rows = tuple(workbook["Peaks"].iter_rows(values_only=True))
        headers = tuple(peak_rows[0])
        columns = {name: headers.index(name) for name in headers}
        values = {
            (
                row[columns["retention_time"]],
                row[columns["retention_time_unit"]],
                row[columns["area"]],
                row[columns["height"]],
                row[columns["area_unit"]],
                row[columns["height_unit"]],
            )
            for row in peak_rows[1:]
        }
        assert values == {
            (1.25, "min", 10, 2, "pA*s", "pA"),
            (2.5, "min", 20, 4, "pA*s", "pA"),
            (30, "s", 100, None, "AU", None),
            (45, "s", 200, None, "AU", None),
        }
        workbook_text = "\n".join(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        assert "Neutral laboratory pilot" not in workbook_text
        assert str(PILOT.resolve()) not in workbook_text
    finally:
        workbook.close()


def test_committed_xlsx_checksum_matches_fixture_documentation() -> None:
    fixture = FIXTURES / "generic_peaks.xlsx"
    actual = hashlib.sha256(fixture.read_bytes()).hexdigest()
    documentation = (FIXTURES / "README.md").read_text(encoding="utf-8")

    assert actual == EXPECTED_XLSX_SHA256
    assert f"generic_peaks.xlsx  {EXPECTED_XLSX_SHA256}" in documentation


def test_xlsx_generator_reproduces_committed_fixture_bytes(tmp_path: Path) -> None:
    fixture = FIXTURES / "generic_peaks.xlsx"
    generated = tmp_path / "generated" / "generic_peaks.xlsx"

    subprocess.run(
        [
            sys.executable,
            str(FIXTURES / "generate_xlsx.py"),
            "--output",
            str(generated),
        ],
        check=True,
    )

    generated_bytes = generated.read_bytes()
    assert generated_bytes == fixture.read_bytes()
    assert hashlib.sha256(generated_bytes).hexdigest() == EXPECTED_XLSX_SHA256
    with zipfile.ZipFile(generated) as archive:
        members = archive.infolist()
        assert [member.filename for member in members] == sorted(
            member.filename for member in members
        )
        assert all(member.compress_type == zipfile.ZIP_STORED for member in members)
        assert all(member.date_time == (2026, 1, 1, 0, 0, 0) for member in members)
        assert all(member.create_system == 3 for member in members)
        assert all(member.external_attr == 0o600 << 16 for member in members)


def test_committed_synthetic_fixtures_match_verified_support() -> None:
    expected = {
        "generic_peaks.csv": "generic_csv",
        "generic_peaks.tsv": "generic_tsv",
        "generic_peaks_semicolon.txt": "generic_semicolon",
        "generic_peaks.xlsx": "generic_xlsx",
    }
    for name, adapter_id in expected.items():
        inspected = inspect_file(FIXTURES / name)
        assert inspected.file.adapter_id == adapter_id
        assert inspected.file.bundle is not None
        assert len(inspected.file.bundle.peaks) == 2
        assert len(inspected.file.bundle.signals) == 1


def test_committed_fixtures_convert_to_one_ordered_workbook(tmp_path: Path) -> None:
    inputs = tuple(
        FIXTURES / name
        for name in (
            "generic_peaks.xlsx",
            "generic_peaks_semicolon.txt",
            "generic_peaks.tsv",
            "generic_peaks.csv",
        )
    )
    result = convert(inputs, tmp_path / "fixtures.xlsx", include_signals=True)
    assert result.success_count == 4
    assert result.failure_count == 0
    assert result.sort.effective.value == "acquired_at"

    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        samples = list(workbook["Samples"].values)
        sample_id_column = samples[0].index("sample_id")
        assert [row[sample_id_column] for row in samples[1:]] == [
            "synthetic_csv",
            "synthetic_tsv",
            "synthetic_txt",
            "synthetic_xlsx",
        ]
        assert any(name.startswith("Signals_FID") for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_basic_example_demonstrates_natural_filename_sort(tmp_path: Path) -> None:
    result = convert(EXAMPLES, tmp_path / "example.xlsx", sort="filename")
    assert result.success_count == 3
    assert result.sort.effective.value == "filename"

    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        samples = list(workbook["Samples"].values)
        source_file_column = samples[0].index("source_file")
        assert [row[source_file_column] for row in samples[1:]] == [
            "sample_1.csv",
            "sample_2.csv",
            "sample_10.csv",
        ]
    finally:
        workbook.close()


UNICODE_EXAMPLE = Path(__file__).parents[2] / "examples" / "unicode"


def test_unicode_filename_example_preserves_name_and_peak_order(tmp_path: Path) -> None:
    result = convert(
        (UNICODE_EXAMPLE,),
        tmp_path / "unicode.xlsx",
        extensions=(".csv",),
        sort="filename",
    )
    assert result.success_count == 1
    assert result.failure_count == 0
    assert result.sort.effective.value == "filename"

    workbook = load_workbook(result.output_path, read_only=True, data_only=False)
    try:
        samples = list(workbook["Samples"].values)
        header = list(samples[0])
        source_file_column = header.index("source_file")
        status_column = header.index("status")
        sample_rows = samples[1:]
        assert len(sample_rows) == 1
        assert sample_rows[0][source_file_column] == "시료_신호.csv"
        assert sample_rows[0][status_column] == "success"

        peaks = list(workbook["Peaks"].values)
        peak_header = list(peaks[0])
        sample_id_column = peak_header.index("sample_id")
        compound_column = peak_header.index("compound")
        assert [row[sample_id_column] for row in peaks[1:]] == [
            "unicode_demo",
            "unicode_demo",
        ]
        assert [row[compound_column] for row in peaks[1:]] == ["methanol", "ethanol"]
    finally:
        workbook.close()
