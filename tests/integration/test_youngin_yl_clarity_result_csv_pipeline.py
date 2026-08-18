from __future__ import annotations

import hashlib
import sys
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.api import convert, inspect_file
from ordifile.core.detection import SOURCE_IDENTITY_PROBE_REASON
from ordifile.core.models import FileStatus

PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_agilent_chemstation_result_xml import (  # noqa: E402
    synthetic_result_xml_bytes,
)
from generate_shimadzu_labsolutions_result_ascii import (  # noqa: E402
    synthetic_result_ascii_bytes,
)
from generate_youngin_yl_clarity_result_csv import (  # noqa: E402
    synthetic_result_csv_bytes,
)


def _four_peaks() -> tuple[tuple[str, str, str, str, str, str], ...]:
    return (
        ("1", "10", "2", "10", "20", "0.1"),
        ("2", "20", "3", "20", "30", "0.2"),
        ("3", "30", "4", "30", "40", "0.3"),
        ("4", "40", "5", "40", "10", "0.4"),
    )


def _workbook_values(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return tuple(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()


def test_inspect_and_bundle_use_private_source_alias(tmp_path: Path) -> None:
    private_name = "private-person-yl-clarity-result.csv"
    data = synthetic_result_csv_bytes()
    source = tmp_path / private_name
    source.write_bytes(data)
    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.WARNING
    assert inspected.file.adapter_id == "youngin_yl_clarity_result_csv"
    assert inspected.file.source.public_reference == expected_reference
    assert inspected.file.source.path == Path(expected_reference)
    assert inspected.file.bundle is not None
    assert len(inspected.file.bundle.peaks) == 2
    assert all(reason == SOURCE_IDENTITY_PROBE_REASON for _, _, reason in inspected.probes)
    assert private_name not in repr(inspected)


def test_standalone_result_workbook_preserves_units_order_and_no_compound(
    tmp_path: Path,
) -> None:
    private_name = "private-result.csv"
    source = tmp_path / private_name
    source.write_bytes(synthetic_result_csv_bytes())
    output = tmp_path / "result.xlsx"

    result = convert(source, output)

    assert result.failure_count == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].values)
        headers = peaks[0]
        rows = peaks[1:]
        assert len(rows) == 2
        assert [row[headers.index("peak_number")] for row in rows] == [1, 2]
        assert [row[headers.index("observation_order")] for row in rows] == [1, 2]
        assert {row[headers.index("manufacturer")] for row in rows} == {"YoungIn"}
        assert {row[headers.index("detector")] for row in rows} == {None}
        assert {row[headers.index("channel")] for row in rows} == {"Signal 1: TCD"}
        assert {row[headers.index("retention_time_unit")] for row in rows} == {"min"}
        assert {row[headers.index("area_unit")] for row in rows} == {"mV.s"}
        assert {row[headers.index("height_unit")] for row in rows} == {"mV"}
        assert {row[headers.index("compound")] for row in rows} == {None}

        order = list(workbook["Peak_Order_Matrix"].values)
        assert len(order) == 2
        assert order[1][2:7] == ("YoungIn", None, "Signal 1: TCD", "min", "mV.s")
        assert order[1][7:] == (1.25, 100.5, 2.5, 150.75)
        assert workbook["Peak_Matrix"].max_column == 1
        assert private_name not in _workbook_values(output)
    finally:
        workbook.close()


def test_two_runs_preserve_six_peaks_and_two_populated_signal_streams(
    tmp_path: Path,
) -> None:
    source_a = tmp_path / "private-a.csv"
    source_a.write_bytes(synthetic_result_csv_bytes())
    source_b = tmp_path / "private-b.csv"
    source_b.write_bytes(
        synthetic_result_csv_bytes(
            variant="empty_fid_then_tcd",
            tcd_peaks=_four_peaks(),
        )
    )
    output = tmp_path / "two-runs.xlsx"

    result = convert((source_a, source_b), output, sort="input_order")

    assert result.failure_count == 0
    assert sum(len(item.bundle.peaks) for item in result.files if item.bundle is not None) == 6
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Samples"].max_row == 3
        assert workbook["Peaks"].max_row == 7
        assert workbook["Peak_Order_Matrix"].max_row == 3
        order_rows = tuple(workbook["Peak_Order_Matrix"].iter_rows(min_row=2, values_only=True))
        assert {row[4] for row in order_rows} == {"Signal 1: TCD", "Signal 2: TCD"}
        assert {row[6] for row in order_rows} == {"mV.s"}
        assert workbook["Peak_Matrix"].max_column == 1
        assert not any("private-" in value for value in _workbook_values(output))
    finally:
        workbook.close()


def test_generic_csv_keeps_generic_owner_and_relative_provenance(tmp_path: Path) -> None:
    source = tmp_path / "ordinary.csv"
    source.write_text("sample_id,retention_time,area\ngeneric,1,2\n", encoding="utf-8")

    inspected = inspect_file(source)

    assert inspected.file.adapter_id == "generic_csv"
    assert inspected.file.source.public_id is None
    assert inspected.file.source.public_reference == source.name
    probes = {adapter_id: reason for adapter_id, _confidence, reason in inspected.probes}
    assert probes["generic_csv"] != SOURCE_IDENTITY_PROBE_REASON
    assert probes["youngin_yl_clarity_result_csv"] == SOURCE_IDENTITY_PROBE_REASON


def test_malformed_family_never_falls_through_to_generic_csv(tmp_path: Path) -> None:
    data = synthetic_result_csv_bytes().replace(b"\t251.25\t", b"\t999.0\t", 1)
    private_name = "private-malformed.csv"
    source = tmp_path / private_name
    source.write_bytes(data)

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.adapter_id == "youngin_yl_clarity_result_csv"
    assert inspected.file.source.public_reference == f"source-{hashlib.sha256(data).hexdigest()}"
    assert private_name not in repr(inspected)


def test_malformed_family_outranks_generic_claim_and_withholds_private_fields(
    tmp_path: Path,
) -> None:
    private_name = "private-owner-result.csv"
    private_value = "PRIVATE_TRAILER_SENTINEL"
    data = (
        "Signal No.\tSignal Name\tPeak No.\tReten. time [min],"
        "sample_id,retention_time,area\r\n"
        f"{private_value},private-sample,1,2\r\n"
    ).encode()
    source = tmp_path / private_name
    source.write_bytes(data)
    expected_reference = f"source-{hashlib.sha256(data).hexdigest()}"

    inspected = inspect_file(source)

    probes = {adapter_id: confidence for adapter_id, confidence, _reason in inspected.probes}
    assert probes["generic_csv"] == 0.99
    assert probes["youngin_yl_clarity_result_csv"] == 0.70
    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.adapter_id == "youngin_yl_clarity_result_csv"
    assert inspected.file.source.public_reference == expected_reference
    assert private_name not in repr(inspected)
    assert private_value not in repr(inspected)

    output = tmp_path / "private-family-failure.xlsx"
    result = convert(source, output)
    assert result.failure_count == 1
    workbook_values = _workbook_values(output)
    assert expected_reference in workbook_values
    assert private_name not in workbook_values
    assert private_value not in workbook_values


def test_wrong_delimiter_vendor_like_text_is_a_structured_failure(tmp_path: Path) -> None:
    source = tmp_path / "wrong-delimiter.csv"
    source.write_text(
        "Signal No.,Signal Name,Peak No.,Reten. time [min],Area [mV.s]\r\n1,TCD,1,1,2\r\n",
        encoding="ascii",
        newline="",
    )

    inspected = inspect_file(source)

    assert inspected.file.status is FileStatus.FAILED
    assert inspected.file.adapter_id is None


def test_valid_and_corrupt_result_files_are_isolated(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    valid = synthetic_result_csv_bytes()
    (inputs / "private-valid.csv").write_bytes(valid)
    (inputs / "private-corrupt.csv").write_bytes(valid[:-20])
    output = tmp_path / "isolated.xlsx"

    result = convert(inputs, output, extensions=(".csv",))

    assert result.failure_count == 1
    assert sum(item.bundle is not None for item in result.files) == 1
    assert all("private-" not in repr(item) for item in result.files)
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Peaks"].max_row == 3
        assert workbook["Peak_Order_Matrix"].max_row == 2
        assert not any("private-" in value for value in _workbook_values(output))
    finally:
        workbook.close()


def test_three_vendor_synthetic_results_use_one_canonical_workbook(tmp_path: Path) -> None:
    agilent = tmp_path / "private-agilent.xml"
    agilent.write_bytes(synthetic_result_xml_bytes())
    shimadzu = tmp_path / "private-shimadzu.txt"
    shimadzu.write_bytes(synthetic_result_ascii_bytes())
    youngin_a = tmp_path / "private-youngin-a.csv"
    youngin_a.write_bytes(synthetic_result_csv_bytes())
    youngin_b = tmp_path / "private-youngin-b.csv"
    youngin_b.write_bytes(
        synthetic_result_csv_bytes(variant="empty_fid_then_tcd", tcd_peaks=_four_peaks())
    )
    output = tmp_path / "three-vendor.xlsx"

    result = convert(
        (agilent, shimadzu, youngin_a, youngin_b),
        output,
        sort="input_order",
    )

    assert result.failure_count == 0
    assert {item.adapter_id for item in result.files} == {
        "agilent_chemstation_result_xml",
        "shimadzu_labsolutions_result_ascii",
        "youngin_yl_clarity_result_csv",
    }
    expected = 3 + 3 + 6
    assert (
        sum(len(item.bundle.peaks) for item in result.files if item.bundle is not None) == expected
    )
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peaks = list(workbook["Peaks"].values)
        manufacturer = peaks[0].index("manufacturer")
        assert len(peaks) == expected + 1
        assert {row[manufacturer] for row in peaks[1:]} == {"Agilent", "Shimadzu", "YoungIn"}
        assert workbook["Peak_Order_Matrix"].max_row == 5
        matrix_rows = tuple(workbook["Peak_Matrix"].values)
        assert len(matrix_rows[0]) == 3
        assert all(
            all(value is None for value in row[1:])
            for row in matrix_rows[1:]
            if str(row[0]).startswith("YOUNGIN_RESULT_")
        )
    finally:
        workbook.close()
