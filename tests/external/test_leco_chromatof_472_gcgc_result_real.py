from __future__ import annotations

import hashlib
import os
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile.adapters._leco_chromatof_472_gcgc_result_txt import read_gcgc_result
from ordifile.api import convert, inspect_file

EXPECTED_SIZE = 20_040
EXPECTED_SHA256 = "59f336c3e4bb91df32c5111d39a7fa76759a72242a4bd5d873eb623b020af6dd"
EXPECTED_ROWS = 100
EXPECTED_DIGESTS = {
    "name": "4f9f49db2b690cbd13132521dd9eccdf5fbc2789872aff9cec589f77f96b8887",
    "rt1": "98dc62b00caaf69766a273d33a5b3bfdfdaf9cd3834ad76e4b8bc2a2a1c29ec6",
    "rt2": "e2fa11750903d2972ac6a7e63f385c0cfc522f0f3413a8813907ea9bff6a41b5",
    "area": "a2e7d7dc939a4a758c469561fa52d4fd1b4f4f0f6aa0f5bbc0dc24ccb82ee824",
    "height": "9ebab4336100f385e4f7234fc9ec26251cac58cfe09acdea604e35e24a5221a2",
    "spectra": "571592ee9e92a11a3d5974b6484d712a1da425615becce0d00a89d24d522bd06",
    "wb1": "ddeeb15739046638c1bb26403993f4e71e4bc9d00715c8e902c3c82fc1845027",
    "wb2": "516e4008235db926fba16027da5bc4fcc1a879d6d8eccdf28e186bd9442d6211",
    "retention_index": "e2617090c1f4ce5cc365bf48b1a63be30c7a09a22f03352c9bcbd4ba6cf4de7c",
}


def _fixture() -> Path:
    value = os.environ.get("ORDIFILE_LECO_GCXGC_RESULT_FIXTURE")
    if not value:
        raise AssertionError("ORDIFILE_LECO_GCXGC_RESULT_FIXTURE is required")
    return Path(value)


def _text_digest(values: Iterable[str]) -> str:
    return hashlib.sha256("\n".join(values).encode("ascii")).hexdigest()


def test_cc0_model_mixture_full_canonical_and_workbook_comparison(tmp_path: Path) -> None:
    source = _fixture()
    data = source.read_bytes()
    if len(data) != EXPECTED_SIZE or hashlib.sha256(data).hexdigest() != EXPECTED_SHA256:
        raise AssertionError("external LECO GCxGC fixture identity changed")
    decoded = read_gcgc_result(source)
    if len(decoded.peaks) != EXPECTED_ROWS:
        raise AssertionError("external LECO GCxGC peak count changed")
    sequences = {
        "name": (peak.name for peak in decoded.peaks),
        "rt1": (peak.retention_time_text for peak in decoded.peaks),
        "rt2": (peak.secondary_retention_time_text for peak in decoded.peaks),
        "area": (peak.area_text for peak in decoded.peaks),
        "height": (peak.height_text for peak in decoded.peaks),
        "spectra": (peak.spectra for peak in decoded.peaks),
        "wb1": (peak.wb1_text for peak in decoded.peaks),
        "wb2": (peak.wb2_text for peak in decoded.peaks),
        "retention_index": (peak.retention_index_text for peak in decoded.peaks),
    }
    if {key: _text_digest(values) for key, values in sequences.items()} != EXPECTED_DIGESTS:
        raise AssertionError("external LECO GCxGC source sequence changed")

    inspected = inspect_file(source)
    bundle = inspected.file.bundle
    if bundle is None or len(bundle.peaks) != EXPECTED_ROWS:
        raise AssertionError("Ordifile did not preserve every external LECO peak row")
    if inspected.file.source.public_reference != f"source-{EXPECTED_SHA256}":
        raise AssertionError("external LECO source identity is not SHA-derived")
    if any(peak.name.casefold() == "unknown" for peak in decoded.peaks):
        raise AssertionError("external LECO profile unexpectedly contains an unknown name sentinel")
    for source_peak, canonical in zip(decoded.peaks, bundle.peaks, strict=True):
        if (
            canonical.peak_number is not None
            or canonical.observation_order != source_peak.observation_order
            or canonical.retention_time != source_peak.retention_time
            or canonical.secondary_retention_time != source_peak.secondary_retention_time
            or canonical.area != source_peak.area
            or canonical.height != source_peak.height
            or canonical.compound != source_peak.name
            or canonical.compound_source != "canonical:leco_chromatof_gcxgc_result_txt.name"
            or canonical.retention_time_unit != "s"
            or canonical.secondary_retention_time_unit != "s"
            or canonical.area_unit != "AU"
            or canonical.height_unit != "AU"
            or canonical.detector is not None
            or canonical.channel is not None
            or canonical.start_time is not None
            or canonical.end_time is not None
        ):
            raise AssertionError("external LECO canonical mapping changed")
    metadata = {entry.key: entry for entry in bundle.metadata}
    for source_peak in decoded.peaks:
        prefix = f"peak_{source_peak.observation_order:06d}"
        expected_metadata = {
            f"{prefix}_name": (source_peak.name, None),
            f"{prefix}_spectra": (source_peak.spectra, None),
            f"{prefix}_wb1": (source_peak.wb1_text, "s"),
            f"{prefix}_wb2": (source_peak.wb2_text, "s"),
            f"{prefix}_retention_index": (source_peak.retention_index_text, None),
        }
        for key, (value, unit) in expected_metadata.items():
            entry = metadata.get(key)
            if entry is None or entry.value != value or entry.unit != unit:
                raise AssertionError("external LECO row metadata changed")
            if entry.source is None or f":row:{source_peak.source_row}:column:" not in entry.source:
                raise AssertionError("external LECO row metadata provenance changed")

    before = hashlib.sha256(source.read_bytes()).hexdigest()
    output = tmp_path / "leco-gcxgc-result.xlsx"
    result = convert(source, output)
    if result.failure_count:
        raise AssertionError("external LECO result-only conversion failed")
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        peak_rows = tuple(workbook["Peaks"].values)
        headers = peak_rows[0]
        rows = peak_rows[1:]
        if len(rows) != EXPECTED_ROWS:
            raise AssertionError("external LECO workbook rows were lost or duplicated")
        for row, source_peak in zip(rows, decoded.peaks, strict=True):
            if (
                row[headers.index("retention_time")] != source_peak.retention_time
                or row[headers.index("secondary_retention_time")]
                != source_peak.secondary_retention_time
                or row[headers.index("area")] != source_peak.area
                or row[headers.index("height")] != source_peak.height
                or row[headers.index("compound")] != source_peak.name
                or row[headers.index("observation_order")] != source_peak.observation_order
            ):
                raise AssertionError("external LECO workbook scientific mapping changed")
        order = tuple(workbook["Peak_Order_Matrix_2D"].iter_rows(min_row=2, values_only=True))
        if len(order) != 1 or order[0][2:8] != ("LECO", None, None, "s", "s", "AU"):
            raise AssertionError("external LECO two-dimensional stream identity changed")
        expected_order = tuple(
            value
            for peak in decoded.peaks
            for value in (peak.retention_time, peak.secondary_retention_time, peak.area)
        )
        if order[0][8:] != expected_order:
            raise AssertionError("external LECO two-dimensional order matrix changed")
        if "Peak_Order_Matrix" in workbook.sheetnames:
            raise AssertionError("external LECO 2D rows were projected into the 1D matrix")
        workbook_metadata = {
            row[3]: (row[4], row[5], row[6])
            for row in tuple(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        }
        for source_peak in decoded.peaks:
            prefix = f"peak_{source_peak.observation_order:06d}"
            for suffix, expected_value, expected_unit in (
                ("name", source_peak.name, None),
                ("spectra", source_peak.spectra, None),
                ("wb1", source_peak.wb1_text, "s"),
                ("wb2", source_peak.wb2_text, "s"),
                ("retention_index", source_peak.retention_index_text, None),
            ):
                value, unit, provenance = workbook_metadata[f"{prefix}_{suffix}"]
                if value != expected_value or unit != expected_unit:
                    raise AssertionError("external LECO workbook row metadata changed")
                if f":row:{source_peak.source_row}:column:" not in str(provenance):
                    raise AssertionError("external LECO workbook metadata provenance changed")
    finally:
        workbook.close()
    if hashlib.sha256(source.read_bytes()).hexdigest() != before:
        raise AssertionError("external LECO fixture changed during conversion")
