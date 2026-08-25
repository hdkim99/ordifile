# Copyright 2026 hdkim99
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import socket
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from ordifile import (
    ColumnSelector,
    ConversionPlanProblem,
    ConversionPlanReadiness,
    ConversionPlanRoute,
    ConversionRecipe,
    PeakTableFormat,
    PeakTableMapping,
    PeakTableMappingProfile,
    PeakTableMappingSet,
)
from ordifile.api import convert
from ordifile.core.models import BatchOutcome
from ordifile.desktop.models import DesktopRequest
from ordifile.desktop.services import (
    convert_preflight_plan,
    convert_selection,
    inspect_selection,
    preflight_selection,
)

FIXTURE = Path("tests/fixtures/synthetic/generic_peaks.csv")
PROJECT_ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "tests" / "fixtures" / "synthetic"))
from generate_youngin_yl_clarity_prm import synthetic_prm_bytes  # noqa: E402
from generate_youngin_yl_clarity_result_csv import (  # noqa: E402
    synthetic_result_csv_bytes,
)


def _sheet_values(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()


def test_desktop_service_discovers_folder_and_reports_detected_adapter(tmp_path: Path) -> None:
    input_folder = tmp_path / "inputs"
    input_folder.mkdir()
    (input_folder / "peaks.csv").write_bytes(FIXTURE.read_bytes())

    report = inspect_selection((input_folder,), sort="auto")

    assert report.outcome is BatchOutcome.SUCCESS
    assert len(report.files) == 1
    assert report.files[0].adapter_id == "generic_csv"
    assert "Verified" in report.files[0].format_name


def test_desktop_preflight_executes_the_exact_immutable_plan(tmp_path: Path) -> None:
    output = tmp_path / "planned.xlsx"
    request = DesktopRequest((FIXTURE,), output, "input_order")

    preflight = preflight_selection(request)

    assert preflight.plan is not None and preflight.plan.is_executable
    assert preflight.files[0].plan_route is ConversionPlanRoute.GENERIC_INPUT
    assert not output.exists()

    converted = convert_preflight_plan(preflight.plan)

    assert converted.outcome is BatchOutcome.SUCCESS
    assert output.is_file()
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Peaks" in workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.researcher_acceptance
def test_desktop_converts_9_1_prm_to_scientific_signals_without_extra_controls(
    tmp_path: Path,
) -> None:
    source = tmp_path / "scientific.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.1.0.76 FULL, SN: SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )
    output = tmp_path / "scientific.xlsx"

    preflight = preflight_selection(DesktopRequest((source,), output))

    assert preflight.plan is not None
    assert preflight.plan.readiness is ConversionPlanReadiness.READY
    assert preflight.plan.options.include_signals is True
    assert preflight.files[0].plan_route is ConversionPlanRoute.EXACT_ADAPTER
    assert preflight.files[0].plan_problem is ConversionPlanProblem.NONE

    converted = convert_preflight_plan(preflight.plan)

    assert converted.outcome is BatchOutcome.SUCCESS
    assert converted.summary is not None
    assert converted.summary.scientific_signal_series == 2
    assert converted.summary.peak_records == 0
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert {"Signals_FID", "Signals_TCD"}.issubset(workbook.sheetnames)
        assert not any(name.startswith("Signals_Records_") for name in workbook.sheetnames)
        fid_rows = tuple(workbook["Signals_FID"].iter_rows(min_row=2, values_only=True))
        tcd_rows = tuple(workbook["Signals_TCD"].iter_rows(min_row=2, values_only=True))
        assert {row[6] for row in fid_rows} == {"min"}
        assert {row[9] for row in fid_rows} == {"pA"}
        assert {row[9] for row in tcd_rows} == {"mV"}
        assert workbook["Peaks"].max_row == 1
    finally:
        workbook.close()


@pytest.mark.researcher_acceptance
def test_desktop_keeps_9_1_prm_signals_and_result_peaks_as_independent_sources(
    tmp_path: Path,
) -> None:
    prm = tmp_path / "scientific.prm"
    prm.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.1.0.76 FULL, SN: SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )
    result_csv = tmp_path / "result.csv"
    result_csv.write_bytes(synthetic_result_csv_bytes())
    output = tmp_path / "mixed-youngin.xlsx"

    preflight = preflight_selection(DesktopRequest((prm, result_csv), output))

    assert preflight.plan is not None and preflight.plan.is_executable
    assert [item.plan_route for item in preflight.files] == [
        ConversionPlanRoute.EXACT_ADAPTER,
        ConversionPlanRoute.EXACT_ADAPTER,
    ]
    converted = convert_preflight_plan(preflight.plan)
    assert converted.summary is not None
    assert converted.summary.scientific_signal_series == 2
    assert converted.summary.peak_records == 2

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        signal_rows = list(workbook["Signals_FID"].iter_rows(min_row=2, values_only=True))
        peak_rows = list(workbook["Peaks"].iter_rows(min_row=2, values_only=True))
        assert len(signal_rows) == 2
        assert len(peak_rows) == 2
        assert {row[0] for row in signal_rows}.isdisjoint({row[0] for row in peak_rows})
        assert {row[1] for row in signal_rows}.isdisjoint({row[1] for row in peak_rows})
    finally:
        workbook.close()


@pytest.mark.researcher_acceptance
def test_desktop_preflight_converts_validated_9_0_and_compatible_unknown_profiles(
    tmp_path: Path,
) -> None:
    validated = tmp_path / "validated.prm"
    validated.write_bytes(synthetic_prm_bytes())
    unknown = tmp_path / "unknown.prm"
    unknown.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.2.0.0 FULL, SN: SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )

    validated_output = tmp_path / "validated.xlsx"
    validated_preflight = preflight_selection(DesktopRequest((validated,), validated_output))
    assert validated_preflight.plan is not None
    assert validated_preflight.plan.readiness is ConversionPlanReadiness.READY
    assert "YOUNGIN_PRM_VALIDATED_SCIENTIFIC_SIGNAL" in (
        validated_preflight.plan.entries[0].issue_codes
    )
    converted = convert_preflight_plan(validated_preflight.plan)
    assert converted.summary is not None
    assert converted.summary.scientific_signal_series == 1
    assert converted.summary.structural_record_series == 0
    workbook = load_workbook(validated_output, read_only=True, data_only=False)
    try:
        assert "Signals_TCD" in workbook.sheetnames
        units = {row[9] for row in workbook["Signals_TCD"].iter_rows(min_row=2, values_only=True)}
        assert units == {"mV"}
    finally:
        workbook.close()

    unknown_preflight = preflight_selection(DesktopRequest((unknown,), tmp_path / "unknown.xlsx"))
    assert unknown_preflight.plan is not None
    assert unknown_preflight.plan.readiness is ConversionPlanReadiness.READY
    assert unknown_preflight.plan.summary.routable == 1
    assert "YOUNGIN_PRM_FAMILY_COMPATIBLE_SCIENTIFIC_UNIT_UNRESOLVED" in (
        unknown_preflight.plan.entries[0].issue_codes
    )
    assert unknown_preflight.files[0].plan_route is ConversionPlanRoute.EXACT_ADAPTER
    assert unknown_preflight.files[0].plan_problem is ConversionPlanProblem.NONE

    unknown_converted = convert_preflight_plan(unknown_preflight.plan)
    assert unknown_converted.summary is not None
    assert unknown_converted.summary.scientific_signal_series == 2
    unknown_workbook = load_workbook(tmp_path / "unknown.xlsx", read_only=True, data_only=False)
    try:
        assert {"Signals_FID", "Signals_TCD"}.issubset(unknown_workbook.sheetnames)
        assert {
            row[9] for row in unknown_workbook["Signals_FID"].iter_rows(min_row=2, values_only=True)
        } == {None}
    finally:
        unknown_workbook.close()

    structural = tmp_path / "structural-only.prm"
    structural.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.2.0.0 FULL, SN: SYNTHETIC",
            d_step=2,
        )
    )
    structural_output = tmp_path / "structural-only.xlsx"
    structural_preflight = preflight_selection(DesktopRequest((structural,), structural_output))
    assert structural_preflight.plan is not None
    assert structural_preflight.plan.readiness is ConversionPlanReadiness.READY
    assert "YOUNGIN_PRM_FAMILY_COMPATIBLE_STRUCTURAL_ONLY" in (
        structural_preflight.plan.entries[0].issue_codes
    )
    structural_converted = convert_preflight_plan(structural_preflight.plan)
    assert structural_converted.summary is not None
    assert structural_converted.summary.structural_record_series == 1


def test_desktop_recipe_preflight_executes_embedded_mapping_snapshot(
    tmp_path: Path,
) -> None:
    source = tmp_path / "neutral-result.csv"
    source.write_text("Time,Integrated\n1.25,100\n2.5,250\n", encoding="utf-8")
    output = tmp_path / "recipe-result.xlsx"
    mapping = PeakTableMapping(
        ColumnSelector("Time", 1),
        ColumnSelector("Integrated", 2),
        "min",
        PeakTableFormat.CSV,
    )
    recipe = ConversionRecipe(
        peak_table_mapping=mapping,
        display_label="Daily neutral table",
    )
    request = DesktopRequest((source,), output, recipe=recipe)

    preflight = preflight_selection(request)

    assert preflight.plan is not None and preflight.plan.is_executable
    assert preflight.plan.entries[0].route is ConversionPlanRoute.USER_MAPPING

    converted = convert_preflight_plan(preflight.plan)

    assert converted.outcome is BatchOutcome.SUCCESS
    assert converted.summary is not None
    assert converted.summary.peak_records == 2
    assert output.is_file()
    peaks = _sheet_values(output, "Peaks")
    assert len(peaks) == 3


@pytest.mark.researcher_acceptance
def test_desktop_recipe_preserves_explicit_signal_exclusion(tmp_path: Path) -> None:
    source = tmp_path / "scientific.prm"
    source.write_bytes(
        synthetic_prm_bytes(
            producer_text="YL-Clarity 9.1.0.76 FULL, SN: SYNTHETIC",
            channels=((1.0, 2.0), (3.0, 4.0)),
        )
    )
    output = tmp_path / "recipe-excluded-signals.xlsx"
    recipe = ConversionRecipe(include_signals=False, display_label="Peaks only")

    preflight = preflight_selection(DesktopRequest((source,), output, recipe=recipe))

    assert preflight.plan is not None and preflight.plan.is_executable
    assert preflight.plan.options.include_signals is False
    converted = convert_preflight_plan(preflight.plan)
    assert converted.summary is not None
    assert converted.summary.scientific_signal_series == 2
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert not any(name.startswith("Signals_") for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_desktop_preflight_blocks_a_source_changed_before_conversion(tmp_path: Path) -> None:
    source = tmp_path / "private-source.csv"
    source.write_text("sample_id,area\na,1\n", encoding="utf-8")
    output = tmp_path / "planned.xlsx"
    preflight = preflight_selection(DesktopRequest((source,), output))
    assert preflight.plan is not None
    source.write_text("sample_id,area\na,2\n", encoding="utf-8")

    converted = convert_preflight_plan(preflight.plan)

    assert converted.is_fatal_error
    assert converted.error_code == "CONVERSION_PLAN_STALE"
    assert not output.exists()
    assert source.name not in repr(preflight)


def test_desktop_and_cli_api_create_equivalent_scientific_tables(tmp_path: Path) -> None:
    desktop_output = tmp_path / "desktop.xlsx"
    api_output = tmp_path / "api.xlsx"

    desktop = convert_selection(DesktopRequest((FIXTURE,), desktop_output, "input_order"))
    direct = convert((FIXTURE,), api_output, sort="input_order")

    assert desktop.outcome is BatchOutcome.SUCCESS
    assert desktop.summary is not None
    assert desktop.summary.converted_sources == 1
    assert desktop.summary.sample_records == 1
    assert desktop.summary.peak_records == len(_sheet_values(desktop_output, "Peaks")) - 1
    assert direct.success_count == 1
    for sheet in ("Samples", "Peak_Matrix", "Peaks", "Metadata", "Import_Log"):
        assert _sheet_values(desktop_output, sheet) == _sheet_values(api_output, sheet)


def test_desktop_registry_preview_and_conversion_follow_youngin_result_adapter(
    tmp_path: Path,
) -> None:
    source = tmp_path / "synthetic-youngin-result.csv"
    source.write_bytes(synthetic_result_csv_bytes())

    preview = inspect_selection((source,), sort="input_order")

    assert preview.outcome is BatchOutcome.SUCCESS
    assert preview.failure_count == 0
    assert len(preview.files) == 1
    detected = preview.files[0]
    assert detected.adapter_id == "youngin_yl_clarity_result_csv"
    assert detected.format_name.count("Experimental") == 1
    assert detected.status.value == "Warning"

    desktop_output = tmp_path / "youngin-desktop.xlsx"
    core_output = tmp_path / "youngin-core.xlsx"
    desktop = convert_selection(DesktopRequest((source,), desktop_output, "input_order"))
    core = convert((source,), core_output, sort="input_order")

    assert desktop.outcome is BatchOutcome.SUCCESS
    assert desktop.failure_count == 0
    assert core.failure_count == 0
    assert {item.adapter_id for item in core.files} == {"youngin_yl_clarity_result_csv"}
    for sheet in ("Peaks", "Peak_Order_Matrix", "Metadata", "Import_Log"):
        assert _sheet_values(desktop_output, sheet) == _sheet_values(core_output, sheet)


def test_desktop_conversion_exposes_partial_failure_and_writes_workbook(tmp_path: Path) -> None:
    unsupported = tmp_path / "unsupported.bin"
    unsupported.write_bytes(b"not a supported input")
    output = tmp_path / "partial.xlsx"

    report = convert_selection(DesktopRequest((FIXTURE, unsupported), output))

    assert report.outcome is BatchOutcome.PARTIAL_SUCCESS
    assert report.success_count == 1
    assert report.failure_count == 1
    assert output.is_file()
    assert any(file.status.value == "Failed" for file in report.files)


def test_desktop_conversion_exposes_all_failed_without_hiding_diagnostics(
    tmp_path: Path,
) -> None:
    unsupported = tmp_path / "unsupported.bin"
    unsupported.write_bytes(b"not a supported input")
    output = tmp_path / "failed.xlsx"

    report = convert_selection(DesktopRequest((unsupported,), output))

    assert report.outcome is BatchOutcome.FAILED
    assert report.failure_count == 1
    assert report.files[0].message


def test_desktop_refuses_existing_output_without_overwrite(tmp_path: Path) -> None:
    output = tmp_path / "existing.xlsx"
    original = b"do not replace"
    output.write_bytes(original)

    report = convert_selection(DesktopRequest((FIXTURE,), output))

    assert report.outcome is BatchOutcome.FAILED
    assert report.error_code == "OUTPUT_EXISTS"
    assert output.read_bytes() == original


def test_desktop_conversion_requires_no_network(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    def fail_network(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("network access attempted")

    monkeypatch.setattr(socket, "create_connection", fail_network)
    output = tmp_path / "offline.xlsx"

    report = convert_selection(DesktopRequest((FIXTURE,), output))

    assert report.outcome is BatchOutcome.SUCCESS
    assert output.is_file()


def test_desktop_explicit_mapping_matches_public_api_and_writes_order_matrix(
    tmp_path: Path,
) -> None:
    source = tmp_path / "neutral-result.csv"
    source.write_text(
        "Time coordinate,Integrated response,Unused note\n1.25,100,local\n2.5,250,local\n",
        encoding="utf-8",
    )
    mapping = PeakTableMapping(
        ColumnSelector("Time coordinate", 1),
        ColumnSelector("Integrated response", 2),
        "min",
        PeakTableFormat.CSV,
        area_unit="arbitrary",
        ignored_columns=(ColumnSelector("Unused note", 3),),
    )
    desktop_output = tmp_path / "mapped-desktop.xlsx"
    api_output = tmp_path / "mapped-api.xlsx"

    desktop = convert_selection(
        DesktopRequest(
            (source,),
            desktop_output,
            "input_order",
            mapping,
        )
    )
    direct = convert(
        (source,),
        api_output,
        sort="input_order",
        peak_table_mapping=mapping,
    )

    assert desktop.outcome is BatchOutcome.SUCCESS
    assert direct.success_count == 1
    for sheet in ("Peaks", "Peak_Order_Matrix", "Metadata", "Import_Log"):
        assert _sheet_values(desktop_output, sheet) == _sheet_values(api_output, sheet)


def test_desktop_mapping_set_routes_multiple_tables_without_bypassing_exact_adapter(
    tmp_path: Path,
) -> None:
    first = tmp_path / "neutral-one.csv"
    first.write_text("RT,Area\n1.25,100\n", encoding="utf-8")
    second = tmp_path / "neutral-two.csv"
    second.write_text("Time,Integrated,Note\n2.5,250,local\n", encoding="utf-8")
    exact = tmp_path / "synthetic-youngin-result.csv"
    exact.write_bytes(synthetic_result_csv_bytes())
    first_mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
    )
    second_mapping = PeakTableMapping(
        ColumnSelector("Time", 1),
        ColumnSelector("Integrated", 2),
        "min",
        PeakTableFormat.CSV,
        ignored_columns=(ColumnSelector("Note", 3),),
    )
    mapping_set = PeakTableMappingSet(
        (
            PeakTableMappingProfile(first_mapping, "Template one"),
            PeakTableMappingProfile(second_mapping, "Template two"),
        )
    )
    desktop_output = tmp_path / "desktop-set.xlsx"
    api_output = tmp_path / "api-set.xlsx"

    desktop = convert_selection(
        DesktopRequest(
            (first, second, exact),
            desktop_output,
            "input_order",
            peak_table_mapping_set=mapping_set,
        )
    )
    direct = convert(
        (first, second, exact),
        api_output,
        sort="input_order",
        peak_table_mapping_set=mapping_set,
    )

    assert desktop.outcome is BatchOutcome.SUCCESS
    assert desktop.failure_count == 0
    assert [item.mapping_route for item in desktop.files] == [
        "USER_MAPPING_PROFILE",
        "USER_MAPPING_PROFILE",
        "EXACT_ADAPTER",
    ]
    assert desktop.files[0].mapping_profile_id == mapping_set.profiles[0].profile_id
    assert desktop.files[1].mapping_profile_id == mapping_set.profiles[1].profile_id
    assert desktop.files[2].adapter_id == "youngin_yl_clarity_result_csv"
    assert direct.failure_count == 0
    for sheet in ("Peaks", "Peak_Order_Matrix", "Metadata", "Import_Log"):
        assert _sheet_values(desktop_output, sheet) == _sheet_values(api_output, sheet)


def test_desktop_inspection_exposes_bounded_schema_drift_without_applying_it(
    tmp_path: Path,
) -> None:
    source = tmp_path / "changed.csv"
    source.write_text("RT,Peak Area\n1,2\n", encoding="utf-8")
    mapping = PeakTableMapping(
        ColumnSelector("RT", 1),
        ColumnSelector("Area", 2),
        "min",
        PeakTableFormat.CSV,
    )
    profile = PeakTableMappingProfile(mapping, "Original")
    mapping_set = PeakTableMappingSet((profile,))

    report = inspect_selection(
        (source,),
        sort="input_order",
        peak_table_mapping_set=mapping_set,
    )

    assert report.outcome is BatchOutcome.FAILED
    assert report.files[0].mapping_route == "SCHEMA_DRIFT_CANDIDATE"
    assert report.files[0].mapping_profile_id is None
    assert len(report.files[0].mapping_diagnostics) == 1
    assert report.files[0].mapping_diagnostics[0].profile_id == profile.profile_id
    assert report.files[0].review_input_index == 0
    assert source.name not in repr(report)

    preflight = preflight_selection(
        DesktopRequest(
            (source,),
            tmp_path / "drift-preflight.xlsx",
            "input_order",
            peak_table_mapping_set=mapping_set,
        )
    )
    assert preflight.plan is not None
    assert preflight.files[0].plan_problem is ConversionPlanProblem.MAPPING_SCHEMA_DRIFT
    assert preflight.files[0].mapping_diagnostics[0].profile_id == profile.profile_id
    assert preflight.failure_count == 1
